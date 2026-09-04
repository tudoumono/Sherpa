"""Excelセルの原値・型付き値・数式・表示値を決定的に抽出する。

Microsoft Excelの表示エンジンを利用できないLinux基本経路向け。OOXMLの``<v>``を原値の
権威、openpyxlの型変換結果を型付き値の権威として分離し、対応を明示できる主要書式だけを
表示する。複雑な条件、会計、ロケール、和暦等を推測して近似せず``display_status=unsupported``
にする。LLM・外部通信・再計算は行わない。
"""
from __future__ import annotations

import gc
import json
import posixpath
import re
import zipfile
from dataclasses import dataclass
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET


EXCEL_DISPLAY_PROFILE = "excel-display-linux-v1"

_MAIN = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_DOC_REL_ID = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
_CELL_REF_RE = re.compile(r"^[A-Z]{1,3}[1-9][0-9]*$")
_NUMERIC_TOKEN_RE = re.compile(r"[0#?]")


@dataclass(frozen=True)
class DisplayResult:
    value: str | None
    source: str
    status: str
    reason: str | None


def _number_or_text(value: str | None) -> Any:
    if value in (None, ""):
        return None
    assert value is not None
    try:
        return int(value)
    except ValueError:
        try:
            return float(value)
        except ValueError:
            return value


def _json_scalar(value: Any) -> Any:
    """Python固有型を決定的なJSON値へ変換する。"""
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, datetime):
        return value.isoformat(timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, time):
        return value.isoformat(timespec="seconds")
    if isinstance(value, timedelta):
        return value.total_seconds()
    return str(value)


def _typed_value(value: Any) -> dict[str, Any]:
    if value is None:
        kind = "blank"
    elif isinstance(value, bool):
        kind = "boolean"
    elif isinstance(value, datetime):
        kind = "datetime"
    elif isinstance(value, date):
        kind = "date"
    elif isinstance(value, time):
        kind = "time"
    elif isinstance(value, timedelta):
        kind = "duration_seconds"
    elif isinstance(value, int):
        kind = "integer"
    elif isinstance(value, float):
        kind = "number"
    elif isinstance(value, str):
        kind = "string"
    else:
        kind = value.__class__.__name__.lower()
    return {"type": kind, "value": _json_scalar(value)}


def _resolve_part(base_part: str, target: str) -> str:
    if target.startswith("/"):
        return target.lstrip("/")
    return posixpath.normpath(posixpath.join(posixpath.dirname(base_part), target))


def _sheet_parts(archive: zipfile.ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    by_id = {
        rel.get("Id", ""): _resolve_part("xl/workbook.xml", rel.get("Target", ""))
        for rel in rels.findall(f"{_REL}Relationship")
        if rel.get("Id") and rel.get("Target")
    }
    out: dict[str, str] = {}
    sheets = workbook.find(f"{_MAIN}sheets")
    for sheet in list(sheets) if sheets is not None else []:
        name = sheet.get("name")
        part = by_id.get(sheet.get(_DOC_REL_ID, ""))
        if name and part:
            out[name] = part
    return out


def _shared_strings(archive: zipfile.ZipFile) -> list[str]:
    try:
        root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return ["".join(node.text or "" for node in item.iter(f"{_MAIN}t"))
            for item in root.findall(f"{_MAIN}si")]


def _raw_from_cell(cell: ET.Element, shared: list[str]) -> Any:
    cell_type = cell.get("t", "n")
    if cell_type == "inlineStr":
        return "".join(node.text or "" for node in cell.iter(f"{_MAIN}t"))
    value_node = cell.find(f"{_MAIN}v")
    raw = value_node.text if value_node is not None else None
    if cell_type == "s" and raw is not None:
        try:
            index = int(raw)
        except ValueError:
            return None
        return shared[index] if 0 <= index < len(shared) else None
    if cell_type == "b":
        return raw == "1"
    if cell_type in {"str", "e", "d"}:
        return raw
    return _number_or_text(raw)


def _raw_cells(path: Path, targets: dict[str, set[str]]) -> dict[tuple[str, str], dict[str, Any]]:
    """対象座標だけをworksheet XMLからstream抽出する。"""
    out: dict[tuple[str, str], dict[str, Any]] = {}
    with zipfile.ZipFile(path) as archive:
        parts = _sheet_parts(archive)
        shared = _shared_strings(archive)
        for sheet_name in sorted(targets):
            wanted = targets[sheet_name]
            part = parts.get(sheet_name)
            if not wanted or not part:
                continue
            with archive.open(part) as stream:
                for _event, cell in ET.iterparse(stream, events=("end",)):
                    if cell.tag != f"{_MAIN}c":
                        continue
                    coordinate = cell.get("r", "").upper()
                    if coordinate in wanted:
                        formula_node = cell.find(f"{_MAIN}f")
                        value_node = cell.find(f"{_MAIN}v")
                        formula = None if formula_node is None else f"={formula_node.text or ''}"
                        out[(sheet_name, coordinate)] = {
                            "raw_value": _raw_from_cell(cell, shared),
                            "formula": formula,
                            "has_cache": value_node is not None and value_node.text not in (None, ""),
                        }
                    cell.clear()
    return out


def _split_sections(code: str) -> list[str]:
    sections: list[str] = []
    current: list[str] = []
    quoted = False
    escaped = False
    for char in code:
        if escaped:
            current.append(char)
            escaped = False
        elif char == "\\":
            current.extend(("\\",))
            escaped = True
        elif char == '"':
            quoted = not quoted
            current.append(char)
        elif char == ";" and not quoted:
            sections.append("".join(current))
            current = []
        else:
            current.append(char)
    sections.append("".join(current))
    return sections


def _unquote_literal(text: str) -> str:
    out: list[str] = []
    index = 0
    while index < len(text):
        char = text[index]
        if char == '"':
            end = text.find('"', index + 1)
            if end < 0:
                return ""
            out.append(text[index + 1:end])
            index = end + 1
        elif char == "\\" and index + 1 < len(text):
            out.append(text[index + 1])
            index += 2
        else:
            out.append(char)
            index += 1
    return "".join(out)


def _format_decimal(value: int | float, code: str) -> str | None:
    if any(marker in code for marker in ("[", "]", "_", "*", "?")):
        return None
    if re.search(r"[Ee][+-]?0", code) or "/" in code:
        return None
    placeholders = [match.start() for match in _NUMERIC_TOKEN_RE.finditer(code)]
    if not placeholders:
        return None
    first, last = placeholders[0], placeholders[-1]
    prefix = _unquote_literal(code[:first])
    suffix = _unquote_literal(code[last + 1:])
    core = code[first:last + 1]
    percent_count = code.count("%")
    try:
        number = Decimal(str(value)) * (Decimal(100) ** percent_count)
    except InvalidOperation:
        return None

    decimal_part = core.split(".", 1)[1] if "." in core else ""
    decimal_placeholders = "".join(char for char in decimal_part if char in "0#")
    decimals = len(decimal_placeholders)
    quantum = Decimal(1).scaleb(-decimals)
    number = number.quantize(quantum, rounding=ROUND_HALF_UP)
    grouped = "," in core.split(".", 1)[0]
    rendered = format(number, f",.{decimals}f" if grouped else f".{decimals}f")
    if "#" in decimal_placeholders and "." in rendered:
        mandatory = decimal_placeholders.count("0")
        integer, fraction = rendered.split(".", 1)
        fraction = fraction.rstrip("0")
        fraction += "0" * max(0, mandatory - len(fraction))
        rendered = integer + (("." + fraction) if fraction else "")

    integer_pattern = core.split(".", 1)[0].replace(",", "")
    mandatory_integer = integer_pattern.count("0")
    sign = "-" if rendered.startswith("-") else ""
    unsigned = rendered[1:] if sign else rendered
    integer, dot, fraction = unsigned.partition(".")
    plain_integer = integer.replace(",", "")
    if mandatory_integer > len(plain_integer):
        plain_integer = plain_integer.zfill(mandatory_integer)
        integer = f"{int(plain_integer):,}" if grouped else plain_integer
    rendered = sign + integer + (dot + fraction if dot else "")
    return prefix + rendered + suffix


def _date_tokens(code: str) -> list[tuple[str, bool]] | None:
    """(text, is_token)列へ分解する。quote/escapeはliteralに確定する。"""
    out: list[tuple[str, bool]] = []
    index = 0
    while index < len(code):
        char = code[index]
        if char == '"':
            end = code.find('"', index + 1)
            if end < 0:
                return None
            out.append((code[index + 1:end], False))
            index = end + 1
            continue
        if char == "\\" and index + 1 < len(code):
            out.append((code[index + 1], False))
            index += 2
            continue
        if code[index:index + 5].casefold() == "am/pm":
            out.append((code[index:index + 5], True))
            index += 5
            continue
        if char.casefold() in "ymdhs":
            end = index + 1
            while end < len(code) and code[end].casefold() == char.casefold():
                end += 1
            out.append((code[index:end], True))
            index = end
            continue
        out.append((char, False))
        index += 1
    return out


def _format_date(value: date | datetime | time, code: str) -> str | None:
    if any(marker in code for marker in ("[", "]", "_", "*")) or re.search(r"[ge]+", code, re.I):
        return None
    tokens = _date_tokens(code)
    if tokens is None:
        return None
    dt = value
    year = getattr(dt, "year", 0)
    month = getattr(dt, "month", 0)
    day = getattr(dt, "day", 0)
    hour = getattr(dt, "hour", 0)
    minute = getattr(dt, "minute", 0)
    second = getattr(dt, "second", 0)
    has_ampm = any(is_token and token.casefold() == "am/pm" for token, is_token in tokens)
    rendered: list[str] = []
    for index, (token, is_token) in enumerate(tokens):
        if not is_token:
            rendered.append(token)
            continue
        lower = token.casefold()
        if lower == "am/pm":
            rendered.append("AM" if hour < 12 else "PM")
        elif lower[0] == "y":
            rendered.append(f"{year % 100:02d}" if len(lower) <= 2 else f"{year:04d}")
        elif lower[0] == "d":
            rendered.append(f"{day:02d}" if len(lower) >= 2 else str(day))
        elif lower[0] == "h":
            display_hour = hour % 12 or 12 if has_ampm else hour
            rendered.append(f"{display_hour:02d}" if len(lower) >= 2 else str(display_hour))
        elif lower[0] == "s":
            rendered.append(f"{second:02d}" if len(lower) >= 2 else str(second))
        elif lower[0] == "m":
            previous = tokens[index - 1][0] if index else ""
            following = tokens[index + 1][0] if index + 1 < len(tokens) else ""
            is_minute = previous == ":" or following == ":" or any(
                is_t and text.casefold().startswith("h") for text, is_t in tokens[:index])
            value_m = minute if is_minute else month
            rendered.append(f"{value_m:02d}" if len(lower) >= 2 else str(value_m))
    return "".join(rendered)


def format_display(value: Any, number_format: str, *, formula_cache_missing: bool = False) -> DisplayResult:
    """主要Excel書式だけを表示し、未対応は推測しない。"""
    if formula_cache_missing:
        return DisplayResult(None, EXCEL_DISPLAY_PROFILE, "unsupported", "formula_cache_missing")
    if value is None:
        return DisplayResult("", "raw_value", "rendered", None)
    if isinstance(value, str):
        return DisplayResult(value, "raw_value", "rendered", None)
    if isinstance(value, bool):
        return DisplayResult("TRUE" if value else "FALSE", EXCEL_DISPLAY_PROFILE, "rendered", None)

    sections = _split_sections(number_format or "General")
    if len(sections) != 1:
        return DisplayResult(None, EXCEL_DISPLAY_PROFILE, "unsupported", "multi_section_format")
    code = sections[0].strip() or "General"
    if code.casefold() == "general":
        return DisplayResult(str(_json_scalar(value)), "raw_value", "rendered", None)
    if isinstance(value, (datetime, date, time)):
        rendered = _format_date(value, code)
        reason = None if rendered is not None else "complex_date_format"
        return DisplayResult(rendered, EXCEL_DISPLAY_PROFILE, "rendered" if rendered is not None else "unsupported", reason)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        rendered = _format_decimal(value, code)
        reason = None if rendered is not None else "complex_number_format"
        return DisplayResult(rendered, EXCEL_DISPLAY_PROFILE, "rendered" if rendered is not None else "unsupported", reason)
    return DisplayResult(None, EXCEL_DISPLAY_PROFILE, "unsupported", "unsupported_value_type")


def extract_cell_metadata(path: str | Path, targets: dict[str, set[str]]) -> dict[tuple[str, str], dict[str, Any]]:
    """対象セルの10項目契約を返す。targetsは``{sheet: {A1, ...}}``。

    `wb_formula`（data_only=False）と `wb_values`（data_only=True）を**逐次**ロードする（MEM-1）。
    両方を同時にオブジェクトツリーへ展開するとピークメモリが約2倍になるため、1冊読み終えて
    閉じてからもう1冊を開く＝ピーク時に生きているブックを常に1冊に保つ。
    """
    import openpyxl

    source = Path(path)
    normalized_targets = {
        sheet: {coordinate.upper() for coordinate in coordinates if _CELL_REF_RE.fullmatch(coordinate.upper())}
        for sheet, coordinates in targets.items()
    }
    raw = _raw_cells(source, normalized_targets)

    def _load_formula_pass() -> dict[tuple[str, str], dict[str, Any]]:
        """第1パス（wb_formula）: シート/セルごとに必要な情報だけを中間dictへ抽出してから閉じる。
        シート欠落チェック（旧・両ブック存在チェック）は、ここに無ければ第2パスで自然にスキップ
        される形で再現する。

        独立した関数スコープに閉じ込めるのが要点——`wb_formula`だけでなく、ループ内で束縛される
        `formula_sheet`/`formula_cell`（Cellはparentツリー経由でWorkbook全体を参照する）もこの
        関数が返った時点でフレームごと消える。呼び出し元のローカル変数として残さないことで、
        `wb_formula.close()`（循環参照ゆえ即解放しない）に加えてGCが確実に回収できる状態を作る
        （MEM-1のRV是正#1・close()だけでは不十分）。
        """
        intermediate: dict[tuple[str, str], dict[str, Any]] = {}
        wb_formula = openpyxl.load_workbook(source, data_only=False, read_only=False, keep_links=False)
        try:
            for sheet_name in sorted(normalized_targets):
                if sheet_name not in wb_formula.sheetnames:
                    continue
                formula_sheet = wb_formula[sheet_name]
                for coordinate in sorted(normalized_targets[sheet_name]):
                    formula_cell = formula_sheet[coordinate]
                    xml = raw.get((sheet_name, coordinate), {})
                    formula = xml.get("formula")
                    # shared formulaのfollowerはOOXML上``<f t="shared" si="…"/>``となり、
                    # XML直読だけでは式が単独の``=``になる。openpyxlがmasterから復元した式を
                    # 権威として補い、存在する数式を空式として保存しない。
                    if (formula in {None, "="} and isinstance(formula_cell.value, str)
                            and formula_cell.value.startswith("=")):
                        formula = formula_cell.value
                    has_cache = bool(xml.get("has_cache")) if formula else False
                    intermediate[(sheet_name, coordinate)] = {
                        "raw_value": xml.get("raw_value", _json_scalar(formula_cell.value)),
                        "formula": formula,
                        "has_cache": has_cache,
                        "style_id": int(formula_cell.style_id),
                        "number_format": formula_cell.number_format or "General",
                        "formula_cell_value": formula_cell.value,
                    }
        finally:
            wb_formula.close()
        return intermediate

    intermediate = _load_formula_pass()
    # 関数フレームが消えても、openpyxlのWorkbookは内部循環参照（cell⇄parent等）を持つため
    # CPythonの参照カウント方式だけでは即解放されない（次の周期的GCまで生き残る）。第2パスを
    # 開く前にGCを1回回し、ピーク時に生きているブックを常に1冊に保つ。
    gc.collect()

    # 第2パス（wb_values）: formulaの場合のみ再計算値を要するため、第1パスの中間dictと合流する。
    out: dict[tuple[str, str], dict[str, Any]] = {}
    wb_values = openpyxl.load_workbook(source, data_only=True, read_only=False, keep_links=False)
    try:
        for sheet_name in sorted(normalized_targets):
            if sheet_name not in wb_values.sheetnames:
                continue
            values_sheet = wb_values[sheet_name]
            for coordinate in sorted(normalized_targets[sheet_name]):
                key = (sheet_name, coordinate)
                mid = intermediate.get(key)
                if mid is None:
                    continue
                formula = mid["formula"]
                has_cache = mid["has_cache"]
                values_cell = values_sheet[coordinate]
                semantic_value = values_cell.value if formula else mid["formula_cell_value"]
                calculation_status = "cached" if formula and has_cache else "cache_missing" if formula else "not_formula"
                display = format_display(
                    semantic_value,
                    mid["number_format"],
                    formula_cache_missing=bool(formula and not has_cache),
                )
                out[key] = {
                    "raw_value": mid["raw_value"],
                    "typed_value": _typed_value(semantic_value),
                    "formula": formula,
                    "cached_value": _json_scalar(values_cell.value) if formula and has_cache else None,
                    "calculation_status": calculation_status,
                    "style_id": mid["style_id"],
                    "number_format": mid["number_format"],
                    "display_value": display.value,
                    "display_source": display.source,
                    "display_status": display.status,
                    "display_reason": display.reason,
                }
    finally:
        wb_values.close()
    return out


def enrich_evidence(ir, path: str | Path) -> None:
    """XLSX Evidenceのcell/formula extensionへ表示契約をin-placeで追加する。"""
    targets: dict[str, set[str]] = {}
    elements = []
    for element in ir.elements:
        if element.type not in {"cell", "formula"} or not element.locator.sheet or not element.locator.cell_range:
            continue
        coordinate = element.locator.cell_range.upper()
        if ":" in coordinate or not _CELL_REF_RE.fullmatch(coordinate):
            continue
        targets.setdefault(element.locator.sheet, set()).add(coordinate)
        elements.append((element, coordinate))
    if not elements:
        return
    metadata = extract_cell_metadata(path, targets)
    missing = {
        "raw_value": None,
        "typed_value": {"type": "blank", "value": None},
        "formula": None,
        "cached_value": None,
        "calculation_status": "not_formula",
        "style_id": 0,
        "number_format": "General",
        "display_value": None,
        "display_source": EXCEL_DISPLAY_PROFILE,
        "display_status": "unsupported",
        "display_reason": "cell_not_found",
    }
    for element, coordinate in elements:
        element.extension.update(metadata.get((element.locator.sheet, coordinate), missing))


def metadata_json(metadata: dict[str, Any]) -> str:
    """テスト・診断向けの決定的直列化。"""
    return json.dumps(metadata, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
