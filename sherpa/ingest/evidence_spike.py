"""4形式を同じEvidence IR契約へ通すE2aの限定spike。

現行Document IRを共通coreへ写し、現行IRが落としているDrawingML/PDF object inventoryを原本packageから
補う。画像の内容、overlayの業務状態、図の意味は推測せず、存在・位置・hash・幾何関係と未解釈状態だけを
残す。E2b以降の本番adapterを固定する前に、共通coreと形式別extensionの境界を検証するための実装である。
"""
from __future__ import annotations

import hashlib
import json
import mimetypes
import posixpath
import zipfile
from dataclasses import asdict, dataclass, replace as _dataclass_replace
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree as ET

from . import document_ir, evidence_ir


_REL = "{http://schemas.openxmlformats.org/package/2006/relationships}"
_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
_W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
_WP = "{http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing}"
_PIC = "{http://schemas.openxmlformats.org/drawingml/2006/picture}"
_P = "{http://schemas.openxmlformats.org/presentationml/2006/main}"
_X = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"
_C = "{http://schemas.openxmlformats.org/drawingml/2006/chart}"
_DGM = "{http://schemas.openxmlformats.org/drawingml/2006/diagram}"
_V = "{urn:schemas-microsoft-com:vml}"
_O = "{urn:schemas-microsoft-com:office:office}"
_CT = "{http://schemas.openxmlformats.org/package/2006/content-types}"
# v3（L3・可視性・廃止表現の全形式展開）: `_xlsx_objects` に図形/画像によるセルの覆い判定
# （`_occlusion_ratio`・`mark_hidden`）を追加し、`_adapt_document_ir` が非表示行/列・取り消し線を
# `table:N`/`strike:N` から対応する `cell` 要素の `visibility`/`extension["visibility_reason"]` へ
# 反映するようになった。
# v4: `_xlsx_objects` が `xdr:grpSp`（グループ）の中を pptx の `walk()` と同じ考え方で再帰し、
# 子図形を個別要素として出すようになった（従来はグループ直下しか見ておらず、中の図形は
# 1個の連結文字列に潰れて検索に出なかった）。`start_object_id`/`end_object_id` はコネクタ要素
# だけに付ける（従来は子孫コネクタの接続先をグループ自身の属性として誤って持っていた）。
# v5（L9・R3）: 図形/コネクタの`extension["prst"]`（DrawingML `prstGeom/@prst`）を追加し、
# コネクタの`start_object_id`/`end_object_id`をpptxと対称に`connects_to`関係へ解決するようになった。
XLSX_ADAPTER_VERSION = "xlsx-evidence-adapter-v5"
# v3（L3）: docxの`strike:N`（`_build_docx_ir`）が`visibility_reason="strike"`を持つようになった
# （`_adapt_document_ir`は既存の`extension["visibility_reason"]`コピー経路をそのまま使うため
# `_adapt_document_ir`自体は変更していないが、docxが出すEvidence要素の中身が変わるため版を上げる）。
# v4（S1・D-2）: `_build_docx_ir`が段落のsource_mapへ`floating_anchors`（浮動図形の関連付けの事実・
# 幾何断定なし）を持つようになり、`_adapt_document_ir`がそれをEvidence extensionへ運ぶようになった。
DOCX_ADAPTER_VERSION = "docx-evidence-adapter-v4"
# v3（L3）: `_pptx_objects`/`_adapt_document_ir` が `covered_by_text`（前面テキストによる重ね）を
# 前面要素の生テキストごと `extension` へ直接持たせるようになった（従来は `document_ir_source_map` に
# 前面要素の生ID参照だけが埋もれており、Evidence IR単体からは前面テキストへ辿れなかった）。
# v4（L9・R3）: 図形の`extension["prst"]`（DrawingML `prstGeom/@prst`）を追加した。
PPTX_ADAPTER_VERSION = "pptx-evidence-adapter-v4"
PDF_ADAPTER_VERSION = "pdf-evidence-adapter-v2"
PDF_XOBJECT_MAX_DEPTH = 16
PDF_XOBJECT_MAX_RESOLUTIONS = 4096


@dataclass(frozen=True)
class _ContentHost:
    """OOXML内で独立したrelationshipを持つcontent-bearing XML part。"""

    part: str
    host_kind: str
    host_visibility: str
    host_lifecycle: str = "active"
    slide: int | None = None
    visibility: str = "visible"
    lifecycle: str = "active"

    def metadata(self) -> dict[str, Any]:
        return {
            "host_part": self.part,
            "host_kind": self.host_kind,
            "host_visibility": self.host_visibility,
            "host_lifecycle": self.host_lifecycle,
        }


def _canonical(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _stable_id(prefix: str, *parts: Any) -> str:
    payload = "\n".join(_canonical(part) for part in parts)
    return f"{prefix}:" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:24]


def _source(path: Path) -> evidence_ir.EvidenceSource:
    return evidence_ir.EvidenceSource(
        file_type=path.suffix.lower().lstrip("."),
        content_hash="sha256:" + hashlib.sha256(path.read_bytes()).hexdigest(),
    )


class _Builder:
    def __init__(self, source: evidence_ir.EvidenceSource):
        self.ir = evidence_ir.EvidenceIR(
            schema_version=evidence_ir.EVIDENCE_IR_SCHEMA_VERSION,
            parser_profile=evidence_ir.EVIDENCE_PARSER_PROFILE,
            source=source,
        )
        self._coverage_by_id: dict[str, evidence_ir.CoverageItem] = {}
        self._element_coverage: dict[str, str] = {}
        self._element_index: dict[str, int] = {}

    def add_coverage(
        self,
        scope: str,
        locator: evidence_ir.Locator,
        status: str,
        reason: str | None = None,
        *,
        detected_kind: str,
        content_basis: str | None = None,
        reason_code: str | None = None,
        parser_id: str = evidence_ir.EVIDENCE_PARSER_PROFILE,
        detail: dict[str, Any] | None = None,
    ) -> str:
        reason_code = reason_code or reason or ("content_extracted" if status == "extracted" else "unspecified")
        if content_basis is None:
            if status == "intentionally_ignored":
                content_basis = "none"
            elif detected_kind in {"picture", "image_xobject", "image_asset", "package_thumbnail"}:
                content_basis = "pixel_only"
            elif detected_kind in {"text_object", "positioned_text"}:
                content_basis = "text_layer"
            else:
                content_basis = "structured"
        coverage_id = evidence_ir.make_coverage_id(scope, detected_kind, locator)
        item = evidence_ir.CoverageItem(
            coverage_id=coverage_id,
            scope=scope,
            detected_kind=detected_kind,
            locator=locator,
            status=status,
            content_basis=content_basis,
            reason_code=reason_code,
            parser_id=parser_id,
            detail=detail or {},
        )
        existing = self._coverage_by_id.get(coverage_id)
        if existing is not None and existing.status != status:
            raise ValueError(
                f"coverage status conflict: {coverage_id}: {existing.status} != {status}"
            )
        if existing is not None and existing != item:
            raise ValueError(f"coverage classification conflict: {coverage_id}")
        if existing is None:
            self._coverage_by_id[coverage_id] = item
            self.ir.coverage.append(item)
        return coverage_id

    def add_element(
        self,
        element_type: str,
        locator: evidence_ir.Locator,
        *,
        parent_id: str | None,
        order: int,
        value: Any,
        coverage_status: str = "extracted",
        coverage_reason: str | None = None,
        coverage_content_basis: str | None = None,
        coverage_detail: dict[str, Any] | None = None,
        coverage_id: str | None = None,
        visibility: str = "visible",
        lifecycle: str = "active",
        extension: dict[str, Any] | None = None,
    ) -> str:
        element_id = _stable_id("evidence", element_type, asdict(locator), order)
        if coverage_id is None:
            coverage_id = self.add_coverage(
                "object",
                locator,
                coverage_status,
                coverage_reason,
                detected_kind=element_type,
                content_basis=coverage_content_basis,
                detail=coverage_detail,
            )
        self.ir.elements.append(evidence_ir.EvidenceElement(
            element_id=element_id,
            type=element_type,
            parent_id=parent_id,
            order=order,
            value=value,
            locator=locator,
            coverage_id=coverage_id,
            visibility=visibility,
            lifecycle=lifecycle,
            extension=extension or {},
        ))
        self._element_coverage[element_id] = coverage_id
        self._element_index[element_id] = len(self.ir.elements) - 1
        return element_id

    def coverage_for(self, element_id: str) -> str:
        return self._element_coverage[element_id]

    def mark_hidden(self, element_id: str, *, reason: str, extra_extension: dict[str, Any] | None = None) -> None:
        """既存要素の`visibility`を`"hidden"`へ差し替える（xlsxの図形/画像による覆いなど、要素追加後に
        幾何交差が判明するケース用）。`EvidenceElement`はfrozen dataclassのため`dataclasses.replace`で
        差し替える。`value`/`locator`/`coverage_id`は変えない（意味は断定せず可視性と理由だけ更新する・
        意味の断定はしない共通思想）。呼び出し順で複数の前面図形が同じ要素へ一致した場合は
        最後の呼び出し（=z_order降順で最も手前）が勝つ（先勝ちで固定しない）。
        """
        index = self._element_index[element_id]
        element = self.ir.elements[index]
        merged_extension = {**element.extension, "visibility_reason": reason}
        if extra_extension:
            merged_extension.update(extra_extension)
        self.ir.elements[index] = _dataclass_replace(element, visibility="hidden", extension=merged_extension)

    def add_relation(
        self,
        relation_type: str,
        source_id: str,
        target_id: str,
        *,
        evidence_ids: list[str],
        confidence: float = 1.0,
        extension: dict[str, Any] | None = None,
    ) -> None:
        relation_id = _stable_id("relation", relation_type, source_id, target_id, extension or {})
        self.ir.relations.append(evidence_ir.EvidenceRelation(
            relation_id=relation_id,
            type=relation_type,
            source_id=source_id,
            target_id=target_id,
            evidence_ids=evidence_ids,
            confidence=confidence,
            extension=extension or {},
        ))


def _part_for_element(file_type: str, element: document_ir.Element, slide_parts: dict[int, str]) -> str:
    source_map = element.source_map
    if file_type == "docx":
        fixed_parts = {
            "comment": "word/comments.xml",
            "footnote": "word/footnotes.xml",
            "endnote": "word/endnotes.xml",
        }
        if element.type in fixed_parts:
            return fixed_parts[element.type]
        part = source_map.get("part")
        if part:
            return part if "/" in part else f"word/{part}"
        return "word/document.xml"
    if file_type == "pptx":
        return source_map.get("part") or slide_parts.get(source_map.get("slide"), "ppt/presentation.xml")
    if file_type == "xlsx":
        return "xl/workbook.xml"
    return f"{file_type}:document"


def _locator_for_legacy_element(
    file_type: str,
    element: document_ir.Element,
    slide_parts: dict[int, str],
) -> evidence_ir.Locator:
    sm = element.source_map
    cell_range = sm.get("range") or sm.get("cell")
    return evidence_ir.Locator(
        part=_part_for_element(file_type, element, slide_parts),
        slide=sm.get("slide") if isinstance(sm.get("slide"), int) else None,
        sheet=sm.get("sheet") if isinstance(sm.get("sheet"), str) else None,
        cell_range=cell_range if isinstance(cell_range, str) else None,
        object_id=element.element_id,
        bbox=_bounds_to_bbox(sm.get("bounds")),
        extension={"document_ir_source_map": sm},
    )


def _bounds_to_bbox(bounds: Any) -> list[int | float] | None:
    if not isinstance(bounds, list) or len(bounds) != 4 or not all(isinstance(item, (int, float)) for item in bounds):
        return None
    x, y, cx, cy = bounds
    return [x, y, x + cx, y + cy]


def _adapt_document_ir(
    builder: _Builder,
    legacy: document_ir.DocumentIR | None,
    *,
    consume_legacy: bool = False,
) -> dict[str, str]:
    """現行IRを共通coreへ写す。cellは独立element、parent_idでtable所属を明示する。

    ``consume_legacy``は通常取り込みの巨大XLSX用。既存IR成果物とchunkを書き終えた後に限って使い、
    tableごとのcellをEvidenceへ移し終えた時点で元listを空にして二重保持のpeakを抑える。

    L3（可視性・廃止表現の全形式展開）: xlsxのcell要素を組み立てる際、親`table:N`の`source_map`が持つ
    `hidden_rows`/`hidden_columns`（この表の範囲に絞り込んだ非表示行/列一覧）と、同じ`legacy.elements`内の
    `strike_text`要素（`(sheet, cell)`で対応付け）を突き合わせ、対応するcellの`visibility`/
    `extension["visibility_reason"]`へ直接反映する（`ooxml_arm._build_xlsx_ir`docstring「非表示行/列・
    取り消し線のcell単位への反映」参照・`_field_piece`が読むのはcell要素自身の状態のため）。
    """
    if legacy is None:
        return {}
    file_type = legacy.source.file_type
    slide_parts = {
        element.source_map["slide"]: element.source_map["part"]
        for element in legacy.elements
        if element.type == "slide"
        and isinstance(element.source_map.get("slide"), int)
        and isinstance(element.source_map.get("part"), str)
    }
    ids: dict[str, str] = {}
    legacy_by_id: dict[str, document_ir.Element] = {}
    # parentは現行IRで必ず先に現れるが、将来の並び替えに依存しないよう2段でIDを確定する。
    for element in legacy.elements:
        locator = _locator_for_legacy_element(file_type, element, slide_parts)
        ids[element.element_id] = _stable_id("evidence", element.type, asdict(locator), element.order)
        legacy_by_id[element.element_id] = element

    # L3（可視性・廃止表現の全形式展開）: xlsxの取り消し線（`_build_xlsx_ir`が独立の`strike_text`要素として
    # 抽出済み）を、対応するcell要素へも構造的に反映する（`strike_text`要素だけでは`_field_piece`が読む
    # cell自身のvisibility/extensionに触れないため）。`(sheet, cell)`の集合だけを先に作っておき、
    # cell構築時にO(1)で引ける形にする（table当たりcell数が多い場合の走査量を増やさない）。
    xlsx_strike_cells: set[tuple[str | None, str | None]] = (
        {(el.source_map.get("sheet"), el.source_map.get("cell"))
         for el in legacy.elements if el.type == "strike_text"}
        if file_type == "xlsx" else set()
    )

    for element in legacy.elements:
        locator = _locator_for_legacy_element(file_type, element, slide_parts)
        truncated = bool(element.source_map.get("truncated"))
        extension: dict[str, Any] = {
            "origin": "document-ir-v2-adapter",
            "extraction": asdict(element.extraction),
            "visibility_reason": element.visibility_reason,
        }
        # L3（可視性・廃止表現の全形式展開）: pptxの前面テキストによる重ね（`_build_pptx_ir`の
        # covered_by_text＝取り消し線的表現）は`source_map`に前面**要素ID（document_ir内部ID）**しか
        # 持たないため、そのままではEvidence側から前面テキストへ辿れない。ここで前面要素を解決し、
        # Evidence element_idと前面テキストをこの要素自身のextensionへ直接持たせる（意味の断定＝「廃止」等は
        # しない・前面テキストをそのまま埋めるだけ＝意味の断定はしない共通思想）。
        covered_by_text_id = element.source_map.get("covered_by_text")
        if isinstance(covered_by_text_id, str):
            front_legacy = legacy_by_id.get(covered_by_text_id)
            front_evidence_id = ids.get(covered_by_text_id)
            if front_legacy is not None and front_evidence_id is not None:
                extension["covered_by_text"] = {"element_id": front_evidence_id, "text": front_legacy.text}
        # D-2（docxの浮動図形と段落の関連付け・`ooxml_arm._docx_floating_anchor_facts`docstring参照）:
        # 段落側のsource_mapに載った関連付けの事実（幾何断定なし）をそのままEvidence extensionへ運ぶ。
        # `_docx_host_objects`のEMU座標`overlaps`（浮動図形どうしの幾何関係）とは別物であることを
        # キー名（"floating_anchors"）で区別する。
        floating_anchors = element.source_map.get("floating_anchors")
        if isinstance(floating_anchors, list) and floating_anchors:
            extension["floating_anchors"] = floating_anchors
        element_id = builder.add_element(
            element.type,
            locator,
            parent_id=ids.get(element.parent_id) if element.parent_id else None,
            order=element.order,
            value=element.text,
            coverage_status="unsupported" if truncated else "extracted",
            coverage_reason="scan_cap_reached" if truncated else None,
            visibility=element.visibility,
            lifecycle=element.status,
            extension=extension,
        )
        if element_id != ids[element.element_id]:
            raise ValueError("legacy element ID calculation drift")
        parent_coverage_id = builder.coverage_for(element_id)
        # L3: xlsxはこの表要素のsource_mapが「この矩形範囲に関係する非表示行/列」を既に持っている
        # （`_build_xlsx_ir`参照・シート全体ではなく表の範囲に絞り込み済み）。cellごとに判定するための
        # 集合へ先に変換する（毎cellでリスト内包しない）。
        hidden_rows = set(element.source_map.get("hidden_rows") or []) if file_type == "xlsx" else set()
        hidden_columns = set(element.source_map.get("hidden_columns") or []) if file_type == "xlsx" else set()
        if file_type == "xlsx":
            from openpyxl.utils import get_column_letter
        for cell_order, cell in enumerate(element.cells or [], start=1):
            cell_range = _a1(cell.row, cell.column) if file_type == "xlsx" else None
            cell_locator = evidence_ir.Locator(
                part=locator.part,
                sheet=locator.sheet,
                cell_range=cell_range,
                # XLSXはsheet+cell_rangeで一意。DOCX等はcell_rangeが無いためobject IDを残す。
                object_id=(
                    f"{element.element_id}/cell:{cell.row}:{cell.column}"
                    if file_type != "xlsx" else None
                ),
                extension={
                    "row": cell.row,
                    "column": cell.column,
                    "row_span": cell.row_span,
                    "column_span": cell.column_span,
                },
            )
            cell_visibility = element.visibility
            cell_extension: dict[str, Any] = {"role": cell.role}
            if file_type == "xlsx":
                # 優先順位: 非表示行 → 非表示列 → 取り消し線（非表示行/列は実際に画面から消える幾何的
                # 事実＝hidden_sheetと同じ扱い。取り消し線は本文が読める状態のまま＝visibilityは変えず
                # reasonだけ残す＝「意味の断定はしない」共通設計）。
                if cell.row in hidden_rows:
                    cell_visibility = "hidden"
                    cell_extension["visibility_reason"] = "hidden_row"
                elif get_column_letter(cell.column) in hidden_columns:
                    cell_visibility = "hidden"
                    cell_extension["visibility_reason"] = "hidden_column"
                elif (locator.sheet, cell_range) in xlsx_strike_cells:
                    cell_extension["visibility_reason"] = "strike"
            builder.add_element(
                "cell",
                cell_locator,
                parent_id=element_id,
                order=cell_order,
                value=cell.text,
                coverage_id=parent_coverage_id,
                visibility=cell_visibility,
                lifecycle=element.status,
                extension=cell_extension,
            )
            # 親子所属はparent_idで一意に表せるため、cellごとの冗長belongs_to relationは作らない。
            # 30万cell級でrelation/coverageを二重に持つメモリ増幅を避ける。
        if consume_legacy and element.cells:
            element.cells.clear()
    return ids


def _a1(row: int, column: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(column)}{row}"


def _rels_name(part: str) -> str:
    path = PurePosixPath(part)
    return str(path.parent / "_rels" / f"{path.name}.rels")


def _relationships(entries: dict[str, bytes], part: str) -> dict[str, str]:
    try:
        root = ET.fromstring(entries[_rels_name(part)])
    except (KeyError, ET.ParseError):
        return {}
    out: dict[str, str] = {}
    for rel in root.findall(f"{_REL}Relationship"):
        rel_id = rel.get("Id")
        target = rel.get("Target")
        if not rel_id or not target or rel.get("TargetMode") == "External":
            continue
        out[rel_id] = (
            posixpath.normpath(target.lstrip("/"))
            if target.startswith("/")
            else posixpath.normpath(posixpath.join(posixpath.dirname(part), target))
        )
    return out


def _relationship_records(entries: dict[str, bytes], part: str) -> dict[str, dict[str, str]]:
    """relationshipのtargetだけでなくtypeも失わずに返す。外部targetは原本内assetとして扱わない。"""
    try:
        root = ET.fromstring(entries[_rels_name(part)])
    except (KeyError, ET.ParseError):
        return {}
    records: dict[str, dict[str, str]] = {}
    for relation in root.findall(f"{_REL}Relationship"):
        rel_id = relation.get("Id")
        target = relation.get("Target")
        if not rel_id or not target or relation.get("TargetMode") == "External":
            continue
        resolved = (
            posixpath.normpath(target.lstrip("/"))
            if target.startswith("/")
            else posixpath.normpath(posixpath.join(posixpath.dirname(part), target))
        )
        records[rel_id] = {"target": resolved, "type": relation.get("Type", "")}
    return records


def _image_relationship_records(entries: dict[str, bytes], part: str) -> dict[str, dict[str, str]]:
    """画像relationshipを、外部targetの生値を保持せずに列挙する。"""
    try:
        root = ET.fromstring(entries[_rels_name(part)])
    except (KeyError, ET.ParseError):
        return {}
    records: dict[str, dict[str, str]] = {}
    for relation in root.findall(f"{_REL}Relationship"):
        relationship_id = relation.get("Id")
        target = relation.get("Target")
        if not relationship_id or not target:
            continue
        target_mode = relation.get("TargetMode", "Internal")
        record = {
            "relationship_type": relation.get("Type", ""),
            "target_mode": target_mode,
        }
        if target_mode.casefold() == "external":
            record["external_target_sha256"] = "sha256:" + hashlib.sha256(target.encode("utf-8")).hexdigest()
        else:
            record["media_part"] = (
                posixpath.normpath(target.lstrip("/"))
                if target.startswith("/")
                else posixpath.normpath(posixpath.join(posixpath.dirname(part), target))
            )
        records[relationship_id] = record
    return records


def _content_type(entries: dict[str, bytes], part: str) -> str | None:
    root = _xml(entries, "[Content_Types].xml")
    if root is None:
        return None
    overrides = {
        node.get("PartName", "").lstrip("/"): node.get("ContentType", "")
        for node in root.findall(f"{_CT}Override")
    }
    if part in overrides:
        return overrides[part] or None
    defaults = {
        node.get("Extension", "").lower(): node.get("ContentType", "")
        for node in root.findall(f"{_CT}Default")
    }
    return defaults.get(PurePosixPath(part).suffix.lower().lstrip(".")) or None


def _package_entries(path: Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as archive:
        return {info.filename: archive.read(info) for info in archive.infolist() if not info.is_dir()}


_DOCUMENT_PROPERTY_NAMES = frozenset({
    "AppVersion", "Application", "Characters", "CharactersWithSpaces", "Company", "DocSecurity",
    "HiddenSlides", "HyperlinksChanged", "Lines", "LinksUpToDate", "Manager", "MMClips", "Notes",
    "Pages", "Paragraphs", "PresentationFormat", "ScaleCrop", "SharedDoc", "Slides", "Template",
    "TotalTime", "Words", "category", "contentStatus", "created", "creator", "description",
    "identifier", "keywords", "language", "lastModifiedBy", "lastPrinted", "modified", "revision",
    "subject", "title", "version",
})
_DOCUMENT_PROPERTY_CANONICAL_NAMES = {
    "Application": "application",
    "Company": "company",
    "creator": "author",
}


def _package_metadata(entries: dict[str, bytes]) -> dict[str, dict[str, Any]]:
    """標準docPropsの原値を検索本文へ水増しせずpart単位のcoverage metadataへ写す。"""
    metadata: dict[str, dict[str, Any]] = {}
    for part in ("docProps/core.xml", "docProps/app.xml"):
        root = _xml(entries, part)
        if root is None:
            continue
        properties: list[dict[str, Any]] = []
        for node in root.iter():
            name = node.tag.split("}")[-1]
            if name not in _DOCUMENT_PROPERTY_NAMES or node.text is None:
                continue
            value = node.text
            if not value.strip():
                continue
            properties.append({
                "ordinal": len(properties) + 1,
                "xml_tag": node.tag,
                "native_property_name": name,
                "property_name": _DOCUMENT_PROPERTY_CANONICAL_NAMES.get(name, name),
                "raw_value": value,
            })
        metadata[part] = {
            "metadata_kind": "standard_document_properties",
            "part_sha256": hashlib.sha256(entries[part]).hexdigest(),
            "part_size": len(entries[part]),
            "properties": properties,
        }
    return metadata


_PACKAGE_CLASSIFIER_ID = "ooxml-package-classifier-v2"
_IMAGE_PART_SUFFIXES = frozenset({
    ".apng", ".avif", ".bmp", ".emf", ".gif", ".heic", ".jpeg", ".jpg", ".png",
    ".svg", ".tif", ".tiff", ".webp", ".wmf",
})
_PACKAGE_SUPPORT_EXACT = frozenset({
    "[Content_Types].xml",
    "xl/calcChain.xml",
    "xl/styles.xml",
    "word/fontTable.xml",
    "word/numbering.xml",
    "word/settings.xml",
    "word/styles.xml",
    "word/stylesWithEffects.xml",
    "word/webSettings.xml",
    "ppt/presProps.xml",
    "ppt/tableStyles.xml",
    "ppt/viewProps.xml",
})
_PACKAGE_SUPPORT_PREFIXES = (
    "ppt/printersettings/",
    "xl/printersettings/",
    "xl/theme/",
    "word/theme/",
    "ppt/theme/",
)


def _is_xml_part(part: str) -> bool:
    lower = part.lower()
    return lower.endswith((".xml", ".rels", ".vml"))


def _package_part_kind(part: str) -> str:
    """pathだけから、抽出成否とは独立したpart種別を決定する。"""
    lower = part.lower()
    suffix = PurePosixPath(lower).suffix
    if lower == "[content_types].xml":
        return "package_content_types"
    if lower.endswith(".rels"):
        return "package_relationships"
    if lower.startswith("docprops/"):
        return "package_thumbnail" if suffix in _IMAGE_PART_SUFFIXES else "package_metadata"
    if "/media/" in lower:
        return "image_asset" if suffix in _IMAGE_PART_SUFFIXES else "binary_asset"
    if "/diagrams/" in lower:
        return "smartart_part"
    if "/charts/" in lower:
        return "chart_part"
    if "/embeddings/" in lower or "oleobject" in lower:
        return "ole_embedding"
    if lower.endswith(".vml") or "vmldrawing" in lower:
        return "vml_part"
    if lower.startswith("customxml/"):
        return "custom_xml_part"
    if part in _PACKAGE_SUPPORT_EXACT or lower.startswith(_PACKAGE_SUPPORT_PREFIXES):
        return "package_support_part"
    return "xml_content_part" if _is_xml_part(part) else "binary_content_part"


def _is_empty_office_custom_xml(entries: dict[str, bytes], part: str) -> bool:
    if PurePosixPath(part).name.lower().startswith("itemprops"):
        return True
    root = _xml(entries, part)
    if root is None or root.tag.split("}")[-1] != "Sources":
        return False
    return not any(child.tag.split("}")[-1] == "Source" for child in root)


def _package_coverage(
    builder: _Builder,
    entries: dict[str, bytes],
    content_parts: set[str],
    *,
    part_details: dict[str, dict[str, Any]] | None = None,
) -> None:
    """ZIP内の全partを狭いallowlistとcontent-bearing分類へ振り分ける。

    ``content_parts``または要素locatorで実際に扱ったpartだけをextractedとする。未知partや
    SmartArt/OLE/VML等を補助part扱いにせず、明示的なunsupportedとして残す。
    """
    represented_parts = {element.locator.part for element in builder.ir.elements}
    extracted_parts = content_parts | represented_parts
    for part in sorted(entries):
        locator = evidence_ir.Locator(part=part)
        detected_kind = _package_part_kind(part)
        common = {
            "detected_kind": detected_kind,
            "parser_id": _PACKAGE_CLASSIFIER_ID,
            "detail": {"part": part, **(part_details or {}).get(part, {})},
        }
        if _is_xml_part(part):
            try:
                ET.fromstring(entries[part])
            except ET.ParseError as exc:
                builder.add_coverage(
                    "part",
                    locator,
                    "failed",
                    detected_kind=detected_kind,
                    content_basis="structured",
                    reason_code="parse_failed",
                    parser_id=_PACKAGE_CLASSIFIER_ID,
                    detail={"part": part, "error_type": exc.__class__.__name__},
                )
                continue
        if part in extracted_parts:
            builder.add_coverage(
                "part", locator, "extracted", content_basis="structured",
                reason_code="content_extracted", **common,
            )
        elif detected_kind == "image_asset":
            builder.add_coverage(
                "part", locator, "metadata_only", content_basis="pixel_only",
                reason_code="binary_asset_content_uninterpreted", **common,
            )
        elif detected_kind == "package_thumbnail":
            builder.add_coverage(
                "part", locator, "intentionally_ignored", content_basis="none",
                reason_code="package_preview_thumbnail", **common,
            )
        elif detected_kind == "binary_asset":
            builder.add_coverage(
                "part", locator, "metadata_only", content_basis="binary_only",
                reason_code="binary_asset_content_uninterpreted", **common,
            )
        elif detected_kind == "package_metadata":
            builder.add_coverage(
                "part", locator, "metadata_only", content_basis="structured",
                reason_code="package_metadata_uninterpreted", **common,
            )
        elif detected_kind == "custom_xml_part" and _is_empty_office_custom_xml(entries, part):
            builder.add_coverage(
                "part", locator, "intentionally_ignored", content_basis="none",
                reason_code="empty_office_custom_xml_support", **common,
            )
        elif detected_kind in {"package_content_types", "package_relationships", "package_support_part"}:
            builder.add_coverage(
                "part", locator, "intentionally_ignored", content_basis="none",
                reason_code="package_support_part", **common,
            )
        else:
            reason_code = {
                "smartart_part": "unsupported_smartart",
                "chart_part": "unsupported_chart",
                "ole_embedding": "unsupported_ole",
                "vml_part": "unsupported_vml",
                "custom_xml_part": "unsupported_custom_xml",
                "xml_content_part": "unsupported_content_part",
                "binary_content_part": "unsupported_binary_part",
            }[detected_kind]
            builder.add_coverage(
                "part",
                locator,
                "unsupported",
                content_basis="binary_only" if detected_kind in {"ole_embedding", "binary_content_part"} else "structured",
                reason_code=reason_code,
                **common,
            )


def _xml(entries: dict[str, bytes], part: str) -> ET.Element | None:
    try:
        return ET.fromstring(entries[part])
    except (KeyError, ET.ParseError):
        return None


def _xlsx_sheet_parts(entries: dict[str, bytes]) -> dict[str, str]:
    root = _xml(entries, "xl/workbook.xml")
    if root is None:
        return {}
    rels = _relationships(entries, "xl/workbook.xml")
    out: dict[str, str] = {}
    for sheet in root.findall(f"{_X}sheets/{_X}sheet"):
        name, rel_id = sheet.get("name"), sheet.get(f"{_R}id")
        if name and rel_id in rels:
            out[name] = rels[rel_id]
    return out


def _marker(anchor: ET.Element, tag: str) -> tuple[int, int, int, int] | None:
    node = anchor.find(f"{_XDR}{tag}")
    if node is None:
        return None
    values = []
    for name in ("col", "row", "colOff", "rowOff"):
        child = node.find(f"{_XDR}{name}")
        try:
            values.append(int(child.text or 0) if child is not None else 0)
        except ValueError:
            return None
    return tuple(values)  # type: ignore[return-value]


def _xlsx_anchor(anchor: ET.Element) -> tuple[str | None, dict[str, Any]]:
    start, end = _marker(anchor, "from"), _marker(anchor, "to")
    kind = PurePosixPath(anchor.tag).name.split("}")[-1]
    extension: dict[str, Any] = {"anchor_type": kind, "from_marker": start, "to_marker": end}
    if kind == "twoCellAnchor" and start is not None and end is not None:
        # `to`は次のセル境界を指すため、range終端は0-based marker値をそのまま1-based座標として使う。
        return f"{_a1(start[1] + 1, start[0] + 1)}:{_a1(max(end[1], 1), max(end[0], 1))}", extension
    if kind == "oneCellAnchor" and start is not None:
        return _a1(start[1] + 1, start[0] + 1), extension
    return None, extension


def _xlsx_embedded_anchor(node: ET.Element) -> tuple[str | None, dict[str, Any]]:
    """oleObject/objectPr内のnamespace差を許容してfrom/to cell anchorを読む。"""
    markers: dict[str, tuple[int, int] | None] = {"from": None, "to": None}
    for marker in node.iter():
        local = marker.tag.split("}")[-1]
        if local not in markers:
            continue
        values: dict[str, int] = {}
        for child in list(marker):
            child_local = child.tag.split("}")[-1]
            if child_local not in {"col", "row"}:
                continue
            try:
                values[child_local] = int(child.text or 0)
            except ValueError:
                continue
        if "col" in values and "row" in values:
            markers[local] = (values["col"], values["row"])
    start, end = markers["from"], markers["to"]
    if start is None:
        return None, {"from_cell": None, "to_cell": None}
    start_cell = _a1(start[1] + 1, start[0] + 1)
    end_cell = _a1(end[1] + 1, end[0] + 1) if end is not None else start_cell
    return f"{start_cell}:{end_cell}" if end_cell != start_cell else start_cell, {
        "from_cell": start,
        "to_cell": end,
    }


def _cell_rect(cell_range: str | None) -> list[float] | None:
    if not cell_range:
        return None
    from openpyxl.utils.cell import range_boundaries
    try:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    except ValueError:
        return None
    if not all(isinstance(value, int) for value in (min_col, min_row, max_col, max_row)):
        return None
    # Excel locatorの幾何は列/行境界を0-basedの正規化cell座標で表す。
    return [float(min_col - 1), float(min_row - 1), float(max_col), float(max_row)]


def _area(rect: list[int | float]) -> float:
    return max(0.0, float(rect[2] - rect[0])) * max(0.0, float(rect[3] - rect[1]))


def _drawing_kind(node: ET.Element) -> str | None:
    local = node.tag.split("}")[-1]
    if local == "sp":
        props = node.find(f"{_XDR}nvSpPr/{_XDR}cNvSpPr")
        return "textbox" if props is not None and props.get("txBox") == "1" else "shape"
    return {
        "pic": "picture",
        "cxnSp": "connector",
        "grpSp": "group",
        "graphicFrame": "graphic_frame",
    }.get(local)


def _occlusion_ratio() -> float:
    """覆い判定のしきい値（交差面積 ÷ 対象面積・0.0-1.0）。**定義は `office_md._OCCLUSION_RATIO` 1箇所**。

    pptx は `_adapt_document_ir`→`ooxml_arm`→`office_md` の経路で同じ定数へ到達する。xlsx はここで
    直接判定するため、値を複製せず呼び出しのたびに読む（複製すると片方だけ変えたときに pptx と
    xlsx で判定が黙って食い違う）。import は関数内で行う——`office_md` 側が `evidence_spike` を
    関数内 import しており、モジュール階層で相互参照すると循環するため。
    """
    from . import office_md
    return office_md._OCCLUSION_RATIO


def _xlsx_shape_has_solid_fill(node: ET.Element) -> bool:
    """xlsx図形（`xdr:sp`）が`xdr:spPr/a:solidFill`を持つか（`a:noFill`ならFalse）。

    `office_md._pptx_has_solid_fill`と同じ判定をxlsxのspreadsheetDrawing名前空間へ適用したもの
    （DrawingMLの塗り語彙自体はpptx/xlsxで共通のため二重の判定基準を作らない）。
    """
    sp_pr = node.find(f"{_XDR}spPr")
    if sp_pr is None:
        return False
    if sp_pr.find(f"{_A}noFill") is not None:
        return False
    return sp_pr.find(f"{_A}solidFill") is not None


def _c_nv_pr(node: ET.Element, prefix: str, kind: str) -> ET.Element | None:
    paths = {
        "textbox": f"{prefix}nvSpPr/{prefix}cNvPr",
        "shape": f"{prefix}nvSpPr/{prefix}cNvPr",
        "picture": f"{prefix}nvPicPr/{prefix}cNvPr",
        "connector": f"{prefix}nvCxnSpPr/{prefix}cNvPr",
        "group": f"{prefix}nvGrpSpPr/{prefix}cNvPr",
        "graphic_frame": f"{prefix}nvGraphicFramePr/{prefix}cNvPr",
    }
    return node.find(paths[kind])


def _chart_payload(entries: dict[str, bytes], chart_part: str) -> tuple[str | None, dict[str, Any]]:
    """chart XMLの文字・参照式・cached値だけを推測せず共通抽出する。"""
    root = _xml(entries, chart_part)
    if root is None:
        return None, {}
    texts = list(dict.fromkeys(
        node.text for node in root.iter(f"{_A}t") if node.text and node.text.strip()
    ))
    references = list(dict.fromkeys(
        node.text for node in root.iter(f"{_C}f") if node.text and node.text.strip()
    ))
    cached_values = [
        node.text for node in root.iter(f"{_C}v") if node.text is not None
    ]
    lines = [
        *(f"グラフ内文字: {value}" for value in texts),
        *(f"参照範囲: {value}" for value in references),
        *(f"キャッシュ値: {value}" for value in cached_values),
    ]
    return ("\n".join(lines) or None), {
        "chart_texts": texts,
        "chart_references": references,
        "chart_cached_values": cached_values,
    }


def _xlsx_chart_payload(entries: dict[str, bytes], chart_part: str) -> tuple[str | None, dict[str, Any]]:
    """旧private名を共通chart抽出へ委譲する。"""
    return _chart_payload(entries, chart_part)


def _add_unparsed_notice(
    builder: _Builder,
    locator: evidence_ir.Locator,
    *,
    detected_kind: str,
    reason_code: str,
    detail: dict[str, Any] | None = None,
) -> None:
    builder.add_coverage(
        "region",
        locator,
        "unsupported",
        detected_kind=detected_kind,
        content_basis="structured",
        reason_code=reason_code,
        detail=detail or {},
    )


def _drawing_asset(
    entries: dict[str, bytes],
    host_part: str,
    node: ET.Element,
    *,
    role: str,
) -> dict[str, Any]:
    """pictureとshape fillの全blip参照を、外部参照も黙って捨てずasset契約へ写す。"""
    relationship_records = _image_relationship_records(entries, host_part)
    references: list[dict[str, Any]] = []
    relationship_ids: list[str] = []
    unresolved: list[str] = []
    external: list[str] = []
    for blip in node.findall(f".//{_A}blip"):
        identifiers = [
            (attribute, relationship_id)
            for attribute in ("embed", "link")
            if isinstance((relationship_id := blip.get(f"{_R}{attribute}")), str) and relationship_id
        ]
        if not identifiers:
            references.append({
                "asset_role": role,
                "relationship_attribute": "missing",
                "binding_status": "relationship_id_missing",
            })
            unresolved.append("missing_relationship_id")
            continue
        for attribute, relationship_id in identifiers:
            if relationship_id not in relationship_ids:
                relationship_ids.append(relationship_id)
            reference: dict[str, Any] = {
                "relationship_id": relationship_id,
                "relationship_attribute": attribute,
                "asset_role": role,
            }
            relationship = relationship_records.get(relationship_id)
            if relationship is None:
                reference["binding_status"] = "relationship_missing"
                unresolved.append(relationship_id)
            elif relationship.get("target_mode", "Internal").casefold() == "external":
                reference.update(relationship)
                reference["binding_status"] = "external_reference"
                external.append(relationship_id)
            else:
                media_part = relationship.get("media_part")
                reference.update(relationship)
                if media_part not in entries:
                    reference["binding_status"] = "target_missing"
                    unresolved.append(relationship_id)
                else:
                    media = entries[media_part]
                    reference.update({
                        "asset_sha256": hashlib.sha256(media).hexdigest(),
                        "binding_status": "resolved",
                    })
                    if (pixel_size := _image_pixel_size(media)) is not None:
                        reference["pixel_size"] = pixel_size
            if reference not in references:
                references.append(reference)
    if not references:
        return {}
    resolved = [reference for reference in references if reference.get("binding_status") == "resolved"]
    primary = resolved[0] if resolved else references[0]
    extension: dict[str, Any] = {
        **primary,
        "asset_role": role,
        "image_relationship_ids": relationship_ids,
        "image_references": references,
        # extract_assets/routeは参照ごとの全候補を扱う。失敗候補もhash無しのまま残す。
        "assets": references,
    }
    if unresolved:
        extension["unresolved_image_relationship_ids"] = list(dict.fromkeys(unresolved))
    if external:
        extension["external_image_relationship_ids"] = list(dict.fromkeys(external))
    return extension


def _add_raster_reference_coverage(
    builder: _Builder,
    locator: evidence_ir.Locator,
    extension: dict[str, Any],
    *,
    failed_kind: str,
    external_kind: str,
) -> None:
    """解決不能/外部画像参照をCoverage noticeへ安全に搬送する。"""
    references = extension.get("image_references")
    if not isinstance(references, list):
        return
    failed = [
        reference for reference in references
        if isinstance(reference, dict)
        and reference.get("binding_status") in {"relationship_id_missing", "relationship_missing", "target_missing"}
    ]
    external = [
        reference for reference in references
        if isinstance(reference, dict) and reference.get("binding_status") == "external_reference"
    ]
    if failed:
        builder.add_coverage(
            "object",
            locator,
            "failed",
            detected_kind=failed_kind,
            content_basis="pixel_only",
            reason_code="asset_binding_failed",
            detail={
                "relationship_ids": [
                    reference.get("relationship_id") or "missing_relationship_id"
                    for reference in failed
                ],
                "references": failed,
            },
        )
    if external:
        builder.add_coverage(
            "object",
            locator,
            "unsupported",
            detected_kind=external_kind,
            content_basis="pixel_only",
            reason_code="external_image_reference_not_fetched",
            detail={"references": external},
        )


def _smartart_payload(
    entries: dict[str, bytes], host_part: str, node: ET.Element,
) -> tuple[str | None, dict[str, Any], set[str]]:
    rels = _relationships(entries, host_part)
    relationship_ids: dict[str, str] = {}
    targets: dict[str, str] = {}
    for rel_ids in node.findall(f".//{_DGM}relIds"):
        for role in ("dm", "lo", "qs", "cs"):
            relationship_id = rel_ids.get(f"{_R}{role}")
            if not relationship_id:
                continue
            relationship_ids[role] = relationship_id
            target = rels.get(relationship_id)
            if target in entries:
                targets[role] = target
    data_part = targets.get("dm")
    texts = []
    if data_part:
        data = _xml(entries, data_part)
        if data is not None:
            texts = list(dict.fromkeys(
                text.text for text in data.iter(f"{_A}t") if text.text and text.text.strip()
            ))
    value = "\n".join(f"SmartArt内文字: {text}" for text in texts) or None
    return value, {
        "relationship_ids": relationship_ids,
        "diagram_parts": targets,
        "diagram_texts": texts,
        "semantic_graph_status": "unsupported",
    }, {data_part} if data_part else set()


def _ole_metadata(
    entries: dict[str, bytes], host_part: str, relationship_id: str, node: ET.Element,
) -> dict[str, Any]:
    relation = _relationship_records(entries, host_part).get(relationship_id, {})
    embedded_part = relation.get("target")
    extension: dict[str, Any] = {
        "relationship_id": relationship_id,
        "relationship_type": relation.get("type"),
        "embedded_part": embedded_part,
        "prog_id": node.get("ProgID") or node.get("progId"),
        "object_id": node.get("ObjectID") or node.get("spid"),
        "embedded_content_status": "unsupported",
    }
    if embedded_part in entries:
        extension.update({
            "binary_sha256": hashlib.sha256(entries[embedded_part]).hexdigest(),
            "content_type": _content_type(entries, embedded_part),
            "binary_size": len(entries[embedded_part]),
        })
    return extension


def _add_chart_data(
    builder: _Builder,
    entries: dict[str, bytes],
    *,
    chart_part: str,
    parent_id: str,
    object_token: str,
    sheet: str | None = None,
    slide: int | None = None,
) -> None:
    locator = evidence_ir.Locator(
        part=chart_part,
        sheet=sheet,
        slide=slide,
        object_id=f"chart-data:{object_token}",
    )
    if _xml(entries, chart_part) is None:
        builder.add_coverage(
            "object",
            locator,
            "failed",
            detected_kind="chart_data",
            content_basis="structured",
            reason_code="parse_failed",
            detail={"chart_part": chart_part, "error_type": "ParseError"},
        )
        return
    value, extension = _chart_payload(entries, chart_part)
    builder.add_element(
        "chart_data",
        locator,
        parent_id=parent_id,
        order=1,
        value=value,
        extension={"chart_part": chart_part, **extension},
    )
    _add_unparsed_notice(
        builder,
        locator,
        detected_kind="chart_unparsed_attributes",
        reason_code="unsupported_chart_attributes",
        detail={"chart_part": chart_part},
    )


def _add_smartart_data(
    builder: _Builder,
    entries: dict[str, bytes],
    *,
    parent_id: str,
    value: str | None,
    extension: dict[str, Any],
    host_part: str,
    object_token: str,
    sheet: str | None = None,
    slide: int | None = None,
) -> None:
    data_part = extension.get("diagram_parts", {}).get("dm")
    locator = evidence_ir.Locator(
        part=data_part or host_part,
        sheet=sheet,
        slide=slide,
        object_id=f"smartart-data:{object_token}",
    )
    if not data_part:
        builder.add_coverage(
            "object",
            locator,
            "unsupported",
            detected_kind="smartart_data",
            content_basis="structured",
            reason_code="smartart_data_part_missing",
            detail={"host_part": host_part},
        )
        return
    if _xml(entries, data_part) is None:
        builder.add_coverage(
            "object",
            locator,
            "failed",
            detected_kind="smartart_data",
            content_basis="structured",
            reason_code="parse_failed",
            detail={"diagram_data_part": data_part, "error_type": "ParseError"},
        )
        return
    builder.add_element(
        "smartart_data",
        locator,
        parent_id=parent_id,
        order=1,
        value=value,
        extension=extension,
    )
    _add_unparsed_notice(
        builder,
        locator,
        detected_kind="smartart_semantic_graph",
        reason_code="unsupported_smartart_semantics",
        detail={"diagram_data_part": data_part},
    )


def _integer(value: str | None) -> int | None:
    try:
        return int(value) if value is not None else None
    except ValueError:
        return None


def _image_pixel_size(data: bytes) -> list[int] | None:
    """任意ライブラリに依存せず、OOXMLで頻出するPNG/GIF/JPEGのピクセル寸法を読む。"""
    if len(data) >= 24 and data.startswith(b"\x89PNG\r\n\x1a\n"):
        return [int.from_bytes(data[16:20], "big"), int.from_bytes(data[20:24], "big")]
    if len(data) >= 10 and data[:6] in {b"GIF87a", b"GIF89a"}:
        return [int.from_bytes(data[6:8], "little"), int.from_bytes(data[8:10], "little")]
    if len(data) >= 4 and data[:2] == b"\xff\xd8":
        offset = 2
        while offset + 9 < len(data):
            if data[offset] != 0xFF:
                offset += 1
                continue
            marker = data[offset + 1]
            if marker in {0xD8, 0xD9}:
                offset += 2
                continue
            if offset + 4 > len(data):
                break
            length = int.from_bytes(data[offset + 2:offset + 4], "big")
            if length < 2 or offset + 2 + length > len(data):
                break
            if marker in {0xC0, 0xC1, 0xC2, 0xC3, 0xC5, 0xC6, 0xC7, 0xC9, 0xCA, 0xCB, 0xCD, 0xCE, 0xCF}:
                height = int.from_bytes(data[offset + 5:offset + 7], "big")
                width = int.from_bytes(data[offset + 7:offset + 9], "big")
                return [width, height]
            offset += 2 + length
    return None


def _vml_length(value: str | None) -> float | None:
    if not value:
        return None
    lowered = value.strip().lower()
    for suffix in ("pt", "px", "in", "cm", "mm"):
        if lowered.endswith(suffix):
            lowered = lowered[:-len(suffix)]
            break
    try:
        return float(lowered)
    except ValueError:
        return None


def _vml_geometry(shape: ET.Element) -> tuple[str | None, list[float] | None, dict[str, Any], bool]:
    style_text = shape.get("style", "")
    style = {
        name.strip().lower(): value.strip()
        for declaration in style_text.split(";") if ":" in declaration
        for name, value in [declaration.split(":", 1)]
    }
    left = _vml_length(style.get("margin-left") or style.get("left"))
    top = _vml_length(style.get("margin-top") or style.get("top"))
    width = _vml_length(style.get("width"))
    height = _vml_length(style.get("height"))
    bbox = [left, top, left + width, top + height] if None not in {left, top, width, height} else None
    anchor_values: list[int] | None = None
    row = column = None
    for child in shape.iter():
        local = child.tag.split("}")[-1]
        if local == "Anchor" and child.text:
            try:
                values = [int(value.strip()) for value in child.text.split(",")]
            except ValueError:
                values = []
            if len(values) >= 8:
                anchor_values = values[:8]
        elif local == "Row" and child.text:
            row = _integer(child.text)
        elif local == "Column" and child.text:
            column = _integer(child.text)
    cell_range = None
    if anchor_values is not None:
        start_col, start_row, end_col, end_row = (
            anchor_values[0], anchor_values[2], anchor_values[4], anchor_values[6]
        )
        cell_range = f"{_a1(start_row + 1, start_col + 1)}:{_a1(end_row + 1, end_col + 1)}"
    elif row is not None and column is not None:
        cell_range = _a1(row + 1, column + 1)
    extension = {
        "style": style_text,
        "style_fields": style,
        "anchor": anchor_values,
        "coordinate_system": "vml-style" if bbox is not None else "xlsx-vml-anchor" if cell_range else "unknown",
    }
    unparsed_geometry = bbox is None and cell_range is None or any(
        shape.get(attribute) is not None for attribute in ("path", "coordsize", "coordorigin")
    )
    return cell_range, bbox, extension, unparsed_geometry


_VML_OBJECT_TAGS = frozenset({
    f"{_V}shape",
    f"{_V}group",
    f"{_V}rect",
    f"{_V}roundrect",
    f"{_V}oval",
    f"{_V}line",
    f"{_V}polyline",
    f"{_V}curve",
    f"{_V}arc",
    f"{_V}image",
})


def _owned_vml_nodes(shape: ET.Element):
    """shape/groupに属するノードを原文順で返す。子VML objectの内容は子側で別に抽出する。"""
    yield shape
    for child in shape:
        if child.tag in _VML_OBJECT_TAGS:
            continue
        yield from _owned_vml_nodes(child)


def _vml_texts(shape: ET.Element) -> tuple[list[str], list[str]]:
    """VML textboxのWordprocessingML文字とWordArt textpathを原文順で保持する。"""
    texts: list[str] = []
    sources: list[str] = []
    owned = list(_owned_vml_nodes(shape))
    consumed_text_nodes: set[int] = set()
    for node in owned:
        value = None
        source = None
        if node.tag == f"{_W}p":
            runs = [child for child in _owned_vml_nodes(node) if child.tag == f"{_W}t"]
            value = "".join(child.text or "" for child in runs)
            consumed_text_nodes.update(id(child) for child in runs)
            source = "w:p/w:t"
        elif node.tag == f"{_W}t" and id(node) not in consumed_text_nodes:
            value, source = node.text, "w:t"
        elif node.tag == f"{_V}textpath":
            value, source = node.get("string"), "v:textpath@string"
        if not isinstance(value, str) or not value.strip():
            continue
        texts.append(value)
        sources.append(source or "unknown")
    return texts, sources


def _vml_assets(entries: dict[str, bytes], part: str, shape: ET.Element) -> dict[str, Any]:
    """VML shapeの画像参照をDrawingMLと同じasset契約へ写す。

    VMLは``v:imagedata``のほか、画像塗りを``v:fill``へ保持する原本がある。実際のOffice文書では
    relationship属性が``r:id``または旧Office名前空間の``o:relid``になるため、どちらもpart自身の
    relationshipから解決する。参照単位を維持し、同じbytesを複数箇所が参照してもここでは統合しない。
    """
    relationships = _image_relationship_records(entries, part)
    references: list[dict[str, Any]] = []
    relationship_ids: list[str] = []
    unresolved: list[str] = []
    external: list[str] = []
    for node in _owned_vml_nodes(shape):
        local_name = node.tag.split("}")[-1]
        if node.tag not in {f"{_V}imagedata", f"{_V}fill", f"{_V}image"}:
            continue
        identifier = next((
            (attribute, value)
            for attribute, key in (
                ("id", f"{_R}id"),
                ("embed", f"{_R}embed"),
                ("link", f"{_R}link"),
                ("relid", f"{_O}relid"),
                ("relid", "relid"),
            )
            if isinstance((value := node.get(key)), str) and value
        ), None)
        relationship_id = identifier[1] if identifier is not None else None
        role = "picture_content" if node.tag == f"{_V}image" else "shape_fill"
        if relationship_id is None:
            # 単色/gradientだけのv:fillは画像assetではない。imagedata、またはsrcを持つfillだけを
            # content-bearingな未解決画像として分類する。
            if node.tag in {f"{_V}imagedata", f"{_V}image"} or node.get("src"):
                unresolved.append(f"{local_name}:missing_relationship_id")
                references.append({
                    "asset_role": role,
                    "relationship_attribute": "missing",
                    "binding_status": "relationship_id_missing",
                    "vml_element": local_name,
                })
            continue
        relationship_ids.append(relationship_id)
        reference: dict[str, Any] = {
            "relationship_id": relationship_id,
            "relationship_attribute": identifier[0],
            "asset_role": role,
            "vml_element": local_name,
        }
        relationship = relationships.get(relationship_id)
        if relationship is None:
            reference["binding_status"] = "relationship_missing"
            unresolved.append(relationship_id)
        elif relationship.get("target_mode", "Internal").casefold() == "external":
            reference.update(relationship)
            reference["binding_status"] = "external_reference"
            external.append(relationship_id)
        else:
            media_part = relationship.get("media_part")
            reference.update(relationship)
            if media_part not in entries:
                reference["binding_status"] = "target_missing"
                unresolved.append(relationship_id)
            else:
                media = entries[media_part]
                reference.update({
                    "asset_sha256": hashlib.sha256(media).hexdigest(),
                    "binding_status": "resolved",
                })
                if (pixel_size := _image_pixel_size(media)) is not None:
                    reference["pixel_size"] = pixel_size
        references.append(reference)
    extension: dict[str, Any] = {"image_references": references, "assets": references} if references else {}
    if relationship_ids:
        extension["image_relationship_ids"] = relationship_ids
    if unresolved:
        extension["unresolved_image_relationship_ids"] = unresolved
    if external:
        extension["external_image_relationship_ids"] = external
    if references:
        resolved = [reference for reference in references if reference.get("binding_status") == "resolved"]
        extension.update(resolved[0] if resolved else references[0])
    return extension


def _extract_vml_shapes(
    builder: _Builder,
    entries: dict[str, bytes],
    root: ET.Element,
    *,
    part: str,
    parent_id: str | None = None,
    sheet: str | None = None,
    slide: int | None = None,
    order_base: int = 0,
    visibility: str = "visible",
    lifecycle: str = "active",
    host_metadata: dict[str, Any] | None = None,
) -> None:
    shapes = (node for node in root.iter() if node.tag in _VML_OBJECT_TAGS)
    for index, shape in enumerate(shapes, start=1):
        primitive = shape.tag.split("}")[-1]
        object_id = shape.get("id") or f"vml:{primitive}:{index}"
        cell_range, bbox, geometry, unparsed_geometry = _vml_geometry(shape)
        asset_extension = _vml_assets(entries, part, shape)
        texts, text_sources = _vml_texts(shape)
        locator = evidence_ir.Locator(
            part=part,
            slide=slide,
            sheet=sheet,
            cell_range=cell_range,
            object_id=object_id,
            bbox=bbox,
            extension={
                "coordinate_system": geometry["coordinate_system"],
                **(host_metadata or {}),
            },
        )
        builder.add_element(
            "vml_shape",
            locator,
            parent_id=parent_id,
            order=order_base + index,
            value="\n".join(texts) or None,
            visibility=visibility,
            lifecycle=lifecycle,
            extension={
                **geometry,
                "vml_primitive": primitive,
                "shape_type": shape.get("type"),
                "title": shape.get("title"),
                "text_sources": text_sources,
                **(host_metadata or {}),
                **asset_extension,
            },
        )
        _add_raster_reference_coverage(
            builder,
            locator,
            asset_extension,
            failed_kind="vml_image_asset",
            external_kind="vml_external_image_reference",
        )
        if unparsed_geometry:
            _add_unparsed_notice(
                builder,
                locator,
                detected_kind="vml_unparsed_geometry",
                reason_code="unsupported_vml_geometry",
                detail={"shape_id": object_id},
            )


def _connection(node: ET.Element | None) -> int | None:
    return _integer(node.get("id")) if node is not None else None


def _prst_geom(node: ET.Element, prefix: str, kind: str) -> str | None:
    """DrawingMLの`prstGeom/@prst`（プリセット図形種）をpptx/xlsx共通で取り出す。

    Mermaidフローチャート化（L9）のノード形状マッピングの入力。閉じた語彙として扱わず、
    未知の`prst`値もそのまま保持する（マッピング側でフォールバックする）。
    """
    if kind not in {"shape", "textbox", "connector"}:
        return None
    sp_pr = node.find(f"{prefix}spPr")
    if sp_pr is None:
        return None
    geom = sp_pr.find(f"{_A}prstGeom")
    return geom.get("prst") if geom is not None else None


def _xlsx_objects(builder: _Builder, entries: dict[str, bytes], legacy_ids: dict[str, str]) -> set[str]:
    content_parts = {"xl/workbook.xml"}
    content_parts.update(part for part in ("xl/sharedStrings.xml", "xl/styles.xml") if part in entries)
    sheet_ids = {
        element.locator.sheet: element.element_id
        for element in builder.ir.elements
        if element.type == "sheet" and element.locator.sheet
    }
    for sheet_name, sheet_part in _xlsx_sheet_parts(entries).items():
        content_parts.add(sheet_part)
        root = _xml(entries, sheet_part)
        if root is None:
            continue
        sheet_rels = _relationships(entries, sheet_part)
        content_parts.update(
            target for target in sheet_rels.values()
            if target in entries and target.startswith("xl/comments") and target.endswith(".xml")
        )
        for drawing_ref in root.findall(f"{_X}drawing"):
            drawing_part = sheet_rels.get(drawing_ref.get(f"{_R}id", ""))
            drawing = _xml(entries, drawing_part or "")
            if drawing is None or drawing_part is None:
                continue
            content_parts.add(drawing_part)
            drawing_rels = _relationships(entries, drawing_part)
            # L9: コネクタのrelation化はpptx（`_pptx_objects.walk`）と同じ二段パス——
            # drawing part全体（全anchor）を先に歩いてobject_idを集め、その後で解決する。
            pending_connectors: list[tuple[str, int | None, int | None]] = []
            ids_by_object: dict[int, str] = {}
            for z_order, anchor in enumerate(list(drawing), start=1):
                cell_range, anchor_ext = _xlsx_anchor(anchor)
                source_rect = _cell_rect(cell_range)

                def walk(nodes: list[ET.Element], parent_id: str | None, order_base: int) -> None:
                    """`xdr:grpSp`の子を個別要素として辿る（`_pptx_objects.walk`と同じ再帰形）。

                    グループの中身を1本の連結文字列へ潰さず、ノード名で検索できるようにする
                    （xlsxはpptxと違いグループの中を歩いていなかったため個別要素が出ていなかった）。
                    xlsxのdrawing座標系はanchor（from/to marker）単位でしかセル範囲を持たない
                    （pptxのEMU座標＋累積変換とは異なる）ため、group子孫の`cell_range`は親anchorの
                    ものをそのまま継承する（個々の子のオフセットまでは追わない＝図形種別
                    `prstGeom`の抽出と同じくMermaid化レーンのスコープ）。
                    """
                    local_index = 0
                    for node in nodes:
                        kind = _drawing_kind(node)
                        if kind is None:
                            continue
                        local_index += 1
                        order = z_order if order_base == 0 else order_base + local_index
                        props = _c_nv_pr(node, _XDR, kind)
                        object_id = _integer(props.get("id")) if props is not None else None
                        text = "".join(part.text or "" for part in node.findall(f".//{_A}t"))
                        extension: dict[str, Any] = {
                            **anchor_ext,
                            "z_order": z_order,
                            "name": props.get("name", "") if props is not None else "",
                            "description": props.get("descr", "") if props is not None else "",
                        }
                        if (prst := _prst_geom(node, _XDR, kind)) is not None:
                            extension["prst"] = prst
                        chart_part = None
                        chart_extension: dict[str, Any] = {}
                        smartart_value = None
                        smartart_extension: dict[str, Any] = {}
                        smartart_parts: set[str] = set()
                        if kind == "graphic_frame":
                            chart = node.find(f".//{_C}chart")
                            relationship_id = chart.get(f"{_R}id", "") if chart is not None else ""
                            candidate = drawing_rels.get(relationship_id)
                            if candidate in entries:
                                kind = "chart"
                                chart_part = candidate
                                content_parts.add(candidate)
                                _, chart_extension = _xlsx_chart_payload(entries, candidate)
                                extension.update({
                                    "relationship_id": relationship_id,
                                    "chart_part": candidate,
                                    **chart_extension,
                                })
                            else:
                                smartart_value, smartart_extension, smartart_parts = _smartart_payload(
                                    entries, drawing_part, node,
                                )
                                if smartart_extension.get("relationship_ids"):
                                    kind = "smartart"
                                    content_parts.update(smartart_parts)
                                    extension.update(smartart_extension)
                        asset_role = "picture_content" if kind == "picture" else "shape_fill"
                        extension.update(_drawing_asset(entries, drawing_part, node, role=asset_role))
                        if kind == "connector":
                            extension["start_object_id"] = _connection(node.find(f".//{_A}stCxn"))
                            extension["end_object_id"] = _connection(node.find(f".//{_A}endCxn"))
                        locator = evidence_ir.Locator(
                            part=drawing_part,
                            sheet=sheet_name,
                            cell_range=cell_range,
                            object_id=object_id,
                            extension={"coordinate_system": "xlsx-cell-anchor"},
                        )
                        element_id = builder.add_element(
                            kind,
                            locator,
                            parent_id=parent_id,
                            order=order,
                            value=text or None,
                            coverage_status="metadata_only" if kind == "picture" else "extracted",
                            coverage_reason="image_content_uninterpreted" if kind == "picture" else None,
                            extension=extension,
                        )
                        if object_id is not None:
                            ids_by_object[object_id] = element_id
                        if kind == "connector":
                            pending_connectors.append(
                                (element_id, extension.get("start_object_id"), extension.get("end_object_id")))
                        _add_raster_reference_coverage(
                            builder,
                            locator,
                            extension,
                            failed_kind="drawing_image_asset",
                            external_kind="drawing_external_image_reference",
                        )
                        if chart_part is not None:
                            _add_chart_data(
                                builder,
                                entries,
                                chart_part=chart_part,
                                parent_id=element_id,
                                object_token=str(object_id if object_id is not None else order),
                                sheet=sheet_name,
                            )
                        if kind == "smartart":
                            _add_smartart_data(
                                builder,
                                entries,
                                parent_id=element_id,
                                value=smartart_value,
                                extension=smartart_extension,
                                host_part=drawing_part,
                                object_token=str(object_id if object_id is not None else order),
                                sheet=sheet_name,
                            )
                        if source_rect is not None and order_base == 0:
                            # 状態は確定せず、drawingと抽出済みcell regionの幾何交差だけをrelationにする。
                            # **グループ子孫（order_base != 0）は対象外**: 子の cell_range は親 anchor の
                            # 継承（近似）であり、その矩形で overlaps/覆いを主張すると「グループ内の
                            # 小さな図形がグループ全域を覆う」という事実でない幾何を relation にしてしまう
                            # （「推測しない・幾何交差だけ」の契約に反する）。子要素は検索向けの個別
                            # 要素としてだけ出し、幾何関係はトップレベル（anchor 矩形が自分自身のもの）に限る。
                            same_sheet = [target for target in builder.ir.elements if target.locator.sheet == sheet_name]
                            exact_tables = [target for target in same_sheet
                                            if target.type == "table" and _cell_rect(target.locator.cell_range) == source_rect]
                            targets = exact_tables or [target for target in same_sheet if target.type == "cell"]
                            for target in targets:
                                target_rect = _cell_rect(target.locator.cell_range)
                                if target_rect is None:
                                    continue
                                intersection = _bbox_intersection(source_rect, target_rect)
                                if intersection is None:
                                    continue
                                source_area, target_area = _area(source_rect), _area(target_rect)
                                intersection_area = _area(intersection)
                                builder.add_relation(
                                    "overlaps",
                                    element_id,
                                    target.element_id,
                                    evidence_ids=[element_id, target.element_id],
                                    extension={
                                        "coordinate_system": "normalized-cell",
                                        "intersection_bbox": intersection,
                                        "intersection_area": intersection_area,
                                        "source_overlap_ratio": intersection_area / source_area if source_area else 0.0,
                                        "target_overlap_ratio": intersection_area / target_area if target_area else 0.0,
                                    },
                                )
                            # L3（可視性・廃止表現の全形式展開）: 不透明な前面図形/画像（塗りつぶし図形または画像）が
                            # セルを閾値以上覆うなら、そのセル要素を hidden へ差し替える（`_field_piece` は cell 自身の
                            # `visibility` を見るため table 要素ではなく cell 要素を対象にする）。意味の断定
                            # （「廃止」等）はしない＝前面図形の種別・テキスト・名前だけを `occluded_by` へ残す
                            # （意味の断定はしない共通思想）。
                            is_opaque_occluder = kind == "picture" or (kind == "shape" and _xlsx_shape_has_solid_fill(node))
                            if is_opaque_occluder:
                                occluded_by: dict[str, Any] = {"kind": kind, "element_id": element_id, "z_order": z_order}
                                if text:
                                    occluded_by["text"] = text
                                elif extension.get("name"):
                                    occluded_by["name"] = extension["name"]
                                for target in same_sheet:
                                    if target.type != "cell":
                                        continue
                                    target_rect = _cell_rect(target.locator.cell_range)
                                    if target_rect is None:
                                        continue
                                    intersection = _bbox_intersection(source_rect, target_rect)
                                    if intersection is None:
                                        continue
                                    target_area = _area(target_rect)
                                    ratio = _area(intersection) / target_area if target_area else 0.0
                                    if ratio >= _occlusion_ratio():
                                        builder.mark_hidden(
                                            target.element_id,
                                            reason="occluded_by_picture" if kind == "picture" else "occluded_by_shape",
                                            extra_extension={"occluded_by": occluded_by},
                                        )
                        if kind == "group":
                            child_nodes = [child for child in list(node) if _drawing_kind(child) is not None]
                            walk(child_nodes, element_id, order * 1000)

                walk(list(anchor), sheet_ids.get(sheet_name), 0)
            for connector_id, start_object_id, end_object_id in pending_connectors:
                start_id, end_id = ids_by_object.get(start_object_id), ids_by_object.get(end_object_id)
                if start_id and end_id:
                    builder.add_relation(
                        "connects_to",
                        start_id,
                        end_id,
                        evidence_ids=[start_id, connector_id, end_id],
                        extension={"connector_element_id": connector_id, "directed": True},
                    )
        for ole_order, ole in enumerate(root.findall(f".//{_X}oleObject"), start=1):
            relationship_id = ole.get(f"{_R}id", "")
            if not relationship_id:
                continue
            cell_range, anchor_extension = _xlsx_embedded_anchor(ole)
            extension = {
                **anchor_extension,
                **_ole_metadata(entries, sheet_part, relationship_id, ole),
                "shape_id": ole.get("shapeId"),
            }
            builder.add_element(
                "ole_object",
                evidence_ir.Locator(
                    part=sheet_part,
                    sheet=sheet_name,
                    cell_range=cell_range,
                    object_id=ole.get("shapeId") or relationship_id,
                    extension={"coordinate_system": "xlsx-cell-anchor"},
                ),
                parent_id=sheet_ids.get(sheet_name),
                order=100_000 + ole_order,
                value=None,
                coverage_status="metadata_only",
                coverage_reason="ole_binary_metadata_only",
                coverage_content_basis="binary_only",
                extension=extension,
            )
        for vml_part in sorted({
            target for target in sheet_rels.values()
            if target in entries and (target.lower().endswith(".vml") or "vmldrawing" in target.lower())
        }):
            vml_root = _xml(entries, vml_part)
            if vml_root is None:
                continue
            content_parts.add(vml_part)
            _extract_vml_shapes(
                builder,
                entries,
                vml_root,
                part=vml_part,
                parent_id=sheet_ids.get(sheet_name),
                sheet=sheet_name,
                order_base=200_000,
            )
    return content_parts


def _bbox_intersection(a: list[int | float], b: list[int | float]) -> list[int | float] | None:
    x0, y0, x1, y1 = max(a[0], b[0]), max(a[1], b[1]), min(a[2], b[2]), min(a[3], b[3])
    return [x0, y0, x1, y1] if x1 > x0 and y1 > y0 else None


def _docx_content_hosts(entries: dict[str, bytes]) -> list[_ContentHost]:
    """本文と補助本文partを、各part固有のrelationship境界を保って列挙する。"""
    hosts = [
        _ContentHost("word/document.xml", "document_body", "document_body"),
    ]
    kinds = {
        "word/comments.xml": ("comments", "annotation"),
        "word/footnotes.xml": ("footnotes", "reference_note"),
        "word/endnotes.xml": ("endnotes", "reference_note"),
    }
    for part in sorted(entries):
        if part.startswith("word/header") and part.endswith(".xml"):
            hosts.append(_ContentHost(part, "header", "page_repeat"))
        elif part.startswith("word/footer") and part.endswith(".xml"):
            hosts.append(_ContentHost(part, "footer", "page_repeat"))
        elif part in kinds:
            kind, host_visibility = kinds[part]
            hosts.append(_ContentHost(part, kind, host_visibility))
    return hosts


def _docx_host_objects(
    builder: _Builder,
    entries: dict[str, bytes],
    host: _ContentHost,
    *,
    order_base: int,
) -> set[str]:
    """1つのWord content host内のDrawingML/VML/OLEを、そのhost自身のrelsで解決する。"""
    content_parts = {host.part}
    root = _xml(entries, host.part)
    if root is None:
        return content_parts
    rels = _relationships(entries, host.part)
    # 本文の既存locator/element IDは不変にし、追加hostだけに明示的なhost metadataを刻む。
    host_metadata = {} if host.host_kind == "document_body" else host.metadata()
    positioned: list[tuple[str, list[int | float], int]] = []
    for order, anchor in enumerate(root.findall(f".//{_WP}anchor") + root.findall(f".//{_WP}inline"), start=1):
        props = anchor.find(f"{_WP}docPr")
        object_id = _integer(props.get("id")) if props is not None else None
        extent = anchor.find(f"{_WP}extent")
        x_node = anchor.find(f"{_WP}positionH/{_WP}posOffset")
        y_node = anchor.find(f"{_WP}positionV/{_WP}posOffset")
        try:
            cx, cy = int(extent.get("cx")), int(extent.get("cy"))
            extent_emu: list[int] | None = [cx, cy]
        except (AttributeError, TypeError, ValueError):
            extent_emu = None
        try:
            x, y = int(x_node.text or 0), int(y_node.text or 0)
            if extent_emu is None:
                raise ValueError
            cx, cy = extent_emu
            bbox: list[int | float] | None = [x, y, x + cx, y + cy]
        except (AttributeError, TypeError, ValueError):
            bbox = None
        is_picture = anchor.find(f".//{_PIC}pic") is not None
        chart = anchor.find(f".//{_C}chart")
        chart_relationship_id = chart.get(f"{_R}id", "") if chart is not None else ""
        chart_part = rels.get(chart_relationship_id)
        smartart_value, smartart_extension, smartart_parts = _smartart_payload(entries, host.part, anchor)
        text = "".join(node.text or "" for node in anchor.findall(f".//{_W}t"))
        kind = (
            "picture" if is_picture else
            "chart" if chart_part in entries else
            "smartart" if smartart_extension.get("relationship_ids") else
            "floating_textbox" if text else "floating_object"
        )
        extension: dict[str, Any] = {
            "anchor_type": anchor.tag.split("}")[-1],
            "name": props.get("name", "") if props is not None else "",
            "description": props.get("descr", "") if props is not None else "",
            "z_order": _integer(anchor.get("relativeHeight")),
            **host_metadata,
        }
        alpha = next((_integer(node.get("val")) for node in anchor.findall(f".//{_A}alpha")
                      if _integer(node.get("val")) is not None), None)
        if alpha is not None:
            extension["alpha"] = alpha
        asset_role = "picture_content" if is_picture else "shape_fill"
        extension.update(_drawing_asset(entries, host.part, anchor, role=asset_role))
        if chart_part in entries:
            content_parts.add(chart_part)
            _, chart_extension = _chart_payload(entries, chart_part)
            extension.update({
                "relationship_id": chart_relationship_id,
                "chart_part": chart_part,
                **chart_extension,
            })
        if kind == "smartart":
            content_parts.update(smartart_parts)
            extension.update(smartart_extension)
        locator = evidence_ir.Locator(
            part=host.part,
            object_id=object_id,
            bbox=bbox,
            extension={
                "coordinate_system": "word-page-emu",
                **({"host_object_index": order} if host_metadata else {}),
                **({"extent_emu": extent_emu} if host_metadata and extent_emu is not None else {}),
                **host_metadata,
            },
        )
        element_id = builder.add_element(
            kind,
            locator,
            parent_id=None,
            order=order_base + order,
            value=text or None,
            coverage_status="metadata_only" if is_picture else "extracted",
            coverage_reason="image_content_uninterpreted" if is_picture else None,
            visibility=host.visibility,
            lifecycle=host.lifecycle,
            extension=extension,
        )
        _add_raster_reference_coverage(
            builder,
            locator,
            extension,
            failed_kind="drawing_image_asset",
            external_kind="drawing_external_image_reference",
        )
        if chart_part in entries:
            _add_chart_data(
                builder,
                entries,
                chart_part=chart_part,
                parent_id=element_id,
                object_token=str(object_id if object_id is not None else order_base + order),
            )
        if kind == "smartart":
            _add_smartart_data(
                builder,
                entries,
                parent_id=element_id,
                value=smartart_value,
                extension=smartart_extension,
                host_part=host.part,
                object_token=str(object_id if object_id is not None else order_base + order),
            )
        if bbox is not None:
            positioned.append((element_id, bbox, extension.get("z_order") or order))

    _extract_vml_shapes(
        builder,
        entries,
        root,
        part=host.part,
        slide=host.slide,
        order_base=order_base + 100_000,
        visibility=host.visibility,
        lifecycle=host.lifecycle,
        host_metadata=host_metadata,
    )
    vml_shapes = {
        shape.get("id"): shape for shape in root.iter(f"{_V}shape") if shape.get("id")
    }
    for ole_order, ole in enumerate(root.iter(f"{_O}OLEObject"), start=1):
        relationship_id = ole.get(f"{_R}id", "")
        if not relationship_id:
            continue
        shape_id = ole.get("ShapeID") or ole.get("shapeId")
        host_shape = vml_shapes.get(shape_id)
        cell_range = None
        bbox = None
        geometry: dict[str, Any] = {}
        if host_shape is not None:
            cell_range, bbox, geometry, _unparsed = _vml_geometry(host_shape)
        builder.add_element(
            "ole_object",
            evidence_ir.Locator(
                part=host.part,
                cell_range=cell_range,
                object_id=ole.get("ObjectID") or shape_id or relationship_id,
                bbox=bbox,
                extension={
                    "coordinate_system": geometry.get("coordinate_system", "word-flow"),
                    **host_metadata,
                },
            ),
            parent_id=None,
            order=order_base + 200_000 + ole_order,
            value=None,
            coverage_status="metadata_only",
            coverage_reason="ole_binary_metadata_only",
            coverage_content_basis="binary_only",
            visibility=host.visibility,
            lifecycle=host.lifecycle,
            extension={
                "shape_id": shape_id,
                **host_metadata,
                **geometry,
                **_ole_metadata(entries, host.part, relationship_id, ole),
            },
        )

    # 幾何関係だけを確定し、「廃止」の適用や状態はSemantic viewまで保留する。
    for index, (left_id, left_bbox, left_z) in enumerate(positioned):
        for right_id, right_bbox, right_z in positioned[index + 1:]:
            intersection = _bbox_intersection(left_bbox, right_bbox)
            if intersection is None:
                continue
            source_id, target_id = (left_id, right_id) if left_z >= right_z else (right_id, left_id)
            builder.add_relation(
                "overlaps",
                source_id,
                target_id,
                evidence_ids=[source_id, target_id],
                extension={"intersection_bbox": intersection, "coordinate_system": "word-page-emu"},
            )
    return content_parts


def _external_vml_host_context(
    entries: dict[str, bytes],
    hosts: list[_ContentHost],
    vml_part: str,
) -> tuple[dict[str, Any], str, str, int | None]:
    owners = [host for host in hosts if vml_part in _relationships(entries, host.part).values()]
    if not owners:
        return (
            {"host_part": vml_part, "host_kind": "detached_vml", "host_visibility": "unknown", "host_lifecycle": "active"},
            "visible",
            "active",
            None,
        )
    primary = owners[0]
    return (
        {**primary.metadata(), "host_parts": [owner.part for owner in owners]},
        primary.visibility,
        primary.lifecycle,
        primary.slide,
    )


def _docx_objects(builder: _Builder, entries: dict[str, bytes]) -> set[str]:
    hosts = _docx_content_hosts(entries)
    content_parts: set[str] = set()
    main = hosts[0]
    content_parts.update(_docx_host_objects(builder, entries, main, order_base=0))

    # 既存本文objectの要素順を変えないため、外部VMLと補助hostは本文objectの後へ追加する。
    for vml_index, vml_part in enumerate(sorted(
        candidate for candidate in entries
        if candidate.startswith("word/") and (candidate.lower().endswith(".vml") or "vmldrawing" in candidate.lower())
    ), start=1):
        vml_root = _xml(entries, vml_part)
        if vml_root is None:
            continue
        content_parts.add(vml_part)
        metadata, visibility, lifecycle, slide = _external_vml_host_context(entries, hosts, vml_part)
        _extract_vml_shapes(
            builder,
            entries,
            vml_root,
            part=vml_part,
            slide=slide,
            order_base=300_000 + (vml_index - 1) * 10_000,
            visibility=visibility,
            lifecycle=lifecycle,
            host_metadata=metadata,
        )

    for host_index, host in enumerate(hosts[1:], start=1):
        content_parts.update(_docx_host_objects(
            builder,
            entries,
            host,
            order_base=1_000_000 * host_index,
        ))
    return content_parts


def _ppt_props(node: ET.Element, kind: str) -> ET.Element | None:
    return _c_nv_pr(node, _P, kind)


def _ppt_kind(node: ET.Element) -> str | None:
    local = node.tag.split("}")[-1]
    return {"sp": "shape", "pic": "picture", "grpSp": "group", "cxnSp": "connector",
            "graphicFrame": "graphic_frame"}.get(local)


def _raw_xfrm(node: ET.Element, kind: str) -> tuple[float, float, float, float] | None:
    if kind == "group":
        xfrm = node.find(f"{_P}grpSpPr/{_A}xfrm")
    elif kind in {"graphic_frame", "table"}:
        xfrm = node.find(f"{_P}xfrm")
    else:
        props_name = "spPr"
        xfrm = node.find(f"{_P}{props_name}/{_A}xfrm")
    if xfrm is None:
        return None
    off, ext = xfrm.find(f"{_A}off"), xfrm.find(f"{_A}ext")
    try:
        return float(off.get("x")), float(off.get("y")), float(ext.get("cx")), float(ext.get("cy"))
    except (AttributeError, TypeError, ValueError):
        return None


def _transform_bbox(
    raw: tuple[float, float, float, float] | None,
    transform: tuple[float, float, float, float],
) -> list[int | float] | None:
    if raw is None:
        return None
    sx, sy, tx, ty = transform
    x, y, cx, cy = raw
    values = [sx * x + tx, sy * y + ty, sx * (x + cx) + tx, sy * (y + cy) + ty]
    return [int(value) if value.is_integer() else value for value in values]


def _child_transform(
    node: ET.Element,
    parent: tuple[float, float, float, float],
) -> tuple[float, float, float, float]:
    xfrm = node.find(f"{_P}grpSpPr/{_A}xfrm")
    if xfrm is None:
        return parent
    off, ext, child_off, child_ext = (
        xfrm.find(f"{_A}off"),
        xfrm.find(f"{_A}ext"),
        xfrm.find(f"{_A}chOff"),
        xfrm.find(f"{_A}chExt"),
    )
    try:
        sx0, sy0, tx0, ty0 = parent
        scale_x = float(ext.get("cx")) / float(child_ext.get("cx"))
        scale_y = float(ext.get("cy")) / float(child_ext.get("cy"))
        offset_x = float(off.get("x")) - float(child_off.get("x")) * scale_x
        offset_y = float(off.get("y")) - float(child_off.get("y")) * scale_y
        return sx0 * scale_x, sy0 * scale_y, sx0 * offset_x + tx0, sy0 * offset_y + ty0
    except (AttributeError, TypeError, ValueError, ZeroDivisionError):
        return parent


def _ppt_text(node: ET.Element) -> str:
    """DrawingMLの段落境界を保って文字列化する。run境界は連結し、段落間だけLFにする。"""
    paragraphs = [
        "".join(part.text or "" for part in paragraph.findall(f".//{_A}t"))
        for paragraph in node.findall(f".//{_A}p")
    ]
    return "\n".join(text for text in paragraphs if text)


def _ppt_alpha(node: ET.Element) -> int | None:
    return next(
        (_integer(alpha.get("val")) for alpha in node.findall(f".//{_A}alpha")
         if _integer(alpha.get("val")) is not None),
        None,
    )


def _ppt_table_cells(
    builder: _Builder,
    table_node: ET.Element,
    *,
    table_id: str,
    table_locator: evidence_ir.Locator,
    visibility: str,
    coverage_id: str,
) -> None:
    """PPTXの``a:tbl``をrow/column/span付きcellへ展開する。

    ``hMerge``/``vMerge`` continuationは値を持つ独立cellにせず、先頭cellのspanで表す。壊れた表で
    continuationに文字が入っている場合だけ、欠落防止のため通常cellとして残す。
    """
    table = table_node.find(f".//{_A}tbl")
    if table is None:
        return
    order = 0
    for row_index, row in enumerate(table.findall(f"{_A}tr"), start=1):
        column = 1
        for cell in row.findall(f"{_A}tc"):
            text = _ppt_text(cell)
            horizontal_merge = cell.get("hMerge") in {"1", "true"}
            vertical_merge = cell.get("vMerge") in {"1", "true"}
            try:
                column_span = max(1, int(cell.get("gridSpan", "1")))
            except ValueError:
                column_span = 1
            try:
                row_span = max(1, int(cell.get("rowSpan", "1")))
            except ValueError:
                row_span = 1
            if (horizontal_merge or vertical_merge) and not text:
                # span起点が既に消費した論理列なので、continuation自体ではcolumnを進めない。
                continue
            order += 1
            cell_locator = evidence_ir.Locator(
                part=table_locator.part,
                slide=table_locator.slide,
                object_id=f"{table_locator.object_id}/cell:{row_index}:{column}",
                extension={
                    "row": row_index,
                    "column": column,
                    "row_span": row_span,
                    "column_span": column_span,
                    "coordinate_system": "pptx-table-grid",
                },
            )
            builder.add_element(
                "cell",
                cell_locator,
                parent_id=table_id,
                order=order,
                value=text,
                coverage_id=coverage_id,
                visibility=visibility,
                extension={"role": "unknown"},
            )
            column += column_span


def _ppt_template_elements(builder: _Builder, entries: dict[str, bytes]) -> set[str]:
    """slide/master/layout/notes masterを検査済みpartにし、実在するtemplate文字も保持する。"""
    prefixes = ("ppt/slideLayouts/", "ppt/slideMasters/", "ppt/notesMasters/")
    content_parts: set[str] = set()
    order = 0
    for part in sorted(
        candidate for candidate in entries
        if candidate.startswith(prefixes) and candidate.endswith(".xml")
    ):
        root = _xml(entries, part)
        if root is None:
            continue
        content_parts.add(part)
        texts = list(dict.fromkeys(
            node.text for node in root.iter(f"{_A}t") if node.text and node.text.strip()
        ))
        for text_index, text in enumerate(texts, start=1):
            order += 1
            builder.add_element(
                "presentation_template_text",
                evidence_ir.Locator(part=part, object_id=f"text:{text_index}"),
                parent_id=None,
                order=order,
                value=text,
                extension={"template_part": part},
            )
    return content_parts


def _ppt_auxiliary_hosts(
    entries: dict[str, bytes],
    slide_parts: list[str],
    hidden_slides: set[int],
) -> list[_ContentHost]:
    """master/layout/notes系partを、対応slideと用途を失わずcontent host化する。"""
    notes_to_slide: dict[str, int] = {}
    for slide_number, slide_part in enumerate(slide_parts, start=1):
        for target in _relationships(entries, slide_part).values():
            if target.startswith("ppt/notesSlides/notesSlide") and target.endswith(".xml"):
                notes_to_slide.setdefault(target, slide_number)

    hosts: list[_ContentHost] = []
    prefixes = (
        ("ppt/slideMasters/slideMaster", "slide_master", "template_inherited", "template_definition"),
        ("ppt/slideLayouts/slideLayout", "slide_layout", "template_inherited", "template_definition"),
        ("ppt/notesMasters/notesMaster", "notes_master", "notes_template", "template_definition"),
        ("ppt/notesSlides/notesSlide", "notes_slide", "notes_only", "active"),
    )
    for part in sorted(entries):
        if not part.endswith(".xml"):
            continue
        descriptor = next((item for item in prefixes if part.startswith(item[0])), None)
        if descriptor is None:
            continue
        _prefix, host_kind, host_visibility, host_lifecycle = descriptor
        slide = notes_to_slide.get(part) if host_kind == "notes_slide" else None
        visibility = "hidden" if slide in hidden_slides else "visible"
        hosts.append(_ContentHost(
            part,
            host_kind,
            host_visibility,
            host_lifecycle=host_lifecycle,
            slide=slide,
            visibility=visibility,
        ))
    return hosts


def _ppt_auxiliary_rasters(
    builder: _Builder,
    entries: dict[str, bytes],
    host: _ContentHost,
    *,
    order_base: int,
) -> set[str]:
    """非slide hostのpicture/画像fillをhost固有relsでEvidenceへ運ぶ。"""
    root = _xml(entries, host.part)
    if root is None:
        return {host.part}
    sp_tree = root.find(f"{_P}cSld/{_P}spTree")
    if sp_tree is None:
        return {host.part}
    host_metadata = host.metadata()
    sequence = 0

    def walk(
        nodes: list[ET.Element],
        transform=(1.0, 1.0, 0.0, 0.0),
        paint_base: int = 0,
    ) -> None:
        nonlocal sequence
        local_order = 0
        for node in nodes:
            raw_kind = _ppt_kind(node)
            if raw_kind is None:
                continue
            local_order += 1
            paint_order = local_order * 1_000_000 if paint_base == 0 else paint_base + local_order
            if raw_kind == "group":
                child_nodes = [child for child in list(node) if _ppt_kind(child) is not None]
                walk(child_nodes, _child_transform(node, transform), paint_order)
                continue
            if raw_kind not in {"picture", "shape"}:
                continue
            has_blip = node.find(f".//{_A}blip") is not None
            if raw_kind == "shape" and not has_blip:
                continue
            sequence += 1
            props = _ppt_props(node, raw_kind)
            object_id = _integer(props.get("id")) if props is not None else None
            asset_role = "picture_content" if raw_kind == "picture" else "shape_fill"
            asset_extension = _drawing_asset(entries, host.part, node, role=asset_role)
            if has_blip and not asset_extension:
                # unresolved shape fillもrouterでfailed_bindingへ搬送できるよう役割だけは保持する。
                asset_extension = {"asset_role": asset_role}
            extension: dict[str, Any] = {
                "name": props.get("name", "") if props is not None else "",
                "description": props.get("descr", "") if props is not None else "",
                "z_order": local_order,
                "paint_order": paint_order,
                **host_metadata,
                **asset_extension,
            }
            if (alpha := _ppt_alpha(node)) is not None:
                extension["alpha"] = alpha
            locator = evidence_ir.Locator(
                part=host.part,
                slide=host.slide,
                object_id=object_id,
                bbox=_transform_bbox(_raw_xfrm(node, raw_kind), transform),
                extension={
                    "coordinate_system": "pptx-slide-emu",
                    "host_object_index": sequence,
                    **host_metadata,
                },
            )
            builder.add_element(
                raw_kind,
                locator,
                parent_id=None,
                order=order_base + sequence,
                # template/notes文字は既存template/Document IR要素が保持する。画像要素へ重複転記しない。
                value=None,
                coverage_status="metadata_only" if raw_kind == "picture" else "extracted",
                coverage_reason="image_content_uninterpreted" if raw_kind == "picture" else None,
                visibility=host.visibility,
                lifecycle=host.lifecycle,
                extension=extension,
            )
            _add_raster_reference_coverage(
                builder,
                locator,
                extension,
                failed_kind="drawing_image_asset",
                external_kind="drawing_external_image_reference",
            )

    walk(list(sp_tree))
    _extract_vml_shapes(
        builder,
        entries,
        root,
        part=host.part,
        slide=host.slide,
        order_base=order_base + 100_000,
        visibility=host.visibility,
        lifecycle=host.lifecycle,
        host_metadata=host_metadata,
    )
    return {host.part}


def _pptx_objects(builder: _Builder, entries: dict[str, bytes]) -> set[str]:
    content_parts = {"ppt/presentation.xml"}
    content_parts.update(_ppt_template_elements(builder, entries))
    content_parts.update(
        part for part in entries if part.startswith("ppt/notesSlides/notesSlide") and part.endswith(".xml")
    )
    # presentation.xmlのsldIdLstとrelationshipを辿り、ファイル名順ではなく表示順を使う。
    presentation = _xml(entries, "ppt/presentation.xml")
    rels = _relationships(entries, "ppt/presentation.xml")
    slide_parts: list[str] = []
    if presentation is not None:
        for slide_id in presentation.findall(f"{_P}sldIdLst/{_P}sldId"):
            part = rels.get(slide_id.get(f"{_R}id", ""))
            if part:
                slide_parts.append(part)
    if not slide_parts:
        slide_parts = sorted(part for part in entries if part.startswith("ppt/slides/slide") and part.endswith(".xml"))

    slide_ids = {
        element.locator.slide: element.element_id
        for element in builder.ir.elements
        if element.type == "slide" and element.locator.slide is not None
    }
    hidden_slides = {
        element.locator.slide
        for element in builder.ir.elements
        if element.type == "slide" and element.visibility == "hidden" and element.locator.slide is not None
    }
    auxiliary_hosts = _ppt_auxiliary_hosts(entries, slide_parts, hidden_slides)
    drawing_hosts = [
        _ContentHost(
            part,
            "slide",
            "slide_canvas",
            slide=slide_number,
            visibility="hidden" if slide_number in hidden_slides else "visible",
        )
        for slide_number, part in enumerate(slide_parts, start=1)
    ] + auxiliary_hosts
    adapted_by_slot: dict[tuple[int, int], evidence_ir.EvidenceElement] = {}
    adapted_raw: list[tuple[evidence_ir.EvidenceElement, int, int]] = []
    for element in builder.ir.elements:
        source_map = element.locator.extension.get("document_ir_source_map", {})
        slide = element.locator.slide
        z_index = source_map.get("z_index") if isinstance(source_map, dict) else None
        if (element.extension.get("origin") == "document-ir-v2-adapter"
                and element.type == "shape" and isinstance(slide, int) and isinstance(z_index, int)):
            adapted_raw.append((element, slide, z_index))
    first_z = {
        slide: min(z_index for _element, candidate_slide, z_index in adapted_raw if candidate_slide == slide)
        for slide in {slide for _element, slide, _z_index in adapted_raw}
    }
    for element, slide, z_index in adapted_raw:
        adapted_by_slot[(slide, z_index - first_z[slide] + 1)] = element
    for slide_number, part in enumerate(slide_parts, start=1):
        content_parts.add(part)
        root = _xml(entries, part)
        if root is None:
            continue
        sp_tree = root.find(f"{_P}cSld/{_P}spTree")
        if sp_tree is None:
            continue
        slide_rels = _relationships(entries, part)
        pending_connectors: list[tuple[str, int | None, int | None]] = []
        ids_by_object: dict[int, str] = {}
        positioned: list[tuple[str, list[int | float], int]] = []

        def walk(
            nodes: list[ET.Element],
            parent_id: str | None,
            transform=(1.0, 1.0, 0.0, 0.0),
            paint_base: int = 0,
        ) -> None:
            local_order = 0
            for node in nodes:
                raw_kind = _ppt_kind(node)
                if raw_kind is None:
                    continue
                local_order += 1
                frame_kind = "table" if raw_kind == "graphic_frame" and node.find(f".//{_A}tbl") is not None else raw_kind
                kind = frame_kind
                chart_part = None
                chart_relationship_id = ""
                smartart_value = None
                smartart_extension: dict[str, Any] = {}
                smartart_parts: set[str] = set()
                ole_node = None
                ole_relationship_id = ""
                if raw_kind == "graphic_frame" and frame_kind != "table":
                    chart = node.find(f".//{_C}chart")
                    chart_relationship_id = chart.get(f"{_R}id", "") if chart is not None else ""
                    chart_part = slide_rels.get(chart_relationship_id)
                    if chart_part in entries:
                        kind = "chart"
                    else:
                        smartart_value, smartart_extension, smartart_parts = _smartart_payload(entries, part, node)
                        if smartart_extension.get("relationship_ids"):
                            kind = "smartart"
                        else:
                            ole_node = node.find(f".//{_P}oleObj")
                            ole_relationship_id = ole_node.get(f"{_R}id", "") if ole_node is not None else ""
                            if ole_relationship_id and slide_rels.get(ole_relationship_id) in entries:
                                kind = "ole_object"
                props = (
                    node.find(f"{_P}nvGraphicFramePr/{_P}cNvPr")
                    if raw_kind == "graphic_frame"
                    else _ppt_props(node, raw_kind)
                )
                object_id = _integer(props.get("id")) if props is not None else None
                bbox = _transform_bbox(_raw_xfrm(node, frame_kind), transform)
                text = _ppt_text(node)
                # top-levelのzを大きい桁、group子をその内側の小さい桁に置き、次のtop-levelより前に保つ。
                paint_order = local_order * 1_000_000 if paint_base == 0 else paint_base + local_order
                extension: dict[str, Any] = {
                    "name": props.get("name", "") if props is not None else "",
                    "description": props.get("descr", "") if props is not None else "",
                    "z_order": local_order,
                    "paint_order": paint_order,
                }
                if (prst := _prst_geom(node, _P, kind)) is not None:
                    extension["prst"] = prst
                if (alpha := _ppt_alpha(node)) is not None:
                    extension["alpha"] = alpha
                asset_role = (
                    "picture_content" if kind == "picture" else
                    "ole_preview" if kind == "ole_object" else
                    "shape_fill"
                )
                extension.update(_drawing_asset(entries, part, node, role=asset_role))
                if chart_part in entries:
                    content_parts.add(chart_part)
                    _, chart_extension = _chart_payload(entries, chart_part)
                    extension.update({
                        "relationship_id": chart_relationship_id,
                        "chart_part": chart_part,
                        **chart_extension,
                    })
                if kind == "smartart":
                    content_parts.update(smartart_parts)
                    extension.update(smartart_extension)
                if kind == "ole_object" and ole_node is not None:
                    extension.update(_ole_metadata(entries, part, ole_relationship_id, ole_node))
                start_id = _connection(node.find(f".//{_A}stCxn")) if kind == "connector" else None
                end_id = _connection(node.find(f".//{_A}endCxn")) if kind == "connector" else None
                if kind == "connector":
                    extension.update({"start_object_id": start_id, "end_object_id": end_id})
                adapted = adapted_by_slot.get((slide_number, local_order)) if parent_id is None else None
                visibility = "hidden" if slide_number in hidden_slides else (
                    adapted.visibility if adapted is not None else "visible"
                )
                reason = (
                    "hidden_slide_inherited" if slide_number in hidden_slides else
                    adapted.extension.get("visibility_reason") if adapted is not None else None
                )
                if reason:
                    extension["visibility_reason"] = reason
                # L3: covered_by_text（前面テキストによる重ね）もadapted要素から引き継ぐ（visibility_reasonと
                # 同じ経路・二重に判定し直さない）。
                if adapted is not None and adapted.extension.get("covered_by_text") is not None:
                    extension["covered_by_text"] = adapted.extension["covered_by_text"]
                locator = evidence_ir.Locator(
                    part=part,
                    slide=slide_number,
                    object_id=object_id,
                    bbox=bbox,
                    extension={"coordinate_system": "pptx-slide-emu"},
                )
                element_id = builder.add_element(
                    kind,
                    locator,
                    parent_id=parent_id or slide_ids.get(slide_number),
                    order=local_order,
                    value=None if kind in {"table", "ole_object"} else (text or None),
                    coverage_status="metadata_only" if kind in {"picture", "ole_object"} else "extracted",
                    coverage_reason=(
                        "image_content_uninterpreted" if kind == "picture" else
                        "ole_binary_metadata_only" if kind == "ole_object" else None
                    ),
                    coverage_content_basis="binary_only" if kind == "ole_object" else None,
                    visibility=visibility,
                    extension=extension,
                )
                _add_raster_reference_coverage(
                    builder,
                    locator,
                    extension,
                    failed_kind="drawing_image_asset",
                    external_kind="drawing_external_image_reference",
                )
                if object_id is not None:
                    ids_by_object[object_id] = element_id
                if bbox is not None:
                    positioned.append((element_id, bbox, paint_order))
                if kind == "table":
                    _ppt_table_cells(
                        builder,
                        node,
                        table_id=element_id,
                        table_locator=locator,
                        visibility=visibility,
                        coverage_id=builder.coverage_for(element_id),
                    )
                if chart_part in entries:
                    _add_chart_data(
                        builder,
                        entries,
                        chart_part=chart_part,
                        parent_id=element_id,
                        object_token=str(object_id if object_id is not None else local_order),
                        slide=slide_number,
                    )
                if kind == "smartart":
                    _add_smartart_data(
                        builder,
                        entries,
                        parent_id=element_id,
                        value=smartart_value,
                        extension=smartart_extension,
                        host_part=part,
                        object_token=str(object_id if object_id is not None else local_order),
                        slide=slide_number,
                    )
                if kind == "connector":
                    pending_connectors.append((element_id, start_id, end_id))
                if kind == "group":
                    child_nodes = [child for child in list(node) if _ppt_kind(child) is not None]
                    walk(child_nodes, element_id, _child_transform(node, transform), paint_order)

        walk(list(sp_tree), None)
        for connector_id, start_object_id, end_object_id in pending_connectors:
            start_id, end_id = ids_by_object.get(start_object_id), ids_by_object.get(end_object_id)
            if start_id and end_id:
                builder.add_relation(
                    "connects_to",
                    start_id,
                    end_id,
                    evidence_ids=[start_id, connector_id, end_id],
                    extension={"connector_element_id": connector_id, "directed": True},
                )
        elements_by_id = {element.element_id: element for element in builder.ir.elements}
        for index, (left_id, left_bbox, left_z) in enumerate(positioned):
            for right_id, right_bbox, right_z in positioned[index + 1:]:
                left_element, right_element = elements_by_id.get(left_id), elements_by_id.get(right_id)
                if left_element is None or right_element is None:
                    continue
                # groupは子の包含bboxであり、業務上のoverlayではない。子との重複relationを水増ししない。
                if left_element.type == "group" or right_element.type == "group":
                    continue
                intersection = _bbox_intersection(left_bbox, right_bbox)
                if intersection is None:
                    continue
                source_id, target_id = (left_id, right_id) if left_z >= right_z else (right_id, left_id)
                builder.add_relation(
                    "overlaps",
                    source_id,
                    target_id,
                    evidence_ids=[source_id, target_id],
                    extension={"intersection_bbox": intersection, "coordinate_system": "pptx-slide-emu"},
                )
    for vml_index, vml_part in enumerate(sorted(
        candidate for candidate in entries
        if candidate.startswith("ppt/") and (candidate.lower().endswith(".vml") or "vmldrawing" in candidate.lower())
    ), start=1):
        vml_root = _xml(entries, vml_part)
        if vml_root is None:
            continue
        content_parts.add(vml_part)
        metadata, visibility, lifecycle, slide = _external_vml_host_context(entries, drawing_hosts, vml_part)
        _extract_vml_shapes(
            builder,
            entries,
            vml_root,
            part=vml_part,
            slide=slide,
            order_base=300_000 + (vml_index - 1) * 10_000,
            visibility=visibility,
            lifecycle=lifecycle,
            host_metadata=metadata,
        )
    for host_index, host in enumerate(auxiliary_hosts, start=1):
        content_parts.update(_ppt_auxiliary_rasters(
            builder,
            entries,
            host,
            order_base=10_000_000 + host_index * 1_000_000,
        ))
    return content_parts


def _matrix_multiply(left: list[float], right: list[float]) -> list[float]:
    a, b, c, d, e, f = left
    g, h, i, j, k, l = right
    return [a * g + c * h, b * g + d * h, a * i + c * j, b * i + d * j,
            a * k + c * l + e, b * k + d * l + f]


def _matrix_bbox(matrix: list[float]) -> list[float]:
    a, b, c, d, e, f = matrix
    points = [(e, f), (a + e, b + f), (c + e, d + f), (a + c + e, b + d + f)]
    return [min(p[0] for p in points), min(p[1] for p in points),
            max(p[0] for p in points), max(p[1] for p in points)]


def _matrix_rect_bbox(matrix: list[float], rect: list[float]) -> list[float]:
    """任意のPDF矩形をmatrixでuser spaceへ写したaxis-aligned bboxを返す。"""
    left, bottom, right, top = rect
    points = [
        _pdf_point(matrix, left, bottom),
        _pdf_point(matrix, right, bottom),
        _pdf_point(matrix, right, top),
        _pdf_point(matrix, left, top),
    ]
    return [
        min(point[0] for point in points),
        min(point[1] for point in points),
        max(point[0] for point in points),
        max(point[1] for point in points),
    ]


def _pdf_point(matrix: list[float], x: float, y: float) -> tuple[float, float]:
    a, b, c, d, e, f = matrix
    return a * x + c * y + e, b * x + d * y + f


def _pdf_image_payload(xobject: Any, object_name: str) -> tuple[str, bytes, Any]:
    """既に安全に解決したImage XObjectをpypdfの固定decoderで通常画像bytesへ変換する。"""
    from pypdf.generic._image_xobject import _xobj_to_image

    extension, image_bytes, image = _xobj_to_image(xobject)
    return f"{object_name.lstrip('/')}{extension}", image_bytes, image


def _pdf_xobject_at_path(page: Any, xobject_path: list[str]) -> Any:
    """Evidenceへ固定したresource pathだけを辿る。page全XObjectの再列挙は行わない。"""

    def dereference(value: Any) -> Any:
        return value.get_object() if hasattr(value, "get_object") else value

    resources = dereference(page.get("/Resources"))
    current = None
    for index, raw_name in enumerate(xobject_path):
        pdf_name = "/" + str(raw_name).lstrip("/")
        xobjects = dereference(resources.get("/XObject")) if resources is not None else None
        reference = xobjects.get(pdf_name) if xobjects is not None else None
        if reference is None:
            raise KeyError(pdf_name)
        current = dereference(reference)
        if index < len(xobject_path) - 1:
            nested_resources = current.get("/Resources")
            if nested_resources is not None:
                resources = dereference(nested_resources)
    if current is None or str(current.get("/Subtype", "")) != "/Image":
        raise KeyError("resource path does not end at an Image XObject")
    return current


def _pdf_xobject_specs(
    builder: _Builder,
    *,
    reader: Any,
    page_number: int,
    page_bbox: list[float],
    operations: list[tuple[list[Any], bytes]],
    resources: Any,
) -> list[dict[str, Any]]:
    """page/Formの``Do``を再帰し、実際に描画されたImage XObjectだけを列挙する。

    ``page.images``の平坦なfile nameは、異なるForm内で同名の画像を区別できず、Form自体を画像と
    誤認する。そこでresource pathを保持したままSubtypeを判定し、Form Matrixと各content streamの
    graphics stateを合成する。深さ・循環・解決件数の境界はCoverageへ明示し、黙って打ち切らない。
    """
    from pypdf.generic import ContentStream

    identity = [1.0, 0.0, 0.0, 1.0, 0.0, 0.0]
    image_specs: list[dict[str, Any]] = []
    resolution_count = 0

    def dereference(value: Any) -> Any:
        return value.get_object() if hasattr(value, "get_object") else value

    def path_locator(
        path: tuple[str, ...],
        operator_path: tuple[int, ...],
        *,
        bbox: list[float] | None = None,
        suffix: str = "",
    ) -> evidence_ir.Locator:
        clean_path = [name.lstrip("/") for name in path]
        joined = "/".join(clean_path)
        return evidence_ir.Locator(
            part=f"page:{page_number}/xobject:{joined}{suffix}",
            page=page_number,
            object_id=f"{joined}{suffix}",
            bbox=bbox,
            extension={
                "coordinate_system": "pdf-user-space",
                "xobject_path": clean_path,
                "operator_path": list(operator_path),
            },
        )

    def add_issue(
        path: tuple[str, ...],
        operator_path: tuple[int, ...],
        *,
        status: str,
        detected_kind: str,
        content_basis: str,
        reason_code: str,
        bbox: list[float] | None,
        detail: dict[str, Any] | None = None,
        suffix: str = "",
    ) -> None:
        builder.add_coverage(
            "object",
            path_locator(path, operator_path, bbox=bbox, suffix=suffix),
            status,
            detected_kind=detected_kind,
            content_basis=content_basis,
            reason_code=reason_code,
            parser_id=PDF_ADAPTER_VERSION,
            detail={
                "xobject_path": [name.lstrip("/") for name in path],
                "operator_path": list(operator_path),
                **(detail or {}),
            },
        )

    def object_token(reference: Any, resolved: Any) -> tuple[Any, ...]:
        indirect = reference if hasattr(reference, "idnum") else getattr(resolved, "indirect_reference", None)
        if indirect is not None and hasattr(indirect, "idnum"):
            return ("indirect", int(indirect.idnum), int(getattr(indirect, "generation", 0)))
        return ("direct", id(resolved))

    def walk(
        current_operations: list[tuple[list[Any], bytes]],
        current_resources: Any,
        *,
        initial_ctm: list[float],
        path_prefix: tuple[str, ...],
        operator_prefix: tuple[int, ...],
        active_forms: frozenset[tuple[Any, ...]],
        form_depth: int,
        clip_bbox: list[float] | None,
    ) -> None:
        nonlocal resolution_count
        stack: list[list[float]] = []
        ctm = list(initial_ctm)
        for operator_index, (operands, operator) in enumerate(current_operations):
            if operator == b"q":
                stack.append(list(ctm))
                continue
            if operator == b"Q":
                ctm = stack.pop() if stack else list(initial_ctm)
                continue
            if operator == b"cm" and len(operands) == 6:
                try:
                    ctm = _matrix_multiply(ctm, [float(value) for value in operands])
                except (TypeError, ValueError):
                    pass
                continue
            if operator != b"Do" or not operands:
                continue

            pdf_name = str(operands[0])
            if not pdf_name.startswith("/"):
                pdf_name = "/" + pdf_name
            path = (*path_prefix, pdf_name)
            operator_path = (*operator_prefix, operator_index)
            reference_bbox = _matrix_bbox(ctm)
            if resolution_count >= PDF_XOBJECT_MAX_RESOLUTIONS:
                add_issue(
                    path,
                    operator_path,
                    status="unsupported",
                    detected_kind="pdf_xobject_reference",
                    # 上限によりSubtype未解決でも、Do参照は表示内容を指す。
                    # ``none``にするとrendererがnoticeを出さず、内容が静かに落ちる。
                    content_basis="binary_only",
                    reason_code="xobject_resource_limit_exceeded",
                    bbox=reference_bbox,
                    detail={"max_resolutions": PDF_XOBJECT_MAX_RESOLUTIONS},
                )
                continue
            resolution_count += 1

            try:
                resource_dict = dereference(current_resources)
                xobjects = dereference(resource_dict.get("/XObject")) if resource_dict is not None else None
                reference = xobjects.get(pdf_name) if xobjects is not None else None
            except Exception as exc:
                add_issue(
                    path,
                    operator_path,
                    status="failed",
                    detected_kind="pdf_xobject_reference",
                    # Do演算子は表示内容への参照であり、resource辞書を解決できなくても
                    # content-bearing である。none にするとcoverage noticeが描画されず
                    # 未解決の表示内容がsilent-dropする。
                    content_basis="binary_only",
                    reason_code="xobject_resource_resolution_failed",
                    bbox=reference_bbox,
                    detail={"error_type": exc.__class__.__name__},
                )
                continue
            if reference is None:
                add_issue(
                    path,
                    operator_path,
                    status="unsupported",
                    detected_kind="pdf_xobject_reference",
                    content_basis="binary_only",
                    reason_code="xobject_resource_missing",
                    bbox=reference_bbox,
                )
                continue
            try:
                xobject = dereference(reference)
                subtype = str(xobject.get("/Subtype", ""))
            except Exception as exc:
                add_issue(
                    path,
                    operator_path,
                    status="failed",
                    detected_kind="pdf_xobject",
                    content_basis="binary_only",
                    reason_code="xobject_parse_failed",
                    bbox=reference_bbox,
                    detail={"error_type": exc.__class__.__name__},
                )
                continue

            if subtype == "/Image":
                try:
                    image_name, image_bytes, decoded_image = _pdf_image_payload(xobject, path[-1])
                    pixel_size = list(decoded_image.size)
                    source_stream = xobject.get_data()
                except Exception as exc:
                    add_issue(
                        path,
                        operator_path,
                        status="failed",
                        detected_kind="image_xobject",
                        content_basis="pixel_only",
                        reason_code="image_decode_failed",
                        bbox=reference_bbox,
                        detail={"error_type": exc.__class__.__name__},
                    )
                    continue
                asset_extension = Path(image_name).suffix.lower() or ".bin"
                raw_filters = xobject.get("/Filter")
                source_filters = (
                    [str(item) for item in raw_filters]
                    if isinstance(raw_filters, (list, tuple))
                    else ([str(raw_filters)] if raw_filters is not None else [])
                )
                visible_bbox = _bbox_intersection(reference_bbox, clip_bbox) if clip_bbox is not None else reference_bbox
                image_specs.append({
                    "name": "/".join(name.lstrip("/") for name in path),
                    "bbox": reference_bbox,
                    "effective_bbox": visible_bbox,
                    "extension": {
                        "operator_index": operator_path[0],
                        "operator_path": list(operator_path),
                        "matrix": list(ctm),
                        "xobject_path": [name.lstrip("/") for name in path],
                        "asset_sha256": hashlib.sha256(image_bytes).hexdigest(),
                        "source_stream_sha256": hashlib.sha256(source_stream).hexdigest(),
                        "source_stream_basis": "pypdf_decoded_stream",
                        "source_filters": source_filters,
                        "pixel_size": pixel_size,
                        "image_name": image_name,
                        "asset_extension": asset_extension,
                        "media_type": mimetypes.guess_type(image_name)[0] or "application/octet-stream",
                        **({"visible_bbox": visible_bbox} if visible_bbox != reference_bbox else {}),
                    },
                })
                continue

            if subtype != "/Form":
                add_issue(
                    path,
                    operator_path,
                    status="unsupported",
                    detected_kind="pdf_xobject",
                    content_basis="binary_only",
                    reason_code="xobject_subtype_unsupported",
                    bbox=reference_bbox,
                    detail={"subtype": subtype or None},
                )
                continue

            try:
                raw_matrix = xobject.get("/Matrix")
                form_matrix = list(identity) if raw_matrix is None else [float(value) for value in raw_matrix]
                if len(form_matrix) != 6:
                    raise ValueError("Form Matrix must contain 6 numbers")
            except (TypeError, ValueError) as exc:
                add_issue(
                    path,
                    operator_path,
                    status="failed",
                    detected_kind="pdf_form_xobject",
                    content_basis="structured",
                    reason_code="form_matrix_parse_failed",
                    bbox=reference_bbox,
                    detail={"error_type": exc.__class__.__name__},
                )
                continue
            form_ctm = _matrix_multiply(ctm, form_matrix)
            form_bbox = reference_bbox
            child_clip = clip_bbox
            raw_form_bbox = xobject.get("/BBox")
            if raw_form_bbox is not None:
                try:
                    rect = [float(value) for value in raw_form_bbox]
                    if len(rect) != 4:
                        raise ValueError("Form BBox must contain 4 numbers")
                    form_bbox = _matrix_rect_bbox(form_ctm, rect)
                    child_clip = _bbox_intersection(clip_bbox, form_bbox) if clip_bbox is not None else form_bbox
                except (TypeError, ValueError) as exc:
                    add_issue(
                        path,
                        operator_path,
                        status="unsupported",
                        detected_kind="pdf_form_bbox",
                        content_basis="structured",
                        reason_code="form_bbox_parse_failed",
                        bbox=reference_bbox,
                        detail={"error_type": exc.__class__.__name__},
                        suffix="/bbox",
                    )

            next_depth = form_depth + 1
            if next_depth > PDF_XOBJECT_MAX_DEPTH:
                add_issue(
                    path,
                    operator_path,
                    status="unsupported",
                    detected_kind="pdf_form_xobject",
                    content_basis="structured",
                    reason_code="xobject_recursion_depth_exceeded",
                    bbox=form_bbox,
                    detail={"max_depth": PDF_XOBJECT_MAX_DEPTH},
                )
                continue
            token = object_token(reference, xobject)
            if token in active_forms:
                add_issue(
                    path,
                    operator_path,
                    status="failed",
                    detected_kind="pdf_form_xobject",
                    content_basis="structured",
                    reason_code="xobject_cycle_detected",
                    bbox=form_bbox,
                )
                continue
            try:
                form_operations = list(ContentStream(xobject, reader).operations)
                raw_resources = xobject.get("/Resources")
                form_resources = current_resources if raw_resources is None else dereference(raw_resources)
            except Exception as exc:
                add_issue(
                    path,
                    operator_path,
                    status="failed",
                    detected_kind="pdf_form_xobject",
                    content_basis="structured",
                    reason_code="form_content_stream_parse_failed",
                    bbox=form_bbox,
                    detail={"error_type": exc.__class__.__name__},
                )
                continue
            builder.add_coverage(
                "object",
                path_locator(path, operator_path, bbox=form_bbox),
                "extracted",
                detected_kind="pdf_form_xobject",
                content_basis="structured",
                reason_code="content_extracted",
                parser_id=PDF_ADAPTER_VERSION,
                detail={
                    "xobject_path": [name.lstrip("/") for name in path],
                    "operator_path": list(operator_path),
                },
            )
            walk(
                form_operations,
                form_resources,
                initial_ctm=form_ctm,
                path_prefix=path,
                operator_prefix=operator_path,
                active_forms=active_forms | {token},
                form_depth=next_depth,
                clip_bbox=child_clip,
            )

    walk(
        operations,
        resources,
        initial_ctm=identity,
        path_prefix=(),
        operator_prefix=(),
        active_forms=frozenset(),
        form_depth=0,
        clip_bbox=page_bbox,
    )
    return image_specs


def _pdf_axis_segments(operations: list[tuple[list[Any], bytes]]) -> list[tuple[float, float, float, float]]:
    """content streamの直交線分をuser spaceへ写す。曲線や斜線を表として扱わない。"""
    stack: list[list[float]] = []
    ctm = [1.0, 0.0, 0.0, 1.0, 0.0, 0.0]
    current: tuple[float, float] | None = None
    segments: list[tuple[float, float, float, float]] = []
    for operands, operator in operations:
        if operator == b"q":
            stack.append(list(ctm))
        elif operator == b"Q":
            ctm = stack.pop() if stack else [1.0, 0.0, 0.0, 1.0, 0.0, 0.0]
        elif operator == b"cm" and len(operands) == 6:
            try:
                ctm = _matrix_multiply(ctm, [float(value) for value in operands])
            except (TypeError, ValueError):
                continue
        elif operator == b"m" and len(operands) >= 2:
            try:
                current = _pdf_point(ctm, float(operands[0]), float(operands[1]))
            except (TypeError, ValueError):
                current = None
        elif operator == b"l" and len(operands) >= 2 and current is not None:
            try:
                target = _pdf_point(ctm, float(operands[0]), float(operands[1]))
            except (TypeError, ValueError):
                continue
            if abs(current[0] - target[0]) <= 0.5 or abs(current[1] - target[1]) <= 0.5:
                segments.append((*current, *target))
            current = target
        elif operator == b"re" and len(operands) >= 4:
            try:
                x, y, width, height = (float(value) for value in operands[:4])
            except (TypeError, ValueError):
                continue
            corners = [_pdf_point(ctm, x, y), _pdf_point(ctm, x + width, y),
                       _pdf_point(ctm, x + width, y + height), _pdf_point(ctm, x, y + height)]
            for left, right in zip(corners, corners[1:] + corners[:1], strict=True):
                if abs(left[0] - right[0]) <= 0.5 or abs(left[1] - right[1]) <= 0.5:
                    segments.append((*left, *right))
        elif operator in {b"n", b"S", b"s", b"f", b"F", b"f*", b"B", b"B*", b"b", b"b*"}:
            current = None
    return segments


def _pdf_table_grids(
    segments: list[tuple[float, float, float, float]],
) -> list[tuple[list[float], list[float]]]:
    """交差する縦横線のconnected componentから表gridを得る。独立した上下・左右の表を混ぜない。"""
    axis = [
        segment for segment in segments
        if (abs(segment[0] - segment[2]) <= 0.5 and abs(segment[1] - segment[3]) >= 5)
        or (abs(segment[1] - segment[3]) <= 0.5 and abs(segment[0] - segment[2]) >= 5)
    ]
    if not axis:
        return []
    adjacency: list[set[int]] = [set() for _ in axis]
    vertical = [i for i, item in enumerate(axis) if abs(item[0] - item[2]) <= 0.5]
    horizontal = [i for i, item in enumerate(axis) if abs(item[1] - item[3]) <= 0.5]
    for vertical_index in vertical:
        vx = axis[vertical_index][0]
        vlow, vhigh = sorted((axis[vertical_index][1], axis[vertical_index][3]))
        for horizontal_index in horizontal:
            hy = axis[horizontal_index][1]
            hlow, hhigh = sorted((axis[horizontal_index][0], axis[horizontal_index][2]))
            if hlow - 0.75 <= vx <= hhigh + 0.75 and vlow - 0.75 <= hy <= vhigh + 0.75:
                adjacency[vertical_index].add(horizontal_index)
                adjacency[horizontal_index].add(vertical_index)
    seen: set[int] = set()
    grids: list[tuple[list[float], list[float]]] = []
    for start in range(len(axis)):
        if start in seen or not adjacency[start]:
            continue
        component: set[int] = set()
        pending = [start]
        while pending:
            index = pending.pop()
            if index in component:
                continue
            component.add(index)
            pending.extend(adjacency[index] - component)
        seen.update(component)
        xs = sorted({round(axis[index][0], 3) for index in component if index in vertical})
        ys = sorted({round(axis[index][1], 3) for index in component if index in horizontal}, reverse=True)
        if len(xs) >= 2 and len(ys) >= 2:
            grids.append((xs, ys))
    return sorted(grids, key=lambda grid: (-grid[1][0], grid[0][0]))


def _pdf_text_groups(page: Any, operations: list[tuple[list[Any], bytes]], page_bbox: list[float]) -> list[dict[str, Any]]:
    """pypdfのdecoded textとlayout座標を安全に突き合わせる。ずれたPDFはpage textへ縮退する。"""
    decoded: list[tuple[str, float, float]] = []

    def visitor(text: str, _cm: list[float], tm: list[float], _font: Any, _size: float) -> None:
        value = text.strip()
        if value:
            decoded.append((value, float(tm[4]), float(tm[5])))

    page_text = page.extract_text(visitor_text=visitor) or ""
    try:
        from pypdf._text_extraction._layout_mode._fixed_width_page import text_show_operations
        groups = [item for item in text_show_operations(iter(operations), page._layout_mode_fonts())
                  if str(item.get("text", "")).strip()]
    except Exception:
        groups = []
    if len(groups) != len(decoded):
        return ([{"text": page_text.strip(), "bbox": page_bbox, "position_confidence": 0.0}]
                if page_text.strip() else [])
    offsets = [tm_x - float(group["tx"]) for (_text, tm_x, _tm_y), group in zip(decoded, groups, strict=True) if tm_x]
    x_offset = sorted(offsets)[len(offsets) // 2] if offsets else 0.0
    result: list[dict[str, Any]] = []
    for index, ((text, _tm_x, _tm_y), group) in enumerate(zip(decoded, groups, strict=True), start=1):
        left = float(group["tx"]) + x_offset
        right = float(group["displaced_tx"]) + x_offset
        baseline = float(group["ty"])
        height = max(float(group["font_height"]), float(group["font_size"]), 1.0)
        result.append({
            "text": text,
            "bbox": [round(left, 3), round(baseline - height * 0.25, 3), round(right, 3), round(baseline + height, 3)],
            "baseline_y": round(baseline, 3),
            "font_size": float(group["font_size"]),
            "group_index": index,
            "position_confidence": 1.0,
        })
    return result


def _pdf_cell_slot(group: dict[str, Any], grid: tuple[list[float], list[float]]) -> tuple[int, int] | None:
    xs, ys = grid
    bbox = group["bbox"]
    x = float(bbox[0])
    y = float(group.get("baseline_y", (bbox[1] + bbox[3]) / 2))
    column = next((index for index, (left, right) in enumerate(zip(xs, xs[1:], strict=False), start=1)
                   if left - 0.75 <= x < right + 0.75), None)
    row = next((index for index, (top, bottom) in enumerate(zip(ys, ys[1:], strict=False), start=1)
                if bottom - 0.75 <= y <= top + 0.75), None)
    return (row, column) if row is not None and column is not None else None


def _pdf_objects(builder: _Builder, path: Path) -> None:
    from pypdf import PdfReader
    from pypdf.generic import ContentStream

    reader = PdfReader(path)
    for page_number, page in enumerate(reader.pages, start=1):
        media = page.mediabox
        page_bbox = [float(media.left), float(media.bottom), float(media.right), float(media.top)]
        page_locator = evidence_ir.Locator(part=f"page:{page_number}", page=page_number, bbox=page_bbox,
                                           extension={"coordinate_system": "pdf-user-space"})
        page_id = builder.add_element("page", page_locator, parent_id=None, order=page_number, value=None)
        builder.add_coverage(
            "part",
            page_locator,
            "extracted",
            detected_kind="pdf_page",
            content_basis="structured",
            reason_code="content_extracted",
            parser_id=PDF_ADAPTER_VERSION,
        )
        try:
            content = ContentStream(page.get_contents(), reader)
            operations = content.operations
        except Exception as exc:
            failure_locator = evidence_ir.Locator(part=f"page:{page_number}/contents", page=page_number)
            builder.add_coverage(
                "region",
                failure_locator,
                "failed",
                detected_kind="pdf_content_stream",
                content_basis="structured",
                reason_code="content_stream_parse_failed",
                parser_id=PDF_ADAPTER_VERSION,
                detail={"error_type": exc.__class__.__name__},
            )
            continue

        text_indexes = [index for index, (_args, operator) in enumerate(operations)
                        if operator in {b"Tj", b"TJ", b"'", b'"'}]
        text_groups = _pdf_text_groups(page, operations, page_bbox) if text_indexes else []
        image_specs = _pdf_xobject_specs(
            builder,
            reader=reader,
            page_number=page_number,
            page_bbox=page_bbox,
            operations=operations,
            resources=page.get("/Resources"),
        )

        page_area = max(0.0, page_bbox[2] - page_bbox[0]) * max(0.0, page_bbox[3] - page_bbox[1])
        covering_specs = []
        for spec in image_specs:
            intersection = _bbox_intersection(page_bbox, spec.get("effective_bbox") or spec["bbox"])
            overlap_area = 0.0 if intersection is None else \
                (intersection[2] - intersection[0]) * (intersection[3] - intersection[1])
            ratio = overlap_area / page_area if page_area else 0.0
            spec["page_overlap_ratio"] = ratio
            if text_indexes and max(text_indexes) < spec["extension"]["operator_index"] and ratio >= 0.95:
                covering_specs.append(spec)

        text_id = None
        if text_indexes:
            text_locator = evidence_ir.Locator(
                part=f"page:{page_number}/contents",
                page=page_number,
                object_id="text-layer",
                bbox=page_bbox,
                extension={"operator_indexes": text_indexes, "coordinate_system": "pdf-user-space"},
            )
            text_id = builder.add_element(
                "text_object",
                text_locator,
                parent_id=page_id,
                order=1,
                value=page.extract_text() or "",
                visibility="hidden" if covering_specs else "visible",
                extension={
                    "layer": "text",
                    "visual_currentness": "unknown",
                    "aggregate": True,
                    **({"visibility_reason": "covered_by_full_page_image", "use_for_current_answer": False}
                       if covering_specs else {}),
                },
            )

        if not covering_specs and text_groups and all(item.get("position_confidence") == 1.0 for item in text_groups):
            grids = _pdf_table_grids(_pdf_axis_segments(operations))
            assigned: set[int] = set()
            for table_index, grid in enumerate(grids, start=1):
                xs, ys = grid
                table_groups: dict[tuple[int, int], list[dict[str, Any]]] = {}
                for group_index, group in enumerate(text_groups):
                    if (slot := _pdf_cell_slot(group, grid)) is not None:
                        table_groups.setdefault(slot, []).append(group)
                        assigned.add(group_index)
                if not table_groups:
                    continue
                table_bbox = [xs[0], ys[-1], xs[-1], ys[0]]
                title_candidates = [
                    group for index, group in enumerate(text_groups)
                    if index not in assigned and float(group["bbox"][1]) >= ys[0]
                ]
                title = min(title_candidates, key=lambda item: float(item["bbox"][1]) - ys[0])["text"] \
                    if title_candidates else f"PDF表{table_index}"
                table_locator = evidence_ir.Locator(
                    part=f"page:{page_number}/contents",
                    page=page_number,
                    object_id=f"table:{table_index}",
                    bbox=table_bbox,
                    extension={"coordinate_system": "pdf-user-space"},
                )
                table_id = builder.add_element(
                    "table", table_locator, parent_id=page_id, order=1000 + table_index, value=None,
                    extension={"name": title, "column_boundaries": xs, "row_boundaries": ys,
                               "geometry_source": "vector_grid"},
                )
                table_coverage = builder.coverage_for(table_id)
                for (row, column), groups_in_cell in sorted(table_groups.items()):
                    value = " ".join(group["text"] for group in sorted(groups_in_cell, key=lambda item: item["bbox"][0]))
                    cell_locator = evidence_ir.Locator(
                        part=f"page:{page_number}/contents",
                        page=page_number,
                        object_id=f"table:{table_index}/cell:{row}:{column}",
                        bbox=[xs[column - 1], ys[row], xs[column], ys[row - 1]],
                        extension={"row": row, "column": column, "row_span": 1, "column_span": 1,
                                   "coordinate_system": "pdf-user-space"},
                    )
                    builder.add_element(
                        "cell", cell_locator, parent_id=table_id, order=row * 1000 + column, value=value,
                        coverage_id=table_coverage,
                    )
            for group_index, group in enumerate(text_groups):
                if group_index in assigned:
                    continue
                locator = evidence_ir.Locator(
                    part=f"page:{page_number}/contents",
                    page=page_number,
                    object_id=f"text:{group['group_index']}",
                    bbox=group["bbox"],
                    extension={"coordinate_system": "pdf-user-space", "text_group_index": group["group_index"]},
                )
                builder.add_element(
                    "positioned_text", locator, parent_id=page_id, order=2000 + group_index,
                    value=group["text"], extension={"font_size": group["font_size"], "position_confidence": 1.0},
                )

        for image_order, spec in enumerate(image_specs, start=1):
            raw_name = spec["name"]
            image_bbox = spec["bbox"]
            image_id = builder.add_element(
                "image_xobject",
                evidence_ir.Locator(
                    part=f"page:{page_number}/xobject:{raw_name}",
                    page=page_number,
                    object_id=raw_name,
                    bbox=image_bbox,
                    extension={
                        "coordinate_system": "pdf-user-space",
                        "xobject_path": spec["extension"]["xobject_path"],
                        "operator_path": spec["extension"]["operator_path"],
                    },
                ),
                parent_id=page_id,
                order=image_order,
                value=None,
                coverage_status="metadata_only",
                coverage_reason="image_content_uninterpreted",
                extension=spec["extension"],
            )
            operator_index = spec["extension"]["operator_index"]
            if text_id is not None and text_indexes and max(text_indexes) < operator_index:
                builder.add_relation(
                    "painted_after",
                    text_id,
                    image_id,
                    evidence_ids=[text_id, image_id],
                    extension={"operator_order": [max(text_indexes), operator_index]},
                )
                if spec["page_overlap_ratio"] >= 0.95:
                    builder.add_relation(
                        "covers",
                        image_id,
                        text_id,
                        evidence_ids=[image_id, text_id],
                        extension={"page_overlap_ratio": spec["page_overlap_ratio"]},
                    )
                    builder.add_relation(
                        "covered_by",
                        text_id,
                        image_id,
                        evidence_ids=[text_id, image_id],
                        extension={"page_overlap_ratio": spec["page_overlap_ratio"], "use_for_current_answer": False},
                    )


def _legacy_ir(path: Path) -> document_ir.DocumentIR | None:
    from .arms import ooxml_arm

    builders = {
        ".docx": ooxml_arm._build_docx_ir,
        ".pptx": ooxml_arm._build_pptx_ir,
        ".xlsx": ooxml_arm._build_xlsx_ir,
    }
    builder = builders.get(path.suffix.lower())
    return builder(path) if builder is not None else None


def extract(
    path: str | Path,
    *,
    legacy_ir: document_ir.DocumentIR | None = None,
    consume_legacy: bool = False,
) -> evidence_ir.EvidenceIR:
    """XLSX/DOCX/PPTX/PDFを共通Evidence IRへ通す。外部通信・LLMは使用しない。"""
    source_path = Path(path)
    suffix = source_path.suffix.lower()
    if suffix not in {".xlsx", ".docx", ".pptx", ".pdf"}:
        raise ValueError(f"unsupported Evidence IR spike input: {suffix}")
    builder = _Builder(_source(source_path))
    if suffix == ".pdf":
        _pdf_objects(builder, source_path)
    else:
        legacy_ids = _adapt_document_ir(
            builder, legacy_ir if legacy_ir is not None else _legacy_ir(source_path), consume_legacy=consume_legacy)
        entries = _package_entries(source_path)
        package_metadata = _package_metadata(entries)
        if suffix == ".xlsx":
            content_parts = _xlsx_objects(builder, entries, legacy_ids)
        elif suffix == ".docx":
            content_parts = _docx_objects(builder, entries)
        else:
            content_parts = _pptx_objects(builder, entries)
        content_parts.update(package_metadata)
        _package_coverage(builder, entries, content_parts, part_details=package_metadata)
    errors = evidence_ir.validation_errors(builder.ir)
    if errors:
        raise ValueError("invalid Evidence IR spike: " + ",".join(errors))
    return builder.ir


def extract_assets(path: str | Path, ir: evidence_ir.EvidenceIR, destination: str | Path) -> list[Path]:
    """OOXML/PDF内のEvidence画像をcontent hash名でgeneration stageへ取り出す。

    Evidence IRに記録済みの``media_part``だけを対象とし、書込前にbytesのSHA-256を再照合する。
    destinationは非公開stage内を呼び出し側が指定するため、ここでは同名hash assetを決定的に上書きする。
    PDFはpypdfが列挙したpage image bytesを同じhashで再照合する。
    """
    source_path = Path(path)
    if source_path.suffix.lower() not in {".xlsx", ".docx", ".pptx", ".pdf"}:
        return []
    if source_path.suffix.lower() == ".pdf":
        from pypdf import PdfReader
        requested = [
            element for element in ir.elements
            if element.type == "image_xobject" and isinstance(element.extension.get("asset_sha256"), str)
        ]
        if not requested:
            return []
        target_dir = Path(destination)
        target_dir.mkdir(parents=True, exist_ok=True)
        written: dict[str, Path] = {}
        reader = PdfReader(source_path)
        expected: set[str] = set()
        for element in requested:
            expected_digest = str(element.extension["asset_sha256"])
            expected.add(expected_digest)
            page_number = element.locator.page
            xobject_path = element.extension.get("xobject_path")
            if not isinstance(page_number, int) or not isinstance(xobject_path, list) or not xobject_path:
                raise ValueError("PDF asset locator is incomplete")
            try:
                xobject = _pdf_xobject_at_path(reader.pages[page_number - 1], xobject_path)
                _image_name, raw, _decoded_image = _pdf_image_payload(xobject, xobject_path[-1])
            except (IndexError, KeyError) as exc:
                raise ValueError("PDF asset locator cannot be resolved") from exc
            actual_digest = hashlib.sha256(raw).hexdigest()
            if actual_digest != expected_digest:
                raise ValueError("PDF asset hash mismatch")
            suffix = element.extension.get("asset_extension", ".bin")
            suffix = suffix if isinstance(suffix, str) and suffix.startswith(".") else ".bin"
            target = target_dir / f"{actual_digest}{suffix}"
            if actual_digest not in written:
                target.write_bytes(raw)
                written[actual_digest] = target
        if set(written) != expected:
            raise ValueError("PDF asset inventory mismatch")
        return [written[digest] for digest in sorted(written)]
    requested: dict[str, str] = {}
    for element in ir.elements:
        candidates = element.extension.get("assets")
        if not isinstance(candidates, list):
            candidates = [element.extension]
        for candidate in candidates:
            if not isinstance(candidate, dict):
                continue
            media_part = candidate.get("media_part")
            digest = candidate.get("asset_sha256")
            if isinstance(media_part, str) and isinstance(digest, str):
                requested[media_part] = digest
    if not requested:
        return []
    target_dir = Path(destination)
    target_dir.mkdir(parents=True, exist_ok=True)
    written: list[Path] = []
    with zipfile.ZipFile(source_path) as package:
        for media_part, expected_digest in sorted(requested.items()):
            raw = package.read(media_part)
            actual_digest = hashlib.sha256(raw).hexdigest()
            if actual_digest != expected_digest:
                raise ValueError(f"asset hash mismatch: {media_part}")
            suffix = Path(media_part).suffix.lower() or ".bin"
            target = target_dir / f"{actual_digest}{suffix}"
            target.write_bytes(raw)
            written.append(target)
    return written
