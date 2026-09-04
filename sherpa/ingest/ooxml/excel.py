"""Excel（.xlsx）の生 OOXML 抽出層（DOC-IR-004・パッケージ docstring＝`sherpa/ingest/ooxml/__init__.py` 参照）。

`arms/ooxml_arm._build_xlsx_ir` が消費する純関数群。H2（`docs/proposals/2026-08-28-人間向けMDの刷新.md`）
以降、`office_md._xlsx_md`（人間向け MD・`human_md.render_xlsx` 経由）も本モジュールの `regions()` を
document-ir と共有する共通土台として消費する（旧・シート丸ごと1枚の打切り付きパイプ表は撤去済み）。

すべて純関数・決定的（走査・辞書・リスト順を必ずソートで固定）・LLM 不使用。壊れた/欠落した OOXML パート
（`xl/externalLinks/*` 等）は例外を投げず空へ縮退する（`ooxml/word.py`／`ooxml/powerpoint.py` と同じ
fail-safe 方針）。

**`load_two` の値用ロード（`wb_values`）は read_only=False（通常ロード）を使う**（設計判断）:
`openpyxl` の read_only ワークシート（`ReadOnlyWorksheet`）は `merged_cells`／`row_dimensions`／
`column_dimensions`／セル単位の `hyperlink`／`comment` を持たない（ストリーミング読み取りのため）。
本モジュールは結合セル・非表示行列・ハイパーリンク・コメントをすべて抽出する必要があるため、
値用だけは通常ロード（全セルをメモリに展開）を選ぶ。トレードオフ: 巨大シートでの通常ロードは
`read_only=True` より重い。ここは `regions()`／`formulas()`／`cell_hyperlinks()`／`cell_comments()`
側のcap（行・列上限と`DEFAULT_CAP_CELLS`の総セル予算）が上限を守る（詳細は各関数 docstring）。
数式用ロード（`wb_formula`）は上記のリッチな属性を必要としないため `read_only=True`（軽量な
ストリーミング読み込み）を使う（`load_two` docstring 参照）——openpyxl は同一ワークブックを
「表示値（キャッシュ）」と「数式文字列」を同時に返す API を持たないため2回のロード自体は残るが、
うち1回（値用）だけが「重い」通常ロードで足りる。

**「表・連続領域・設定欄」の意味分類はしない**（孤立セルも小さな `Region` として出す）。意味分類（ヘッダ/
データ/設定欄の判別）は検索用表現生成層（RAG-REP）の責務（`table_semantics.py` が表のヘッダ解釈を担うのと
同じ位置付け）。
"""
from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass
from xml.etree import ElementTree as ET

_RELS = "{http://schemas.openxmlformats.org/package/2006/relationships}"
# HM1（人間向けMDの画像存在注記）専用の最小限のネームスペース。`evidence_spike.py` の `_xlsx_objects`
# （z_order・図形種別・アセット解決・覆い判定まで含むフル実装）とは目的が異なる——ここでは「シートに
# 画像が何枚あるか」という存在だけの事実を得れば足りるため、独立の軽量スキャナとして持つ（意図的に
# 二重実装を避けなかった＝完全な図形解析を人間向けMD側へ持ち込まない設計判断）。
_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
_SML = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
_XDR = "{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}"

# 巨大シートの安全弁。行数だけを固定値へ縛ると狭い長大表を途中で失うため、Excel の実際の上限
# （行 1,048,576・列 16,384＝XFD）に加えて総走査セル数 5,000,000 を安全弁として持つ（人間向け MD の
# 「全量方針」＝正典 §10 裁定#1 の「セル数の安全弁付き」はこの `DEFAULT_CAP_CELLS` を指す）。
# 列上限は Excel の実仕様値（16,384）を使う（独自の追加制限を持たず全量方針＝正典 §10 裁定#1と
# 整合させる）。狭い表ほど多くの行を読める一方、列数が多いシートほど `effective_cap_rows` が行数側を
# 絞ってセル数予算を守る。`regions()` は有効上限を呼び出し側から明示的に受け取る（テストで
# 小さい値へ差し替えて truncated を固定しやすくするため）。`formulas`/`cell_hyperlinks`/`cell_comments` は
# 同じ定数をモジュール属性として**呼び出し時に**参照する（関数のデフォルト引数値ではなく関数本体で
# `DEFAULT_CAP_ROWS`/`DEFAULT_CAP_COLS` を直接読む）ため、`monkeypatch.setattr(excel, "DEFAULT_CAP_ROWS", …)`
# 一箇所で全スキャン系関数の上限を差し替えられる（デフォルト引数値は def 時に束縛されて monkeypatch が
# 効かない、という定石の罠を避けるための設計＝`_build_xlsx_ir` もこの2定数を都度モジュール属性として読む）。
DEFAULT_CAP_ROWS = 1_048_576
DEFAULT_CAP_COLS = 16_384
DEFAULT_CAP_CELLS = 5_000_000


def effective_cap_rows(sheet_max_column: int | None) -> int:
    """列数に応じた行走査上限。総セル予算を超えず、固定10,000行による狭い長大表の欠落を避ける。"""
    columns = max(1, min(sheet_max_column or 1, DEFAULT_CAP_COLS))
    return max(1, min(DEFAULT_CAP_ROWS, DEFAULT_CAP_CELLS // columns))


@dataclass
class Region:
    """連続領域（非空セルの4連結成分を、隣接表の癒着解消のため最大矩形へ分割したもの）の外接矩形
    （`regions()` の戻り値要素）。

    `min_row`/`max_row`/`min_col`/`max_col` は1-based の絶対シート座標。`range` は Excel A1形式
    （例 `"A1:C10"`）。`truncated` は cap（行/列の走査上限）に到達した領域だけ True（cap 超過の黙認防止・
    「打切りの結果このサイズになった」ことを明示するフラグであり、それ以上続きがあるとは限定しないが
    可能性を示す）。
    """
    min_row: int
    max_row: int
    min_col: int
    max_col: int
    range: str
    truncated: bool = False
    # この領域が**実際に所有する**セル座標（外接矩形どうしが重なった時、他領域の非空セルを bbox 走査で
    # 重複出力したり数式を誤った親へ紐付けたりしないための正本）。背景色付き・結合セルの継続セルなど
    # 値を持たない座標も、連結性維持のために占有扱いされたものは含みうる（`value_cell_count` は値を
    # 持つセルだけを数える）。
    cells: frozenset = frozenset()
    # 表候補スコア算出の元になった値セル数・密度（`_region_score` 参照）。抑制（表示/索引から外す）は
    # 行わず、判断材料として保持するだけ＝閾値判断は消費側（renderer 等）の責務。
    value_cell_count: int = 0
    density: float = 0.0
    score: float = 0.0
    # True の場合、この外接矩形は「範囲内は全セル占有」を保証するヒストグラム法の抽出結果ではなく、
    # 面積上限（`_MAX_RECT_DECOMPOSE_CELLS`）・シート全体の反復回数/領域数予算（`_MAX_RECT_SPLITS_PER_SHEET`/
    # `_MAX_REGIONS_PER_SHEET`）のいずれかに達したことで単純な外接矩形へ縮退した結果を示す。この場合
    # bbox の内部に非占有セルを含みうる（`cells` だけが所有座標の正本＝重複出力防止のownership解決は
    # 従来どおり `cells` を使えば安全）。消費側（human MD レンダラ・rag.md レンダラ等）が表の見た目の
    # 精度（隣接表の癒着解消がどこまで効いたか）を判断する材料として利用する。
    split_budget_exhausted: bool = False


# 表候補スコア（`score`）算出時、値セル数がこの件数に達するまでは密度を按分して割り引く（孤立した1〜数
# セルが常に score=density の高評価になるのを防ぐ、緩やかな下駄＝ハード閾値ではなく連続的な補正）。
_SCORE_MIN_CELLS = 4

# ヒストグラム法による最大矩形の反復抽出（`_split_component`）の計算量安全弁。連結成分の外接矩形面積が
# これを超える場合は分割を行わず単一の外接矩形のまま返す（巨大な単一表を割っても得るものが無い一方、
# 計算量だけが増えるため）。
#
# HM1（`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §8.4 の L1 実測での「excel2md に
# 負けた」唯一の箇所）: 旧値 5,000 では JPX-021.xlsx の「統合設計」シートの2連結成分（A30:Z289＝面積
# 6,760・A291:N654＝面積5,096）がどちらも分割されず単一外接矩形へ縮退していた。5,000→20,000へ引き上げ
# （選択肢(a)。選択肢(b)＝上限超過成分を再帰的に半分割してから分割する案は「崖を無くせる」利点は
# あるが、`_split_component`/`regions()` のシート全体予算（`_MAX_RECT_SPLITS_PER_SHEET`/
# `_MAX_REGIONS_PER_SHEET`）との整合を再帰全体で保つ実装が重く、上限自体は今回も撤廃しない
# （`split_budget_exhausted` の自己申告は残る）以上、値を上げるだけで実害のあるケースを解消できるなら
# 複雑さに見合わないと判断した）。実測（JPX-021.xlsx 実データ・`_build_xlsx_ir`＋`human_md.render_xlsx`
# フル実行）: 旧値5,000＝0.856秒（縮退2件）→ 新値20,000＝0.733秒（縮退0件・退行なし、僅かな高速化は
# 縮退回避で巨大単一表の描画コストが減ったため）。新値は依然として崖を上へ動かしただけだが
# （選択肢(a)の性質上残る既知の限界）、実務のExcelファイルで単一連結成分が20,000セルを超える
# ケースは非常に稀という判断。合成の最悪ケース（櫛形パターン・連結成分が密ではなく細長い）で
# 面積40,000・`_MAX_RECT_SPLITS_PER_SHEET`一杯の32回反復だと0.72秒程度（実測）——シート全体の
# 反復予算（64回）で頭打ちになるため、1シートあたりの最悪コストは新値でも数秒以内に収まる。
_MAX_RECT_DECOMPOSE_CELLS = 20_000
# 1連結成分あたりの最大分割反復回数。到達時点で残りのセルは4連結成分ごとの外接矩形へまとめて必ず出力する
# （分割の粒度が粗くなるだけで、セルが失われることはない＝完全性は反復回数の上限とは独立に保たれる）。
_MAX_RECT_SPLITS = 32

# シート全体での安全弁。`_MAX_RECT_DECOMPOSE_CELLS`/`_MAX_RECT_SPLITS` は連結成分1つあたりの上限のため、
# 分割を要する連結成分（例: 櫛形の孤立成分）が同一シートに多数あると合計コストは頭打ちにならない。
# `regions()` は連結成分を処理するたびにこの2つの予算を消費し、どちらかを使い切った時点以降の連結成分は
# ヒストグラム分割を行わず単一の外接矩形にフォールバックする（セルが失われることはない＝完全性はシート
# 全体の予算とも独立に保たれる）。
_MAX_RECT_SPLITS_PER_SHEET = 64
_MAX_REGIONS_PER_SHEET = 256


def _region_score(value_cell_count: int, area: int) -> tuple[float, float]:
    """`(density, score)` を返す（`Region.density`/`Region.score` と `expand_regions_for_merges` の
    再計算で共有する単一の算出ロジック）。"""
    density = value_cell_count / area if area else 0.0
    taper = min(1.0, value_cell_count / _SCORE_MIN_CELLS) if _SCORE_MIN_CELLS else 1.0
    return density, density * taper


def _connected_subcomponents(cells: set[tuple[int, int]]) -> list[set[tuple[int, int]]]:
    """`cells`（1-based 座標集合）を4連結成分（上下左右のみ）に分割する（`(min_row, min_col)` が
    最小の座標から見つかる順・決定的）。"""
    visited: set[tuple[int, int]] = set()
    out: list[set[tuple[int, int]]] = []
    for start in sorted(cells):
        if start in visited:
            continue
        stack = [start]
        visited.add(start)
        component: set[tuple[int, int]] = {start}
        while stack:
            cr, cc = stack.pop()
            for dr, dc in ((1, 0), (-1, 0), (0, 1), (0, -1)):
                nb = (cr + dr, cc + dc)
                if nb in cells and nb not in visited:
                    visited.add(nb)
                    component.add(nb)
                    stack.append(nb)
        out.append(component)
    return out


def _extract_max_rectangle(occ: set[tuple[int, int]], min_row: int, max_row: int,
                           min_col: int, max_col: int) -> tuple[int, int, int, int]:
    """`occ` のうち `[min_row,max_row]×[min_col,max_col]` 範囲内で、全セルが `occ` に含まれる
    最大面積の矩形を1つ `(top, left, bottom, right)` で返す（ヒストグラム法・モノトニックスタック・
    O(範囲の行数×列数)）。`occ` はこの範囲内に必ず1セル以上を含む前提（空なら呼び出し側の誤り）。

    同面積の解が複数あれば `(top, left)` が最小の物を選ぶ（決定的なタイブレーク・同じ入力に対し
    常に同じ分割結果になることを保証する）。
    """
    n_cols = max_col - min_col + 1
    heights = [0] * n_cols
    best_area = 0
    best: tuple[int, int, int, int] | None = None
    for r in range(min_row, max_row + 1):
        for ci in range(n_cols):
            heights[ci] = heights[ci] + 1 if (r, min_col + ci) in occ else 0
        stack: list[tuple[int, int]] = []          # (開始列index, 高さ)
        for ci in range(n_cols + 1):
            h = heights[ci] if ci < n_cols else 0
            start = ci
            while stack and stack[-1][1] >= h:
                s, sh = stack.pop()
                area = sh * (ci - s)
                candidate = (r - sh + 1, min_col + s, r, min_col + ci - 1)
                if area > 0 and (area > best_area
                                 or (area == best_area and candidate[:2] < best[:2])):
                    best_area, best = area, candidate
                start = s
            stack.append((start, h))
    assert best is not None
    return best


def _split_component(component: set[tuple[int, int]], min_row: int, max_row: int,
                     min_col: int, max_col: int, *,
                     max_splits: int, max_regions: int
                     ) -> tuple[list[tuple[frozenset, int, int, int, int, bool]], int]:
    """1つの4連結成分をヒストグラム法の最大矩形反復抽出で分割し、`([(所有セル集合, min_row, min_col,
    max_row, max_col, split_budget_exhausted), ...], 実際に使った反復回数)` を返す（隣接する複数表の
    癒着解消）。呼び出し側で `(min_row, min_col)` 順に並べ直すこと。

    `max_splits`は呼び出し側が渡す**この呼び出し1回あたりの**反復回数上限、`max_regions`は
    **この呼び出し1回が出力してよい領域数の上限**（いずれも `regions()` がシート全体の予算の残量を
    計算して渡す・唯一の呼び出し元＝`regions()` が常に整数で渡すため両方必須引数）。

    アルゴリズム: まず `max_splits` にだけ従って「自然な」分割（ヒストグラム抽出を反復し、打ち切った
    残りは4連結成分ごとの断片に分ける）を最後まで計算する。この時点では `max_regions` を一切考慮
    しない——**分割結果の件数を先に確定させてから予算と突き合わせる**（部分的に分割を進めてから
    途中で日和ると、既に出力済みの件数＋残りをまとめた1件が `max_regions` を超えてしまう罠がある。
    例: `max_regions=1` の3セルL字成分に対し「まず1回だけ抽出→まだ2セル残っている→それを1件に
    まとめて追加」とすると合計2件になり予算を超える）。自然な分割の件数が `max_regions` に収まれば
    そのまま採用する。収まらなければ**分割そのものを行わず**、成分全体を1つの外接矩形へ丸ごと畳む
    （保証された1件だけを使う＝予算を超えることは絶対にない）。外接矩形面積が
    `_MAX_RECT_DECOMPOSE_CELLS` を超える成分・`max_splits`/`max_regions` が0（呼び出し時点で予算切れ）
    の場合も同様に単一の外接矩形のまま返す。

    完全性（silent-drop ゼロ）はどちらの経路でも保たれる（採用する分割結果・畳んだ単一矩形のいずれも
    `cells`（`Region.cells` 相当）の和集合は元の成分全体と一致する）。

    `split_budget_exhausted`（各出力要素の6番目）は、その領域が「範囲内は全セル占有」を保証する
    ヒストグラム抽出そのものではなく、上記いずれかの上限到達によるフォールバックで生成されたかを示す
    （`Region.split_budget_exhausted` docstring 参照）。
    """
    area = (max_row - min_row + 1) * (max_col - min_col + 1)
    if area > _MAX_RECT_DECOMPOSE_CELLS or max_splits <= 0 or max_regions <= 0:
        return [(frozenset(component), min_row, min_col, max_row, max_col, True)], 0

    remaining = set(component)
    natural: list[tuple[frozenset, int, int, int, int, bool]] = []
    splits = 0
    while remaining and splits < max_splits:
        r_min = min(r for r, _ in remaining)
        r_max = max(r for r, _ in remaining)
        c_min = min(c for _, c in remaining)
        c_max = max(c for _, c in remaining)
        top, left, bottom, right = _extract_max_rectangle(remaining, r_min, r_max, c_min, c_max)
        rect_cells = frozenset((r, c) for r in range(top, bottom + 1) for c in range(left, right + 1))
        natural.append((rect_cells, top, left, bottom, right, False))
        remaining -= rect_cells
        splits += 1
    if remaining:                                          # 反復回数上限で打ち切った残りを断片ごとに
        for sub in _connected_subcomponents(remaining):
            sr = [r for r, _ in sub]
            sc = [c for _, c in sub]
            natural.append((frozenset(sub), min(sr), min(sc), max(sr), max(sc), True))

    if len(natural) <= max_regions:
        return natural, splits
    # 自然な分割結果が予算に収まらない: 部分的な結果を採用すると合計が max_regions を超えてしまうため、
    # 分割そのものを諦めて成分全体を1つの外接矩形へ畳む（保証された1件のみ使用・完全性は cells で保つ）。
    # 反復に使った splits（計算コスト）はシート全体の予算から差し引くため、そのまま正直に返す。
    return [(frozenset(component), min_row, min_col, max_row, max_col, True)], splits


def load_two(p):
    """openpyxl で `p` を2回ロードする（`(wb_values, wb_formula)`）。

    `wb_values`=`data_only=True`（表示=キャッシュ済み計算値）・`read_only=False`（モジュール docstring
    の判断根拠を参照＝結合セル/非表示行列/ハイパーリンク/コメントの取得に通常ロードが必要）。

    `wb_formula`=`data_only=False`（数式文字列）は `read_only=True`（ストリーミング）で読む:
    消費側（`formulas()`）は `iter_rows()`/`.value`/`.row`/`.column`/`.coordinate` しか使わず、
    `read_only` ワークシートが持たない `merged_cells`／`row_dimensions`／セル単位の `hyperlink`／
    `comment` を必要としない（それらは全て `wb_values` 側から取る）。1ファイルにつき「重い」通常
    ロードを2回払う必要は無く、数式抽出専用の軽量ストリーミング読み込み1回で足りる（openpyxl は
    同一ワークブックを「表示値（キャッシュ）」と「数式文字列」を同時に返す API を持たないため
    2回のロード自体は残るが、うち1回を大幅に軽くする＝`_build_xlsx_ir` docstring にも記載）。
    呼び出し側は使用後に両方を `close()` すること（fail-safe な後始末は呼び出し側の責務＝
    `_build_xlsx_ir` が `finally` で行う）。
    """
    import openpyxl
    wb_values = openpyxl.load_workbook(p, data_only=True, read_only=False)
    wb_formula = openpyxl.load_workbook(p, data_only=False, read_only=True)
    return wb_values, wb_formula


def sheet_states(wb) -> list[dict]:
    """ワークブック内の全シートを**ブック内順**で `{"name": <タイトル>, "state": "visible"|"hidden"|"veryHidden"}`
    のリストにする（`wb.worksheets` の順序＝ファイル内のシート順をそのまま使う＝決定的）。
    """
    return [{"name": ws.title, "state": ws.sheet_state} for ws in wb.worksheets]


def _range_str(min_r: int, min_c: int, max_r: int, max_c: int) -> str:
    from openpyxl.utils import get_column_letter
    return f"{get_column_letter(min_c)}{min_r}:{get_column_letter(max_c)}{max_r}"


def regions(ws_values: list[list], cap_rows: int, cap_cols: int, *,
           merged: dict[tuple[int, int], dict] | None = None,
           filled: set[tuple[int, int]] | None = None) -> list["Region"]:
    """非空セルの4連結成分（上下左右のみ・斜めは繋がない）を、隣接する複数表の癒着を解消するため最大矩形へ
    分割し、`Region` のリストで返す。

    `ws_values` は**既に呼び出し側が cap+1 まで有界化した**値グリッド（行のリストのリスト・
    `None`＝空セル）: `_build_xlsx_ir` が `ws.max_row`/`ws.max_column`（実使用範囲）と `cap_rows+1`/
    `cap_cols+1` の小さい方まで `iter_rows(values_only=True)` で読み取ったものを渡す契約（本関数自体は
    渡された grid の総量を制限しない＝有界化は呼び出し側の責務。理由: openpyxl の通常ワークシートは
    `max_row`/`max_col` を明示指定した `iter_rows` に対し、実データが疎でも指定範囲まるごとの空セルを
    生成するため、**シートの実使用範囲を超えて**大きな cap をいきなり指定すると小さなシートでも遅くなる
    ＝呼び出し側が実使用範囲と cap の小さい方で読むことで通常サイズのシートは高速なまま、cap は
    「本当に巨大/巨大と自称するシート」だけを頭打ちにする安全弁として働く）。

    `merged`（`merged_map()` の戻り値と同じ形）・`filled`（背景色付きセルの1-based座標集合・
    `filled_cells()` 参照）は**値を持たなくても占有として扱う**追加情報（いずれも省略可＝値のみの
    占有判定にフォールバック）:

    - **背景色**: `filled` に含まれる座標は値の有無に関わらず占有マスにする（値が空でも塗り分け
      レイアウトの区画を取りこぼさない）。「背景色」の定義自体（単色/縞模様/グラデーション塗り・
      白／自動／テーマ背景1の除外・条件付き書式は対象外）は `filled_cells()`/`_fill_is_colored()`
      docstring 参照（`regions()` 自体は `filled` の中身の妥当性を判定しない＝呼び出し側の責務）。
    - **結合セル**: 1つの結合範囲内の**いずれかのセルが占有済み**（値がある、または背景色付き）なら、
      その範囲の全セル（anchor・非anchor 継続セルとも）を占有にする。逆に範囲内が完全に空かつ無地なら
      占有にしない（罫線・spacer 目的の空結合まで表候補化するのを避ける）。これにより、値の無い結合の
      広がりだけで隣接する2つの値クラスタを繋いでいる特殊なレイアウトでも、正しく1つの連結成分として
      検出できる。

    アルゴリズム: `ws_values` を `cap_rows`×`cap_cols` に切り詰めた上で占有マス（非空セル＋上記の
    背景色/結合による追加占有）を求め、行優先で未訪問の占有マスを見つけるたびに BFS で4連結成分を1つ
    確定する。各連結成分は、そのままでは外接矩形が隣接する別の表を巻き込みうるため、ヒストグラム法に
    よる最大矩形の反復抽出（`_split_component`）でさらに分割する（1連結成分から複数の `Region` が
    生まれうる＝表が2つ癒着していたケースは通常2つの `Region` に分かれる）。分割コストはシート全体で
    `_MAX_RECT_SPLITS_PER_SHEET`（反復回数）・`_MAX_REGIONS_PER_SHEET`（出力 `Region` 数）の2つの予算を
    共有する（1シートに分割を要する連結成分が多数ある場合の合計コストを頭打ちにする安全弁）。

    **`Region` 数の契約**: `len(regions(...)) <= max(連結成分数, _MAX_REGIONS_PER_SHEET)`。各連結成分は
    silent-drop ゼロのため最低でも1つの `Region`（分割できなければ外接矩形1件）を出す＝連結成分数を
    下回ることはできない（例: 市松模様のように孤立した1セル成分が500個あれば、`_MAX_REGIONS_PER_SHEET`
    が256でも500件になる。これはバグではなく契約どおりの挙動）。逆に連結成分数が予算未満であれば、
    合計は `_MAX_REGIONS_PER_SHEET` を超えない（1つの連結成分だけが予算を独占して他の成分の分まで
    使い切ることはない＝各成分に「保証された1件」を割り当てた上で、余りの予算だけを早く処理された
    成分から順に分け合う）。予算（保証枠を超えた追加の分割）を使い切った成分は単一の外接矩形へ
    フォールバックし（`split_budget_exhausted=True`）、セルが失われることはない。

    `Region.value_cell_count`/`density`/`score`: 各領域について、値を持つセル数（背景色/結合継続などの
    占有専用セルを除く）・その密度（値セル数／矩形面積）・両者を合成した表候補スコアを付与する
    （`_region_score` 参照）。**スコアによる抑制（表示/索引からの除外）はここでは行わない**＝閾値判断は
    消費側（人間向け MD レンダラ・rag.md レンダラ等）の責務。

    `truncated`: `ws_values` の行数が `cap_rows` を超えている（＝呼び出し側が cap+1 まで読み cap 超過を検知
    できた）、または列数が `cap_cols` を超えている行がある場合に、その **cap 境界に接する領域だけ** True。
    境界に接しない領域（cap よりずっと手前で完結している領域）は cap 超過の影響を受けていないため False。
    領域どうしの外接矩形が（連結していない孤立セルを内包する形で）幾何的に重なる稀なケースがありうるが、
    本関数は矩形を返すだけで重複排除はしない（呼び出し側がセル配列を組む際に同じ座標が複数の table 要素へ
    空セルとして重複出力されうる＝実害の無い冗長性として許容する設計判断）。

    戻り値は `(min_row, min_col)` 昇順（シート内の左上優先＝「領域順」）でソート済み。同じ入力に対し
    常に同じ順序・同じ分割結果を返す（`_extract_max_rectangle` のタイブレークが決定的なため）。
    """
    total_rows = len(ws_values)
    row_trunc = total_rows > cap_rows
    rows = ws_values[:cap_rows]
    col_trunc = any(len(r) > cap_cols for r in rows)

    value_occupied: set[tuple[int, int]] = set()
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row[:cap_cols], start=1):
            if v is not None and str(v).strip() != "":
                value_occupied.add((r, c))

    occupied: set[tuple[int, int]] = set(value_occupied)
    if filled:
        occupied.update((r, c) for (r, c) in filled if r <= cap_rows and c <= cap_cols)
    if merged:
        spans: dict[tuple[int, int], list[tuple[int, int]]] = {}
        for coord, info in merged.items():
            r, c = coord
            if r > cap_rows or c > cap_cols:
                continue
            spans.setdefault(info["anchor"], []).append(coord)
        for span_cells in spans.values():
            if any(cell in occupied for cell in span_cells):
                occupied.update(span_cells)

    components = list(_connected_subcomponents(occupied))
    # 各連結成分に「保証された1件」を割り当てた上で、余りの予算（`_MAX_REGIONS_PER_SHEET` が連結成分数
    # より大きい場合の差分）だけを早く処理された成分から順に分け合う（`Region` 数の契約は docstring
    # 参照）。連結成分数自体が予算を超える場合は保証枠すら削れない＝合計は連結成分数のまま。
    extra_region_budget = max(0, _MAX_REGIONS_PER_SHEET - len(components))

    out: list[Region] = []
    splits_budget = _MAX_RECT_SPLITS_PER_SHEET
    for component in components:
        comp_min_r = min(r for r, _ in component)
        comp_max_r = max(r for r, _ in component)
        comp_min_c = min(c for _, c in component)
        comp_max_c = max(c for _, c in component)
        effective_max_splits = min(_MAX_RECT_SPLITS, splits_budget) if splits_budget > 0 else 0
        effective_max_regions = 1 + extra_region_budget          # 保証1件 + 残りの共有予算
        sub_regions, used = _split_component(
            component, comp_min_r, comp_max_r, comp_min_c, comp_max_c,
            max_splits=effective_max_splits, max_regions=effective_max_regions)
        splits_budget -= used
        extra_region_budget -= max(0, len(sub_regions) - 1)      # この成分が消費した「追加分」だけ減らす
        for cells, top, left, bottom, right, budget_exhausted in sub_regions:
            touches_cap = (row_trunc and bottom == cap_rows) or (col_trunc and right == cap_cols)
            value_count = len(cells & value_occupied)
            area = (bottom - top + 1) * (right - left + 1)
            density, score = _region_score(value_count, area)
            out.append(Region(min_row=top, max_row=bottom, min_col=left, max_col=right,
                              range=_range_str(top, left, bottom, right), truncated=touches_cap,
                              cells=cells, value_cell_count=value_count, density=density, score=score,
                              split_budget_exhausted=budget_exhausted))
    out.sort(key=lambda rg: (rg.min_row, rg.min_col))
    return out


def sheet_truncated(ws_values: list[list], cap_rows: int, cap_cols: int,
                    sheet_max_row: int | None = None, sheet_max_col: int | None = None) -> bool:
    """走査が**シート全域をカバーしていない**（＝cap で打ち切った）可能性を示すか（RV Med #2・RV2巡是正）。

    厳密な「cap 外に非空セルが在るか」は cap を超えて全域走査しないと判定できず、cap の目的（巨大シートの
    暴走防止）と矛盾する。そこで意味論を**保守的**に定義する: シートの申告範囲（openpyxl の
    `max_row`/`max_column`＝dimension 由来）が cap を超えていれば「見ていない領域がある」として True
    （申告範囲が過大で実は空だった場合の偽陽性は許容＝「走査していない」事実を正直に言うフラグ）。
    加えて、読み込んだ番兵グリッド（cap+1 まで）内に実際の非空セルがあれば dimension が過小申告でも True
    （精度向上の補助）。`regions()` の `truncated`（cap 境界に接した領域）と合わせ、cap の完全外側で完結する
    領域（例 cap=3 で A8 のみ）が**無警告で消える**ことを防ぐ（黙認しない契約の残り半分）。
    """
    if sheet_max_row is not None and sheet_max_row > cap_rows:
        return True
    if sheet_max_col is not None and sheet_max_col > cap_cols:
        return True
    for r, row in enumerate(ws_values, start=1):
        for c, v in enumerate(row, start=1):
            if (r > cap_rows or c > cap_cols) and v is not None and str(v).strip() != "":
                return True
    return False


def expand_regions_for_merges(region_list: list["Region"], merged: dict[tuple[int, int], dict],
                              cap_rows: int, cap_cols: int) -> list["Region"]:
    """各領域の外接矩形を、所有セル中の**結合 anchor の span** まで広げた Region リストを返す（RV Low #3）。

    値を持つ anchor（例 A1 で `A1:C1` 結合）の継続セルは非占有のため bbox に入らず、`source_map.range` が
    実セル範囲（span 含む）より狭くなる＝範囲スコープの hidden_columns 判定も漏れていた。拡張は cap で
    クランプし、クランプが起きた領域は `truncated=True` にする（黙認しない）。所有セル集合（`cells`）は
    変えない（継続セルは値を持たない＝出力対象でないため）。矩形拡張で面積が変わるため、`density`/`score`
    は `value_cell_count`（不変）と新しい面積から再計算する（`regions()` と同じ `_region_score`）。
    """
    out: list[Region] = []
    for rg in region_list:
        max_r, max_c, clamped = rg.max_row, rg.max_col, False
        for coord in rg.cells:
            info = merged.get(coord)
            if info is None or info["anchor"] != coord:
                continue
            span_r = coord[0] + info["row_span"] - 1
            span_c = coord[1] + info["column_span"] - 1
            if span_r > cap_rows:
                span_r, clamped = cap_rows, True
            if span_c > cap_cols:
                span_c, clamped = cap_cols, True
            max_r, max_c = max(max_r, span_r), max(max_c, span_c)
        if (max_r, max_c) == (rg.max_row, rg.max_col):
            out.append(rg)
        else:
            area = (max_r - rg.min_row + 1) * (max_c - rg.min_col + 1)
            density, score = _region_score(rg.value_cell_count, area)
            out.append(Region(min_row=rg.min_row, max_row=max_r, min_col=rg.min_col, max_col=max_c,
                              range=_range_str(rg.min_row, rg.min_col, max_r, max_c),
                              truncated=rg.truncated or clamped, cells=rg.cells,
                              value_cell_count=rg.value_cell_count, density=density, score=score,
                              split_budget_exhausted=rg.split_budget_exhausted))
    return out


def _clip_merge_enumeration_bounds(min_row: int, min_col: int, max_row: int, max_col: int,
                                   cap_rows: int, cap_cols: int) -> tuple[int, int]:
    """結合範囲 `[min_row,max_row]×[min_col,max_col]` を**座標展開する前に** `cap_rows`×`cap_cols`、かつ
    総面積 `DEFAULT_CAP_CELLS` 以内へクリップした `(max_row, max_col)` を返す（`min_row`/`min_col` は
    変えない）。

    `w:mergeCell`（Excel の結合セル）は宣言上 Excel の絶対上限（A1:XFD1048576）まで指定できるため、
    座標展開の**前**にこれを行わないと、1つの結合範囲だけで最大 `1,048,576 × 16,384 ≈ 172億` 座標を
    辞書へ書き込もうとしてメモリ/時間が破綻する。`cap_rows`/`cap_cols` は呼び出し側が渡す既存の
    走査上限をそのまま使い、さらに**両軸の積**（面積）を `DEFAULT_CAP_CELLS` 以内へ追加でクリップする
    （`cap_rows`×`cap_cols` の単純な積では、例えば列幅の狭いシートに `cap_rows` が大きく振られた
    状態で列側だけ巨大な結合があると、両軸クリップだけでは依然として億単位になりうるため）。
    行方向を優先して削る（既存の cap 系関数が行優先で走査する方針と揃える）。
    `row_span`/`column_span` 自体（`merged_map()` が返す値）はクリップしない——これは
    `expand_regions_for_merges()` 側が cap と突き合わせて別途クランプする値であり、ここで削ると
    その判定が狂う。
    """
    max_row = min(max_row, cap_rows)
    max_col = min(max_col, cap_cols)
    area = (max_row - min_row + 1) * (max_col - min_col + 1)
    if area > DEFAULT_CAP_CELLS:
        width = max_col - min_col + 1
        max_row = min_row + max(1, DEFAULT_CAP_CELLS // width) - 1
    return max_row, max_col


def merged_map(ws, cap_rows: int, cap_cols: int) -> dict[tuple[int, int], dict]:
    """結合セル範囲を座標展開した辞書 `{(row, col): {"anchor": (ar, ac), "row_span", "column_span"}}`。

    範囲内の**全座標**（anchor 自身も含む）をキーにする。呼び出し側は座標を引いて: キーが無ければ通常
    セル（span 1,1）、`info["anchor"] == (row, col)` なら anchor（span はそのまま使う）、それ以外なら
    非anchor の継続セル（`cells` に出さない＝DOCX の `w:vMerge` 継続セルと同じ規約）と判定する。

    **展開は `cap_rows`/`cap_cols` でクリップしてから行う**（`_clip_merge_enumeration_bounds` 参照）:
    宣言上 Excel の絶対上限まで指定できる結合範囲（A1:XFD1048576 等）をそのまま座標展開すると
    座標数が桁違いに膨れ上がるため。クリップは**座標展開の範囲だけ**に効き、辞書の値に入る
    `row_span`/`column_span` は結合の宣言どおりの値のまま返す（`expand_regions_for_merges()` が
    その値と cap を突き合わせて改めてクランプする設計のため）。

    `ws.merged_cells.ranges` を `(min_row, min_col)` 順にソートしてから展開する（決定的な処理順・
    正常な OOXML では結合範囲は重ならないため出力に実質的な影響は無いが、走査順を安定させる）。

    **既知の限界（受容記録）**: 結合範囲の非anchorセルだけに値がある異常な OOXML（Excel の仕様上、
    非anchorセルの値は無効・Excel 自体が UI 上そのような値を作らない）は、`load_two()` の通常ロードの
    時点で openpyxl がその値を破棄する（実測済み＝raw XML で非anchorセルへ値を注入して保存→再読込しても
    値は復元されない）。本モジュールより手前でデータが失われるため、`regions()`/`_build_xlsx_ir` 側では
    修復しようがない。クラッシュはしない（`filled_cells()` が anchor の塗りを見るため、anchor 自体に
    値または塗りがあれば結合範囲全体が占有として扱われる＝表候補としては維持される）。
    """
    out: dict[tuple[int, int], dict] = {}
    for mc in sorted(ws.merged_cells.ranges, key=lambda m: (m.min_row, m.min_col)):
        anchor = (mc.min_row, mc.min_col)
        row_span = mc.max_row - mc.min_row + 1
        col_span = mc.max_col - mc.min_col + 1
        clip_max_row, clip_max_col = _clip_merge_enumeration_bounds(
            mc.min_row, mc.min_col, mc.max_row, mc.max_col, cap_rows, cap_cols)
        for r in range(mc.min_row, clip_max_row + 1):
            for c in range(mc.min_col, clip_max_col + 1):
                out[(r, c)] = {"anchor": anchor, "row_span": row_span, "column_span": col_span}
    return out


_DRAWINGML_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


_HEX6_RE = re.compile(r"^[0-9A-Fa-f]{6}$")
_HEX8_RE = re.compile(r"^[0-9A-Fa-f]{8}$")


def _is_white_hex(value) -> bool:
    """`value`（RGB6桁または ARGB8桁の16進文字列を想定）が厳密に白（`FFFFFF`）と確認できるか。

    文字列全体の形式を検証してから比較する（`str.endswith("FFFFFF")` だけの判定は
    `"garbageFFFFFF"` のような、6桁 HEX として不正な値まで白として誤受理してしまう＝実際に
    踏んだ回帰）。6桁（テーマ/`srgbClr`/`sysClr` の `lastClr` は無アルファの6桁）はそのまま比較、
    8桁（`wb._colors`/`Color.rgb` の ARGB）はアルファ成分を除いた末尾6桁を比較する。どちらの形式にも
    一致しない・16進以外の文字が混じる場合は白と断定せず False を返す（解決不能＝占有側の一部）。
    """
    if not isinstance(value, str):
        return False
    if _HEX6_RE.fullmatch(value):
        return value.upper() == "FFFFFF"
    if _HEX8_RE.fullmatch(value):
        return value[2:].upper() == "FFFFFF"
    return False


def _resolve_indexed_rgb(wb, indexed: int) -> str | None:
    """`wb`（openpyxl `Workbook`）のインデックスパレット（`wb._colors`）から `indexed` が指す実際の
    RGB 文字列を取得する。`wb._colors` はブック固有のカスタム上書き（`<indexedColors>`）を反映済み
    （openpyxl がロード時に読み、未カスタマイズなら標準64色パレットのまま）。`wb` が無い・パレットが
    無い・インデックスが範囲外なら解決不能として `None` を返す（呼び出し側は安全側＝占有として扱う）。
    """
    colors = getattr(wb, "_colors", None) if wb is not None else None
    if not colors or not (0 <= indexed < len(colors)):
        return None
    return colors[indexed]


def _resolve_theme_lt1_rgb(wb) -> str | None:
    """`wb.loaded_theme`（テーマ part の生 XML バイト列。実ファイルから読み込んだブックには
    OOXML 仕様上必ず入っている）から背景1（`lt1`）の実際の RGB を取り出す。`xml.etree` で正規に
    パースする（正規表現の部分一致だと、閉じタグを欠くなどの壊れた XML でもそれらしい断片にマッチして
    誤って白を受理しうるため）。直接色（`a:srgbClr`）・システム色参照（`a:sysClr` の `lastClr`
    フォールバック値）のどちらにも対応する。`wb` が無い・テーマ未ロード・パース失敗
    （`ET.ParseError`）・`lt1` 要素が見つからない・色情報が想定外の構造の場合は解決不能として `None`
    を返す（呼び出し側は安全側＝占有として扱う。実ファイル起点では通常発生しないが、保存前の
    ワークブックを直接渡す経路や壊れたテーマ part では起こりうる＝受容記録）。
    """
    theme_bytes = getattr(wb, "loaded_theme", None) if wb is not None else None
    if not theme_bytes:
        return None
    try:
        root = ET.fromstring(theme_bytes)
    except ET.ParseError:
        return None
    lt1 = root.find(f".//{_DRAWINGML_NS}lt1")
    if lt1 is None:
        return None
    srgb = lt1.find(f"{_DRAWINGML_NS}srgbClr")
    if srgb is not None:
        return srgb.get("val")
    sys_clr = lt1.find(f"{_DRAWINGML_NS}sysClr")
    if sys_clr is not None:
        return sys_clr.get("lastClr")
    return None


class ColorResolver:
    """1ワークブック分の色解決コンテキスト。複数シートを横断して走査する呼び出し側（`_build_xlsx_ir`
    等）はワークブックあたり1つだけ構築してシートごとの `filled_cells()` 呼び出しへ使い回すこと
    （`filled_cells()` の `resolver` 引数省略時は自前で1つ構築するが、それは**呼び出しごと**に
    新規構築される＝複数シートに渡って使い回したい場合は明示的に渡す必要がある）。

    `wb.loaded_theme` の XML パース（`_resolve_theme_lt1_rgb`）は同一ブック内であれば何度呼んでも
    結果が変わらないため、初回だけ実行してキャッシュする（テーマ色のセルが多数あっても、また
    シートが複数あっても decode/パースは1ワークブックにつき1回きり）。インデックスパレット
    （`_resolve_indexed_rgb`）は `wb._colors` の単純な添字参照でコストが無視できるためキャッシュしない。
    """
    def __init__(self, wb):
        self._wb = wb
        self._lt1_rgb: str | None = None
        self._lt1_resolved = False

    def indexed_rgb(self, indexed: int) -> str | None:
        return _resolve_indexed_rgb(self._wb, indexed)

    def theme_lt1_rgb(self) -> str | None:
        if not self._lt1_resolved:
            self._lt1_rgb = _resolve_theme_lt1_rgb(self._wb)
            self._lt1_resolved = True
        return self._lt1_rgb


def _is_white_color(color, resolver: "ColorResolver") -> bool:
    """openpyxl の `Color`（`None` も許容）が「白／自動＝背景色として占有扱いしない色」とみなせるか。
    `resolver`（`ColorResolver`）はインデックスパレット・テーマの実際の値を引くのに使う。

    判定順序（`auto` は tint 判定より**先に**白とみなす＝`auto` に tint が付いていても無視する。
    Excel の「自動」色は具体的な基準色を持たないため tint で暗色化する対象がそもそも無い）:
    1. 自動色（`auto`）→ 無条件に白。
    2. tint が負（Excel の tint モデルで暗色化を意味する）→ 白ではない（RGB/indexed/テーマの
       いずれの基準色にも一様に適用する。暗くグレー化した見た目はもはや「白」ではないため）。
    3. RGB が白（`_is_white_hex` で厳密検証・アルファ成分は無視）。
    4. インデックスパレット（`indexed`）が `resolver` 固有のパレットで白に解決する。
    5. テーマの背景1（`theme == 0`）が `resolver` 固有のテーマ定義で白に解決する。

    `color` が `None`、インデックス/テーマが `resolver` から解決できない（`None` を返す）、値の形式が
    不正（`_is_white_hex` 参照）、または上記のどれにも一致しない場合は白と断定せず False を返す
    （`filled_cells()` 全体の方針＝取りこぼし（表の分裂）を過検出より避ける側へ倒す：不明な色・
    白と確認できない色は占有側に倒す）。
    """
    if color is None:
        return False
    # openpyxl の `Color.auto`/`.theme`/`.indexed` は未設定だと（`None` ではなく）Typed 記述子オブジェクト
    # 自身を返す実装上の癖があり、真偽値としての単純な truthy 判定（`if getattr(...):`）や `isinstance`
    # を使わない同値比較は誤判定しうる（実測済みの罠）。`is True`/`isinstance(..., int)`/`== 値` で、
    # 未設定時に記述子オブジェクトが紛れ込んでも誤判定しないようにする（`tint` は未設定時も実測で
    # 通常の `float` 既定値 `0.0` を返すため同じ罠は無い）。
    if getattr(color, "auto", None) is True:
        return True
    tint = getattr(color, "tint", 0.0)
    if isinstance(tint, (int, float)) and tint < 0:
        return False
    indexed = getattr(color, "indexed", None)
    if isinstance(indexed, int):
        return _is_white_hex(resolver.indexed_rgb(indexed))
    if getattr(color, "theme", None) == 0:
        return _is_white_hex(resolver.theme_lt1_rgb())
    return _is_white_hex(getattr(color, "rgb", None))


def _fill_is_colored(fill, resolver: "ColorResolver") -> bool:
    """`fill`（openpyxl の `PatternFill`/`GradientFill`、またはそれ以外の未知の型）が、R1 でいう
    「値が無くても占有とみなすべき背景色」を持つか。`resolver` は `_is_white_color` の色解決に使う。

    - 塗りなし（`patternType` が `None`/`"none"`）は対象外。単色塗り（`patternType == "solid"`）は
      前景色（`fgColor`）が白なら対象外とする（`_is_white_color` 参照・既定の白背景と区別が付かない
      過検出を避ける）。単色以外のパターン塗り（縞模様等）は色の組合せまで判別せず一律で占有対象とする。
    - `GradientFill` は `patternType` 属性を持たない（`PatternFill` 専用の属性）ため `getattr` で
      安全に判定する。各ストップの色がすべて白でない限り占有対象とする。
    - 条件付き書式（セル自体の `fill` ではなくシート単位のルールで見た目だけ変わる着色）は対象外
      （本関数はセル自身のスタイル定義しか見ない・静的な OOXML 直パースの方針＝docs/11-Office変換.md
      と同じ「LLM/実行時評価はしない」原則に沿う）。
    """
    if fill is None:
        return False
    pattern_type = getattr(fill, "patternType", None)
    if pattern_type not in (None, "none"):
        if pattern_type == "solid" and _is_white_color(getattr(fill, "fgColor", None), resolver):
            return False
        return True
    stops = getattr(fill, "stop", None)                # GradientFill のみが持つ（PatternFill には無い）
    if stops:
        return any(not _is_white_color(getattr(stop, "color", None), resolver) for stop in stops)
    return False


def filled_cells(ws, cap_rows: int, cap_cols: int, *, resolver: "ColorResolver | None" = None
                 ) -> set[tuple[int, int]]:
    """背景色（単色/縞模様パターン塗り・グラデーション塗り）が設定されているセルの1-based座標集合
    （`regions()` の `filled` 引数用・`cap_rows`×`cap_cols` 以内・色の判定は `_fill_is_colored` 参照）。
    白判定はセルが属するワークブック（`ws.parent`）固有のインデックスパレット・テーマ定義を、
    `resolver`（`ColorResolver`）を使って行う。

    `resolver` を省略すると `ws.parent` から自前で1つ構築する（単一シートだけを走査する場合はこれで
    十分）。**複数シートを持つワークブックを走査する呼び出し側**（`_build_xlsx_ir` 等）は、ワークブック
    レベルの入口で `ColorResolver(wb)` を1つだけ構築し、シートごとの呼び出しに明示的に渡すこと
    （省略した場合はシートの数だけ新規構築され、`wb.loaded_theme` の XML パースがシート数だけ
    繰り返されてしまう＝実際に踏んだ非効率）。

    結合範囲の非anchorセル（openpyxl の `MergedCell`）は、保存・再読込を経ても常に既定（塗りなし）の
    `fill` しか持たない（openpyxl は anchor 以外へスタイルを伝播しない＝実測済みの挙動）。そのため
    非anchorセルは自身の `fill` ではなく、結合範囲の anchor セルの `fill` を見る（Excel の見た目上、
    結合範囲全体が anchor の書式で塗られることに対応）。

    `anchor_of` の座標展開は `cap_rows`/`cap_cols` でクリップしてから行う（`merged_map()` と同じ
    `_clip_merge_enumeration_bounds` を使う）: 宣言上 Excel の絶対上限まで指定できる結合範囲を
    そのまま座標展開すると座標数が桁違いに膨れ上がるため。
    """
    from openpyxl.cell.cell import MergedCell

    if resolver is None:
        resolver = ColorResolver(ws.parent)
    anchor_of: dict[tuple[int, int], tuple[int, int]] = {}
    for mc in ws.merged_cells.ranges:
        anchor = (mc.min_row, mc.min_col)
        clip_max_row, clip_max_col = _clip_merge_enumeration_bounds(
            mc.min_row, mc.min_col, mc.max_row, mc.max_col, cap_rows, cap_cols)
        for r in range(mc.min_row, clip_max_row + 1):
            for c in range(mc.min_col, clip_max_col + 1):
                anchor_of[(r, c)] = anchor

    max_row = min(ws.max_row or 1, cap_rows)
    max_col = min(ws.max_column or 1, cap_cols)
    out: set[tuple[int, int]] = set()
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            fill = cell.fill
            if isinstance(cell, MergedCell):
                anchor = anchor_of.get((cell.row, cell.column))
                if anchor is not None:
                    fill = ws.cell(row=anchor[0], column=anchor[1]).fill
            if _fill_is_colored(fill, resolver):
                out.add((cell.row, cell.column))
    return out


def hidden_rows(ws) -> list[int]:
    """非表示行の1-based行番号（昇順）。`ws.row_dimensions` は明示的に設定された行だけを持つ辞書
    （openpyxl は既定値の行は書き出さない）ため、`hidden` が真の行だけを抽出すれば足りる。"""
    return sorted(r for r, dim in ws.row_dimensions.items() if dim.hidden)


def hidden_cols(ws) -> list[str]:
    """非表示列の列文字（例 `"D"`・列インデックス昇順）。`hidden_rows` の列版。

    グループ化された非表示（`ColumnDimension` が `min..max` の範囲を1エントリで表す・例 `B:D`）は
    範囲を**全列へ展開**する（RV Low #4: 辞書 key（先頭列）だけ返すと `["B"]` になり C/D が漏れる）。
    """
    from openpyxl.utils import column_index_from_string, get_column_letter
    idxs: set[int] = set()
    for key, dim in ws.column_dimensions.items():
        if not dim.hidden:
            continue
        lo = dim.min or column_index_from_string(key)
        hi = dim.max or lo
        idxs.update(range(lo, hi + 1))
    return [get_column_letter(i) for i in sorted(idxs)]


def formulas(ws_formula, ws_values) -> list[dict]:
    """`ws_formula`（`data_only=False`）内の数式セル（`=` で始まる値）を走査し、`ws_values`
    （`data_only=True`）の同座標を突き合わせて `{"cell", "row", "column", "formula", "has_cached"}` の
    リストを返す（`(row, column)` 昇順）。

    `has_cached`＝`ws_values` 側の同座標の値が `None` でないか（未計算式／キャッシュ破棄済みは `None` に
    なる openpyxl の挙動を利用）。行・列上限と`DEFAULT_CAP_CELLS`の総セル予算（モジュール属性・呼び出し時に
    読む＝テストで`monkeypatch.setattr(excel, "DEFAULT_CAP_ROWS", …)`すれば効く）で走査範囲を
    `min(ws.max_row, effective_cap+1)`に有界化する（`regions()`と同じ理由＝疎な大きいシートでの暴走防止）。
    """
    max_row = min(ws_formula.max_row or 1, effective_cap_rows(ws_formula.max_column) + 1)
    max_col = min(ws_formula.max_column or 1, DEFAULT_CAP_COLS + 1)
    out: list[dict] = []
    for row in ws_formula.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                cached = ws_values.cell(row=cell.row, column=cell.column).value
                out.append({"cell": cell.coordinate, "row": cell.row, "column": cell.column,
                           "formula": v, "has_cached": cached is not None})
    out.sort(key=lambda e: (e["row"], e["column"]))
    return out


def defined_names(wb) -> list[dict]:
    """名前付き範囲（ブック全体＝global／シート限定＝local）を
    `{"name", "value": <参照先文字列>, "scope": "workbook" | <シート名>}` のリストで返す。

    決定的な順序: まずブック全体スコープ（`wb.defined_names`）を名前昇順、続けてシート限定スコープを
    **ブック内シート順**で辿り、各シート内は名前昇順（`defined_names(wb)` は `ws.defined_names` も
    合わせて辿るため、シート単位の関数ではなく wb を受け取る設計にしている）。
    """
    out: list[dict] = []
    for name in sorted(wb.defined_names):
        dn = wb.defined_names[name]
        out.append({"name": name, "value": dn.value or "", "scope": "workbook"})
    for ws in wb.worksheets:
        for name in sorted(ws.defined_names):
            dn = ws.defined_names[name]
            out.append({"name": name, "value": dn.value or "", "scope": ws.title})
    return out


def cell_hyperlinks(ws) -> list[dict]:
    """`ws` 内のセル単位ハイパーリンクを `{"cell", "row", "column", "target", "text"}` のリストで返す
    （`(row, column)` 昇順）。

    **`ws` は `data_only=True` でロードした値ワークシートを渡すこと**（`text`＝セルの表示値＝キャッシュ済み
    計算値を使うため。数式ワークシートを渡すと数式文字列が `text` に入ってしまう）。

    `target` の解決順は `word.py hyperlinks()` と同じ考え方: 外部 URL（`Hyperlink.target`）優先、無ければ
    文書内参照（`Hyperlink.location`・`"#" + location`）。どちらも無ければそのセルは結果から省略する
    （遷移先の無いリンクはノイズという `word.py` と同じ判断）。走査範囲は `formulas()` と同じ cap 契約。
    """
    max_row = min(ws.max_row or 1, effective_cap_rows(ws.max_column) + 1)
    max_col = min(ws.max_column or 1, DEFAULT_CAP_COLS + 1)
    out: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            h = cell.hyperlink
            if h is None:
                continue
            target = h.target or (("#" + h.location) if h.location else None)
            if not target:
                continue
            v = cell.value
            out.append({"cell": cell.coordinate, "row": cell.row, "column": cell.column,
                       "target": target, "text": "" if v is None else str(v)})
    out.sort(key=lambda e: (e["row"], e["column"]))
    return out


def cell_comments(ws) -> list[dict]:
    """`ws` 内のセルコメントを `{"cell", "row", "column", "text", "author"}` のリストで返す
    （`(row, column)` 昇順）。本文が空のコメントは出さない（`word.py comments()` と同じ方針）。走査範囲は
    `formulas()` と同じ cap 契約。
    """
    max_row = min(ws.max_row or 1, effective_cap_rows(ws.max_column) + 1)
    max_col = min(ws.max_column or 1, DEFAULT_CAP_COLS + 1)
    out: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            c = cell.comment
            if c is None:
                continue
            text = (c.text or "").strip()
            if not text:
                continue
            out.append({"cell": cell.coordinate, "row": cell.row, "column": cell.column,
                       "text": text, "author": c.author or ""})
    out.sort(key=lambda e: (e["row"], e["column"]))
    return out


def strike_cells(ws) -> list[dict]:
    """取り消し線（`cell.font.strike`＝OOXML の `s:font/strike`）が設定されたセルを
    `{"cell", "row", "column", "text"}` のリストで返す（`(row, column)` 昇順）。

    値が無いセル（`cell.value is None`）は取り消し線があっても出さない（意味を持たない書式だけの
    水増しを避ける・`cell_comments` と同じ方針）。結合範囲の非anchorセル（`MergedCell`）は
    値を持たないため（`filled_cells` docstring 参照＝openpyxl は非anchorへスタイルを伝播しない）
    この `value is None` チェックだけで自然に除外される（`filled_cells` のような anchor 逆引きは不要）。
    走査範囲は `formulas()` と同じ cap 契約。
    """
    max_row = min(ws.max_row or 1, effective_cap_rows(ws.max_column) + 1)
    max_col = min(ws.max_column or 1, DEFAULT_CAP_COLS + 1)
    out: list[dict] = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            font = cell.font
            if font is None or not font.strike:
                continue
            v = cell.value
            if v is None:
                continue
            out.append({"cell": cell.coordinate, "row": cell.row, "column": cell.column, "text": str(v)})
    out.sort(key=lambda e: (e["row"], e["column"]))
    return out


_EXTERNAL_LINK_RELS_RE = re.compile(r"xl/externalLinks/_rels/[^/]+\.rels")


def external_link_targets(zf: zipfile.ZipFile) -> list[str]:
    """zip 内 `xl/externalLinks/_rels/*.rels` の `Target` 属性値一覧（ソート済み・重複も保持）。

    外部ブック参照（他ファイルへのリンク）の存在を示す来歴情報。パート欠落（外部リンク無し）は空リスト。
    壊れた rels パートは無視して続行する（fail-safe＝`word.py`/`powerpoint.py` と同じ方針）。
    """
    out: list[str] = []
    names = sorted(n for n in zf.namelist() if _EXTERNAL_LINK_RELS_RE.fullmatch(n))
    for n in names:
        try:
            root = ET.fromstring(zf.read(n))
        except (KeyError, ET.ParseError):
            continue
        for r in root.iter(f"{_RELS}Relationship"):
            target = r.get("Target")
            if target:
                out.append(target)
    return sorted(out)


def _resolve_rel_target(part: str, target: str) -> str:
    """rels の `Target` 属性値を zip 内の絶対パートパスへ解決する（`part` は Target を持つ側のパート
    自身のパス）。先頭 `/` は zip ルートからの絶対パス、それ以外は `part` のディレクトリからの相対
    パスという OOXML パッケージ関係の解決規則どおり（`evidence_spike._relationships` と同じロジック・
    どちらも `xl/worksheets/_rels/sheet1.xml.rels` 等の実ファイルで先頭 `/` 形式が普通に使われるため
    相対解決だけでは足りない）。"""
    from posixpath import dirname, join, normpath
    if target.startswith("/"):
        return normpath(target.lstrip("/"))
    return normpath(join(dirname(part), target))


def _load_rels(zf: zipfile.ZipFile, part: str) -> dict[str, str]:
    """`part` に対応する `_rels/*.rels` から `{Id: 解決済み絶対パートパス}` を返す（外部リンク
    ＝`TargetMode="External"` は zip 内パートではないため除外）。パート・rels 欠落/破損は空 dict
    （fail-safe＝`external_link_targets` と同じ方針）。"""
    from posixpath import basename, dirname, join
    rels_name = join(dirname(part), "_rels", basename(part) + ".rels")
    try:
        root = ET.fromstring(zf.read(rels_name))
    except (KeyError, ET.ParseError):
        return {}
    out: dict[str, str] = {}
    for rel in root.iter(f"{_RELS}Relationship"):
        rel_id, target = rel.get("Id"), rel.get("Target")
        if rel_id and target and rel.get("TargetMode") != "External":
            out[rel_id] = _resolve_rel_target(part, target)
    return out


def picture_counts_by_sheet(zf: zipfile.ZipFile) -> dict[str, int]:
    """ブック内の各シート名 → そのシートの drawing part に含まれる画像（`xdr:pic`）の枚数。

    画像が0枚のシートはキー自体を持たない（人間向けMDの注記は「画像がある」場合だけ出すため、
    消費側は `.get(name)` で判定できれば十分）。壊れた/欠落したパート（workbook.xml・シート
    パート・drawing パートのいずれか）はそのシート分だけ黙って除外する（fail-safe＝本モジュールの
    他の関数と同じ方針・IR 構築自体を失敗させない）。図形（`xdr:sp`）やチャート・SmartArt は数えない
    （「画像」に限定＝人間向け注記の文言と一致させる）。
    """
    try:
        wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    except (KeyError, ET.ParseError):
        return {}
    wb_rels = _load_rels(zf, "xl/workbook.xml")
    out: dict[str, int] = {}
    for sheet in wb_root.findall(f"{_SML}sheets/{_SML}sheet"):
        name, rid = sheet.get("name"), sheet.get(f"{_R}id")
        sheet_part = wb_rels.get(rid) if rid else None
        if not name or not sheet_part:
            continue
        try:
            sheet_root = ET.fromstring(zf.read(sheet_part))
        except (KeyError, ET.ParseError):
            continue
        drawing_ref = sheet_root.find(f"{_SML}drawing")
        if drawing_ref is None:
            continue
        drawing_rid = drawing_ref.get(f"{_R}id")
        drawing_part = _load_rels(zf, sheet_part).get(drawing_rid) if drawing_rid else None
        if not drawing_part:
            continue
        try:
            drawing_root = ET.fromstring(zf.read(drawing_part))
        except (KeyError, ET.ParseError):
            continue
        # `xdr:pic` は必ず anchor（`oneCellAnchor`/`twoCellAnchor`）の子孫（グループ化図形
        # `xdr:grpSp` の中はさらに孫以下）で、drawing part 直下の子ではない——`.iter()` で深さに
        # 関係なく木全体から数える（anchor/group の階層を自前でたどる再帰は不要）。
        count = sum(1 for _ in drawing_root.iter(f"{_XDR}pic"))
        if count:
            out[name] = count
    return out
