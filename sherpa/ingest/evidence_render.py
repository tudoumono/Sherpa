"""Evidence IRとContext IRから内容中心のRAG用Markdown・構造chunkを同時生成する。

値や関係をLLMで言い換えず、`〈項目〉: 「値」` の key-value 形式で決定的に直列化する（「である」文の
自然文テンプレートは使わない＝値の境界を明示し断定文特有の誤り検出不能を避けるため）。検索文章は
文書・論理領域・業務record key・field名と値を中心とし、セル番地やbboxはcitation metadataへ分離する。
recordが文字数上限に収まる限り1chunkとし、超える場合だけ同じrecord keyを持つfield groupへ分割する。

Markdownは人間向け正本ではなくAI/RAG向けの搬送表示である。画像は実assetへの相対Markdown記法と、
位置・未解釈状態の自然文を併記する。

**RAG正本はMarkdown側（D1・`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md`§8.1）**:
各chunkの本文直前に決定的なアンカー行 `<!-- chunk:{chunk_id} -->` を1行だけ出す（`_markdown`）。
索引時（`es_index._validate_rag_chunks`）はこのアンカーでMarkdownを分割し、その本文をES索引対象に
する。JSONL（`{rel}.rag_chunks.jsonl`）はもう検索本文（旧`search_text`）を持たない——citation・
locator・隣接キー（B1）等の証跡専業サイドカーへ降格した。chunk_idはJSONLとMarkdownアンカーの
双方を`_chunk_id_for_record`という同じ式で計算するため、生成時点で必ず1:1になる。
"""
from __future__ import annotations

import hashlib
import json
import os
import re
import uuid
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from . import ai_observation, context_ir, evidence_ir, mermaid_render


RAG_RENDERER_VERSION = "evidence-rag-renderer-v1alpha10"
RAG_CHUNKER_VERSION = "evidence-rag-chunker-v1alpha10"
MAX_GROUP_CHARS = 1800
MAX_TEXT_SPAN_CHARS = 1200
_BREAK_CHARS = frozenset({"\n", "。", "！", "？", ";", "；"})
_CONTAINER_TYPES = frozenset({"sheet", "slide", "page", "table", "group"})
_EXCEL_VALUE_KEYS = (
    "raw_value", "typed_value", "formula", "cached_value", "calculation_status", "style_id",
    "number_format", "display_value", "display_source", "display_status", "display_reason",
)


@dataclass(frozen=True)
class RenderedEvidence:
    markdown: str
    chunks: list[dict]
    coverage_summary: dict[str, int]


def _identifier_metadata_fields(batch: context_ir.IdentifierMentionBatch) -> dict[str, Any]:
    """識別子objectと完全性envelopeを同じpieceへ載せる。"""
    return {
        "identifier_mentions": [asdict(mention) for mention in batch.mentions],
        "identifier_metadata_complete": batch.complete,
        "identifier_mention_count": batch.mention_count,
        "identifier_mention_overflow_count": batch.overflow_count,
        "identifier_mention_limit": context_ir.IDENTIFIER_MAX_MENTIONS_PER_CHUNK,
    }


def _extract_identifier_metadata(
    value: Any,
    *,
    field_label: str,
    evidence_id: str,
    locator: evidence_ir.Locator | dict[str, Any],
    text_offset: int = 0,
    allow_record_identity: bool = False,
    match_start: int | None = None,
    match_end: int | None = None,
) -> dict[str, Any]:
    return _identifier_metadata_fields(context_ir.identifier_mentions_with_metadata(
        value,
        field_label=field_label,
        evidence_id=evidence_id,
        locator=locator,
        text_offset=text_offset,
        allow_record_identity=allow_record_identity,
        match_start=match_start,
        match_end=match_end,
    ))


def _stable_id(prefix: str, *parts: Any) -> str:
    raw = json.dumps(parts, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return f"{prefix}:" + hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]


def _value_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _excel_value_metadata(cell: evidence_ir.EvidenceElement) -> dict[str, Any] | None:
    if not any(key in cell.extension for key in _EXCEL_VALUE_KEYS):
        return None
    return {key: cell.extension.get(key) for key in _EXCEL_VALUE_KEYS}


def _cell_render_text(cell: evidence_ir.EvidenceElement) -> str:
    metadata = _excel_value_metadata(cell)
    if metadata is not None and metadata.get("display_status") in {"rendered", "supported"}:
        display = metadata.get("display_value")
        if display is not None:
            return _value_text(display)
    return _value_text(cell.value)


def _excel_value_note(metadata: dict[str, Any] | None) -> str:
    """material な Excel メタデータを改行区切りの key-value 行として返す（呼び出し側は `+=` 直結）。

    非空時は先頭に `\n` を含める（呼び出し側が改行を挟まないため）。material でなければ `""`
    （改行も付けない＝余分な空行を作らない）。
    """
    if metadata is None or not _excel_value_is_material(metadata):
        return ""
    raw = _value_text(metadata.get("raw_value"))
    display = _value_text(metadata.get("display_value"))
    number_format = _value_text(metadata.get("number_format"))
    formula = _value_text(metadata.get("formula"))
    status = _value_text(metadata.get("display_status"))
    reason = _value_text(metadata.get("display_reason"))
    lines = [f"Excel原値: 「{raw}」", f"Excel書式: 「{number_format}」", f"Excel表示状態: {status}"]
    if display:
        lines.append(f"Excel表示値: 「{display}」")
    if formula:
        lines.append(f"Excel数式: 「{formula}」")
    if reason:
        lines.append(f"Excel表示理由: {reason}")
    return "\n" + "\n".join(lines)


def _excel_value_is_material(metadata: dict[str, Any]) -> bool:
    raw = _value_text(metadata.get("raw_value"))
    display = _value_text(metadata.get("display_value"))
    number_format = _value_text(metadata.get("number_format"))
    return bool(
        metadata.get("formula") is not None
        or metadata.get("display_status") == "unsupported"
        or number_format.casefold() != "general"
        or (display and display != raw)
    )


# L4a（可視性・廃止表現の表示側・提案書§2.3）: L3/S1/Z が Evidence IR へ載せた
# `extension["visibility_reason"]` の平文ラベル対応表。ここに無い reason（None・未知値）は
# `_field_piece`/`_element_piece` 側の既存フォールバック（生の `visibility`/`lifecycle` 値）に譲る
# ＝この表に載っている reason だけ、生の enum 値の代わりにこのラベルを出す。
_VISIBILITY_REASON_LABELS: dict[str, str] = {
    "occluded_by_picture": "画像に覆われている",
    "occluded_by_shape": "図形に覆われている",
    "occluded": "図形に覆われている",
    "hidden_sheet": "シートが非表示",
    "very_hidden": "シートが完全非表示",
    "hidden_row": "行が非表示",
    "hidden_column": "列が非表示",
    "hidden_run": "非表示文字",
    "hidden_slide": "スライドが非表示",
    "hidden_slide_inherited": "非表示スライド内の要素",
    "off_slide": "スライド範囲外",
}


def _occlusion_kv_lines(extension: dict[str, Any]) -> list[str]:
    """`extension`（cell/element いずれも同じキー語彙）から可視性・廃止表現のkey-value行を作る。

    意味の断定はしない（`occluded_by`/`covered_by_text` の前面テキストをそのまま写すだけ・
    `docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §2.3）。生の reason 文字列
    （`occluded_by_picture` 等）は `_VISIBILITY_REASON_LABELS` で平文へ写像し、rag.md には出さない。
    """
    lines: list[str] = []
    reason = extension.get("visibility_reason")
    label = _VISIBILITY_REASON_LABELS.get(reason) if isinstance(reason, str) else None
    if label:
        lines.append(f"可視性: 「{label}」")
    overlap_text = None
    for key in ("occluded_by", "covered_by_text"):
        candidate = extension.get(key)
        if isinstance(candidate, dict) and isinstance(candidate.get("text"), str) and candidate["text"]:
            overlap_text = candidate["text"]
            break
    if overlap_text:
        lines.append(f"重なり: 「{overlap_text}」")
    if reason == "strike":
        lines.append("取り消し線: 「あり」")
    floating_anchors = extension.get("floating_anchors")
    if isinstance(floating_anchors, list):
        for anchor in floating_anchors:
            if not isinstance(anchor, dict):
                continue
            value = anchor.get("text") or anchor.get("name")
            if not value:
                continue
            anchor_label = "背面図形" if anchor.get("behind_doc") else "前面図形"
            lines.append(f"{anchor_label}: 「{value}」")
    return lines


def _split_exact(text: str) -> list[tuple[int, int, str]]:
    """長大な1セル/1object本文を文字offsetつきで欠落なく分割する。"""
    if len(text) <= MAX_TEXT_SPAN_CHARS:
        return [(0, len(text), text)]
    spans: list[tuple[int, int, str]] = []
    start = 0
    while start < len(text):
        hard_end = min(len(text), start + MAX_TEXT_SPAN_CHARS)
        end = hard_end
        if hard_end < len(text):
            floor = start + (MAX_TEXT_SPAN_CHARS * 3 // 5)
            for pos in range(hard_end - 1, floor - 1, -1):
                if text[pos] in _BREAK_CHARS:
                    end = pos + 1
                    break
        spans.append((start, end, text[start:end]))
        start = end
    return spans


def _section(locator: evidence_ir.Locator) -> list[str]:
    if locator.sheet:
        return [f"シート「{locator.sheet}」"]
    if locator.slide is not None:
        return [f"スライド{locator.slide}"]
    if locator.page is not None:
        return [f"ページ{locator.page}"]
    return [f"文書part「{locator.part}」"]


def _compact_locator(locator: evidence_ir.Locator) -> dict:
    """citationに必要な原本位置だけを残し、cellごとの巨大extension複製を避ける。"""
    out: dict[str, Any] = {}
    if not locator.sheet:
        out["part"] = locator.part
    for key in ("page", "slide", "sheet", "cell_range", "object_id", "bbox"):
        value = getattr(locator, key)
        if value is not None:
            out[key] = value
    source_map = locator.extension.get("document_ir_source_map")
    if isinstance(source_map, dict):
        kept = {
            key: source_map[key]
            for key in (
                "paragraph_index", "table_index", "cell_paragraph_index", "level", "note_id", "comment_id",
                "author", "date", "target",
            )
            if key in source_map
        }
        if kept:
            out["source_map"] = kept
    grid = {
        key: locator.extension[key]
        for key in ("row", "column", "row_span", "column_span")
        if key in locator.extension
    }
    if grid:
        out["grid"] = grid
    return out


def _compact_cell_locator(locator: evidence_ir.Locator) -> dict:
    """cell citationはchunk-levelのsheet/region文脈を参照し、セル範囲だけを繰り返す。"""
    return {"cell_range": locator.cell_range} if locator.cell_range else _compact_locator(locator)


def _cell_row_column(element: evidence_ir.EvidenceElement) -> tuple[int | None, int | None]:
    extension = element.locator.extension
    row = extension.get("row")
    column = extension.get("column")
    return (row if isinstance(row, int) else None, column if isinstance(column, int) else None)


def _left_label_map(
    cells: list[evidence_ir.EvidenceElement],
) -> dict[str, evidence_ir.EvidenceElement]:
    candidates: dict[tuple[int, int], list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for cell in cells:
        row, column = _cell_row_column(cell)
        row_span = cell.locator.extension.get("row_span", 1)
        row_span = row_span if isinstance(row_span, int) and row_span > 0 else 1
        if row is None or column is None or not _value_text(cell.value):
            continue
        if column + 1 != 2 and row_span <= 1:
            continue
        for covered_row in range(row, row + row_span):
            candidates[(covered_row, column + 1)].append(cell)
    result: dict[str, evidence_ir.EvidenceElement] = {}
    for target in cells:
        row, column = _cell_row_column(target)
        if row is None or column is None:
            continue
        labels = candidates.get((row, column), [])
        if labels:
            result[target.element_id] = min(labels, key=lambda item: _cell_row_column(item)[1] or 0)
    return result


def _table_layout(
    ir: evidence_ir.EvidenceIR,
    table_id: str,
    cells: list[evidence_ir.EvidenceElement],
) -> tuple[dict[int, str], dict[str, list[dict]]]:
    """縦結合ラベルによる論理blockと、連続/分離の配置関係を決定的に作る。"""
    rows = sorted({row for cell in cells if (row := _cell_row_column(cell)[0]) is not None})
    block_by_row: dict[int, str] = {}
    merged_anchor_by_row: dict[int, str | None] = {}
    anchors_by_row: dict[int, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for cell in cells:
        cell_row, column = _cell_row_column(cell)
        row_span = cell.locator.extension.get("row_span", 1)
        row_span = row_span if isinstance(row_span, int) and row_span > 0 else 1
        if cell_row is None or column is None or row_span <= 1:
            continue
        for covered_row in range(cell_row, cell_row + row_span):
            anchors_by_row[covered_row].append(cell)
    for row in rows:
        anchors = anchors_by_row.get(row, [])
        anchor = min(anchors, key=lambda item: _cell_row_column(item)[1] or 0) if anchors else None
        merged_anchor_by_row[row] = anchor.element_id if anchor else None
        block_by_row[row] = _stable_id(
            "logical-block", ir.source.content_hash, table_id, anchor.element_id if anchor else f"row:{row}")

    relations_by_evidence: dict[str, list[dict]] = defaultdict(list)
    left_labels = _left_label_map(cells)
    by_column: dict[int, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for cell in cells:
        row, column = _cell_row_column(cell)
        if row is not None and column is not None and _value_text(cell.value):
            by_column[column].append(cell)
    for column_cells in by_column.values():
        ordered = sorted(column_cells, key=lambda item: (_cell_row_column(item)[0] or 0, item.order))
        for source, target in zip(ordered, ordered[1:], strict=False):
            source_row, _ = _cell_row_column(source)
            target_row, _ = _cell_row_column(target)
            if source_row is None or target_row != source_row + 1:
                continue
            relation_type = None
            basis: list[str] = []
            if (merged_anchor_by_row[source_row] is not None
                    and merged_anchor_by_row[source_row] == merged_anchor_by_row[target_row]):
                relation_type = "continues"
                basis = ["same_table", "same_column", "shared_vertical_merged_label"]
            else:
                source_label = left_labels.get(source.element_id)
                target_label = left_labels.get(target.element_id)
                if source_label is not None and target_label is not None and source_label.element_id != target_label.element_id:
                    relation_type = "separate"
                    basis = ["same_table", "same_column", "distinct_left_labels"]
            if relation_type is None:
                continue
            relation = {
                "type": relation_type,
                "source_id": source.element_id,
                "target_id": target.element_id,
                "source_locator": _compact_locator(source.locator),
                "target_locator": _compact_locator(target.locator),
                "confidence": 1.0,
                "basis": basis,
            }
            relations_by_evidence[source.element_id].append(relation)
            relations_by_evidence[target.element_id].append(relation)
    return block_by_row, dict(relations_by_evidence)


def _field_piece(
    cell: evidence_ir.EvidenceElement,
    label: str,
    span_index: int,
    span_count: int,
    start: int,
    end: int,
    exact: str,
    layout_relations: list[dict],
    *,
    allow_record_identity: bool = False,
    identifier_metadata: dict[str, Any] | None = None,
) -> dict:
    row, column = _cell_row_column(cell)
    field_label = label or "記載値"
    excel_metadata = _excel_value_metadata(cell)
    if span_count > 1:
        semantic = f"{field_label}（{span_index + 1}/{span_count}）:\n{exact}"
    else:
        semantic = f"{field_label}: 「{exact}」"
    for relation in layout_relations:
        if relation["type"] == "continues":
            semantic += "（同じ縦結合ラベル内の前後記載と連続）"
        elif relation["type"] == "separate":
            semantic += "（隣接行と左ラベルが異なる別項目）"
    occlusion_lines = _occlusion_kv_lines(cell.extension)
    if span_index == 0:
        semantic += _excel_value_note(excel_metadata)
        if occlusion_lines:
            semantic += "\n" + "\n".join(occlusion_lines)
    exact_lines = [f"{field_label}: {exact}"]
    if span_index == 0 and excel_metadata is not None and _excel_value_is_material(excel_metadata):
        exact_lines.extend([
            f"{field_label}のExcel原値: {_value_text(excel_metadata.get('raw_value'))}",
            f"{field_label}のExcel書式: {_value_text(excel_metadata.get('number_format'))}",
            f"{field_label}のExcel表示状態: {_value_text(excel_metadata.get('display_status'))}",
        ])
        if excel_metadata.get("formula") is not None:
            exact_lines.append(f"{field_label}のExcel数式: {_value_text(excel_metadata.get('formula'))}")
    if span_index == 0 and occlusion_lines:
        exact_lines.extend(occlusion_lines)
    if identifier_metadata is None:
        identifier_metadata = _extract_identifier_metadata(
            exact,
            field_label=label,
            evidence_id=cell.element_id,
            locator=cell.locator,
            text_offset=start,
            allow_record_identity=allow_record_identity,
        )
    piece = {
        "semantic": semantic,
        "markdown": semantic,
        "exact": "\n".join(exact_lines),
        "evidence_id": cell.element_id,
        "locator": _compact_cell_locator(cell.locator),
        "header_path": label.split(" > ") if label else [],
        "row": row,
        "column": column,
        "text_span": (
            {"start": start, "end": end, "index": span_index + 1, "count": span_count}
            if span_count > 1 else None
        ),
        "relations": layout_relations,
        **identifier_metadata,
    }
    if excel_metadata is not None:
        piece["object_metadata"] = {"excel_value": excel_metadata}
    return piece


def _pack_row(pieces: list[dict], context_prefix: str) -> list[list[dict]]:
    groups: list[list[dict]] = []
    current: list[dict] = []
    current_chars = len(context_prefix)
    for piece in pieces:
        piece_chars = len(piece["semantic"])
        if current and current_chars + 1 + piece_chars > MAX_GROUP_CHARS:
            groups.append(current)
            current = []
            current_chars = len(context_prefix)
        current.append(piece)
        current_chars += piece_chars + (1 if current_chars else 0)
    if current:
        groups.append(current)
    return groups


def _base_record(
    *,
    source_hash: str,
    kind: str,
    logical_record_id: str,
    group_index: int,
    group_count: int,
    section_path: list[str],
    pieces: list[dict],
    sort_key: tuple,
    context_prefix: str = "",
    parent_id: str | None = None,
    source_name: str | None = None,
    document_context: dict | None = None,
    region_context: dict | None = None,
    record_keys: list[dict] | None = None,
    related_record_ids: list[str] | None = None,
    context_confidence: float | None = None,
) -> dict:
    evidence_ids = list(dict.fromkeys(
        evidence_id
        for piece in pieces
        for evidence_id in (
            piece.get("evidence_id"),
            *(citation["evidence_id"] for citation in piece.get("extra_citations", [])),
        )
        if isinstance(evidence_id, str) and evidence_id
    ))
    citations = []
    for piece in pieces:
        citations.append(_citation(piece))
        citations.extend(piece.get("extra_citations", []))
    citations = list({
        json.dumps(citation, ensure_ascii=False, sort_keys=True): citation
        for citation in citations
    }.values())
    facts = "\n".join(piece["semantic"] for piece in pieces)
    semantic = "\n".join(part for part in (context_prefix, facts) if part)
    markdown = "\n\n".join(part for part in (context_prefix, "\n\n".join(
        piece["markdown"] for piece in pieces)) if part)
    exact = "\n".join(piece["exact"] for piece in pieces)
    record_id = _stable_id(
        "rag-record", source_hash, kind, logical_record_id, group_index, group_count,
        evidence_ids, [piece["text_span"] for piece in pieces], semantic,
    )
    all_identifier_mentions = list({
        json.dumps(mention, ensure_ascii=False, sort_keys=True, separators=(",", ":")): mention
        for piece in pieces
        for mention in piece.get("identifier_mentions", [])
        if isinstance(mention, dict)
    }.values())
    piece_overflow_count = sum(
        count if isinstance((count := piece.get("identifier_mention_overflow_count")), int) and not isinstance(count, bool)
        else 0
        for piece in pieces
    )
    aggregate_overflow_count = max(
        0, len(all_identifier_mentions) - context_ir.IDENTIFIER_MAX_MENTIONS_PER_CHUNK,
    )
    identifier_mentions = all_identifier_mentions[:context_ir.IDENTIFIER_MAX_MENTIONS_PER_CHUNK]
    overflow_count = piece_overflow_count + aggregate_overflow_count
    mention_count = len(identifier_mentions) + overflow_count
    metadata_complete = (
        overflow_count == 0
        and all(piece.get("identifier_metadata_complete", True) is True for piece in pieces)
    )
    exact_identifier_values = {
        role: list(dict.fromkeys(
            mention["value"]
            for mention in identifier_mentions
            if mention.get("role") == role and isinstance(mention.get("value"), str)
        ))
        for role in ("record_identity", "reference", "unclassified")
    }
    return {
        "record_id": record_id,
        "logical_record_id": logical_record_id,
        "kind": kind,
        "field_group_index": group_index,
        "field_group_count": group_count,
        "section_path": section_path,
        "semantic_text": semantic,
        "markdown_text": markdown,
        "exact_text": exact,
        "evidence_ids": evidence_ids,
        "citations": citations,
        "relations": list({
            json.dumps(relation, ensure_ascii=False, sort_keys=True): relation
            for piece in pieces for relation in piece.get("relations", [])
        }.values()),
        "_sort_key": sort_key,
        "parent_id": parent_id,
        "source_name": source_name,
        "document_context": document_context or {},
        "region_context": region_context or {},
        "record_keys": record_keys or [],
        "identifier_mentions": identifier_mentions,
        "identifier_metadata_complete": metadata_complete,
        "identifier_mention_count": mention_count,
        "identifier_mention_overflow_count": overflow_count,
        "identifier_mention_limit": context_ir.IDENTIFIER_MAX_MENTIONS_PER_CHUNK,
        "record_identity_values_exact": exact_identifier_values["record_identity"],
        "reference_values_exact": exact_identifier_values["reference"],
        "unclassified_identifier_values_exact": exact_identifier_values["unclassified"],
        "related_record_ids": related_record_ids or [],
        "context_confidence": context_confidence,
    }


def _citation(piece: dict) -> dict:
    citation = {"locator": piece["locator"]}
    if isinstance(piece.get("evidence_id"), str) and piece["evidence_id"]:
        citation["evidence_id"] = piece["evidence_id"]
    for key in ("coverage_id", "coverage_status", "content_basis", "reason_code"):
        if piece.get(key) is not None:
            citation[key] = piece[key]
    # 単純な一段見出しはbodyに同じfield名がある。結合見出しの階層だけcitationにも残す。
    header_path = piece.get("header_path", [])
    if len(header_path) > 1:
        citation["header_path"] = header_path
    if piece.get("text_span") is not None:
        citation["text_span"] = piece["text_span"]
    if piece.get("evidence_state"):
        citation["evidence_state"] = piece["evidence_state"]
    if piece.get("object_metadata"):
        citation["object_metadata"] = piece["object_metadata"]
    return citation


def _coverage_reason(item: evidence_ir.CoverageItem) -> str | None:
    """v1alpha1/v1alpha2のcoverage理由をrenderer内で一意に扱う。"""
    value = getattr(item, "reason_code", None)
    if isinstance(value, str) and value:
        return value
    legacy = getattr(item, "reason", None)
    return legacy if isinstance(legacy, str) and legacy else None


def _coverage_basis(item: evidence_ir.CoverageItem) -> str:
    value = getattr(item, "content_basis", None)
    if isinstance(value, str) and value:
        return value
    reason = _coverage_reason(item)
    if reason in {"image_content_uninterpreted", "binary_asset_content_uninterpreted"}:
        return "pixel_only"
    if item.status == "intentionally_ignored":
        return "none"
    return "structured"


def _coverage_payload(item: evidence_ir.CoverageItem) -> dict:
    return {
        "coverage_id": item.coverage_id,
        "coverage_status": item.status,
        "content_basis": _coverage_basis(item),
        "reason_code": _coverage_reason(item),
    }


def _enrich_record_citations(records: list[dict], ir: evidence_ir.EvidenceIR) -> None:
    """Evidence参照へcoverage結果を付与し、検索後も未解釈/未対応を判別可能にする。"""
    elements = {element.element_id: element for element in ir.elements}
    coverage = {item.coverage_id: item for item in ir.coverage}
    for record in records:
        for citation in record.get("citations", []):
            item = None
            evidence_id = citation.get("evidence_id")
            if isinstance(evidence_id, str) and evidence_id in elements:
                item = coverage.get(elements[evidence_id].coverage_id)
            coverage_id = citation.get("coverage_id")
            if item is None and isinstance(coverage_id, str):
                item = coverage.get(coverage_id)
            if item is not None:
                citation.update(_coverage_payload(item))


def _coverage_notice_records(ir: evidence_ir.EvidenceIR, source_name: str) -> list[dict]:
    """content-bearingな未対応/失敗を、件数summaryだけでなくlocator付きrecordへする。"""
    records: list[dict] = []
    for index, item in enumerate(ir.coverage, start=1):
        if item.status not in {"unsupported", "failed"}:
            continue
        basis = _coverage_basis(item)
        if basis == "none":
            continue
        reason = _coverage_reason(item) or "unspecified"
        detected_kind = getattr(item, "detected_kind", None) or item.scope
        locator = _compact_locator(item.locator)
        location = json.dumps(locator, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
        status_label = "未対応" if item.status == "unsupported" else "抽出失敗"
        safe_reference_detail: list[dict[str, Any]] = []
        if reason in {"asset_binding_failed", "external_image_reference_not_fetched"}:
            references = item.detail.get("references") if isinstance(item.detail, dict) else None
            if isinstance(references, list):
                for reference in references:
                    if not isinstance(reference, dict):
                        continue
                    safe_reference_detail.append({
                        key: reference[key]
                        for key in (
                            "relationship_id", "relationship_attribute", "relationship_type", "target_mode",
                            "external_target_sha256", "binding_status", "asset_role", "vml_element",
                        )
                        if reference.get(key) not in (None, "")
                    })
        reference_text = (
            ", references=" + json.dumps(safe_reference_detail, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
            if safe_reference_detail else ""
        )
        exact = f"{status_label}: kind={detected_kind}, reason={reason}, locator={location}{reference_text}"
        semantic = (
            f"原本「{source_name}」には、{detected_kind}として検知した内容があるが、"
            f"変換結果は{status_label}である。理由コードは{reason}、原本位置は{location}である。"
            "内容を抽出済みとは扱わない。"
        )
        if safe_reference_detail:
            semantic += f" 画像relationshipの安全化済み情報は{reference_text.removeprefix(', ')}である。"
        # MEM-2: サイズ/セル数系ガードは実測値を平文に含める（利用者向けの一報の質・入口ガードの
        # 上限値と実測値が両方分かって初めて「分割する」等の対処判断ができる）。
        detail = item.detail if isinstance(item.detail, dict) else {}
        if reason == "cell_count_exceeded":
            measured, cap = detail.get("measured_cells"), detail.get("cap_cells")
            if isinstance(measured, int) and isinstance(cap, int):
                semantic += (
                    f" セル数が多すぎるため取り込み対象外である（上限{cap}セル・このファイル約{measured}セル）。")
        elif reason == "uncompressed_size_exceeded":
            measured, cap = detail.get("measured_bytes"), detail.get("cap_bytes")
            if isinstance(measured, int) and isinstance(cap, int):
                semantic += (
                    " 展開後サイズが大きすぎるため取り込み対象外である"
                    f"（上限{cap // (1024 * 1024)}MiB・このファイル約{measured // (1024 * 1024)}MiB）。")
        piece = {
            "semantic": semantic,
            "markdown": exact + "\n\n" + semantic,
            "exact": exact,
            "evidence_id": None,
            "coverage_id": item.coverage_id,
            "coverage_status": item.status,
            "content_basis": basis,
            "reason_code": reason,
            "locator": locator,
            "header_path": [],
            "text_span": None,
            "relations": [],
            "identifier_mentions": [],
            "identifier_metadata_complete": True,
            "identifier_mention_count": 0,
            "identifier_mention_overflow_count": 0,
            "identifier_mention_limit": context_ir.IDENTIFIER_MAX_MENTIONS_PER_CHUNK,
        }
        if safe_reference_detail:
            piece["object_metadata"] = {"image_references": safe_reference_detail}
        records.append(_base_record(
            source_hash=ir.source.content_hash,
            kind="coverage_notice",
            logical_record_id=_stable_id("logical-coverage-notice", ir.source.content_hash, item.coverage_id),
            group_index=1,
            group_count=1,
            section_path=["未抽出・未対応の原本要素"],
            pieces=[piece],
            sort_key=("~coverage-notice", index, 0, 0, 0, 0),
            source_name=source_name,
            document_context={"source_name": source_name},
        ))
    return records


def _alias_citation(element: evidence_ir.EvidenceElement) -> dict:
    citation = {
        "evidence_id": element.element_id,
        "locator": _compact_locator(element.locator),
        "evidence_state": {"visibility": element.visibility, "lifecycle": element.lifecycle},
    }
    return citation


def _document_context(ctx: context_ir.ContextIR, region: context_ir.ContextRegion) -> dict:
    title_key = (
        region.sheet
        or (f"slide:{region.slide}" if region.slide is not None else None)
        or (f"page:{region.page}" if region.page is not None else "")
    )
    title = ctx.document_titles.get(title_key)
    return {
        "source_name": ctx.source_name,
        "title": title,
        "sheet": region.sheet,
        "slide": region.slide,
        "page": region.page,
        "part": region.part,
    }


def _region_context(region: context_ir.ContextRegion) -> dict:
    out = {
        "region_id": region.region_id,
        "title": region.title,
        "sheet": region.sheet,
        "slide": region.slide,
        "page": region.page,
        "mode": region.mode,
        "confidence": region.confidence,
    }
    if region.sheet:
        from openpyxl.utils import get_column_letter
        out["cell_range"] = (
            f"{get_column_letter(region.start_column)}{region.start_row}:"
            f"{get_column_letter(region.end_column)}{region.end_row}"
        )
    else:
        out.update({
            "part": region.part,
            "table_object_id": region.table_object_id,
            "row_range": [region.start_row, region.end_row],
            "column_range": [region.start_column, region.end_column],
        })
    if region.section_path:
        out["section_path"] = list(region.section_path)
    return out


def _common_header_prefix(region: context_ir.ContextRegion) -> tuple[str, ...]:
    paths = [path for _, path in region.header_paths if path]
    if not paths:
        return ()
    common = list(paths[0])
    for path in paths[1:]:
        while common and tuple(path[:len(common)]) != tuple(common):
            common.pop()
        if not common:
            break
    # 全列が同じ単一見出しだけの場合はfield label自体なので親へ持ち上げない。
    return tuple(common) if any(len(path) > len(common) for path in paths) else ()


def _field_label(
    region: context_ir.ContextRegion,
    column: int,
    common: tuple[str, ...],
) -> str:
    path = region.header_for(column)
    if common and path[:len(common)] == common:
        path = path[len(common):]
    return " > ".join(path)


def _context_prefix(
    ctx: context_ir.ContextIR,
    region: context_ir.ContextRegion,
    record: context_ir.ContextRecord | None,
    *,
    common_header: tuple[str, ...] | None = None,
) -> str:
    document = _document_context(ctx, region)
    parts = [f"原本「{document['source_name']}」"]
    if document.get("title") and document["title"] != Path(document["source_name"]).stem:
        parts.append(f"文書「{document['title']}」")
    if region.sheet:
        parts.append(f"シート「{region.sheet}」")
    elif region.slide is not None:
        parts.append(f"スライド{region.slide}")
    elif region.page is not None:
        parts.append(f"ページ{region.page}")
    for heading in region.section_path:
        if heading and heading != document.get("title"):
            parts.append(f"節「{heading}」")
    if region.title and region.title != document.get("title"):
        parts.append(f"領域「{region.title}」")
    common_text = " > ".join(
        _common_header_prefix(region) if common_header is None else common_header)
    if common_text and common_text != region.title:
        parts.append(f"区分「{common_text}」")
    key_text = "、".join(
        f"{key.label.split(' > ')[-1]}「{key.value}」"
        for key in (record.keys if record else ())
    )
    if key_text:
        parts.append(key_text)
    return "出所: " + " / ".join(parts)


def _region_section_path(
    ctx: context_ir.ContextIR,
    region: context_ir.ContextRegion,
    locator: evidence_ir.Locator,
) -> list[str]:
    if region.part and region.part.startswith("word/"):
        document_title = ctx.document_titles.get("") or Path(ctx.source_name).stem
        section = [f"文書「{document_title}」"]
        section.extend(
            f"節「{heading}」"
            for heading in region.section_path
            if heading != document_title
        )
        if region.table_object_id is not None:
            section.append(f"表「{region.table_object_id}」")
        return section
    return _section(locator)


def _record_sort_key(
    locator: evidence_ir.Locator,
    order: int,
    row: int = 0,
    index: int = 0,
) -> tuple:
    """PPTXはXML/z順ではなく、slide内の上→下・左→右を優先して並べる。"""
    if locator.slide is not None:
        bbox = locator.bbox or [0, 0, 0, 0]
        paint_order = locator.extension.get("paint_order", order)
        return (f"slide:{locator.slide:08d}", float(bbox[1]), float(bbox[0]), paint_order, row, index)
    if locator.page is not None:
        bbox = locator.bbox or [0, 0, 0, 0]
        return (f"page:{locator.page:08d}", -float(bbox[3]), float(bbox[0]), order, row, index)
    return (" / ".join(_section(locator)), order, 0, 0, row, index)


def _context_summary_record(
    ir: evidence_ir.EvidenceIR,
    ctx: context_ir.ContextIR,
    region: context_ir.ContextRegion,
    cells: list[evidence_ir.EvidenceElement],
    table: evidence_ir.EvidenceElement | None,
    aliases: dict[str, list[evidence_ir.EvidenceElement]],
) -> dict | None:
    if region.header_row is None:
        return None
    context_cells = [
        cell for cell in cells
        if (row := _cell_row_column(cell)[0]) is not None and row <= region.header_row and _value_text(cell.value)
    ]
    if not context_cells:
        return None
    pieces = []
    for cell in sorted(context_cells, key=lambda item: (_cell_row_column(item)[0] or 0,
                                                        _cell_row_column(item)[1] or 0)):
        row, column = _cell_row_column(cell)
        label = "領域見出し" if row == region.start_row else (
            "列見出し" if row == region.header_row else "領域情報")
        exact = _value_text(cell.value)
        piece = {
            "semantic": f"{label}: 「{exact}」",
            "markdown": f"{label}: 「{exact}」",
            "exact": f"{label}: {exact}",
            "evidence_id": cell.element_id,
            "locator": _compact_cell_locator(cell.locator),
            "header_path": list(region.header_for(column or 0)),
            "text_span": None,
            "relations": [],
            **_extract_identifier_metadata(
                exact,
                field_label=label,
                evidence_id=cell.element_id,
                locator=cell.locator,
            ),
        }
        if cell.element_id in aliases:
            piece["extra_citations"] = [_alias_citation(alias) for alias in aliases[cell.element_id]]
        pieces.append(piece)
    prefix = _context_prefix(ctx, region, None)
    logical_id = _stable_id("context-summary", ir.source.content_hash, region.region_id)
    return _base_record(
        source_hash=ir.source.content_hash,
        kind="context_summary",
        logical_record_id=logical_id,
        group_index=1,
        group_count=1,
        section_path=_region_section_path(ctx, region, cells[0].locator)
        + ([f"領域「{region.title}」"] if region.title else []),
        pieces=pieces,
        sort_key=_record_sort_key((table or cells[0]).locator, (table.order if table else 0), region.start_row, 0),
        context_prefix=prefix,
        parent_id=region.region_id,
        source_name=ctx.source_name,
        document_context=_document_context(ctx, region),
        region_context=_region_context(region),
        context_confidence=region.confidence,
    )


def _table_records(
    ir: evidence_ir.EvidenceIR,
    ctx: context_ir.ContextIR,
    aliases: dict[str, list[evidence_ir.EvidenceElement]],
) -> list[dict]:
    by_parent: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for element in ir.elements:
        if element.type == "cell" and element.parent_id:
            by_parent[element.parent_id].append(element)
    tables = {element.element_id: element for element in ir.elements if element.type == "table"}
    regions_by_table: dict[str, list[context_ir.ContextRegion]] = defaultdict(list)
    for region in ctx.regions:
        regions_by_table[region.table_id].append(region)
    records_by_region: dict[str, list[context_ir.ContextRecord]] = defaultdict(list)
    for context_record in ctx.records:
        records_by_region[context_record.region_id].append(context_record)
    records: list[dict] = []
    for table_id, cells in sorted(by_parent.items(), key=lambda item: min(cell.order for cell in item[1])):
        table = tables.get(table_id)
        regions = sorted(regions_by_table.get(table_id, []), key=lambda item: item.start_column)
        has_coordinate_fallback = any(region.mode == "coordinate_fallback" for region in regions)
        if has_coordinate_fallback:
            block_by_row, relations_by_evidence = _table_layout(ir, table_id, cells)
            left_labels = _left_label_map(cells)
        else:
            rows = {row for cell in cells if (row := _cell_row_column(cell)[0]) is not None}
            block_by_row = {
                row: _stable_id("logical-block", ir.source.content_hash, table_id, f"row:{row}")
                for row in rows
            }
            relations_by_evidence = {}
            left_labels = {}
        cells_by_row: dict[int, list[evidence_ir.EvidenceElement]] = defaultdict(list)
        for cell in cells:
            row, _ = _cell_row_column(cell)
            if row is not None:
                cells_by_row[row].append(cell)
        table_order = table.order if table else 0
        for region in regions:
            region_cells = [
                cell for cell in cells
                if (column := _cell_row_column(cell)[1]) is not None
                and region.start_column <= column <= region.end_column
            ]
            summary = _context_summary_record(ir, ctx, region, region_cells, table, aliases)
            if summary is not None:
                records.append(summary)
            common_header = _common_header_prefix(region)
            region_records = sorted(records_by_region.get(region.region_id, []), key=lambda item: item.row)
            for context_record in region_records:
                row = context_record.row
                key_evidence_ids = {key.evidence_id for key in context_record.keys}
                row_cells = [
                    cell for cell in cells_by_row.get(row, [])
                    if (column := _cell_row_column(cell)[1]) is not None
                    and region.start_column <= column <= region.end_column
                    and _value_text(cell.value)
                ]
                pieces: list[dict] = []
                for cell in sorted(row_cells, key=lambda item: (_cell_row_column(item)[1] or 0, item.order)):
                    _, column = _cell_row_column(cell)
                    header = _field_label(region, column or 0, common_header)
                    if not header:
                        label = left_labels.get(cell.element_id)
                        header = _value_text(label.value) if label is not None else "記載値"
                    exact = _cell_render_text(cell)
                    spans = _split_exact(exact)
                    for span_index, (start, end, part) in enumerate(spans):
                        identifier_metadata = _extract_identifier_metadata(
                            exact,
                            field_label=header,
                            evidence_id=cell.element_id,
                            locator=cell.locator,
                            allow_record_identity=cell.element_id in key_evidence_ids,
                            match_start=start,
                            match_end=end,
                        )
                        piece = _field_piece(
                            cell, header, span_index, len(spans), start, end, part,
                            (relations_by_evidence.get(cell.element_id, [])
                            if region.mode == "coordinate_fallback" else []),
                            allow_record_identity=cell.element_id in key_evidence_ids,
                            identifier_metadata=identifier_metadata,
                        )
                        if span_index == 0 and cell.element_id in aliases:
                            piece["extra_citations"] = [
                                _alias_citation(alias) for alias in aliases[cell.element_id]
                            ]
                        pieces.append(piece)
                prefix = _context_prefix(ctx, region, context_record, common_header=common_header)
                groups = _pack_row(pieces, prefix)
                logical_record_id = context_record.record_id
                logical_block_id = block_by_row[row]
                section_path = _region_section_path(ctx, region, (table or row_cells[0]).locator)
                if region.title:
                    section_path.append(f"領域「{region.title}」")
                record_keys = [asdict(key) for key in context_record.keys]
                for index, group in enumerate(groups, start=1):
                    records.append(_base_record(
                        source_hash=ir.source.content_hash,
                        kind="table_record" if region.mode == "header_record" else "coordinate_record",
                        logical_record_id=logical_record_id,
                        group_index=index,
                        group_count=len(groups),
                        section_path=section_path,
                        pieces=group,
                        sort_key=_record_sort_key(
                            (table or row_cells[0]).locator, table_order, row, index),
                        context_prefix=prefix,
                        parent_id=region.region_id,
                        source_name=ctx.source_name,
                        document_context=_document_context(ctx, region),
                        region_context=_region_context(region),
                        record_keys=record_keys,
                        related_record_ids=list(ctx.related_records.get(context_record.record_id, ())),
                        context_confidence=context_record.confidence,
                    ))
                    records[-1]["logical_block_id"] = logical_block_id
    return records


def _asset_href(source_name: str, element: evidence_ir.EvidenceElement) -> str | None:
    digest = element.extension.get("asset_sha256")
    media_part = element.extension.get("media_part")
    if not isinstance(digest, str):
        return None
    suffix = (
        Path(media_part).suffix.lower() if isinstance(media_part, str)
        else element.extension.get("asset_extension", ".bin")
    )
    suffix = suffix if isinstance(suffix, str) and suffix.startswith(".") else ".bin"
    return f"{Path(source_name).name}.assets/{digest}{suffix}"


def _shape_fill_assets(source_name: str, element: evidence_ir.EvidenceElement) -> list[dict[str, Any]]:
    """DrawingML/VML画像fillの全参照をcitation可能な形へ正規化する。"""
    raw_assets = element.extension.get("assets")
    candidates = raw_assets if isinstance(raw_assets, list) and raw_assets else [element.extension]
    assets: list[dict[str, Any]] = []
    for index, candidate in enumerate(candidates, start=1):
        if not isinstance(candidate, dict):
            continue
        role = candidate.get("asset_role") or element.extension.get("asset_role")
        digest = candidate.get("asset_sha256")
        if role != "shape_fill" or not isinstance(digest, str) or not digest:
            continue
        media_part = candidate.get("media_part")
        suffix = (
            Path(media_part).suffix.lower() if isinstance(media_part, str)
            else candidate.get("asset_extension", ".bin")
        )
        suffix = suffix if isinstance(suffix, str) and suffix.startswith(".") else ".bin"
        asset = {
            "asset_index": index,
            "asset_role": "shape_fill",
            "asset_sha256": digest,
            "asset_href": f"{Path(source_name).name}.assets/{digest}{suffix}",
        }
        for key in ("media_part", "relationship_id", "pixel_size", "vml_element"):
            value = candidate.get(key)
            if value not in (None, ""):
                asset[key] = value
        assets.append(asset)
    return assets


def _element_piece(
    element: evidence_ir.EvidenceElement,
    elements: dict[str, evidence_ir.EvidenceElement],
    relations: list[evidence_ir.EvidenceRelation],
    source_name: str,
    section_path: list[str] | None = None,
    include_native_metadata: bool = False,
) -> dict | None:
    exact = _value_text(element.value)
    excel_metadata = _excel_value_metadata(element)
    shape_fill_assets = _shape_fill_assets(source_name, element)
    # 基底文に続く付随情報（relation叙述文/ExcelノートKV/状態KV/shape-fill asset説明）。
    # `_non_table_records`の長文分割で本文だけのsemantic/markdownに上書きされた際、
    # 先頭chunkへ引き継ぐための複製（自己完結契約＝断片だけでも要素の全体像が分かる）。
    metadata_tail = ""     # semanticへの追加分（markdownはこれ+markdown_tail）
    markdown_tail = ""     # markdown固有の追加分（画像リンク等・semanticには載せない）
    section_text = "の".join(section_path) if section_path else _section(element.locator)[0]
    document_context = f"原本「{Path(source_name).name}」の{section_text}"
    outgoing = [relation for relation in relations if relation.source_id == element.element_id]
    relation_texts: list[str] = []
    rendered_relations: list[dict] = []
    for relation in outgoing:
        target = elements.get(relation.target_id)
        if target is None:
            continue
        target_text = _value_text(target.value).replace("\n", " ").strip()
        if len(target_text) > 160:
            target_text = target_text[:157] + "..."
        target_label = target_text or _value_text(target.extension.get("name")) or target.type
        if relation.type == "overlaps":
            if element.extension.get("alpha") == 0:
                relation_texts.append(f"この要素は対象「{target_label}」の原本領域に重なっている")
            else:
                relation_texts.append("この要素は別要素の原本領域に重なっている")
        elif relation.type == "connects_to":
            relation_texts.append(f"この要素は対象「{target_label}」へ接続している")
        elif relation.type == "covered_by":
            relation_texts.append("このtext layerは後描画の全面画像に覆われており、現行回答に使用しない")
        elif relation.type in {"covers", "painted_after"}:
            relation_texts.append(f"この要素は対象要素と「{relation.type}」の関係にある")
        rendered_relations.append({
            "type": relation.type,
            "source_id": relation.source_id,
            "target_id": relation.target_id,
            "evidence_ids": relation.evidence_ids,
            "confidence": relation.confidence,
            "target_locator": _compact_locator(target.locator),
            "extension": relation.extension,
        })

    if element.type in {"picture", "image_xobject"}:
        href = _asset_href(source_name, element)
        alt = element.extension.get("name") or f"画像 object {element.locator.object_id}"
        image_md = f"![{alt}（内容未解釈）]({href})" if href else f"画像「{alt}」（内容未解釈）"
        semantic = f"{document_context}に画像「{alt}」が存在する。画像内容は未解釈である。"
        if element.type == "image_xobject" and not any(
            candidate.locator.page == element.locator.page
            and candidate.type in {"text_object", "positioned_text", "cell"}
            and _value_text(candidate.value).strip()
            for candidate in elements.values()
        ):
            semantic += " このページはテキスト層を持たず、内容取得にはOCRまたは画像理解が必要である。"
        if relation_texts:
            semantic += " " + "。".join(relation_texts) + "。"
        markdown = image_md + "\n\n" + semantic
    elif exact:
        type_label = ({
            "heading": "見出し",
            "paragraph": "本文",
            "textbox": "テキストボックス",
            "floating_textbox": "浮動テキストボックス",
            "hidden_text": "非表示文字",
            "deleted_text": "削除履歴文字",
            "comment": "コメント",
            "hyperlink": "ハイパーリンク",
            "header": "ヘッダー",
            "footer": "フッター",
            "footnote": "脚注",
            "endnote": "文末脚注",
            "shape": "図形",
            "notes": "発表者ノート",
            "connector": "コネクタ",
            "graphic_frame": "図表オブジェクト",
            "chart": "グラフ",
            "chart_data": "グラフ原本情報",
            "positioned_text": "配置文字",
            "text_object": "PDFテキスト層",
        }.get(element.type, element.type) if include_native_metadata else f"{element.type}の文字列")
        source_parts = [f"原本「{Path(source_name).name}」", *(section_path or _section(element.locator))]
        base_sentence = "出所: " + " / ".join(source_parts) + f"\n{type_label}: 「{exact}」"
        if relation_texts:
            metadata_tail += "\n" + "。".join(relation_texts) + "。"
        metadata_tail += _excel_value_note(excel_metadata)
        # L4a: `visibility_reason`に平文ラベルの対応があれば、生の`visibility`値の代わりにそれを出す
        # （`occlusion_lines`が可視性行を持つ場合は`state_texts`側の可視性行を出さない＝二重表現の回避）。
        occlusion_lines = _occlusion_kv_lines(element.extension) if include_native_metadata else []
        occlusion_has_visibility = any(line.startswith("可視性:") for line in occlusion_lines)
        state_texts = []
        if element.visibility != "visible" and not occlusion_has_visibility:
            state_texts.append(f"可視性: 「{element.visibility}」")
        if element.lifecycle != "active":
            state_texts.append(f"状態: 「{element.lifecycle}」")
        if include_native_metadata and state_texts:
            metadata_tail += "\n" + " / ".join(state_texts)
        if occlusion_lines:
            metadata_tail += "\n" + "\n".join(occlusion_lines)
        semantic = base_sentence + metadata_tail
        markdown = semantic
    elif relation_texts:
        semantic = "。".join(relation_texts) + "。"
        markdown = semantic
    else:
        # 値も関係もないobjectは存在自体を落とさない。
        semantic = f"{document_context}に{element.type}が存在する。"
        markdown = semantic
    if shape_fill_assets:
        asset_count = len(shape_fill_assets)
        object_label = _value_text(element.extension.get("name")) or _value_text(element.locator.object_id) or element.type
        shape_fill_sentence = (
            f"\nこの要素「{object_label}」には内容未解釈の画像塗りassetが{asset_count}件存在し、"
            "各assetの原本bytesと参照先を保持している。"
        )
        semantic += shape_fill_sentence
        metadata_tail += shape_fill_sentence     # 長文分割後も先頭chunkのsemanticへ復元できるよう複製
        asset_markdown: list[str] = []
        for asset in shape_fill_assets:
            asset_index = asset["asset_index"]
            digest = asset["asset_sha256"]
            asset_markdown.extend([
                f"![{object_label}の画像塗りasset {asset_index}/{asset_count}（内容未解釈）]({asset['asset_href']})",
                f"画像塗りasset {asset_index}/{asset_count}のSHA-256: {digest}",
            ])
        asset_markdown_block = "\n\n" + "\n\n".join(asset_markdown)
        markdown += asset_markdown_block
        markdown_tail += asset_markdown_block    # 長文分割後も先頭chunkのmarkdownへ復元できるよう複製
    object_metadata = {
        key: element.extension[key]
        for key in (
            "anchor_type", "z_order", "alpha", "name", "description", "asset_sha256", "media_part",
            "relationship_id", "pixel_size", "paint_order", "start_object_id", "end_object_id",
            "visibility_reason", "asset_extension", "image_name", "operator_index", "visual_currentness",
            "use_for_current_answer", "chart_part", "chart_texts", "chart_references",
            "host_part", "host_kind", "host_visibility", "host_lifecycle", "host_parts",
            "image_relationship_ids", "unresolved_image_relationship_ids", "external_image_relationship_ids",
            "image_references", "relationship_attribute", "target_mode", "external_target_sha256", "binding_status",
        )
        if element.extension.get(key) not in (None, "")
    } if include_native_metadata else {}
    if shape_fill_assets:
        object_metadata["assets"] = shape_fill_assets
    piece = {
        "semantic": semantic,
        "markdown": markdown,
        "exact": exact,
        "evidence_id": element.element_id,
        "locator": _compact_locator(element.locator),
        "header_path": [],
        "text_span": {"start": 0, "end": len(exact), "index": 1, "count": 1},
        "relations": rendered_relations,
        "metadata_tail": metadata_tail,
        "markdown_tail": markdown_tail,
        **_extract_identifier_metadata(
            exact,
            field_label="",
            evidence_id=element.element_id,
            locator=element.locator,
        ),
    }
    if include_native_metadata:
        piece["evidence_state"] = {"visibility": element.visibility, "lifecycle": element.lifecycle}
    if object_metadata:
        piece["object_metadata"] = object_metadata
    if excel_metadata is not None:
        piece.setdefault("object_metadata", {})["excel_value"] = excel_metadata
    return piece


def _element_section_path(
    ir: evidence_ir.EvidenceIR,
    ctx: context_ir.ContextIR,
    element: evidence_ir.EvidenceElement,
) -> list[str]:
    if ir.source.file_type != "docx":
        return _section(element.locator)
    document_title = ctx.document_titles.get("") or Path(ctx.source_name).stem
    headings = ctx.element_sections.get(element.element_id)
    if headings is None and element.parent_id:
        headings = ctx.element_sections.get(element.parent_id)
    section = [f"文書「{document_title}」"]
    section.extend(f"節「{heading}」" for heading in (headings or ()) if heading != document_title)
    return section


def _docx_textbox_aliases(
    elements: dict[str, evidence_ir.EvidenceElement],
) -> tuple[dict[str, list[evidence_ir.EvidenceElement]], set[str]]:
    """同一浮動textboxのDocument IR/VML多重表現を、幾何情報のある1recordへ束ねる。

    値だけで複数objectを誤同定しないよう、同文の子textboxが1個で、floating/VML候補も
    同じhost partにそれぞれ最大1個の場合に限る。幾何・assetを持つVMLを最優先にし、ホスト段落は
    子textboxと完全同文の場合だけaliasにする。
    """
    floating_by_text: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    child_by_text: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    vml_by_text: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for element in elements.values():
        text = _value_text(element.value)
        if not text:
            continue
        if element.type == "floating_textbox":
            floating_by_text[text].append(element)
        elif element.type == "textbox" and element.parent_id:
            child_by_text[text].append(element)
        elif element.type == "vml_shape":
            vml_by_text[text].append(element)

    aliases: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    suppressed: set[str] = set()
    for text, children in child_by_text.items():
        floating = floating_by_text.get(text, [])
        if len(children) != 1 or len(floating) > 1:
            continue
        child = children[0]
        matching_vml = [candidate for candidate in vml_by_text.get(text, [])
                        if candidate.locator.part == child.locator.part]
        if len(matching_vml) > 1:
            continue
        primary = matching_vml[0] if matching_vml else floating[0] if floating else child
        for alias in [child, *floating]:
            if alias.element_id == primary.element_id:
                continue
            aliases[primary.element_id].append(alias)
            suppressed.add(alias.element_id)
        parent = elements.get(child.parent_id or "")
        if parent is not None and _value_text(parent.value) == text:
            aliases[primary.element_id].append(parent)
            suppressed.add(parent.element_id)
    return dict(aliases), suppressed


def _pptx_object_aliases(
    elements: dict[str, evidence_ir.EvidenceElement],
) -> tuple[dict[str, list[evidence_ir.EvidenceElement]], set[str]]:
    """Document IR由来shapeとnative DrawingML objectをslide/z位置で1つのRAG recordへ束ねる。"""
    slide_ids = {
        element.element_id
        for element in elements.values()
        if element.type == "slide"
    }
    children: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    native_by_slot: dict[tuple[int, int], evidence_ir.EvidenceElement] = {}
    adapted: list[tuple[evidence_ir.EvidenceElement, int]] = []
    for element in elements.values():
        if element.parent_id:
            children[element.parent_id].append(element)
        source_map = element.locator.extension.get("document_ir_source_map", {})
        if element.extension.get("origin") == "document-ir-v2-adapter" and element.type == "shape":
            z_index = source_map.get("z_index") if isinstance(source_map, dict) else None
            if isinstance(z_index, int):
                adapted.append((element, z_index))
            continue
        z_order = element.extension.get("z_order")
        if (element.parent_id in slide_ids and element.locator.slide is not None
                and isinstance(z_order, int)):
            native_by_slot[(element.locator.slide, z_order)] = element

    aliases: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    suppressed: set[str] = set()
    first_z = {
        slide: min(z_index for alias, z_index in adapted if alias.locator.slide == slide)
        for slide in {alias.locator.slide for alias, _z_index in adapted if alias.locator.slide is not None}
    }
    for alias, z_index in adapted:
        slide = alias.locator.slide
        if not isinstance(slide, int):
            continue
        native = native_by_slot.get((slide, z_index - first_z[slide] + 1))
        if native is None:
            continue
        target = native
        if native.type in {"table", "group"}:
            # 束ね先は「実際に RAG レコードとして出力される要素」に限る。
            #   - 入れ子の group など `_CONTAINER_TYPES` は出力されない
            #   - 空セルも表レコードに出力されない
            # どちらを選んでも alias の引用が消え、`element_coverage_missing` で文書ごと
            # RAG 表現を作れなくなる（実測 2026-08-15・pptx 4件）。本文を持つ要素だけを候補にする。
            candidates = [
                item for item in children.get(native.element_id, [])
                if item.type not in _CONTAINER_TYPES and _value_text(item.value)
            ]
            if not candidates:
                # 束ね先が container（table/group）のままだと `_non_table_records` が
                # `_CONTAINER_TYPES` として除外するため、束ねた alias の引用がどこにも出ず
                # `element_coverage_missing` で文書ごと失敗する（実測 2026-08-15: 装飾図形だけの
                # グループを持つ pptx 4件が RAG 表現を作れず failed notice へ縮退していた）。
                # 束ね先が出力されないときは**抑制せず元の要素をそのまま残す**（情報を落とさない）。
                continue
            target = min(candidates, key=lambda item: item.order)
        aliases[target.element_id].append(alias)
        suppressed.add(alias.element_id)
    return dict(aliases), suppressed


def _pdf_object_aliases(
    elements: dict[str, evidence_ir.EvidenceElement],
) -> tuple[dict[str, list[evidence_ir.EvidenceElement]], set[str]]:
    """集約textは監査用に残し、位置付き文字/cellがあるpageではRAG本文の重複だけ抑える。"""
    granular: dict[int, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for element in elements.values():
        if (element.locator.page is not None and element.type in {"positioned_text", "cell"}
                and _value_text(element.value).strip()):
            granular[element.locator.page].append(element)
    aliases: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    suppressed: set[str] = set()
    for element in elements.values():
        page = element.locator.page
        if (element.type != "text_object" or element.extension.get("aggregate") is not True
                or page not in granular):
            continue
        primary = min(granular[page], key=lambda item: item.order)
        aliases[primary.element_id].append(element)
        suppressed.add(element.element_id)
    return dict(aliases), suppressed


def _element_aliases(
    ir: evidence_ir.EvidenceIR,
    elements: dict[str, evidence_ir.EvidenceElement],
) -> tuple[dict[str, list[evidence_ir.EvidenceElement]], set[str]]:
    if ir.source.file_type == "docx":
        return _docx_textbox_aliases(elements)
    if ir.source.file_type == "pptx":
        return _pptx_object_aliases(elements)
    if ir.source.file_type == "pdf":
        return _pdf_object_aliases(elements)
    return {}, set()


def _non_table_records(
    ir: evidence_ir.EvidenceIR,
    ctx: context_ir.ContextIR,
    aliases: dict[str, list[evidence_ir.EvidenceElement]],
    suppressed: set[str],
) -> list[dict]:
    source_name = ctx.source_name
    elements = {element.element_id: element for element in ir.elements}
    cell_ids = {element.element_id for element in ir.elements if element.type == "cell"}
    records: list[dict] = []
    for element in sorted(ir.elements, key=lambda item: item.order):
        if element.element_id in cell_ids or element.element_id in suppressed or element.type in _CONTAINER_TYPES:
            continue
        section_path = _element_section_path(ir, ctx, element)
        raw_piece = _element_piece(
            element,
            elements,
            ir.relations,
            source_name,
            section_path,
            include_native_metadata=ir.source.file_type in {"xlsx", "docx", "pptx", "pdf"},
        )
        if raw_piece is None:
            continue
        extra_citations = []
        for alias in aliases.get(element.element_id, []):
            alias_piece = _element_piece(
                alias,
                elements,
                [],
                source_name,
                _element_section_path(ir, ctx, alias),
                include_native_metadata=True,
            )
            if alias_piece is not None:
                extra_citations.append(_citation(alias_piece))
        if extra_citations:
            raw_piece["extra_citations"] = extra_citations
        exact = raw_piece["exact"]
        spans = _split_exact(exact) if exact else [(0, 0, "")]
        pieces: list[dict] = []
        if len(spans) == 1:
            pieces = [raw_piece]
        else:
            metadata_tail = raw_piece.get("metadata_tail", "")
            markdown_tail = raw_piece.get("markdown_tail", "")
            for index, (start, end, part) in enumerate(spans):
                semantic = (
                    f"原本「{Path(source_name).name}」の{'の'.join(section_path)}にある"
                    f"{element.type}内容（{index + 1}/{len(spans)}）:\n{part}"
                )
                markdown = semantic
                if index == 0:
                    # relation叙述文/ExcelノートKV/状態KV/shape-fill asset説明は要素1件につき
                    # 1回で足りる情報のため、分割後の先頭chunkにだけ引き継ぐ（全chunkへ複製すると
                    # 内容が重複する）。markdown固有の画像リンク等は`markdown_tail`で別途復元する。
                    semantic += metadata_tail
                    markdown += metadata_tail + markdown_tail
                pieces.append({
                    **raw_piece,
                    "semantic": semantic,
                    "markdown": markdown,
                    "exact": part,
                    "text_span": {"start": start, "end": end, "index": index + 1, "count": len(spans)},
                    **_extract_identifier_metadata(
                        exact,
                        field_label="",
                        evidence_id=element.element_id,
                        locator=element.locator,
                        match_start=start,
                        match_end=end,
                    ),
                })
        logical_record_id = _stable_id("logical-element", ir.source.content_hash, element.element_id)
        for index, piece in enumerate(pieces, start=1):
            records.append(_base_record(
                source_hash=ir.source.content_hash,
                kind=element.type,
                logical_record_id=logical_record_id,
                group_index=index,
                group_count=len(pieces),
                section_path=section_path,
                pieces=[piece],
                sort_key=_record_sort_key(element.locator, element.order, 0, index),
                source_name=Path(source_name).name,
                document_context={
                    "source_name": Path(source_name).name,
                    "sheet": element.locator.sheet,
                    "slide": element.locator.slide,
                    "page": element.locator.page,
                },
            ))
    return records


def _ai_observation_records(
    ir: evidence_ir.EvidenceIR,
    ctx: context_ir.ContextIR,
    observation_set: ai_observation.AIObservationSet,
) -> list[dict]:
    """採用済みAI観測をCanonical値と混ぜず、対象画像へcitationできる独立recordにする。"""
    elements = {element.element_id: element for element in ir.elements}
    inputs = {item.input_id: item for item in observation_set.inputs}
    usable = ai_observation.answer_observations(observation_set)
    grouped: dict[str, list[ai_observation.AIObservation]] = defaultdict(list)
    for observation in usable:
        grouped[observation.input_id].append(observation)

    records: list[dict] = []
    for input_id, observations in sorted(grouped.items()):
        observation_input = inputs[input_id]
        target = elements[observation_input.target_evidence_id]
        section_path = _element_section_path(ir, ctx, target)
        section_text = "の".join(section_path)
        logical_record_id = _stable_id(
            "logical-ai-observation",
            ir.source.content_hash,
            observation_set.observation_set_hash,
            input_id,
        )
        for index, observation in enumerate(observations, start=1):
            verification = "数値は検証済み" if observation.numeric_verified else "数値は未検証"
            body = "AI画像観測（原本確定値ではない）\n観測内容: " + observation.text
            semantic = (
                "AI画像観測（原本確定値ではない）。"
                f"原本「{Path(ctx.source_name).name}」の{section_text}にある画像について、"
                f"{observation_set.provider}/{observation_set.model}が観測した候補事実である。"
                f"観測内容は「{observation.text}」。信頼度は{observation.confidence:.2f}で、{verification}である。"
                "Canonical Evidenceの値ではなく、回答時はAI観測であることを明示する。"
            )
            piece = {
                "semantic": semantic,
                "markdown": body + "\n\n" + semantic,
                "exact": body,
                "evidence_id": target.element_id,
                "locator": _compact_locator(target.locator),
                "header_path": [],
                "text_span": None,
                "relations": [],
                "evidence_state": {"visibility": target.visibility, "lifecycle": target.lifecycle},
                "object_metadata": {
                    "asset_sha256": target.extension.get("asset_sha256"),
                    "media_part": target.extension.get("media_part"),
                },
            }
            record = _base_record(
                source_hash=ir.source.content_hash,
                kind="ai_observation",
                logical_record_id=logical_record_id,
                group_index=index,
                group_count=len(observations),
                section_path=section_path,
                pieces=[piece],
                sort_key=_record_sort_key(target.locator, target.order, 0, 10_000 + index),
                parent_id=target.parent_id,
                source_name=Path(ctx.source_name).name,
                document_context={
                    "source_name": Path(ctx.source_name).name,
                    "sheet": target.locator.sheet,
                    "slide": target.locator.slide,
                    "page": target.locator.page,
                },
                context_confidence=observation.confidence,
            )
            record["ai_observation"] = {
                "observation_id": observation.observation_id,
                "observation_set_hash": observation_set.observation_set_hash,
                "schema_version": observation_set.schema_version,
                "resolver_version": ai_observation.AI_OBSERVATION_RESOLVER_VERSION,
                "input_id": input_id,
                "input_asset_sha256": observation_input.asset_sha256,
                "provider": observation_set.provider,
                "model": observation_set.model,
                "model_revision": observation_set.model_revision,
                "execution_mode": observation_set.execution_mode,
                "prompt_schema_version": observation_set.prompt_schema_version,
                "preprocessing_profile": observation_set.preprocessing_profile,
                "response_hash": observation_set.response_hash,
                "kind": observation.kind,
                "pixel_bbox": observation.pixel_bbox,
                "confidence": observation.confidence,
                "numeric_verified": observation.numeric_verified,
                "evidence_status": "ai_observed",
                "attributes": observation.attributes,
            }
            records.append(record)
    return records


# `llm_render._is_ai_observation_body`と同型の本文マーカー——LLM成形（rag.mdの平文しか見えない）が
# 構造化されたrecord kindの代わりにこの行でフロー図recordを識別し、Mermaidコードを書き換え対象から
# 除外できるようにする。L9時点ではllm_render.py側の配線は未着手（他レーンが同ファイルを並行編集中の
# ため触れない）——このマーカーは将来1行の追加で接続できる形にするための備え。
FLOW_DIAGRAM_BODY_MARKER = "フロー図（機械生成・Mermaid）"


def _flow_diagram_records(ir: evidence_ir.EvidenceIR, ctx: context_ir.ContextIR) -> list[dict]:
    """コンテナ（シート/スライド）単位で、図形＋コネクタからMermaid flowchartを1recordにする（L9・R3）。

    既存の要素単位record（`_non_table_records`）は変更しない——ノード/コネクタ要素は既にそちらで
    個別にcitation済みであり、本recordは追加の可視化表現。ただし過去に図形を束ねた表現で
    citationが漏れRAG表現全体が破綻した事故があるため（`_pptx_object_aliases`のコメント参照）、
    図を構成する全要素（ノード＋未接続分を含む全コネクタ）を自己完結でcitationする。
    """
    elements = {element.element_id: element for element in ir.elements}
    containers = [element for element in ir.elements if element.type in {"sheet", "slide"}]
    if not containers:
        return []

    # 要素ごとの直近のsheet/slide祖先をO(n)で求める（groupの深いネストでも全木再走査しない）。
    container_of: dict[str, str | None] = {}

    def _root_container(element_id: str) -> str | None:
        if element_id in container_of:
            return container_of[element_id]
        container_of[element_id] = None  # 循環防止の一時値
        element = elements.get(element_id)
        result: str | None = None
        if element is not None:
            if element.type in {"sheet", "slide"}:
                result = element_id
            elif element.parent_id is not None:
                result = _root_container(element.parent_id)
        container_of[element_id] = result
        return result

    candidate_types = {"shape", "textbox", "connector", "picture", "graphic_frame", "chart", "smartart", "ole_object"}
    nodes_by_container: dict[str, list[evidence_ir.EvidenceElement]] = defaultdict(list)
    for element in ir.elements:
        if element.type not in candidate_types:
            continue
        container_id = _root_container(element.element_id)
        if container_id is not None:
            nodes_by_container[container_id].append(element)

    records: list[dict] = []
    for container in containers:
        container_elements = nodes_by_container.get(container.element_id, [])
        container_ids = {element.element_id for element in container_elements}
        connects_to = [
            relation for relation in ir.relations
            if relation.type == "connects_to"
            and relation.source_id in container_ids and relation.target_id in container_ids
        ]
        overlaps = [
            relation for relation in ir.relations
            if relation.type == "overlaps"
            and (relation.source_id in container_ids or relation.target_id in container_ids)
        ]
        mermaid = mermaid_render.render_flowchart(
            container_elements, connects_to, overlaps=overlaps, elements_by_id=elements)
        if mermaid is None:
            continue

        node_ids = {endpoint for relation in connects_to for endpoint in (relation.source_id, relation.target_id)}
        connector_ids = {element.element_id for element in container_elements if element.type == "connector"}
        participant_ids = node_ids | connector_ids
        participants = sorted(
            (elements[element_id] for element_id in participant_ids),
            key=lambda element: (element.order, element.element_id),
        )
        primary, extra = participants[0], participants[1:]

        section_path = _element_section_path(ir, ctx, container)
        source_name = Path(ctx.source_name).name
        source_line = "出所: " + " / ".join([f"原本「{source_name}」", *section_path])
        prose = f"原本「{source_name}」の{'の'.join(section_path)}にある図形とコネクタから機械的に生成したフロー図である。"
        body = "\n".join([FLOW_DIAGRAM_BODY_MARKER, source_line, prose, "", f"```mermaid\n{mermaid}```"])
        piece = {
            "semantic": body,
            "markdown": body,
            "exact": mermaid,
            "evidence_id": primary.element_id,
            "locator": _compact_locator(primary.locator),
            "header_path": [],
            "text_span": None,
            "relations": [],
            "extra_citations": [_alias_citation(element) for element in extra],
            **_extract_identifier_metadata(
                mermaid,
                field_label="",
                evidence_id=primary.element_id,
                locator=primary.locator,
            ),
        }
        records.append(_base_record(
            source_hash=ir.source.content_hash,
            kind="flow_diagram",
            logical_record_id=_stable_id("logical-flow-diagram", ir.source.content_hash, container.element_id),
            group_index=1,
            group_count=1,
            section_path=section_path,
            pieces=[piece],
            sort_key=_record_sort_key(container.locator, 100_000, 0, 0),
            parent_id=container.element_id,
            source_name=ctx.source_name,
            document_context={
                "source_name": source_name,
                "sheet": container.locator.sheet,
                "slide": container.locator.slide,
            },
        ))
    return records


def _chunk_id_for_record(source_hash: str, record: dict) -> str:
    """recordの決定的chunk_id。`_finalize_chunks`（jsonl）と`_markdown`（rag.mdアンカー）の
    双方が同じ式で計算するため、生成時点で必ず1:1になる（§8.1・D1のアンカー方式）。"""
    return _stable_id("rag-chunk", source_hash, record["record_id"], RAG_CHUNKER_VERSION)


def _apply_hidden_sheet_visibility(records: list[dict], ir: evidence_ir.EvidenceIR) -> None:
    """非表示シート（`sheet` 要素の `hidden_sheet`/`very_hidden`）を、そのシート由来の全レコードへ
    KV 行として伝播する（提案A・§2.3）。

    `sheet` 要素はコンテナ型でレコード化されないため、シート単位の非表示はこの伝播が無いと
    rag.md のどこにも現れない——「廃止した旧版シート」を非表示にする運用が検索で見分けられず、
    可視シートの記述と同格に引用されてしまう。cell/element 自身の `visibility_reason`（覆い・
    取り消し線等）とは独立の行として**各レコードに**足す（チャンクは断片単独で自己完結する契約の
    ため、シート先頭の1レコードだけに付けるのでは足りない）。意味の断定はしない（非表示という
    事実だけを平文で述べる）。
    """
    hidden: dict[str, str] = {}
    for element in ir.elements:
        if element.type == "sheet" and element.visibility == "hidden" and element.locator.sheet:
            reason = element.extension.get("visibility_reason")
            hidden[element.locator.sheet] = (
                "完全非表示のシートにあります（通常の操作では再表示できません）"
                if reason == "very_hidden" else "非表示のシートにあります")
    if not hidden:
        return
    for record in records:
        sheet = ((record.get("region_context") or {}).get("sheet")
                 or (record.get("document_context") or {}).get("sheet"))
        label = hidden.get(sheet) if isinstance(sheet, str) else None
        if not label:
            continue
        line = f"シートの可視性: 「{label}」"
        for key in ("semantic_text", "markdown_text"):
            text = record.get(key)
            if isinstance(text, str) and line not in text:
                record[key] = text + "\n" + line


def _finalize_chunks(
    ir: evidence_ir.EvidenceIR,
    records: list[dict],
    *,
    source_rel_path: str,
) -> list[dict]:
    chunks: list[dict] = []
    ids = [_chunk_id_for_record(ir.source.content_hash, record) for record in records]
    by_logical: dict[str, list[str]] = defaultdict(list)
    for record, chunk_id in zip(records, ids, strict=True):
        by_logical[record["logical_record_id"]].append(chunk_id)
    for index, (record, chunk_id) in enumerate(zip(records, ids, strict=True)):
        siblings = [item for item in by_logical[record["logical_record_id"]] if item != chunk_id]
        coverage_statuses = sorted({
            status
            for citation in record.get("citations", [])
            if isinstance((status := citation.get("coverage_status")), str) and status
        })
        content_bases = {
            basis
            for citation in record.get("citations", [])
            if isinstance((basis := citation.get("content_basis")), str) and basis
        }
        has_raster_asset = any(
            isinstance(asset, dict) and asset.get("asset_role") == "shape_fill"
            for citation in record.get("citations", [])
            if isinstance(citation.get("object_metadata"), dict)
            for asset in citation["object_metadata"].get("assets", [])
            if isinstance(citation["object_metadata"].get("assets"), list)
        )
        chunk = {
            "chunk_id": chunk_id,
            "record_id": record["record_id"],
            "logical_record_id": record["logical_record_id"],
            "logical_block_id": record.get("logical_block_id", record["logical_record_id"]),
            "field_group_index": record["field_group_index"],
            "field_group_count": record["field_group_count"],
            "sibling_chunk_ids": siblings,
            "previous_chunk_id": ids[index - 1] if index > 0 else None,
            "next_chunk_id": ids[index + 1] if index + 1 < len(ids) else None,
            "content_type": record["kind"],
            "section_path": record["section_path"],
            "body": record["exact_text"],
            "citations": record["citations"],
            "relations": record.get("relations", []),
            "source_rel_path": source_rel_path,
            "coverage_statuses": coverage_statuses,
            "has_unresolved_coverage": bool({"metadata_only", "unsupported", "failed"} & set(coverage_statuses)),
            "needs_optional_vision": "pixel_only" in content_bases or has_raster_asset,
            "evidence_tier": "canonical",
            "source_content_hash": ir.source.content_hash,
            "renderer_version": RAG_RENDERER_VERSION,
            "chunker_version": RAG_CHUNKER_VERSION,
            "context_schema_version": context_ir.CONTEXT_IR_SCHEMA_VERSION,
            "context_analyzer_version": (
                context_ir.DOCX_CONTEXT_ANALYZER_VERSION if ir.source.file_type == "docx" else
                context_ir.PPTX_CONTEXT_ANALYZER_VERSION if ir.source.file_type == "pptx" else
                context_ir.PDF_CONTEXT_ANALYZER_VERSION if ir.source.file_type == "pdf" else
                context_ir.CONTEXT_ANALYZER_VERSION
            ),
            "identifier_role_analyzer_version": context_ir.IDENTIFIER_ROLE_ANALYZER_VERSION,
            "identifier_metadata_schema_version": context_ir.IDENTIFIER_METADATA_SCHEMA_VERSION,
            "identifier_mentions": record.get("identifier_mentions", []),
            "identifier_metadata_complete": record.get("identifier_metadata_complete", True),
            "identifier_mention_count": record.get("identifier_mention_count", 0),
            "identifier_mention_overflow_count": record.get("identifier_mention_overflow_count", 0),
            "identifier_mention_limit": record.get(
                "identifier_mention_limit", context_ir.IDENTIFIER_MAX_MENTIONS_PER_CHUNK,
            ),
            "record_identity_values_exact": record.get("record_identity_values_exact", []),
            "reference_values_exact": record.get("reference_values_exact", []),
            "unclassified_identifier_values_exact": record.get("unclassified_identifier_values_exact", []),
        }
        for key in ("parent_id", "source_name", "document_context", "region_context", "record_keys",
                    "related_record_ids", "context_confidence", "ai_observation"):
            value = record.get(key)
            if value not in (None, {}, []):
                chunk[key] = value
        chunks.append(chunk)
    return chunks


def _markdown(
    ir: evidence_ir.EvidenceIR,
    source_name: str,
    records: list[dict],
    summary: dict[str, int],
    observation_set: ai_observation.AIObservationSet | None = None,
) -> str:
    lines = [
        "# AI検索用文書",
        "",
        f"原本: {source_name}",
        f"原本形式: {ir.source.file_type}",
        f"原本SHA-256: {ir.source.content_hash.removeprefix('sha256:')}",
        f"変換プロファイル: {ir.parser_profile} / {RAG_RENDERER_VERSION}",
        ("抽出範囲: " + "、".join(f"{key}={value}" for key, value in sorted(summary.items()))),
        "",
    ]
    if observation_set is not None:
        lines.extend([
            f"採用AI観測Set: {observation_set.observation_set_hash}",
            f"AI観測生成元: {observation_set.provider}/{observation_set.model}（{observation_set.execution_mode}）",
            "AI観測は原本確定値ではない。",
            "",
        ])
    last_section: tuple[str, ...] | None = None
    for record in records:
        # D1（rag.mdを正本にする・§8.1）: このchunkの見出し類より前にアンカーを置く——
        # 索引時（`es_index`）はアンカー間（次のアンカー直前まで）をこのchunkの索引本文にするため、
        # 見出し（## section/### key）もこのchunkの検索文脈として含まれる。
        chunk_id = _chunk_id_for_record(ir.source.content_hash, record)
        lines.append(f"<!-- chunk:{chunk_id} -->")
        section = tuple(record["section_path"])
        if section != last_section:
            lines.extend(["## " + " / ".join(section), ""])
            region_range = record.get("region_context", {}).get("cell_range")
            if region_range:
                lines.extend([f"原本領域: {region_range}", ""])
            last_section = section
        key_heading = " / ".join(
            f"{item['label'].split(' > ')[-1]}「{item['value']}」"
            for item in record.get("record_keys", [])
        )
        if key_heading:
            suffix = (f"（分割 {record['field_group_index']}/{record['field_group_count']}）"
                      if record["field_group_count"] > 1 else "")
            lines.extend([f"### {key_heading}{suffix}", ""])
        elif record["field_group_count"] > 1:
            lines.extend([
                f"### 同一論理レコードの分割 {record['field_group_index']}/{record['field_group_count']}", "",
            ])
        lines.extend([record["markdown_text"], ""])
    return "\n".join(lines).rstrip() + "\n"


def validation_errors(ir: evidence_ir.EvidenceIR, result: RenderedEvidence) -> list[str]:
    errors: list[str] = []
    element_ids = {element.element_id for element in ir.elements}
    coverage_ids = {item.coverage_id for item in ir.coverage}
    cited_ids = {
        citation.get("evidence_id")
        for chunk in result.chunks
        for citation in chunk.get("citations", [])
    }
    expected_cells = {
        element.element_id for element in ir.elements
        if element.type == "cell" and _value_text(element.value)
    }
    missing_cells = expected_cells - cited_ids
    if missing_cells:
        errors.append(f"cell_coverage_missing:{len(missing_cells)}")
    expected_non_table = {
        element.element_id for element in ir.elements
        if element.type not in _CONTAINER_TYPES and element.type != "cell"
    }
    missing_non_table = expected_non_table - cited_ids
    if missing_non_table:
        errors.append(f"element_coverage_missing:{len(missing_non_table)}")
    for chunk in result.chunks:
        if any(
            (
                citation.get("evidence_id") not in element_ids
                and citation.get("coverage_id") not in coverage_ids
            )
            for citation in chunk.get("citations", [])
        ):
            errors.append(f"unknown_evidence:{chunk.get('chunk_id')}")
        if not chunk.get("citations"):
            errors.append(f"citation_missing:{chunk.get('chunk_id')}")
        for relation in chunk.get("relations", []):
            if relation.get("source_id") not in element_ids or relation.get("target_id") not in element_ids:
                errors.append(f"relation_endpoint:{chunk.get('chunk_id')}")
        if not isinstance(chunk.get("source_rel_path"), str) or not chunk["source_rel_path"]:
            errors.append(f"source_rel_path_missing:{chunk.get('chunk_id')}")
        if chunk.get("evidence_tier") != "canonical":
            errors.append(f"canonical_tier_missing:{chunk.get('chunk_id')}")
        if not isinstance(chunk.get("coverage_statuses"), list):
            errors.append(f"coverage_statuses_missing:{chunk.get('chunk_id')}")
        if not isinstance(chunk.get("has_unresolved_coverage"), bool):
            errors.append(f"unresolved_flag_missing:{chunk.get('chunk_id')}")
        if not isinstance(chunk.get("needs_optional_vision"), bool):
            errors.append(f"vision_flag_missing:{chunk.get('chunk_id')}")
    cited_coverage_ids = {
        citation.get("coverage_id")
        for chunk in result.chunks
        for citation in chunk.get("citations", [])
    }
    unresolved_coverage_ids = {
        item.coverage_id
        for item in ir.coverage
        if item.status in {"unsupported", "failed"} and _coverage_basis(item) != "none"
    }
    missing_notices = unresolved_coverage_ids - cited_coverage_ids
    if missing_notices:
        errors.append(f"coverage_notice_missing:{len(missing_notices)}")
    by_logical: dict[str, list[dict]] = defaultdict(list)
    for chunk in result.chunks:
        by_logical[chunk["logical_record_id"]].append(chunk)
    for logical_id, chunks in by_logical.items():
        expected_count = chunks[0]["field_group_count"]
        indexes = sorted(chunk["field_group_index"] for chunk in chunks)
        if len(chunks) != expected_count or indexes != list(range(1, expected_count + 1)):
            errors.append(f"field_group_incomplete:{logical_id}")
        expected_siblings = {chunk["chunk_id"] for chunk in chunks}
        for chunk in chunks:
            if set(chunk["sibling_chunk_ids"]) != expected_siblings - {chunk["chunk_id"]}:
                errors.append(f"sibling_mismatch:{chunk['chunk_id']}")
    # renderer自身がMarkdown pipe tableを作らない。原値中のliteral pipeは保存対象なので禁止しない。
    if any(line.startswith("| ") and line.endswith(" |") for line in result.markdown.splitlines()):
        errors.append("generated_pipe_table")
    # D1: rag.mdのアンカーとjsonlのchunk_idは生成時点で必ず1:1（`_chunk_id_for_record`を双方が
    # 同じ式で呼ぶため）。崩れていたら生成ロジック自体のバグであり、ES側の縮退に任せず生成時に落とす。
    anchor_ids = set(re.findall(r"^<!-- chunk:(\S+) -->$", result.markdown, flags=re.MULTILINE))
    chunk_ids = {chunk["chunk_id"] for chunk in result.chunks}
    if anchor_ids != chunk_ids:
        errors.append("rag_md_anchor_mismatch")
    return sorted(set(errors))


def render(
    ir: evidence_ir.EvidenceIR,
    *,
    source_name: str,
    observation_set: ai_observation.AIObservationSet | None = None,
) -> RenderedEvidence:
    """Evidence IRから同一record集合を経由してMarkdownとchunkを決定的に生成する。"""
    ir_errors = evidence_ir.validation_errors(ir)
    if ir_errors:
        raise ValueError("invalid Evidence IR: " + ",".join(ir_errors))
    incomplete_scans = [
        item for item in ir.coverage
        if item.status == "unsupported" and _coverage_reason(item) == "scan_cap_reached"
    ]
    if incomplete_scans:
        raise ValueError(f"incomplete source scan: {len(incomplete_scans)}")
    if observation_set is not None:
        observation_errors = ai_observation.validation_errors(observation_set, ir=ir)
        if observation_errors:
            raise ValueError("invalid AI Observation Set: " + ",".join(observation_errors))
    context = context_ir.build(ir, source_name=source_name)
    elements = {element.element_id: element for element in ir.elements}
    aliases, suppressed = _element_aliases(ir, elements)
    records = _table_records(ir, context, aliases) + _non_table_records(
        ir, context, aliases, suppressed)
    records.extend(_coverage_notice_records(ir, source_name))
    records.extend(_flow_diagram_records(ir, context))
    if observation_set is not None:
        records.extend(_ai_observation_records(ir, context, observation_set))
    records.sort(key=lambda record: record["_sort_key"])
    _apply_hidden_sheet_visibility(records, ir)
    _enrich_record_citations(records, ir)
    summary = dict(sorted(Counter(item.status for item in ir.coverage).items()))
    chunks = _finalize_chunks(ir, records, source_rel_path=source_name)
    result = RenderedEvidence(
        markdown=_markdown(ir, source_name, records, summary, observation_set),
        chunks=chunks,
        coverage_summary=summary,
    )
    errors = validation_errors(ir, result)
    if errors:
        raise ValueError("invalid rendered Evidence: " + ",".join(errors))
    return result


def chunks_to_jsonl(chunks: list[dict]) -> str:
    return "".join(json.dumps(chunk, ensure_ascii=False, sort_keys=True) + "\n" for chunk in chunks)


def write_chunks_atomic(path: str | Path, chunks: list[dict]) -> Path:
    """chunk配列を巨大な中間文字列へ複製せず、1件ずつJSONLへ原子書込する。"""
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    temporary = target.with_name(f"{target.name}.{os.getpid()}.{uuid.uuid4().hex}.tmp")
    try:
        with temporary.open("w", encoding="utf-8", newline="") as stream:
            for chunk in chunks:
                stream.write(json.dumps(chunk, ensure_ascii=False, sort_keys=True))
                stream.write("\n")
            stream.flush()
            os.fsync(stream.fileno())
        os.replace(temporary, target)
    except BaseException:
        temporary.unlink(missing_ok=True)
        raise
    return target
