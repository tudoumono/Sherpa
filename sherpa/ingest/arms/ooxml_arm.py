"""アーム①（OOXML 直）＝docx/pptx/xlsx を扱う既定アーム（値の権威）。pptx は `office_md.to_markdown`
へ委譲（挙動不変）、docx/xlsx は document-ir を人間向け MD 生成とも共有する（H2・下記参照）。

INGEST-MD §5.6 の「① OOXML 直（値の権威）」。変換ロジック自体は `office_md`／本モジュール（決定的・
LLM 不使用）に残し、本アームは受理判定と来歴メタ（method/confidence/notes）の付与を担う。

DOC-IR-001（docs/archive/proposals/2026-07-20-調査型RAG詳細修正計画.html §6.1・外部レビューの契約修正＝DOC-IR-001.5
で表の格納方式を修正済）: DOCX に限り document-ir-v1（文書標準構造）を構築し `ArmResult.document` に
載せる。IR 構築（`_build_docx_ir`）は `word/document.xml` を歩く。xlsx は今回（DOC-IR-003 時点）も
対象外（`document=None`・DOC-IR-004 以降）。pptx は DOC-IR-003（下記）で対象化済み。

DOC-IR-002（同提案書 §7 フェーズ2「Word」）: 隠し文字・削除本文（変更履歴）・ハイパーリンク先・
テキストボックス・脚注・コメント・ヘッダ/フッタ・ネスト表を要素化する。新規抽出（MD が表示しない情報）は
`sherpa/ingest/ooxml/word.py`（共通生抽出層）に一度だけ実装し、本モジュールはそれを消費して IR 要素を
組み立てるだけ（`_build_docx_ir` docstring に採番/order/rels 解決の設計判断を記す）。

DOC-IR-003（同提案書 §7 フェーズ2「PowerPoint」）: PPTX にも document-ir-v2 を並行構築する。スライド／
shape（テキストを持つもののみ要素化）／発表者ノートを要素化し、既存の A5 幾何オクルージョン判定
（`office_md.py` の `_pptx_bbox` 等）を `sherpa/ingest/ooxml/powerpoint.py` 経由で再利用して
`visibility`/`visibility_reason`/`source_map` へ構造化する（`_build_pptx_ir` docstring 参照）。pptx は
H2（下記）の対象外のため、MD 生成（`office_md._pptx_md`）とは引き続き独立（並行構築のまま）。

DOC-IR-004（同提案書 §7 フェーズ2「Excel」）: XLSX にも document-ir-v2 を構築する。シート／連続領域
（非空セルの4連結成分＝「表」）／数式／名前付き範囲／コメント／ハイパーリンク／外部ブック参照を要素化する。
共通生抽出層 `sherpa/ingest/ooxml/excel.py`（openpyxl・値/数式の2回ロード・cap 付き走査）を消費して IR を
組み立てる。

H2（`docs/proposals/2026-08-28-人間向けMDの刷新.md`）: docx/xlsx の人間向け MD（`{rel}.md`）は、
シート丸ごと1枚の打切り付きパイプ表（旧 `office_md._xlsx_md`）・素朴な `w:tbl` パイプ表（旧
`office_md._table_md` 直呼び）をやめ、本モジュールが構築する document-ir を `human_md.render_xlsx`/
`render_docx` へ渡して生成する。`OoxmlArm.convert()` は docx/xlsx について IR を1回だけ構築し、MD 生成と
`ArmResult.document` の両方に使い回す（xlsx の2回ロードを重ねて払わないための設計判断）。IR 構築に
失敗すれば docx/xlsx は未対応（fail-safe）——「MD 生成ロジックは IR 構築から独立しており IR 失敗の
影響を受けない」という旧来の契約は docx/xlsx について本スライスで終わった（pptx は変わらず独立のまま）。
"""
from __future__ import annotations

import hashlib
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from . import ArmResult
from .. import document_ir

# OOXML＝外部ライブラリ不要（docx/pptx は XML 直読み・xlsx は openpyxl）で常時 MD化できる。
_EXTS = {".docx", ".xlsx", ".pptx"}

# 浮動図形（`wp:anchor`）の関連付け（D-2・下記 `_docx_floating_anchor_facts` 参照）専用の名前空間。
_WP = "{http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing}"
# 画像本体（`pic:pic`・下記 `_docx_picture_count` 参照）の名前空間。インライン（`wp:inline`）/
# 浮動（`wp:anchor`）いずれの配置でも共通してこのタグで現れる。
_PIC = "{http://schemas.openxmlformats.org/drawingml/2006/picture}"

# MS-OFFCRYPTO で暗号化された OOXML は素の zip でなく OLE2/CFB コンテナに包まれる。ただし OLE2 で
# あること自体は旧 Office バイナリ（.doc/.xls/.ppt・非暗号化）とも共通のコンテナ形式のため、
# magic bytes 一致だけでは「暗号化されている」と確定できない。`olefile` で実際に
# `EncryptionInfo` ストリーム（MS-OFFCRYPTO の目印）が存在するかまで確認する。


def _looks_password_protected(p: Path) -> bool:
    """OLE2/CFB コンテナ内に `EncryptionInfo` ストリームがあるか（`olefile` で確認）。

    ストリーム構成まで見ることで、素の OLE2 ファイル（暗号化されていない旧バイナリ等）を
    誤って「パスワード保護」と判定しない。`olefile` 未導入/読めない/OLE2 でない場合は
    判定不能＝False（fail-safe）。
    """
    try:
        import olefile
    except ImportError:
        return False
    try:
        if not olefile.isOleFile(str(p)):
            return False
        with olefile.OleFileIO(str(p)) as ole:
            return ole.exists("EncryptionInfo")
    except Exception:
        return False


def _document_ir_failure_detail(p: Path, exc: Exception) -> str:
    """`document_ir_failed:<detail>` の detail を決める（閉じた語彙 `sherpa.ingest.failure_reasons` 準拠）。

    判別可能な既知パターンだけ意味のあるコードへ寄せる: `EncryptionInfo` ストリームがある
    OLE2/CFB コンテナ＝パスワード保護（暗号化）、zip/XML の構造破損（`BadZipFile`／XML
    `ParseError`／想定した要素が無い `KeyError`）＝ファイル破損。一般 `MemoryError` は
    明示的な入力上限超過とは確認できないため `other` のまま（実際に上限超過を
    検出する仕組みが無いのに size_exceeded と決め付けない）。それ以外は従来どおり例外クラス名の
    まま（`failure_reasons.classify` 側で `other` に分類される）。
    """
    if _looks_password_protected(p):
        return "password_protected"
    if isinstance(exc, (zipfile.BadZipFile, ET.ParseError, KeyError)):
        return "malformed_structure"
    return exc.__class__.__name__


# 抽出不完全の疑い（ING-1・静かな部分抽出検知）: 深いカバレッジ計算はしない・安価な整合チェックのみ。
# 宣言行数（openpyxl `ws.max_row`）に対し、実際に値が入っていた行が極端に少なく、かつ最終非空行が
# 宣言終端の手前で途切れていれば疑いに計上する（cap/予算打ち切り＝正常な自己申告
# 〔`sheet_sm["truncated"]`〕は除く・最終非空行が宣言終端に達している疎シートも除く＝
# `_build_xlsx_ir` 参照）。
_PARTIAL_XLSX_MIN_DECLARED_ROWS = 100     # これ未満の宣言行数は比率のブレが大きく誤検知しやすいため対象外
_PARTIAL_XLSX_MAX_EXTRACTED_RATIO = 0.05  # 抽出行/宣言行がこの比率未満なら疑い

# DOC-IR-001.5（修正2・契約修正 High#2）: JSON 形式の版（`document_ir.DOCUMENT_IR_SCHEMA_VERSION`）と、この
# アームの抽出処理自体の版を分離する。抽出対象（要素種別・採番規則・表の座標解決規則等）を増やしたら
# こちらを上げる＝JSON 形式（schema_version）を変えずとも派生が再生成される。
# DOC-IR-002 で v2 へ bump（隠し文字/削除本文/ハイパーリンク/テキストボックス/脚注/コメント/ヘッダ・フッタ/
# ネスト表を追加抽出＝`.document_ir_sig` drift 経由で `refresh_document_ir` が新要素入り IR を再生成する）。
# v3: `_docx_table_walk` の座標解決規則を変更した（`w:gridSpan`/`w:trPr/gridBefore` を Word 実仕様の
# 表最大列数 63 で打ち切り・`w:vMerge` 継続セルの可視本文を起点セルへ改行連結）。この IR は human_md
# レンダラの共通土台でもあるため、版を上げると document_ir_sig drift 経由で document.json/evidence/rag
# （さらに `office_md._current_human_md_sig()` 経由で人間向け `{rel}.md`）の両方が追随する。
# v4（L3・可視性・廃止表現の全形式展開）: `strike:N`（取り消し線 `w:strike`/`w:dstrike`）を追加抽出。
# v5（HM1 の docx 見送り分）: `DocumentIR.picture_count`（文書全体の画像枚数・`pic:pic` を数える）を追加。
DOCX_EXTRACTOR_VERSION = "docx-ooxml-v5"
# DOC-IR-003（PowerPoint 抽出器の初版）。docx とは独立した版番号（拡張子ごとに抽出器が別モジュール／
# 別関数のため、抽出処理を増やしたときに両方を巻き込まず片方だけ drift させられる＝`_current_document_ir_sig`
# が `docx=`/`pptx=`/`xlsx=` を別々の成分として持つ設計）。
PPTX_EXTRACTOR_VERSION = "pptx-ooxml-v2"
# DOC-IR-004（Excel 抽出器）。docx/pptx と同じく独立した版番号。v2: `excel.regions()` の占有判定に
# 背景色付きセル・結合セルの非anchorを加え、連結成分の分割にヒストグラム法の最大矩形反復抽出を導入
# （隣接する複数表の癒着解消）。`table:N` の座標解決規則が変わるため bump した。この値は
# `_current_document_ir_sig`/`_current_evidence_ir_sig` の `xlsx=` 成分にそのまま入る（`document_ir_sig_drift`/
# `evidence_ir_sig_drift` はこの成分が既存派生のマーカーと不一致なら drift を検知できる＝検知した
# 場合に何を行うかは下記「現在の契約」参照）。
# v3: `excel.DEFAULT_CAP_COLS` を Excel 実仕様値（16,384）へ引き上げ、`merged_map`/`filled_cells` の
# 結合範囲展開を使用範囲・セル数予算でクリップするよう変更した（`ooxml/excel.py` 参照）。座標解決結果が
# 変わるため bump した。
# v4（L3・可視性・廃止表現の全形式展開）: `strike:N`（取り消し線 `font.strike`）を追加抽出。
# v5（HM1）: `excel._MAX_RECT_DECOMPOSE_CELLS`（5,000→20,000）引き上げにより `regions()` の
# `table:N` 境界が変わりうる（従来 `split_budget_exhausted=True` で単一外接矩形へ縮退していた
# 連結成分の一部が、実際の癒着解消済み複数矩形へ分割されるようになる）。加えて `sheet:N` の
# `source_map` へ `picture_count`（画像の存在枚数・0枚なら省略）を追加。
#
# **現在の契約**: `worker.sync`（`_refresh_derived_representations`）がこの成分を含む
# `document_ir_sig_drift` を確認し、drift 時は `refresh_document_ir` が document.json／evidence／rag
# （RAG_ES 有効時は ES 索引）まで連鎖再生成する＝既存 world もこの版を上げるだけで追随する。人間向け
# `{rel}.md`（human_md 生成）は別系統の版管理（`office_md._current_human_md_sig()`／
# `{rel}.derived.json` の `asset_versions.human_md`）で追随する——この `XLSX_EXTRACTOR_VERSION`/
# `DOCX_EXTRACTOR_VERSION` は `_current_human_md_sig()` にも合成されるため、この版を上げると
# `human_md_sig_drift`/`refresh_human_md`（`office_md.py`）も1回だけ `{rel}.md` を作り直す
# （`.rag.md`/ES には触れない・単一 asset の選択的再生成）。
XLSX_EXTRACTOR_VERSION = "xlsx-ooxml-v5"

# document-ir を並行構築する対象拡張子（legacy .xls は対象外＝来歴汚染防止・`_build_xlsx_ir` docstring 参照）。
_IR_EXTS = {".docx", ".pptx", ".xlsx"}


class OoxmlArm:
    """docx/pptx/xlsx を扱う既定アーム（値・構造の権威）。pptx は `office_md.to_markdown` へ委譲
    （バイト一致）。docx/xlsx は document-ir を構築し、人間向け MD（`human_md`）と `ArmResult.document`
    の両方をそこから作る（H2・モジュール docstring 参照）。
    """
    name = "ooxml"

    def accepts(self, path) -> bool:
        return Path(path).suffix.lower() in _EXTS

    def convert(self, path) -> ArmResult | None:
        from .. import office_md
        p = Path(path)
        ext = p.suffix.lower()
        if ext == ".pptx":                         # H2 の対象外＝挙動不変（office_md.to_markdown に委譲）
            md = office_md.to_markdown(path)
            if md is None:
                return None
            notes: list[str] = []
            document = None
            try:
                document = _build_pptx_ir(p)
            except Exception as e:                 # IR 生成の失敗は握りつぶして md 継続（fail-safe）
                notes.append(f"document_ir_failed:{_document_ir_failure_detail(p, e)}")
            return ArmResult(md=md, method="ooxml", confidence=1.0, notes=notes, document=document)
        if ext not in _EXTS:
            return None
        # docx/xlsx（H2）: document-ir を1回だけ構築し、人間向け MD 生成と `ArmResult.document` の
        # 両方に使い回す（旧実装は MD 生成と IR 構築を独立に行っており、xlsx は2回ロードをさらに
        # 倍払っていた）。IR 構築に失敗すれば docx/xlsx はそのまま未対応（fail-safe＝MD だけ生き残る
        # 経路は無い＝旧来の「IR 失敗は MD に影響しない」契約は docx/xlsx について終わった）。
        notes = []
        try:
            document = _build_docx_ir(p) if ext == ".docx" else _build_xlsx_ir(p)
        except Exception as e:
            notes.append(f"document_ir_failed:{_document_ir_failure_detail(p, e)}")
            document = None
        if document is None:
            # fail-closed: bare None ではなく notes 付きの ArmResult を返す。呼び出し元
            # （`office_md._build_derived_into_staging`）がここで bare None を受け取ると
            # 失敗の詳細（どの例外/理由で IR が作れなかったか）が失われ、`document_ir_failed` を
            # 計上できないまま `.document_ir_sig` 等の版マーカーが「全件成功」を装って確定してしまう
            # （次回 sync の drift 検知が働かず再試行の契機を失う）。
            if not notes:                          # 例外は起きず構造的に None（本文/シート欠落等）
                notes.append("document_ir_failed:malformed_structure")
            return ArmResult(md=None, method="ooxml", confidence=0.0, notes=notes, document=None)
        if ext == ".xlsx":
            # DOC-IR-004: 未計算式（has_cached_value=False）が1件でもあれば警告を来歴に残す
            # （`_build_xlsx_ir` の戻り値は docx/pptx と同じ `DocumentIR | None` のまま保つため、
            # ここで要素を走査して数える＝シグネチャを増やさない設計判断）。
            uncached = sum(1 for e in document.elements
                           if e.type == "formula" and e.source_map.get("has_cached_value") is False)
            if uncached:
                notes.append(f"xlsx_uncached_formulas:{uncached}")
            truncated_sheets = sum(1 for e in document.elements
                                   if e.type == "sheet" and e.source_map.get("truncated"))
            if truncated_sheets:                              # RV Med #2: cap 打切りの黙認防止（来歴にも残す）
                notes.append(f"xlsx_truncated_sheets:{truncated_sheets}")
        elif ext == ".docx":
            # `_docx_table_walk` が付けた flags（列/行 span のクランプ・vMerge 継続セルの
            # 本文救済）を持つ表の件数を来歴に残す（xlsx の xlsx_truncated_sheets と同じ黙認防止）。
            flagged = [f for e in document.elements if e.type == "table"
                      for f in e.source_map.get("flags", [])]
            for flag_name in sorted(set(flagged)):
                notes.append(f"{flag_name}_tables:{flagged.count(flag_name)}")
        from .. import human_md
        md = human_md.render_docx(document) if ext == ".docx" else human_md.render_xlsx(document)
        if md is None:
            # IR 自体は構築できた（document is not None）が、レンダラが本文を1つも見つけられなかった
            # 場合（docx の空文書等・xlsx は `human_md.render_xlsx` が常に非 None を返すため実質 docx のみ）。
            # 失敗ではなく「作れたが空」なので document は残したまま返す（`document_ir_failed` は
            # 計上させない＝呼び出し元は `result.document is not None` でこの区別ができる）。
            return ArmResult(md=None, method="ooxml", confidence=1.0, notes=notes, document=document)
        return ArmResult(md=md, method="ooxml", confidence=1.0, notes=notes, document=document)


def _docx_floating_anchor_facts(p_el) -> list[dict]:
    """段落 `p_el` 内の浮動図形（`wp:anchor`）を、幾何断定なしの関連付けの事実として返す
    （D-2・`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §2.2）。

    Word はフロー配置のため、覆われる側の段落の実座標がレイアウト計算前に確定しない
    （xlsx のセル格子・pptx の EMU 座標と異なりここだけ本物の制約がある）。したがって幾何的な
    覆い判定はしない。代わりに「同一アンカー段落に浮動図形がある」という事実と、図形自身の
    name／テキスト／`behindDoc`（背面配置か前面かの別）だけを残す（意味の断定＝「廃止」等は
    しない、という可視性表現の共通思想）。インライン図形（`wp:inline`）はフロー内に
    収まり段落と重なりようがないため対象外。段落テキストが空（本文ゼロ）の場合は呼び出し側で
    そもそも host 要素を作らないため、ここでは判定しない。
    """
    from .. import office_md               # 遅延import（他 _build_*_ir 関数と同じ流儀・循環import回避）

    facts: list[dict] = []
    for anchor in p_el.findall(f".//{_WP}anchor"):
        doc_pr = anchor.find(f"{_WP}docPr")
        name = doc_pr.get("name", "") if doc_pr is not None else ""
        text = "".join(t.text or "" for t in anchor.iter(f"{office_md._W}t"))
        fact: dict = {"behind_doc": anchor.get("behindDoc") in {"1", "true"}}
        if name:
            fact["name"] = name
        if text:
            fact["text"] = text
        facts.append(fact)
    return facts


def _docx_picture_count(root) -> int:
    """文書本文（`root`＝`word/document.xml` のルート）に含まれる画像（`pic:pic`）の総数（HM1）。

    インライン（`wp:inline`）/浮動（`wp:anchor`）いずれの配置も、グループ化図形の中も
    `.iter()` で深さに関係なく数える（`ooxml/excel.picture_counts_by_sheet` と同じ方針）。
    図形・チャート・SmartArt は数えない（「画像」に限定＝人間向け注記の文言と一致させる）。
    ヘッダ/フッタ・脚注/コメント等の別パートは対象外（本文の画像だけを数える）。
    """
    return sum(1 for _ in root.iter(f"{_PIC}pic"))


def _build_docx_ir(p: Path) -> document_ir.DocumentIR | None:
    """DOCX から document-ir-v1 を構築する（DOC-IR-001 の本文/見出し/表に、DOC-IR-002 で以下を追加）。

    新規抽出（MD が表示しない構造）は `sherpa/ingest/ooxml/word.py`（共通生抽出層）に実装済みの純関数を
    消費するだけ（二重実装しない・パッケージ docstring 参照）。表（gridSpan/vMerge/gridBefore 座標解決）は
    引き続き本ファイル内（`_docx_table_walk`）。

    採番規則（**文書版内で決定的な要素ID**。World 再構築内の識別子であり、原本へ要素を前方追加すると後続の
    連番はずれる＝**版をまたぐ安定IDではない**）。DOC-IR-001 の型（`para:N`／`heading:N`／`table:N`）に
    DOC-IR-002 で以下を追加する。全て型ごとに独立したカウンタ（1-based・実際に要素を生成した時だけ増分）:

    - `hidden:N`（隠し文字）／`strike:N`（取り消し線・L3・可視性・廃止表現の全形式展開）／`deleted:N`
      （削除本文）／`link:N`（ハイパーリンク）／`textbox:N`（テキストボックス）: いずれも文書全体を通した
      連番（段落をまたいで増え続ける）。**`order`（本要素の値）はこれと別物**で、「同一段落内での出現順」を
      表す＝段落が変わるたびに 1 から数え直す。ただし5種を跨いだ真の XML 出現順（タグの物理的な前後関係）
      までは追わず、**種別ごとにまとめて**（hidden→strike→deleted→hyperlink→textbox の固定順）連番を振る
      設計にした（1段落に複数種が混在する稀なケースでの実装単純化・決定的である点は変わらない＝設計判断
      として明記）。`strike:N` は `w:strike`/`w:dstrike`（`w:vanish` とは異なり `visibility="visible"`・
      `status="active"` のまま＝取り消し線は本文の可視性を変えない幾何的事実。`visibility_reason="strike"`
      だけを立てて構造的に区別する＝`document_ir.Element` docstring の例外条項参照）。`parent_id` はホスト段落/
      見出しの `element_id`。段落全体が変更履歴の削除で本文ゼロ（`office_md._para_text` が空＝段落/見出し
      要素自体を作らない）の場合は `parent_id=None` とし、`source_map.paragraph_index` だけが原本位置の
      手がかりになる。
    - `hyperlink` の `target` 解決: `r:id`（`word/_rels/document.xml.rels` 経由の外部 URL）優先、無ければ
      `w:anchor`（文書内参照＝`"#"+anchor名`）。**どちらも無い/解決不能なら要素自体を作らない**（`word.py`
      `hyperlinks()` の判断＝遷移先の無いリンクは RAG にとって意味を持たない付随情報のため。表示文字列は
      元々ホスト段落の全文へ含まれておりロストしない）。
    - `footnote:N`／`endnote:N`（`word/footnotes.xml`／`endnotes.xml`）・`comment:N`（`word/comments.xml`）・
      `header:N`／`footer:N`（`word/header*.xml`／`footer*.xml`・zip 名ソート順で走査）: いずれも
      **その型内だけの連番**で、`order` もこの連番と同じ値（part 由来要素は body の読み順に混ぜようがない
      独立した文書パートのため）。`parent_id=None`。区切り線（footnote/endnote の separator／
      continuationSeparator）・空パート（本文ゼロ）は要素を作らない。パート自体の欠落/壊れは例外にせず
      その型を空扱いにする（`word.py` 側の fail-safe）。
    - ネスト表（`w:tc` 内の `w:tbl`）: 外側表の `cells` は従来どおり `tc` 直下の `w:p` のみ（ネスト本文を
      含まない＝現行 MD 実挙動と同じ）。ネスト表は独立した `table:N`（**トップレベル表と同じグローバル連番
      の続き**）とし、`parent_id=<外側 table:N>`、`order=<外側表内でのネスト表の出現順 1..n>`、
      `source_map={"table_index": <トップレベル表の table_index を子孫まで伝播した値>, "host_row": r,
      "host_column": c}`（`host_row`/`host_column` は**直接の親表内での** 1-based 位置＝孫より深い場合は
      その孫の親＝中間表内での位置）。再帰的に何段でも辿る（決定的なら深さ制限は不要という判断＝
      `_append_table` が自身を再帰呼び出しする）。空のネスト表（`office_md._table_md` が偽）は追加しない
      （トップレベル表の空判定と同じ基準）。
    """
    from .. import office_md
    from ..ooxml import word
    with zipfile.ZipFile(p) as z:
        root = ET.fromstring(z.read("word/document.xml"))
        rels = word.load_rels(z, "word/document.xml")
        header_names, footer_names = word.header_footer_names(z)
        header_parts = [(name, word.part_paragraphs(z, name)) for name in header_names]
        footer_parts = [(name, word.part_paragraphs(z, name)) for name in footer_names]
        footnote_list = word.footnotes(z)
        endnote_list = word.endnotes(z)
        comment_list = word.comments(z)
    body = root.find(f"{office_md._W}body")
    if body is None:
        return None

    content_hash = "sha256:" + hashlib.sha256(p.read_bytes()).hexdigest()
    source = document_ir.Source(path=str(p), content_hash=content_hash, file_type="docx")
    elements: list[document_ir.Element] = []

    body_order = 0
    para_seq = heading_seq = table_seq = 0
    hidden_seq = deleted_seq = link_seq = textbox_seq = strike_seq = 0
    para_index = table_index = 0                  # 0-based・生の <w:p>/<w:tbl> 出現順（スキップされても増分）

    def _append_paragraph_extras(p_el, host_id, para_idx, cell_map=None) -> None:
        """段落付随要素（hidden/deleted/link/textbox）を出現順（種別ごとにまとめた固定順）で追加する。

        `cell_map` 指定時＝**表セル内の段落**（RV High #1: セル内の削除本文/リンク先も失わない）:
        `source_map` は `{"table_index", "row", "column", "cell_paragraph_index"}`（row/column は cells と
        同じグリッド座標）になり、`host_id` にはホスト表の `element_id`（`table:N`）が入る。
        """
        nonlocal hidden_seq, deleted_seq, link_seq, textbox_seq, strike_seq

        def _sm(extra=None) -> dict:
            base = dict(cell_map) if cell_map is not None else {"paragraph_index": para_idx}
            if extra:
                base.update(extra)
            return base

        sub_order = 0
        extraction = document_ir.Extraction(method="ooxml", confidence=1.0)
        for text in word.hidden_runs(p_el):
            sub_order += 1
            hidden_seq += 1
            elements.append(document_ir.Element(
                element_id=f"hidden:{hidden_seq}", type="hidden_text", parent_id=host_id,
                order=sub_order, visibility="hidden", visibility_reason="hidden_run",
                status="active", text=text, cells=None,
                source_map=_sm(), extraction=extraction))
        for text in word.strike_runs(p_el):
            # 取り消し線は隠し文字と違い可視性を変えない（本文には残る＝「廃止」運用の幾何的事実だけを
            # 独立要素として構造化する・意味の断定はしない＝CLAUDE.md コメント規律
            # の設計思想を踏襲）。
            sub_order += 1
            strike_seq += 1
            elements.append(document_ir.Element(
                element_id=f"strike:{strike_seq}", type="strike_text", parent_id=host_id,
                order=sub_order, visibility="visible", visibility_reason="strike",
                status="active", text=text, cells=None,
                source_map=_sm(), extraction=extraction))
        for text in word.deleted_runs(p_el):
            sub_order += 1
            deleted_seq += 1
            elements.append(document_ir.Element(
                element_id=f"deleted:{deleted_seq}", type="deleted_text", parent_id=host_id,
                order=sub_order, visibility="visible", status="deleted", text=text, cells=None,
                source_map=_sm(), extraction=extraction))
        for link in word.hyperlinks(p_el, rels):
            sub_order += 1
            link_seq += 1
            elements.append(document_ir.Element(
                element_id=f"link:{link_seq}", type="hyperlink", parent_id=host_id,
                order=sub_order, visibility="visible", status="active", text=link["text"], cells=None,
                source_map=_sm({"target": link["target"]}), extraction=extraction))
        for text in word.textboxes(p_el):
            sub_order += 1
            textbox_seq += 1
            elements.append(document_ir.Element(
                element_id=f"textbox:{textbox_seq}", type="textbox", parent_id=host_id,
                order=sub_order, visibility="visible", status="active", text=text, cells=None,
                source_map=_sm(), extraction=extraction))

    def _append_table(tbl_el, parent_id, order, base_source_map) -> None:
        """1つの `w:tbl`（トップレベル/ネスト共通）を `table:N` として追加し、ネスト表を再帰的に追加する。

        セル内の段落付随要素（hidden/deleted/link/textbox）も `_append_paragraph_extras` の `cell_map` 経由で
        要素化する（RV High #1: 表セル内の削除本文・リンク先が完全に失われていた）。`parent_id` はこの表の
        `element_id`・座標は cells と同じグリッド位置（継続セルの `w:tc` 内も対象＝取りこぼさない）。
        """
        nonlocal table_seq
        cells, nested, cell_paras, flags = _docx_table_walk(tbl_el)
        table_seq += 1
        tid = f"table:{table_seq}"
        table_sm = dict(base_source_map)
        if flags:                                      # クランプ/救済の発生を来歴に残す
            table_sm["flags"] = flags
        elements.append(document_ir.Element(
            element_id=tid, type="table", parent_id=parent_id, order=order,
            visibility="visible", status="active", text=None, cells=cells,
            source_map=table_sm,
            extraction=document_ir.Extraction(method="ooxml", confidence=1.0)))
        table_index_value = base_source_map.get("table_index")
        for row, col, p_els in cell_paras:                 # RV High #1: セル内段落の付随要素も要素化
            for p_i, p_el in enumerate(p_els):
                _append_paragraph_extras(p_el, tid, None, cell_map={
                    "table_index": table_index_value, "row": row, "column": col,
                    "cell_paragraph_index": p_i})
        for n, (row, col, nested_tbl) in enumerate(nested, start=1):
            if not office_md._table_md(nested_tbl):        # 空のネスト表はトップレベルと同じ基準で出さない
                continue
            _append_table(nested_tbl, parent_id=tid, order=n,
                          base_source_map={"table_index": table_index_value, "host_row": row, "host_column": col})

    for el in body:
        if el.tag == f"{office_md._W}p":
            idx = para_index
            para_index += 1
            text = office_md._para_text(el).strip()
            host_id = None
            if text:
                lvl = office_md._heading_level(el)
                body_order += 1
                extraction = document_ir.Extraction(method="ooxml", confidence=1.0)
                # D-2: この段落に浮動図形（`wp:anchor`）がアンカーされていれば、幾何断定なしの
                # 関連付けの事実を source_map へ載せる（`_docx_floating_anchor_facts` docstring参照）。
                floating_anchors = _docx_floating_anchor_facts(el)
                if lvl:
                    heading_seq += 1
                    host_id = f"heading:{heading_seq}"
                    heading_source_map = {"paragraph_index": idx, "level": lvl}
                    if floating_anchors:
                        heading_source_map["floating_anchors"] = floating_anchors
                    elements.append(document_ir.Element(
                        element_id=host_id, type="heading", parent_id=None,
                        order=body_order, visibility="visible", status="active",
                        text=text, cells=None, source_map=heading_source_map,
                        extraction=extraction))
                else:
                    para_seq += 1
                    host_id = f"para:{para_seq}"
                    para_source_map = {"paragraph_index": idx}
                    if floating_anchors:
                        para_source_map["floating_anchors"] = floating_anchors
                    elements.append(document_ir.Element(
                        element_id=host_id, type="paragraph", parent_id=None,
                        order=body_order, visibility="visible", status="active",
                        text=text, cells=None, source_map=para_source_map,
                        extraction=extraction))
            # host_id は本文が空（段落全体が削除等）なら None のまま＝それでも隠し/削除/リンク/テキストボックス
            # は独立に抽出する（削除本文などは text が空でも消えない・parent_id=None＋paragraph_index で位置を示す）。
            _append_paragraph_extras(el, host_id, idx)
        elif el.tag == f"{office_md._W}tbl":
            idx = table_index
            table_index += 1
            if not office_md._table_md(el):        # MD に出ない表（空）は IR にも出さない
                continue
            body_order += 1
            _append_table(el, parent_id=None, order=body_order, base_source_map={"table_index": idx})

    header_seq = 0
    for name, paras in header_parts:
        text = "\n".join(paras)
        if not text:
            continue
        header_seq += 1
        elements.append(document_ir.Element(
            element_id=f"header:{header_seq}", type="header", parent_id=None, order=header_seq,
            visibility="visible", status="active", text=text, cells=None,
            source_map={"part": Path(name).name},
            extraction=document_ir.Extraction(method="ooxml", confidence=1.0)))

    footer_seq = 0
    for name, paras in footer_parts:
        text = "\n".join(paras)
        if not text:
            continue
        footer_seq += 1
        elements.append(document_ir.Element(
            element_id=f"footer:{footer_seq}", type="footer", parent_id=None, order=footer_seq,
            visibility="visible", status="active", text=text, cells=None,
            source_map={"part": Path(name).name},
            extraction=document_ir.Extraction(method="ooxml", confidence=1.0)))

    footnote_seq = 0
    for note in footnote_list:
        footnote_seq += 1
        elements.append(document_ir.Element(
            element_id=f"footnote:{footnote_seq}", type="footnote", parent_id=None, order=footnote_seq,
            visibility="visible", status="active", text=note["text"], cells=None,
            source_map={"note_id": note["note_id"]},
            extraction=document_ir.Extraction(method="ooxml", confidence=1.0)))

    endnote_seq = 0
    for note in endnote_list:
        endnote_seq += 1
        elements.append(document_ir.Element(
            element_id=f"endnote:{endnote_seq}", type="endnote", parent_id=None, order=endnote_seq,
            visibility="visible", status="active", text=note["text"], cells=None,
            source_map={"note_id": note["note_id"]},
            extraction=document_ir.Extraction(method="ooxml", confidence=1.0)))

    comment_seq = 0
    for c in comment_list:
        comment_seq += 1
        elements.append(document_ir.Element(
            element_id=f"comment:{comment_seq}", type="comment", parent_id=None, order=comment_seq,
            visibility="visible", status="active", text=c["text"], cells=None,
            source_map={"comment_id": c["comment_id"], "author": c["author"], "date": c["date"]},
            extraction=document_ir.Extraction(method="ooxml", confidence=1.0)))

    return document_ir.DocumentIR(schema_version=document_ir.DOCUMENT_IR_SCHEMA_VERSION, doc_id="",
                                  source=source, elements=elements,
                                  picture_count=_docx_picture_count(root))


def _build_pptx_ir(p: Path) -> document_ir.DocumentIR | None:
    """PPTX から document-ir-v2 を構築する（DOC-IR-003・提案書 §7 フェーズ2「PowerPoint」）。

    幾何・覆い判定は再実装しない: `office_md.py` の A5 ヘルパ（`_pptx_bbox`／`_pptx_has_solid_fill`／
    `_bbox_intersection_ratio`／`_OCCLUSION_RATIO`）と表示順解決（`_slide_order`）をそのまま消費する。
    新規抽出（非表示スライド・発表者ノート・スライドサイズ・shape 分類）は
    `sherpa/ingest/ooxml/powerpoint.py`（共通生抽出層）に実装済みの純関数を消費するだけ。MD 生成
    （`_pptx_md`／`_pptx_slide_texts*`）には一切触れない＝MD バイト不変。

    採番規則（型ごとに独立したカウンタ・1-based・実際に要素を生成した時だけ増分）:

    - `slide:N`（`N`=`office_md._slide_order` が返す表示順＝MD の `## スライド N` 見出しと同じ番号）。
      `text=None`（本文は shape 側が持つ）・`order=N`・
      `source_map={"slide": N, "part": <slide パート名>}`。非表示スライド（`p:sld @show="0"`）は
      `visibility="hidden"`／`visibility_reason="hidden_slide"`。
    - `shape:N`（`N` はスライドをまたいだグローバル連番・**テキストを持つ shape だけ**要素化する＝
      テキストの無い occluder 候補（無地塗りの空 shape／画像）自体は要素化しない）。`parent_id=<所属
      slide:N>`・`text=<shape 内テキストの "\\n" 結合>`・`order=<スライド内での出現順 1..>`（スライドが
      変わるたびに1から数え直す＝docx の hidden/deleted 系と同じ「グローバル element_id・ローカル
      order」の使い分け）。`source_map={"slide": n, "z_index": i}`（`z_index`＝
      `powerpoint.slide_shapes()` が返すエントリ配列内の 0-based 位置＝spTree 直下の文書順＝z順その
      もの。**テキストの無い occluder 候補にも z_index は振られる**＝`occluded_by` から前面 occluder を
      指すのに使う。テキストを持つ shape 同士は同じ z_index 空間で `covered_by_text` により直接
      element_id で参照し合う）。bbox が取れれば `bounds: [x, y, cx, cy]`（EMU・offset+extent。
      `_pptx_bbox` の `(x0,y0,x1,y1)` から `cx=x1-x0`/`cy=y1-y0` を復元）を追加、取れなければキー自体を
      省略する。グループ内（`p:grpSp` 直下）は幾何判定なし・bbox 不明のまま要素化し（現行 MD と同じ
      「グループはテキストのみ再帰抽出」を踏襲）、`source_map` に `{"group": true}` を追加する（この時
      `bounds` は出さない）。

      **覆い（既存 A5 判定の構造化）**: この shape より後（前面）に、無地塗りの空 shape または画像が
      あり `_bbox_intersection_ratio(自bbox, 前面bbox) >= _OCCLUSION_RATIO` なら
      `visibility="hidden"`／`visibility_reason="occluded"`、
      `source_map["occluded_by"] = {"kind": "solid_shape"|"picture", "z_index": j}`（`j`=前面 occluder の
      z_index。occluder 自体はテキストを持たず要素化しないため element_id ではなく z_index で指す）。
      MD の隠し候補マーカー（`_HIDDEN_MARKER`）と**同じ関数呼び出し**で判定するため、MD 側で
      `［隠し候補］` になる shape は IR でも必ず `occluded` になる（二重実装ではなく同一ロジックの共有）。
      **前面文字による上書き（新規・「非対応」「廃止」等の取り消し線的表現）**: 覆いに該当**しない**
      場合のみ、前面に**テキストを持つ** shape が同じ閾値以上重なっていれば（先着一致で1件だけ）
      `visibility` は `"visible"` のまま `source_map["covered_by_text"] = "shape:M"`
      （`M`=前面 shape の element_id）を追加する。意味の確定（「非対応」の断定等）は検索表現層の責務
      であり、IR は幾何的な前後関係だけを記録する（表側＝offset 順で処理するため、この shape 自身の
      element_id は既に確定済み＝前方参照はしない設計。covered_by_text で参照する側の shape の
      element_id は「このスライド内で先に一括採番」してから本判定を行う2段構え、詳細は実装コメント
      参照）。
      **画面外**: `powerpoint.slide_size()` が取れ、自 bbox がスライド矩形 `(0,0,cx,cy)` と交差しない
      （`_bbox_intersection_ratio(自bbox, スライド矩形) == 0.0`）なら `visibility="hidden"`／
      `visibility_reason="off_slide"`。優先順位は **off_slide → occluded → covered_by_text**
      （最初に成立した1つだけを記録・`visibility="hidden"` が確定したら以降の判定はしない）。
    - `notes:N`（`N` はスライドをまたいだグローバル連番・発表者ノートが**非空**のスライドだけ生成）:
      `parent_id=<所属 slide:N>`・`text=<発表者ノート>`・`order=1`（1スライドに高々1つしか無い子要素と
      いう位置づけで固定値＝`header:N`/`footer:N` の「型内だけの連番＝order もそれと同じ値」パターンとは
      異なり、本文と分離した独立ノートという性質上 order 自体に意味が無いため）。`source_map={"slide": n}`。

    rels 解決（`powerpoint.notes_for_slide`）: スライドの `_rels/slideN.xml.rels` から `notesSlide` の
    Target を辿る。壊れた/欠落した rels・notesSlide パートは例外にせず None へ縮退する（`ooxml/word.py`
    と同じ fail-safe 方針）。

    壊れて読めないスライドパート（`_slide_order` が返した名前が実際には読めない）は 1 スライド分の shape
    抽出だけをスキップする（`slide:N` 要素自体は作る＝IR は「スライドは存在した」事実を残す・fail-safe）。
    presentation.xml 自体が欠落/壊れている等で `_slide_order` が1件もスライドを返さない場合は
    `_build_docx_ir` の body 欠落時と同様 None を返す。
    """
    from .. import office_md
    from ..ooxml import powerpoint as pptx_ooxml

    with zipfile.ZipFile(p) as z:
        slide_names = office_md._slide_order(z)
        if not slide_names:
            return None
        hidden_names = pptx_ooxml.hidden_slide_names(z)
        size = pptx_ooxml.slide_size(z)
        slide_data: list[tuple[str, object, tuple[str, str] | None]] = []
        for name in slide_names:
            try:
                root = ET.fromstring(z.read(name))
            except (KeyError, ET.ParseError):
                root = None
            notes_payload = pptx_ooxml.notes_with_part_for_slide(z, name)
            slide_data.append((name, root, notes_payload))

    content_hash = "sha256:" + hashlib.sha256(p.read_bytes()).hexdigest()
    source = document_ir.Source(path=str(p), content_hash=content_hash, file_type="pptx")
    elements: list[document_ir.Element] = []
    extraction = document_ir.Extraction(method="ooxml", confidence=1.0)
    slide_rect = (0, 0, size[0], size[1]) if size is not None else None

    shape_seq = 0
    notes_seq = 0
    for n, (name, root, notes_payload) in enumerate(slide_data, start=1):
        sid = f"slide:{n}"
        is_hidden_slide = name in hidden_names
        elements.append(document_ir.Element(
            element_id=sid, type="slide", parent_id=None, order=n,
            visibility=("hidden" if is_hidden_slide else "visible"),
            visibility_reason=("hidden_slide" if is_hidden_slide else None),
            status="active", text=None, cells=None,
            source_map={"slide": n, "part": name}, extraction=extraction))

        if root is not None:
            entries = pptx_ooxml.slide_shapes(root)
            # サブパス1: このスライド内でテキストを持つ shape へ先に element_id/order を割り当てる
            # （covered_by_text は前面 shape の element_id を参照するが、その shape は z順で自分より
            # 後に処理されるため、実要素の組み立て前に採番だけ済ませておく必要がある）。
            entry_ids: dict[int, str] = {}
            entry_order: dict[int, int] = {}
            local_order = 0
            for i, entry in enumerate(entries):
                if not entry["has_text"]:
                    continue
                local_order += 1
                shape_seq += 1
                entry_ids[i] = f"shape:{shape_seq}"
                entry_order[i] = local_order

            def _own_state(i, entry):
                """entry 自身の隠れ状態（off_slide → occluded の優先順・covered_by_text は含まない）。

                RV Med #1: `covered_by_text` の参照可否判定にも使うため、要素組み立てとは独立に前計算する
                （前面文字 shape 自身が隠れている＝利用者に見えていないなら「可視の上書き」ではない）。
                """
                bbox = entry["bbox"]
                if bbox is not None and slide_rect is not None and \
                        office_md._bbox_intersection_ratio(bbox, slide_rect) == 0.0:
                    return "hidden", "off_slide", None
                if bbox is not None:
                    for j in range(i + 1, len(entries)):
                        other = entries[j]
                        if not other["occluder"] or other["bbox"] is None:
                            continue
                        if office_md._bbox_intersection_ratio(bbox, other["bbox"]) >= office_md._OCCLUSION_RATIO:
                            return "hidden", "occluded", {
                                "kind": "picture" if other["kind"] == "pic" else "solid_shape", "z_index": j}
                return "visible", None, None

            states = {i: _own_state(i, e) for i, e in enumerate(entries) if e["has_text"]}
            # サブパス2: 実要素を組み立てる（occluded_by/covered_by_text はサブパス1で確定した ID を参照）。
            for i, entry in enumerate(entries):
                if not entry["has_text"]:
                    continue
                text = "\n".join(entry["texts"])
                sm: dict = {"slide": n, "z_index": i}
                if entry["group"]:
                    sm["group"] = True
                elif entry["bbox"] is not None:
                    x0, y0, x1, y1 = entry["bbox"]
                    sm["bounds"] = [x0, y0, x1 - x0, y1 - y0]

                visibility, reason, occluded_by = states[i]
                if occluded_by is not None:
                    sm["occluded_by"] = occluded_by
                bbox = entry["bbox"]
                if visibility == "visible" and bbox is not None:
                    for j in range(i + 1, len(entries)):
                        other = entries[j]
                        if not other["has_text"] or other["bbox"] is None:
                            continue
                        if states[j][0] != "visible":       # RV Med #1: 自身が隠れている前面文字は参照しない
                            continue
                        if office_md._bbox_intersection_ratio(bbox, other["bbox"]) >= office_md._OCCLUSION_RATIO:
                            sm["covered_by_text"] = entry_ids[j]
                            break

                elements.append(document_ir.Element(
                    element_id=entry_ids[i], type="shape", parent_id=sid, order=entry_order[i],
                    visibility=("hidden" if is_hidden_slide else visibility),
                    visibility_reason=("hidden_slide_inherited" if is_hidden_slide else reason), status="active",
                    text=text, cells=None, source_map=sm, extraction=extraction))

        if notes_payload:
            notes_text, notes_part = notes_payload
            notes_seq += 1
            elements.append(document_ir.Element(
                element_id=f"notes:{notes_seq}", type="notes", parent_id=sid, order=1,
                visibility=("hidden" if is_hidden_slide else "visible"),
                visibility_reason=("hidden_slide_inherited" if is_hidden_slide else None),
                status="active", text=notes_text, cells=None,
                source_map={"slide": n, "part": notes_part}, extraction=extraction))

    return document_ir.DocumentIR(schema_version=document_ir.DOCUMENT_IR_SCHEMA_VERSION, doc_id="",
                                  source=source, elements=elements)


# Word の実仕様上のテーブル最大列数（UI で作成できる表の列数上限）。`w:gridSpan`/`w:trPr/gridBefore`
# の `w:val` は理論上どんな整数値も持てる（壊れた/悪意ある OOXML は巨大値を入れうる）ため、この値で
# クランプする（正典 §10 裁定#1「セル数の安全弁付き全量方針」と整合させる）: クランプ無しだと
# `human_md` 側のセル配置が実質無限大の列に広がり、レンダリング時のメモリ/計算量が原本のセル数と
# 無関係に膨張しうる。
_DOCX_MAX_TABLE_COLUMNS = 63


def _docx_table_cells(tbl_el) -> list[document_ir.Cell]:
    """1つの `w:tbl` から位置付きセル配列だけを組む（`_docx_table_walk` の cells 部分・後方互換の薄い別名）。

    座標解決規則の詳細は `_docx_table_walk` docstring を参照。
    """
    return _docx_table_walk(tbl_el)[0]


def _docx_table_walk(tbl_el) -> tuple[list[document_ir.Cell], list[tuple[int, int, object]],
                                      list[tuple[int, int, list]], list[str]]:
    """1つの `w:tbl` から位置付きセル配列（DOC-IR-001.5・修正1・`_build_docx_ir` docstring 参照）**と**、
    各セル直下のネスト表の位置（DOC-IR-002）を1回の走査で同時に組む。戻り値の4番目 `flags` は、
    このテーブルで発生した構造的な異常（クランプ・救済）の種別一覧（重複無し・空リスト＝異常無し）。

    - `column`: 行内の `w:tc` を左から歩き、直前までの `column_span` 累計（`w:trPr/gridBefore` の
      `w:val` があれば列開始をずらす）で実グリッド位置を求める（`w:gridSpan` で複数列を消費するセルの
      後続セルは消費した分だけ列がずれる）。`column_span` は `w:gridSpan` の `w:val`（無ければ 1）。
      列位置の**開始が** `_DOCX_MAX_TABLE_COLUMNS`（Word の実仕様上の表最大列数＝63）を超えるセルは
      **表として出さない**（`cells`/`active_vmerge`/`nested`/`cell_paras` のいずれにも加えない・
      `flags` に `"docx_column_overflow_dropped"` を追加）。壊れた/悪意ある OOXML の
      `w:gridSpan`/`w:trPr/gridBefore` に巨大値が入っていても、63列を超える座標へ複数セルを
      **同じクランプ後座標へ押し込めて衝突させる**（値の上書き消失・`row_span` の二重加算）ことは
      しない——開始位置が範囲内のセルは1列も失わず、範囲外のセルは値を諦める代わりに他のセルと
      座標が絶対に衝突しない、という二者択一にする。開始位置は範囲内だが `column_span` が
      63列を超えて伸びるセルは、63列で止まるよう `column_span` だけをクランプする
      （`flags` に `"docx_column_span_clamped"`・値そのものは失わない＝表示位置を丸めるだけ）。
    - `row_span`: 列位置（開始位置）ごとに進行中の縦マージを `active_vmerge`（列開始位置をキーにした
      辞書）で追跡する。**列上限超過セルを座標クランプではなく除外にしたことで**（上記）、
      `active_vmerge` のキーは常に実際の列位置と1対1対応し、63列を超える複数の raw 座標が同じキーへ
      衝突して1行あたり複数回 `row_span` を加算してしまう不具合は起きない。
      `w:vMerge w:val="restart"` を見つけたら**新規セルを起点として登録**し、以降の行で同じ列位置に現れる
      **継続セル**（`w:vMerge` はあるが `w:val` が無い、または `"continue"`）を見つけるたびに起点セルの
      `row_span` を1つずつ増やす。**継続セルは `cells` に要素を作らない**（内容を持たない merged 継続 tc・
      起点セルの `row_span` で表現するのが規約）。この増分は表の総行数（`w:tr` の総数）を超えない
      （超えようがない＝1行につき最大1回しか増分されないため。念のための防御的クランプを持ち、超過を
      検知したら `flags` に `"docx_row_span_clamped"` を追加する）。`w:vMerge` の無い通常セルに出会った
      列位置は縦マージ連鎖を打ち切る。ある行にその列位置の `w:tc` が現れなかった場合も連鎖を打ち切る
      （穴あき/不整形な表への fail-safe）。**孤児継続セル**（起点 restart が無い列位置での継続＝不整形）は
      黙って捨てず、通常セルとして `cells` へ出す（RV High #1: セルを一切失わない契約の維持・連鎖には
      登録しない）。
    - **継続セル自身の可視本文**: `w:vMerge` 継続セル（内容を持たない想定）が実際には
      `<w:t>` を持つ不整形な OOXML の場合、その本文を起点セルの `text` へ改行連結する（起点セルは
      `cells` に要素があるため、値をどこにも出さず捨てる silent-drop を避ける）。連結が発生したら
      `flags` に `"docx_vmerge_text_merged"` を追加する。
    - 行頭の省略列 `w:trPr/w:gridBefore` は列開始位置に反映する（RV Med #3: 無視すると座標がずれ、
      縦マージ連鎖を誤った列位置で数える）。
    - `role` は全セル `"unknown"`（ヘッダ判定は検索用表現側＝`table_semantics.py` の責務）。
    - セルの出現順（≒ `cells` の並び順）は原本の行→列の走査順（row-major）をそのまま保つ（継続セルは
      要素を作らないため出力に混じらない）。
    - ネスト表（DOC-IR-002）: 各 `w:tc` 直下の `w:tbl`（`tc.findall` の直接子＝1段目のみ検出。孫以降は
      呼び出し側＝`_append_table` がネスト表自身を再度この関数に渡すことで再帰的に見つかる）を
      `(row_idx, col_start, nested_tbl_el)` として `nested` に集める。**継続セルの `w:tc` にネスト表が
      あるような極端な不整形 OOXML でも `w:tc` 自身の row/col として記録する**（継続セルは `cells` に
      要素を作らない一方、ネスト表の位置情報は `w:tc` 自体が持つため独立に出す＝取りこぼさない）。
    """
    from .. import office_md
    cells: list[document_ir.Cell] = []
    nested: list[tuple[int, int, object]] = []
    # `(row, col, [直下の w:p ...])`＝セル内段落の付随要素抽出用（RV High #1）。継続セルの `w:tc` も含める
    # （cells には出さない一方、その tc が段落を持つ不整形でも削除本文等を取りこぼさない）。
    cell_paras: list[tuple[int, int, list]] = []
    flags: list[str] = []
    total_rows = len(tbl_el.findall(f"{office_md._W}tr"))
    active_vmerge: dict[int, document_ir.Cell] = {}     # 列開始位置 -> 進行中の縦マージの起点 Cell
    for row_idx, tr in enumerate(tbl_el.findall(f"{office_md._W}tr"), start=1):
        col = 1
        tr_pr = tr.find(f"{office_md._W}trPr")               # 行頭の省略列（gridBefore）だけ列開始をずらす
        if tr_pr is not None:
            gb = tr_pr.find(f"{office_md._W}gridBefore")
            if gb is not None:
                try:
                    raw_before = max(0, int(gb.get(f"{office_md._W}val", "0")))
                except (TypeError, ValueError):
                    raw_before = 0
                col = 1 + raw_before                          # クランプしない（下のセル単位の判定に委ねる）
        seen_cols: set[int] = set()
        for tc in tr.findall(f"{office_md._W}tc"):
            tc_pr = tc.find(f"{office_md._W}tcPr")
            raw_span = 1
            vmerge_val = None
            has_vmerge = False
            if tc_pr is not None:
                grid_span_el = tc_pr.find(f"{office_md._W}gridSpan")
                if grid_span_el is not None:
                    try:
                        raw_span = max(1, int(grid_span_el.get(f"{office_md._W}val", "1")))
                    except (TypeError, ValueError):
                        raw_span = 1
                vmerge_el = tc_pr.find(f"{office_md._W}vMerge")
                if vmerge_el is not None:
                    has_vmerge = True
                    vmerge_val = vmerge_el.get(f"{office_md._W}val")
            col_start = col
            if col_start > _DOCX_MAX_TABLE_COLUMNS:
                # 開始位置そのものが範囲外＝表として出さない（他のセルとの座標衝突を避けるため
                # 63列へ丸めない・クラス docstring 参照）。この tc 自体は cells/active_vmerge/
                # nested/cell_paras のいずれにも加えず、列位置だけ進めて次の tc へ進む。
                if "docx_column_overflow_dropped" not in flags:
                    flags.append("docx_column_overflow_dropped")
                col += raw_span
                continue
            span = raw_span
            if col_start + raw_span - 1 > _DOCX_MAX_TABLE_COLUMNS:
                span = _DOCX_MAX_TABLE_COLUMNS - col_start + 1     # 63列で止まるようcolumn_spanだけ縮める
                if "docx_column_span_clamped" not in flags:
                    flags.append("docx_column_span_clamped")
            seen_cols.add(col_start)
            origin = active_vmerge.get(col_start) if (has_vmerge and vmerge_val != "restart") else None
            if has_vmerge and vmerge_val != "restart" and origin is not None:
                # 継続セル: 起点セルの row_span を伸ばすだけ（表の総行数は超えない・念のためクランプ）。
                # `active_vmerge` は実座標（63列以内）だけをキーに持つため、複数の raw 座標が同じ
                # キーへ衝突して1行で複数回加算することはない（列上限超過セルは上で除外済み）。
                if origin.row_span < total_rows:
                    origin.row_span += 1
                elif "docx_row_span_clamped" not in flags:
                    flags.append("docx_row_span_clamped")
                # 継続セルが本来持たないはずの可視本文を持つ不整形 OOXML でも、その値を
                # 起点セルへ改行連結して残す（silent-drop 防止＝正典 §10 裁定#1 の全量方針と同じ理由）。
                cont_text = " ".join(office_md._para_text(pp).strip()
                                     for pp in tc.findall(f"{office_md._W}p")).strip()
                if cont_text:
                    origin.text = f"{origin.text}\n{cont_text}" if origin.text else cont_text
                    if "docx_vmerge_text_merged" not in flags:
                        flags.append("docx_vmerge_text_merged")
            else:
                # 通常セル・restart・**孤児継続**（起点無しの continue＝不整形）のいずれもセルとして出す
                # （RV High #1: 孤児継続を捨てると本文つきセルが黙って消える）。
                text = " ".join(office_md._para_text(pp).strip()
                                 for pp in tc.findall(f"{office_md._W}p")).strip()
                cell = document_ir.Cell(row=row_idx, column=col_start, text=text,
                                        row_span=1, column_span=span, role="unknown")
                cells.append(cell)
                if has_vmerge and vmerge_val == "restart":
                    active_vmerge[col_start] = cell         # 新しい縦マージの起点として登録
                else:
                    active_vmerge.pop(col_start, None)      # 通常セル＝この列位置の縦マージ連鎖を打ち切る
            for nested_tbl in tc.findall(f"{office_md._W}tbl"):    # DOC-IR-002: 直下のネスト表を位置付きで収集
                nested.append((row_idx, col_start, nested_tbl))
            p_els = tc.findall(f"{office_md._W}p")          # RV High #1: セル内段落（付随要素の抽出対象）
            if p_els:
                cell_paras.append((row_idx, col_start, p_els))
            col += raw_span
        for c in [c for c in active_vmerge if c not in seen_cols]:
            active_vmerge.pop(c, None)                      # この行に現れなかった列の連鎖も打ち切る（穴あき対策）
    return cells, nested, cell_paras, flags


def _build_xlsx_ir(p: Path) -> document_ir.DocumentIR | None:
    """XLSX から document-ir-v2 を構築する（DOC-IR-004・提案書 §7 フェーズ2「Excel」）。

    共通生抽出層 `sherpa/ingest/ooxml/excel.py` を消費するだけ（純関数群・二重実装しない）。H2
    （`docs/proposals/2026-08-28-人間向けMDの刷新.md`）以降、`office_md._xlsx_md`（人間向け MD）は
    本関数の戻り値をそのまま `human_md.render_xlsx` へ渡して生成する（シート丸ごと1枚の打切り付き
    パイプ表は撤去済み）。安全弁は**2系統**（正典 §10 裁定#1 参照）: (1) `excel.regions()`/`merged_map`
    が持つ cap（`excel.DEFAULT_CAP_CELLS` 等・走査自体をこの段階で頭打ちにする）と、
    (2) `human_md` 側の出力バイト予算（`_MAX_HUMAN_MD_BYTES`・render 時に総出力サイズを頭打ちに
    する）。前者だけでは MD の出力サイズ/メモリまでは守れない（結合セルの展開＝R5 で1セルの値が
    複数セルへ複製されるため、走査セル数を抑えても出力バイト数は別途膨張しうる）。

    **2回ロードのコスト（設計判断）**: `excel.load_two` が `data_only=True`（表示値）と `False`（数式）で
    同一ファイルを2回パースする。openpyxl は「キャッシュ済み表示値」と「数式文字列」を同時に返す API を
    持たないため、正確な `formula:N`（数式文字列＋has_cached_value）を得るには2回ロードが必要（1回のロード
    に妥協すると、数式セルの `text` が数式なのか計算結果なのか選べなくなる＝情報が失われる）。値用
    （`wb_values`）だけ `read_only=False`（通常ロード）を選ぶ（`excel.py` モジュール docstring の判断
    根拠＝結合セル/非表示行列/ハイパーリンク/コメントは read_only ワークシートでは取得できない）。
    数式用（`wb_formula`）はそれらのリッチな属性を必要としないため `read_only=True`（軽量な
    ストリーミング読み込み）にしてある（`excel.load_two` docstring 参照）——「重い」通常ロードは
    1ファイルにつき1回で足りる。行・列の絶対上限に加えて `excel.DEFAULT_CAP_CELLS` の総セル予算で
    実使用範囲を超える巨大シートの走査量を頭打ちにする（呼び出し側はシートの実使用範囲と動的capの
    小さい方までしか読まない）。

    採番規則（型ごとに独立したグローバルカウンタ・1-based・実際に要素を生成した時だけ増分・「文書版内で
    決定的な要素ID」＝`document_ir.py` docstring の一般契約に従う）:

    - `sheet:N`（`N`＝ブック内シート順＝`wb.worksheets` の順）。`text=None`（本文は table/formula 等の
      子要素側が持つ）・`order=N`・`source_map={"sheet": <シート名>}`。非表示シート（`sheet_state`＝
      `"hidden"`/`"veryHidden"`）は `visibility="hidden"`／`visibility_reason="hidden_sheet"`/`"very_hidden"`。
      画像（`xdr:pic`。グループ化図形の中も含む）が1枚以上あるシートは `source_map["picture_count"]`
      に枚数を追加する（HM1・人間向けMDの画像存在注記用＝`excel.picture_counts_by_sheet()` 参照。0枚の
      シートはキー自体を持たない）。図形（テキストボックス等）・チャート・SmartArt は数えない
      （「画像」に限定＝内容の解釈・OCR/VLM 観測は一切含まない存在だけの事実）。
    - `table:N`（シートをまたいだグローバル連番・`excel.regions()` が返す領域ごとに1つ）。`regions()` は
      非空セルに加え背景色付きセル・結合セルの非anchorも占有とみなして4連結成分を求め、ヒストグラム法の
      最大矩形反復抽出で連結成分を分割する（隣接する複数表が1つの巨大な外接矩形に癒着するのを防ぐ）ため、
      **1シートの1箇所の値クラスタから複数の `table:N` が生まれることがある**。`parent_id=<所属 sheet:N>`・
      `order=<シート内での領域出現順 1..>`（`regions()` が `(min_row, min_col)` 順で返すためそのまま採用）。
      `cells`＝矩形内の絶対座標（1-based row/column）を row-major で埋めた `Cell` 配列（空白セルは
      `text=""` で出す）。ただし**他領域が所有するセルは出さない**（外接矩形どうしが重なった場合の重複
      出力防止＝正本は `Region.cells`＝領域の所有座標。所有者の無い空白セルだけが重なり域で双方に `""`
      として現れうる＝実害の無い冗長性として許容）。`excel.merged_map()` の結合anchorには `row_span`/
      `column_span` を反映し、**非anchor（継続セル）は `cells` に出さない**（DOCX の `w:vMerge` 継続セルと
      同じ規約）。矩形は `excel.expand_regions_for_merges()` で結合 span まで拡張済み。`text` は常に
      `None`（値は `cells` 側が持つ・DOCX の表要素と同じ形）。
      `source_map={"sheet", "range": "A1:C10", "hidden_rows": [...], "hidden_columns": [...], "score": 0.0-1.0}`
      （`truncated: true` は cap 到達時のみ追加・`split_budget_exhausted: true` も同様に該当時のみ追加）:
      行/列そのものの非表示は `Cell` にフィールドを足さず（スキーマ変更を避ける）表要素の `source_map`
      に列挙する設計（この表の矩形範囲内に含まれる非表示行/列だけへ絞り込む＝シート全体の非表示行列
      ではなく「この表に関係する分」）。`score`＝`Region.score`（値セル密度から算出した表候補らしさ）
      をそのまま保持する。**分類・抑制はしない**: 領域が実際に意味のある「表」なのか単なる孤立セル・
      設定値なのかの判別はせず、スコアが低くても `table:N` として出す（低スコア領域を落とさない＝抑制は
      消費側の責務。意味分類は検索表現層＝RAG-REP の責務・`document_ir.Cell.role` が常に `"unknown"`
      なのもこの理由）。`split_budget_exhausted`＝`Region.split_budget_exhausted` をそのまま保持する
      （True の場合、この `table:N` の `cells` が埋める矩形範囲は「範囲内は全セル占有」を保証するヒスト
      グラム抽出そのものではなく、面積/反復回数/領域数いずれかの安全弁到達によるフォールバックの
      外接矩形であることを示す＝`excel.py::_split_component`/`Region` docstring 参照）。**消費側の契約**:
      このフラグが立った `table:N` は、隣接する別の表を巻き込んで癒着している可能性がある（`cells` に
      無い座標が矩形内に混在しうる＝表の見た目上の精度が通常より低い）ため、人間向け MD レンダラや
      rag.md レンダラ側で表の分割線・小見出しの精度を保証する用途には使わず、そのまま「1つの大きな表」
      として扱うか、精度が落ちる旨を注記した上で表示すること（値そのもの・セル座標の完全性は
      `cells` により通常どおり保たれる＝silent-drop はしない）。
    - `formula:N`（シートをまたいだグローバル連番・`=` で始まるセルごと）。`text=<数式文字列>`・
      `parent_id`＝**そのセル座標を所有する領域**（`Region.cells` 基準・RV Med #1＝bbox 重複時の誤親子化
      防止）の `table:N` を最優先。所有領域が無い場合（未計算式でキャッシュ値が無く非占有等）は外接矩形
      包含（`regions()` の返却順＝`(min_row, min_col)` 順で最初に一致・決定的なタイブレーク）→ それも
      無ければ `sheet:N` へ縮退。`order`＝**ホスト
      （table または sheet）ごとのローカル連番**（1-based・ホストが変わるたびに独立にカウント＝DOCX の
      `hidden:N`/`deleted:N` が段落ごとにローカル `order` を持つのと同じ考え方）。
      `source_map={"sheet", "cell", "has_cached_value": bool}`。未計算式（`has_cached_value=False`）が
      1件でもあれば `ArmResult.notes` に `"xlsx_uncached_formulas:<件数>"` を追記する（`OoxmlArm.convert()`
      側で本関数の戻り値を走査して行う＝`_build_docx_ir`/`_build_pptx_ir` と同じ `DocumentIR | None` の
      戻り値契約を崩さないための設計判断）。
    - `named_range:N`／`comment:N`／`hyperlink:N`／`strike:N`（取り消し線・L3・可視性・廃止表現の
      全形式展開）／`external_link:N`: いずれも DOCX の `footnote:N`/`header:N` と同じ「型内だけの連番・
      `order` もこの連番と同じ値・`parent_id=None`」パターン（名前付き範囲・コメント・ハイパーリンク・
      取り消し線はセル座標を `source_map` の `"cell"`/`"name"`+`"scope"` で直接指すため、table/sheet を
      親に持たせる木構造上の意味が薄い＝docx の footnote/comment が `parent_id=None` なのと同じ理由）。
      - `named_range:N`: `text=<参照先文字列>`・`source_map={"name", "scope": "workbook"|<シート名>}`。
        `excel.defined_names()` の順（ブック全体スコープを名前昇順→シート限定をシート順×名前昇順）を
        そのまま採用。
      - `comment:N`: `text=<コメント本文>`・`source_map={"sheet", "cell", "author"}`。
      - `hyperlink:N`: `text=<セル表示値>`・`source_map={"sheet", "cell", "target"}`。
      - `strike:N`: `visibility="visible"`／`status="active"`（取り消し線は隠し文字と異なり可視性を
        変えない幾何的事実・docx の `strike:N` と同じ設計方針）。`text=<セル表示値>`・
        `source_map={"sheet", "cell"}`。値の無いセル（`cell.value is None`）は出さない
        （`excel.strike_cells()` 参照）。**この独立要素とは別に**、`evidence_spike._adapt_document_ir`
        が同じ `(sheet, cell)` を持つ `cell` 要素（`table:N` の子）自身にも
        `extension["visibility_reason"]="strike"` を構造的に反映する
        （`_field_piece` が読むのは `cell` 要素自身の状態のため・二重実装ではなく同じ抽出結果を
        2箇所で消費するだけ）。
      - `external_link:N`: `text=None`（`target` はソースが持つ情報の複製になるため source_map のみに
        置く設計判断＝`slide:N`/`table:N` 等「本文は子要素側が持つ」構造要素と同じ扱い）・
        `source_map={"target": <zip 内 rels の Target>}`。`excel.external_link_targets()` の順（ソート済み）
        をそのまま採用。

    要素リスト内の並び順: `external_link:*` → `named_range:*` →（シート順に）`sheet:N` とその
    `table:*`/`formula:*`/`comment:*`/`hyperlink:*`/`strike:*`。前2種はブック全体スコープでどの
    シートにも属さないため先頭にまとめ、以降はシート単位で1シート分の子要素を続けて処理する
    （`_build_pptx_ir` がスライド単位で処理するのと同じ構成方針）。

    **図形/画像による覆い（L3・可視性・廃止表現の全形式展開）**: 本関数はセル値の抽出のみを行い、
    シート内の DrawingML（`xdr:twoCellAnchor` 等）は解析しない——xlsx の drawing 解析（cell_range・
    z_order・shape 種別）は既に `evidence_spike.py` の `_xlsx_objects` が Evidence IR 側で行っているため、
    ここで二重に実装しない。覆い判定（前面の図形/画像が下のセルの `cell_range` と幾何交差するか）は
    `_xlsx_objects` が Evidence IR 構築後に `cell` 要素の `visibility`/`extension["visibility_reason"]`
    を差し替える形で行う（`evidence_spike._Builder.mark_hidden` 参照）。

    **非表示行/列・取り消し線の cell 単位への反映（L3）**: 本関数自体は非表示行/列を `table:N` の
    `source_map`（`hidden_rows`/`hidden_columns`・この表の矩形範囲に絞り込んだ一覧）へ、取り消し線を
    独立 `strike:N` 要素へ、それぞれ格納するだけに留める（cell 単位の反映はしない）。
    `evidence_spike._adapt_document_ir` が `table:N` の子 `cell` 要素を組み立てる際にこの2つの情報を
    突き合わせ、対象セルの `visibility`（非表示行/列＝`"hidden"`）／`extension["visibility_reason"]`
    （`"hidden_row"`／`"hidden_column"`／取り消し線は `"strike"`）へ直接反映する——`_field_piece`
    （検索表現層）が読むのは cell 要素自身の状態であり、`table:N`/`strike:N` 側だけに置いても検索結果へ
    出てこないため。

    壊れて開けない（zip でない/OOXML でない）ファイルは `openpyxl.load_workbook` が例外を投げ、
    `OoxmlArm.convert()` 側の try/except で fail-safe に処理される（docx/pptx と同じ・本関数自体は
    ロード失敗を個別に握らない）。ワークシートが1枚も無い（理論上あり得ないが念のため）場合は None。
    """
    from .. import document_ir
    from ..ooxml import excel

    wb_values, wb_formula = excel.load_two(p)
    try:
        if not wb_values.worksheets:
            return None

        content_hash = "sha256:" + hashlib.sha256(p.read_bytes()).hexdigest()
        source = document_ir.Source(path=str(p), content_hash=content_hash, file_type="xlsx")
        elements: list[document_ir.Element] = []
        extraction = document_ir.Extraction(method="ooxml", confidence=1.0)

        states = {s["name"]: s["state"] for s in excel.sheet_states(wb_values)}
        reason_by_state = {"hidden": "hidden_sheet", "veryHidden": "very_hidden"}
        # ワークブックあたり1つだけ構築し、全シート分の filled_cells() 呼び出しへ使い回す（省略すると
        # シートごとに新規構築され、wb.loaded_theme の XML パースがシート数だけ繰り返されてしまう＝
        # `excel.filled_cells`/`excel.ColorResolver` docstring 参照）。
        color_resolver = excel.ColorResolver(wb_values)

        with zipfile.ZipFile(p) as z:
            ext_targets = excel.external_link_targets(z)
            picture_counts = excel.picture_counts_by_sheet(z)   # HM1: シート名→画像枚数（0枚のシートはキー無し）
        for i, target in enumerate(ext_targets, start=1):
            elements.append(document_ir.Element(
                element_id=f"external_link:{i}", type="external_link", parent_id=None, order=i,
                visibility="visible", status="active", text=None, cells=None,
                source_map={"target": target}, extraction=extraction))

        for i, dn in enumerate(excel.defined_names(wb_values), start=1):
            elements.append(document_ir.Element(
                element_id=f"named_range:{i}", type="named_range", parent_id=None, order=i,
                visibility="visible", status="active", text=dn["value"], cells=None,
                source_map={"name": dn["name"], "scope": dn["scope"]}, extraction=extraction))

        sheet_seq = table_seq = formula_seq = comment_seq = hyperlink_seq = strike_seq = 0
        for ws_v in wb_values.worksheets:
            sheet_seq += 1
            sid = f"sheet:{sheet_seq}"
            name = ws_v.title
            state = states.get(name, "visible")
            reason = reason_by_state.get(state)

            ws_f = wb_formula[name]
            cap_rows = excel.effective_cap_rows(ws_v.max_column)
            max_row = min(ws_v.max_row or 1, cap_rows + 1)
            max_col = min(ws_v.max_column or 1, excel.DEFAULT_CAP_COLS + 1)
            grid = [list(row) for row in
                    ws_v.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True)]
            merges = excel.merged_map(ws_v, cap_rows, excel.DEFAULT_CAP_COLS)
            filled = excel.filled_cells(ws_v, cap_rows, excel.DEFAULT_CAP_COLS, resolver=color_resolver)
            sheet_regions = excel.expand_regions_for_merges(
                excel.regions(grid, cap_rows, excel.DEFAULT_CAP_COLS, merged=merges, filled=filled),
                merges, cap_rows, excel.DEFAULT_CAP_COLS)   # 結合spanまでrangeを拡張
            hidden_r = set(excel.hidden_rows(ws_v))
            hidden_c = set(excel.hidden_cols(ws_v))

            sheet_sm: dict = {"sheet": name}
            if name in picture_counts:
                sheet_sm["picture_count"] = picture_counts[name]  # HM1: 画像の存在（枚数）のみ・内容には触れない
            truncated = excel.sheet_truncated(grid, cap_rows, excel.DEFAULT_CAP_COLS,
                                              sheet_max_row=ws_v.max_row or 1, sheet_max_col=ws_v.max_column or 1)
            if truncated:
                sheet_sm["truncated"] = True                      # RV Med #2: cap 外だけの領域も黙認しない（申告範囲基準）
            declared_rows = ws_v.max_row or 0
            if declared_rows >= _PARTIAL_XLSX_MIN_DECLARED_ROWS and not truncated:
                # cap 打切り（自己申告＝正常）とは独立の整合チェック: `grid` は declared_rows まで
                # 既に読み切っている（cap 未満のため）ので、新たな走査を増やさず判定できる。
                extracted_rows = 0
                last_nonempty_row = 0
                for i, row in enumerate(grid, start=1):
                    if any(v is not None for v in row):
                        extracted_rows += 1
                        last_nonempty_row = i
                # 最終非空行が宣言終端に達している疎シート（例 JPX-007.xlsx＝宣言501行・
                # 実値2行だが2行目が501行目にある）は除外する——抽出が宣言の最後まで正しく到達して
                # おり、内容がまばらなだけの正常な業務ファイルと区別が付かない（黙って途中で
                # 打ち切られた場合との違いは「宣言終端の手前で実データが途切れているか」）。
                if last_nonempty_row < declared_rows and extracted_rows < declared_rows * _PARTIAL_XLSX_MAX_EXTRACTED_RATIO:
                    sheet_sm["partial_extraction_suspected"] = True
                    sheet_sm["declared_rows"] = declared_rows
                    sheet_sm["extracted_rows"] = extracted_rows
            elements.append(document_ir.Element(
                element_id=sid, type="sheet", parent_id=None, order=sheet_seq,
                visibility=("hidden" if reason else "visible"), visibility_reason=reason,
                status="active", text=None, cells=None,
                source_map=sheet_sm, extraction=extraction))

            # RV Med #1: 座標→所有領域。外接矩形どうしが重なっても、非空セルは所有領域の table にだけ出す。
            owner: dict[tuple[int, int], int] = {}
            for ri, rg in enumerate(sheet_regions):
                for coord in rg.cells:
                    owner[coord] = ri

            region_tables: list[tuple[object, str]] = []          # [(Region, table_id)]（formula 親解決用）
            for local_order, rg in enumerate(sheet_regions, start=1):
                table_seq += 1
                tid = f"table:{table_seq}"
                region_tables.append((rg, tid))
                cells: list[document_ir.Cell] = []
                for r in range(rg.min_row, rg.max_row + 1):
                    for c in range(rg.min_col, rg.max_col + 1):
                        own = owner.get((r, c))
                        if own is not None and sheet_regions[own] is not rg:
                            continue                              # 他領域が所有する非空セルは重複出力しない（RV Med #1）
                        info = merges.get((r, c))
                        if info is not None and info["anchor"] != (r, c):
                            continue                              # 非anchor継続セルは出さない
                        row_span = info["row_span"] if info else 1
                        col_span = info["column_span"] if info else 1
                        raw = grid[r - 1][c - 1] if (r - 1 < len(grid) and c - 1 < len(grid[r - 1])) else None
                        cells.append(document_ir.Cell(
                            row=r, column=c, text=("" if raw is None else str(raw)),
                            row_span=row_span, column_span=col_span, role="unknown"))
                from openpyxl.utils import column_index_from_string
                sm = {
                    "sheet": name, "range": rg.range,
                    "hidden_rows": sorted(r for r in hidden_r if rg.min_row <= r <= rg.max_row),
                    "hidden_columns": sorted(
                        (c for c in hidden_c if rg.min_col <= column_index_from_string(c) <= rg.max_col),
                        key=column_index_from_string),
                    "score": round(rg.score, 3),
                }
                if rg.truncated:
                    sm["truncated"] = True
                if rg.split_budget_exhausted:
                    sm["split_budget_exhausted"] = True
                elements.append(document_ir.Element(
                    element_id=tid, type="table", parent_id=sid, order=local_order,
                    visibility="visible", status="active", text=None, cells=cells,
                    source_map=sm, extraction=extraction))

            host_local_order: dict[str, int] = {}
            for f in excel.formulas(ws_f, ws_v):
                formula_seq += 1
                coord = (f["row"], f["column"])
                # RV Med #1: 親解決は**所有領域を最優先**（bbox 重複時の先着 bbox への誤親子化を防ぐ）。
                # 未計算式（キャッシュ無し＝非占有）はどの領域にも属さないため bbox 包含 → sheet の順で縮退。
                own = owner.get(coord)
                if own is not None:
                    host_id = region_tables[own][1]
                else:
                    host_id = next((tid for rg, tid in region_tables
                                    if rg.min_row <= f["row"] <= rg.max_row
                                    and rg.min_col <= f["column"] <= rg.max_col), sid)
                host_local_order[host_id] = host_local_order.get(host_id, 0) + 1
                elements.append(document_ir.Element(
                    element_id=f"formula:{formula_seq}", type="formula", parent_id=host_id,
                    order=host_local_order[host_id], visibility="visible", status="active",
                    text=f["formula"], cells=None,
                    source_map={"sheet": name, "cell": f["cell"], "has_cached_value": f["has_cached"]},
                    extraction=extraction))

            for cm in excel.cell_comments(ws_v):
                comment_seq += 1
                elements.append(document_ir.Element(
                    element_id=f"comment:{comment_seq}", type="comment", parent_id=None,
                    order=comment_seq, visibility="visible", status="active",
                    text=cm["text"], cells=None,
                    source_map={"sheet": name, "cell": cm["cell"], "author": cm["author"]},
                    extraction=extraction))

            for hl in excel.cell_hyperlinks(ws_v):
                hyperlink_seq += 1
                elements.append(document_ir.Element(
                    element_id=f"hyperlink:{hyperlink_seq}", type="hyperlink", parent_id=None,
                    order=hyperlink_seq, visibility="visible", status="active",
                    text=hl["text"], cells=None,
                    source_map={"sheet": name, "cell": hl["cell"], "target": hl["target"]},
                    extraction=extraction))

            for sk in excel.strike_cells(ws_v):
                # 取り消し線は可視性を変えない幾何的事実（docx の strike:N と同じ設計方針）。
                strike_seq += 1
                elements.append(document_ir.Element(
                    element_id=f"strike:{strike_seq}", type="strike_text", parent_id=None,
                    order=strike_seq, visibility="visible", status="active",
                    text=sk["text"], cells=None,
                    source_map={"sheet": name, "cell": sk["cell"]},
                    extraction=extraction))

        return document_ir.DocumentIR(schema_version=document_ir.DOCUMENT_IR_SCHEMA_VERSION, doc_id="",
                                      source=source, elements=elements)
    finally:
        wb_values.close()
        wb_formula.close()
