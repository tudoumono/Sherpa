"""evidence_render.py の v1alpha7 実装と worker.py の分岐②③配線を、実xlsxの往復
（`office_md.build_derived` → `worker.sync()`）で結合検証する。

`worker.sync()`（公開エントリポイント）を通して駆動し、分岐選択そのものも検証対象に含める。
DB/Neo4j には接続しない（`store.get_world`/`store.world_lock`/`worker.world_state`/
`worker._derived_stale`/`worker.run`/`worker._run_locked` を monkeypatch）。ES は
`es_index.rag_es_enabled`/`index_world` を monkeypatch し、実 ES には接続しない。
"""
from __future__ import annotations

import contextlib
import json
import os

os.environ.setdefault("SHERPA_USE_FIXTURES", "1")

import openpyxl
import pytest

from sherpa import es_index, json_io, store, worlds
from sherpa.ingest import evidence_ir as IR
from sherpa.ingest import evidence_render as R
from sherpa.ingest import office_md, worker, world_graph, world_neo4j


def _build_world(tmp_path):
    wd = tmp_path / "world"
    wd.mkdir()
    dmd = tmp_path / "derived"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws["A1"], ws["B1"] = "No", "内容"
    ws["A2"], ws["B2"] = 1, "サンプル内容"
    wb.save(wd / "a.xlsx")
    rep = office_md.build_derived(wd, dmd)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
    return wd, dmd


def _force_rag_drift(dmd):
    """原本不変のまま renderer/chunker 版だけが上がった状況を、marker 内容をずらして再現する。"""
    marker = dmd / ".rag_sig"
    marker.write_text(marker.read_text(encoding="utf-8") + ";simulated-version-bump", encoding="utf-8")


def _force_evidence_ir_drift(dmd):
    marker = dmd / ".evidence_ir_sig"
    marker.write_text(marker.read_text(encoding="utf-8") + ";simulated-version-bump", encoding="utf-8")


def _force_document_ir_drift(dmd):
    """document_ir 版だけが独立に上がった状況（例: XLSX抽出器の版更新）を、marker 内容を
    ずらして再現する（`_force_rag_drift`/`_force_evidence_ir_drift` と同型）。"""
    marker = dmd / ".document_ir_sig"
    marker.write_text(marker.read_text(encoding="utf-8") + ";simulated-version-bump", encoding="utf-8")


def _build_world_with_asset(tmp_path):
    """`_build_world` に、`.assets/` 再生成を直接検証するための raster 画像（PNG）を1点追加した版。
    プレーンな xlsx だけでは shape-fill/picture のような asset を一切生成しない
    （実測確認済み）ため、`.assets` の再生成テストには raster 原本が要る。"""
    wd, dmd = _build_world(tmp_path)
    from PIL import Image
    Image.new("RGB", (4, 4), color="blue").save(wd / "scan.png")
    rep = office_md.build_derived(wd, dmd)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
    return wd, dmd


def _rag(dmd):
    """`.rag.md`／`.rag_chunks.jsonl`／`.assets/` は rag 層（dmd＝md 層の兄弟・§8.1 三階層）。"""
    return dmd.parent / "rag"


def _ir(dmd):
    """`.document.json`／`.evidence.json`／`.derived.json` は ir 層（dmd＝md 層の兄弟・§8.1 三階層）。"""
    return dmd.parent / "ir"


def _no_full_run(world, **kw):
    raise AssertionError("全再構築（run/_run_locked）は呼ばれない前提のテストで呼ばれた")


def _apply_world_monkeypatches(monkeypatch, wd, dmd):
    monkeypatch.setattr(worlds, "world_dir", lambda world: wd)
    monkeypatch.setattr(worlds, "derived_md_dir", lambda world: dmd)
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: False)

    monkeypatch.setattr(worker, "world_state", lambda world: ("sig", {}))
    monkeypatch.setattr(store, "get_world",
                         lambda world: {"last_sig": "sig", "last_manifest": {}, "last_doc_count": 0})
    monkeypatch.setattr(worker, "_derived_stale", lambda world: False)

    @contextlib.contextmanager
    def _noop_lock(world_id):
        yield
    monkeypatch.setattr(store, "world_lock", _noop_lock)
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: False)
    # 全再構築（`run()`／sidecar欠落フォールバックの`_run_locked()`）はどちらも既定で未使用の
    # 前提を明示する（このfixtureを使う各テストが drift/sidecar 状態を個別に設定する）。
    monkeypatch.setattr(worker, "run", _no_full_run)
    monkeypatch.setattr(worker, "_run_locked", _no_full_run)


@pytest.fixture
def _world(tmp_path, monkeypatch):
    wd, dmd = _build_world(tmp_path)
    _apply_world_monkeypatches(monkeypatch, wd, dmd)
    return {"wd": wd, "dmd": dmd}


@pytest.fixture
def _world_with_asset(tmp_path, monkeypatch):
    wd, dmd = _build_world_with_asset(tmp_path)
    _apply_world_monkeypatches(monkeypatch, wd, dmd)
    return {"wd": wd, "dmd": dmd}


def test_rag_refresh_regenerates_once_then_noop(_world_with_asset, monkeypatch):
    """rag drift のみ（renderer/chunker 版だけの更新）は軽量経路 `refresh_rag()` を1回だけ呼び
    （`refresh_evidence_ir()` は呼ばない）、stale 化させた既存の `.rag.md`／`.rag_chunks.jsonl`／
    `.assets` の中身を実際に正しい内容へ置換する（内容不変＝no-opではなく、置換が実際に
    起きたことを直接固定する）。"""
    dmd = _world_with_asset["dmd"]
    correct_md = (_rag(dmd) / "a.xlsx.rag.md").read_text(encoding="utf-8")
    correct_chunks = (_rag(dmd) / "a.xlsx.rag_chunks.jsonl").read_text(encoding="utf-8")
    asset_names_before = sorted(p.name for p in (_rag(dmd) / "scan.png.assets").iterdir())
    assert asset_names_before                            # 前提: PNGのassetが少なくとも1件ある

    _force_rag_drift(dmd)
    assert office_md.rag_sig_drift(dmd) is True

    # stale化: 既存の .rag.md／.rag_chunks.jsonl／.assets 中身を壊れた内容で上書きしておく。
    (_rag(dmd) / "a.xlsx.rag.md").write_text("STALE", encoding="utf-8")
    (_rag(dmd) / "a.xlsx.rag_chunks.jsonl").write_text("STALE", encoding="utf-8")
    for name in asset_names_before:
        (_rag(dmd) / "scan.png.assets" / name).write_bytes(b"STALE")

    evidence_calls: list[int] = []
    rag_calls: list[int] = []
    original_refresh_evidence_ir = office_md.refresh_evidence_ir
    original_refresh_rag = office_md.refresh_rag

    def _tracked_refresh_evidence_ir(*a, **kw):
        evidence_calls.append(1)
        return original_refresh_evidence_ir(*a, **kw)

    def _tracked_refresh_rag(*a, **kw):
        rag_calls.append(1)
        return original_refresh_rag(*a, **kw)
    monkeypatch.setattr(office_md, "refresh_evidence_ir", _tracked_refresh_evidence_ir)
    monkeypatch.setattr(office_md, "refresh_rag", _tracked_refresh_rag)

    res = worker.sync("w")
    assert res["status"] == "unchanged" and res["changed"] is False
    assert office_md.rag_sig_drift(dmd) is False        # RAG_ES無効＝refresh自身が確定する
    assert rag_calls == [1] and evidence_calls == []    # 軽量経路のrefresh_rag()だけが1回呼ばれる

    assert (_rag(dmd) / "a.xlsx.rag.md").read_text(encoding="utf-8") == correct_md              # 置換された
    assert (_rag(dmd) / "a.xlsx.rag_chunks.jsonl").read_text(encoding="utf-8") == correct_chunks
    asset_names_after = sorted(p.name for p in (_rag(dmd) / "scan.png.assets").iterdir())
    assert asset_names_after == asset_names_before      # 同じ内容から再生成＝ファイル名（内容ハッシュ）も同じ
    for name in asset_names_after:
        assert (_rag(dmd) / "scan.png.assets" / name).read_bytes() != b"STALE"                  # 中身は再生成された

    res2 = worker.sync("w")
    assert res2["status"] == "unchanged"                 # 直後の再syncはno-op（backfill/ES確認のみ）


def test_rag_refresh_indexes_es_once_with_content_sig(_world, monkeypatch):
    dmd = _world["dmd"]
    _force_rag_drift(dmd)
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    calls = []

    def _index_world(world, content_sig=None, **kw):
        calls.append(content_sig)
        return {"available": True, "indexed": 1, "chunks": 1}
    monkeypatch.setattr(es_index, "index_world", _index_world)

    res = worker.sync("w")
    assert res["status"] == "unchanged"
    assert calls == ["sig"]                              # content_sig明示・1回だけ
    assert office_md.rag_sig_drift(dmd) is False          # ES成功→workerがwrite_rag_sig_markerで確定


def test_rag_refresh_es_failure_leaves_marker_for_retry(_world, monkeypatch):
    dmd = _world["dmd"]
    _force_rag_drift(dmd)
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    monkeypatch.setattr(es_index, "index_world", lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 0, "chunks": 0, "error": "bulk_failed"})

    worker.sync("w")
    assert office_md.rag_sig_drift(dmd) is True           # 未確定のまま＝次回syncが再試行する

    # 次回 sync で ES が復旧すれば分岐③からやり直されて収束する（無限ループしない範囲で確認）。
    monkeypatch.setattr(es_index, "index_world", lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 1, "chunks": 1})
    worker.sync("w")
    assert office_md.rag_sig_drift(dmd) is False          # 2回目のES成功で収束する


def test_llm_rag_rewrite_reflects_mention_edges_in_graph(_world, monkeypatch):
    """rv-s2-mention #1: `.rag.md` の軽量書換え後（`_reindex_after_rag_rewrite`＝`_llm_render_pass`／
    `regenerate_rag_rule_only` が呼ぶ経路）、Neo4j のグラフ（言及エッジ）が追随することを固定する。
    以前は ES だけ更新し `build_world_graph`→`load_world` を経由しないため、言及エッジが陳腐化した
    まま固定されていた。ここでは実際に `world_graph.build_world` が構築した edges が
    `world_neo4j.load_world` へ渡っていることを、SAMPLE への言及エッジの有無で直接確認する。
    """
    wd = _world["wd"]
    dmd = _world["dmd"]
    # `corpus_docs.iter_world_documents` は `.rag.md` の場所を `worlds.derived_rag_dir` で解決する
    # （`derived_md_dir` とは別の getter・§8.1 三階層）——`_apply_world_monkeypatches` は
    # `derived_md_dir` しか差し替えないため、ここで揃えて差し替える（辞書突合の対象決定に効く
    # ため、事前条件確認より前に揃える）。
    monkeypatch.setattr(worlds, "derived_rag_dir", lambda world: _rag(dmd))
    (wd / "sample.cbl").write_text(
        "       IDENTIFICATION DIVISION.\n       PROGRAM-ID. SAMPLE.\n       PROCEDURE DIVISION.\n",
        encoding="utf-8")

    # 事前条件: 書換え前の rag.md は SAMPLE に言及していない＝言及エッジは無い。
    nodes0, edges0, flags0 = world_graph.build_world(wd, "w")
    assert flags0 == []
    assert not [e for e in edges0 if e["type"] == "DOCUMENTS" and e.get("via") == "mention"]

    # LLM 成形／規則版再生成が rag.md を書き換えた状況を模擬する（本文へ SAMPLE への言及を追加）。
    rag_md_path = _rag(dmd) / "a.xlsx.rag.md"
    rag_md_path.write_text(rag_md_path.read_text(encoding="utf-8") + "\nSAMPLE を使う。\n", encoding="utf-8")

    monkeypatch.setattr(world_neo4j, "_env", lambda: {"uri": "bolt://x", "user": "u", "pw": "p"})
    load_calls = []

    def _fake_load_world(nodes, edges, world, uri, user, pw):
        load_calls.append((nodes, edges))
        return (len(nodes), len(edges))
    monkeypatch.setattr(world_neo4j, "load_world", _fake_load_world)

    ok = worker._reindex_after_rag_rewrite("w")
    assert ok is True                                    # RAG_ES 無効（`_world` fixture の既定）＝無条件成功
    assert len(load_calls) == 1                           # グラフ反映が実際に呼ばれた
    _nodes, edges = load_calls[0]
    dst_cid = world_graph._cid("Module", "w", "sample.cbl", "SAMPLE")
    mentions = [e for e in edges if e["type"] == "DOCUMENTS" and e.get("via") == "mention"
               and e["dst"] == dst_cid]
    assert len(mentions) == 1, "書換え後の rag.md に追随した言及エッジが Neo4j 反映に含まれていない"


def test_document_ir_drift_cascades_to_evidence_and_rag_via_sync(_world):
    """document_ir 版だけが独立に上がった状況（evidence/rag 版は不変）でも、`sync()` 経由で
    document_ir→evidence→rag の順に再生成され、3種の版マーカーがすべて現行値へ更新される
    （`refresh_document_ir` 等の直呼びではなく `worker.sync()` 自身を評価対象にする）。"""
    dmd = _world["dmd"]
    _force_document_ir_drift(dmd)
    assert office_md.document_ir_sig_drift(dmd) is True
    assert office_md.evidence_ir_sig_drift(dmd) is False    # evidence/rag自身の版はまだ上がっていない
    assert office_md.rag_sig_drift(dmd) is False

    # 実ファイルを古い（壊れた）内容で上書きし、sync()がマーカーの帳尻合わせだけでなく
    # 実際に中身を再生成したことを確認できるようにする。
    (_ir(dmd) / "a.xlsx.document.json").write_text("stale-document-ir", encoding="utf-8")
    (_ir(dmd) / "a.xlsx.evidence.json").write_text("stale-evidence-ir", encoding="utf-8")
    (_rag(dmd) / "a.xlsx.rag.md").write_text("stale-rag-md", encoding="utf-8")

    res = worker.sync("w")
    assert res["status"] == "unchanged" and res["changed"] is False

    # document_ir版が上がっただけでも、evidence/rag自身の版判定によらず両方とも連鎖再生成される。
    assert office_md.document_ir_sig_drift(dmd) is False
    assert office_md.evidence_ir_sig_drift(dmd) is False
    assert office_md.rag_sig_drift(dmd) is False
    assert (_ir(dmd) / "a.xlsx.document.json").read_text(encoding="utf-8") != "stale-document-ir"
    assert (_ir(dmd) / "a.xlsx.evidence.json").read_text(encoding="utf-8") != "stale-evidence-ir"
    assert (_rag(dmd) / "a.xlsx.rag.md").read_text(encoding="utf-8") != "stale-rag-md"

    res2 = worker.sync("w")
    assert res2["status"] == "unchanged"                    # 直後の再syncはno-op（1回で収束）


def test_sidecar_missing_falls_back_to_full_run_via_sync(_world, monkeypatch):
    """真の sidecar 欠落（`.evidence.json` 削除）は全再構築（`_run_locked`）フォールバックへ倒れる
    ことを `sync()` 経由で確認する。"""
    dmd = _world["dmd"]
    (_ir(dmd) / "a.xlsx.evidence.json").unlink()
    run_calls = []
    monkeypatch.setattr(worker, "_run_locked", lambda world, **kw: run_calls.append(kw) or
                         {"status": "auto_published", "ledger": 0, "flags": []})
    res = worker.sync("w")
    # `op`（PART-6・Webhook 通知の情報用途のみ）: `sync()` の既定 "sync" がそのまま転送される。
    assert run_calls == [{"reflect": True, "created_by": "admin", "scan_root": None,
                          "run_id": None, "on_run_id": None, "op": "sync"}]
    assert res["changed"] is True


def test_old_schema_manifest_triggers_one_full_rebuild_then_stabilizes(_world, monkeypatch):
    """旧世代マニフェスト（`schema` キーも `assets` キーも持たない・`{"sidecars": [...]}` のみ）は、
    内容が実態と一致していても「要再生成」として全再構築へ一度だけフォールバックする（新形式の
    マニフェストへ移行するまではsidecarの過不足を判定できないため）。移行後の2回目のsyncは
    安定する（過検知ループにならない）。"""
    dmd = _world["dmd"]
    manifest_path = _ir(dmd) / "a.xlsx.derived.json"
    old_manifest = json_io.read_json(manifest_path, default=None)
    assert "schema" in old_manifest                          # 現行形式であることの前提確認
    json_io.write_text_atomic(manifest_path, json.dumps({"sidecars": old_manifest["sidecars"]}))
    assert office_md.rag_sidecars_missing(_world["wd"], dmd) is True

    run_calls = []

    def _simulated_full_rebuild(world, **kw):
        run_calls.append(kw)
        rep = office_md.build_derived(_world["wd"], dmd)      # run()が実際にやり直す派生生成の中核
        assert not rep.get("error")
        return {"status": "auto_published", "ledger": 0, "flags": []}
    monkeypatch.setattr(worker, "_run_locked", _simulated_full_rebuild)

    res1 = worker.sync("w")
    assert len(run_calls) == 1                                # 1回だけ全再構築
    assert res1["changed"] is True
    new_manifest = json_io.read_json(manifest_path, default=None)
    assert new_manifest.get("schema") == office_md._DERIVED_MANIFEST_SCHEMA_VERSION   # 現行形式で書き直された
    assert office_md.rag_sidecars_missing(_world["wd"], dmd) is False

    res2 = worker.sync("w")                                   # 2回目は安定（再度フルリビルドされない）
    assert len(run_calls) == 1
    assert res2["status"] == "unchanged" and res2["changed"] is False


@pytest.mark.parametrize("build_fixture", ["empty_ooxml", "image_only_pdf"])
def test_legitimate_partial_sidecar_case_does_not_trigger_full_rebuild_loop(
        tmp_path, monkeypatch, build_fixture):
    """正当に一部 sidecar が無いケース（空 OOXML の legacy `.md`・image-only PDF の legacy `.md`/
    `.md.meta.json`）を sidecar 欠落として誤検知しないこと。2回連続 sync してもフルリビルドが
    起きないことを固定する（過検知は既定 OFF でも毎 sync の全再取り込みという実害を生むため）。
    """
    wd = tmp_path / "world"
    wd.mkdir()
    dmd = tmp_path / "derived"

    if build_fixture == "empty_ooxml":
        docx_xml = ('<?xml version="1.0"?>'
                    '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                    "<w:body></w:body></w:document>")
        import zipfile
        with zipfile.ZipFile(wd / "empty.docx", "w") as z:
            z.writestr("word/document.xml", docx_xml)
        rep = office_md.build_derived(wd, dmd)
        assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
        assert not (dmd / "empty.docx.md").is_file()
    else:
        from pypdf import PdfWriter
        writer = PdfWriter()
        writer.add_blank_page(width=200, height=200)
        with (wd / "scan.pdf").open("wb") as f:
            writer.write(f)
        orig_backend, orig_pages = office_md._pdf_backend, office_md._pdf_pages
        try:
            office_md._pdf_backend = lambda: "pypdf"
            office_md._pdf_pages = lambda p: [""]   # legacy MD側はtextless扱い（Evidence抽出は実PDFを読む）
            rep = office_md.build_derived(wd, dmd)
        finally:
            office_md._pdf_backend, office_md._pdf_pages = orig_backend, orig_pages
        assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
        assert not (dmd / "scan.pdf.md").is_file()

    monkeypatch.setattr(worlds, "world_dir", lambda world: wd)
    monkeypatch.setattr(worlds, "derived_md_dir", lambda world: dmd)
    monkeypatch.setattr(worker, "world_state", lambda world: ("sig", {}))
    monkeypatch.setattr(store, "get_world",
                         lambda world: {"last_sig": "sig", "last_manifest": {}, "last_doc_count": 0})
    monkeypatch.setattr(worker, "_derived_stale", lambda world: False)

    @contextlib.contextmanager
    def _noop_lock(world_id):
        yield
    monkeypatch.setattr(store, "world_lock", _noop_lock)
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: False)
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: False)
    monkeypatch.setattr(worker, "run", _no_full_run)
    monkeypatch.setattr(worker, "_run_locked", _no_full_run)

    for _ in range(2):                                    # 2回連続でフルリビルドが起きないことを固定
        res = worker.sync("w")
        assert res["status"] == "unchanged" and res["changed"] is False


def test_empty_ooxml_evidence_survives_refresh_evidence_ir_across_two_syncs(tmp_path, monkeypatch):
    """空 OOXML（legacy `.md` を持たない正当なケース）の `.evidence.json`/`.rag.md`/
    `.rag_chunks.jsonl` が、`evidence_ir_sig` drift を契機に `refresh_evidence_ir()` が実際に
    走っても削除されず、2回の sync を経ても内容が保持されることを固定する。"""
    import zipfile

    wd = tmp_path / "world"
    wd.mkdir()
    dmd = tmp_path / "derived"
    docx_xml = ('<?xml version="1.0"?>'
                '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                "<w:body></w:body></w:document>")
    with zipfile.ZipFile(wd / "empty.docx", "w") as z:
        z.writestr("word/document.xml", docx_xml)
    rep = office_md.build_derived(wd, dmd)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
    assert not (dmd / "empty.docx.md").is_file()
    assert (_ir(dmd) / "empty.docx.evidence.json").is_file()
    assert (_rag(dmd) / "empty.docx.rag.md").is_file()

    monkeypatch.setattr(worlds, "world_dir", lambda world: wd)
    monkeypatch.setattr(worlds, "derived_md_dir", lambda world: dmd)
    monkeypatch.setattr(worker, "world_state", lambda world: ("sig", {}))
    monkeypatch.setattr(store, "get_world",
                         lambda world: {"last_sig": "sig", "last_manifest": {}, "last_doc_count": 0})
    monkeypatch.setattr(worker, "_derived_stale", lambda world: False)

    @contextlib.contextmanager
    def _noop_lock(world_id):
        yield
    monkeypatch.setattr(store, "world_lock", _noop_lock)
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: False)
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: False)
    monkeypatch.setattr(worker, "run", _no_full_run)
    monkeypatch.setattr(worker, "_run_locked", _no_full_run)

    _force_evidence_ir_drift(dmd)
    assert office_md.evidence_ir_sig_drift(dmd) is True

    for _ in range(2):                                    # 1回目でrefresh_evidence_irが走り、
        res = worker.sync("w")                            # 2回目はdrift解消済みでno-opになる
        assert res["status"] == "unchanged" and res["changed"] is False
        assert not (dmd / "empty.docx.md").is_file()               # legacy MDは相変わらず正当に無い
        assert (_ir(dmd) / "empty.docx.evidence.json").is_file()        # 削除されていない
        assert (_rag(dmd) / "empty.docx.rag.md").is_file()
        assert (_rag(dmd) / "empty.docx.rag_chunks.jsonl").is_file()
    assert office_md.evidence_ir_sig_drift(dmd) is False   # 1回目のrefreshで解消済み


def _value_text(value) -> str:
    return R._value_text(value)


def test_render_covers_all_evidence_ir_values(tmp_path):
    """Evidence IR の値付き element が citation されること・値が chunk body に実在することを
    ID一致だけでなく本文レベルでも検証する。"""
    _, dmd = _build_world(tmp_path)
    ir = IR.from_json_str((_ir(dmd) / "a.xlsx.evidence.json").read_text(encoding="utf-8"))
    chunks = [json.loads(line) for line in
              (_rag(dmd) / "a.xlsx.rag_chunks.jsonl").read_text(encoding="utf-8").splitlines()]

    cited_ids = {citation["evidence_id"] for chunk in chunks for citation in chunk["citations"]}
    value_element_ids = {el.element_id for el in ir.elements if _value_text(el.value)}
    assert value_element_ids <= cited_ids            # (i) citation coverage（片方向の包含）

    body_all = "\n".join(chunk["body"] for chunk in chunks)
    for element in ir.elements:
        value = _value_text(element.value)
        if value:
            assert value in body_all                 # (ii) body の値保持（IDの一致だけでは検出できない欠落を防ぐ）


def test_render_value_lands_in_specific_chunk_with_expected_source_line(tmp_path):
    """値がどこかの chunk に存在するだけでなく、**特定の1chunk**の body に、期待どおりの
    出所行（D1・rag.md のアンカー間本文の `出所: ...` 行）付きで載っていることを位置レベルで確認する。
    索引本文は jsonl の`search_text`ではなくrag.mdのアンカーから取る（`es_index._parse_rag_md_chunks`）。"""
    _, dmd = _build_world(tmp_path)
    chunks = [json.loads(line) for line in
              (_rag(dmd) / "a.xlsx.rag_chunks.jsonl").read_text(encoding="utf-8").splitlines()]
    assert all("search_text" not in c for c in chunks)   # D1: jsonlはもう索引本文を持たない
    rag_md_bodies, anchor_reason = es_index._parse_rag_md_chunks(
        (_rag(dmd) / "a.xlsx.rag.md").read_text(encoding="utf-8"))
    assert anchor_reason is None

    record_chunks = [c for c in chunks if c["content_type"] in ("table_record", "coordinate_record")]
    target = [c for c in record_chunks if "内容: サンプル内容" in c["body"]]
    assert len(target) == 1                          # 値はちょうど1つのchunkにだけ存在する
    chunk = target[0]
    assert "No: 1" in chunk["body"]                   # 同じレコードの他フィールドも同居している（bodyは「」無し）
    anchor_body = rag_md_bodies[chunk["chunk_id"]]
    assert "内容: 「サンプル内容」" in anchor_body       # rag.md側は値を「」で囲む形
    # 出所行（record-level）が正しい形で載っている（アンカー直後は`### key`見出しが先頭に来るため
    # startswithではなく含有を見る＝D1でアンカーは見出しの前・レコード内容の前に置く設計）。
    assert "出所: 原本「a.xlsx」 / シート「明細」 / No「1」" in anchor_body

    other_chunks = [c for c in chunks if c is not chunk]
    assert all("サンプル内容" not in c["body"] for c in other_chunks)   # 他chunkへの重複混入が無い
