"""world レジストリ 受け入れ（鏡モデル・要 Neo4j＋PG）: 別案件＝追加／参照先変更＝全削除＋再ミラー／削除。

参照元フォルダ（任意パス）を world にバインドし、register→rebind→delete を検証する。
rebind は **旧 world の派生物を完全削除して新パスから作り直す**（差分でなく破棄→再作成）。
"""
from __future__ import annotations

import os
import pathlib
import shutil
import tempfile

import pytest
from _world_setup import driver

from sherpa import store, worlds
from sherpa.impact_service import run_impact

W = "test_world_admin"


@pytest.fixture(autouse=True)
def _compat_mode(monkeypatch):
    """このファイルはログインせず直接叩く前提（compat モード）。"""
    monkeypatch.setenv("SHERPA_AUTH_DISABLED", "1")


from _corpus_helpers import _mk


def _names(drv, wid=W):
    with drv.session() as s:
        return sorted(r["n"] for r in s.run(
            "MATCH (x:Entity {world_id:$w}) RETURN x.name AS n", w=wid))


def _wait_ingest_idle(c, wid, *, timeout=60.0) -> dict:
    """ING-3: POST /worlds・POST /worlds/{wid}/refresh 等は即受付・
    取り込みは背景実行のため、完了（または失敗）まで status をポーリングする（テスト専用）。
    `running_progress is None` かつ `last_run_status` が確定していれば「今は動いていない」。

    受付（run_id 確保）と実処理（`store.upsert_world` 等）が完全に非同期化された
    ため、登録直後の一瞬は world 行自体がまだ存在せず `GET .../status` が 404 を返しうる——
    その窓は「エラー」ではなく「まだ作成中」として扱い、ポーリングを継続する。
    """
    import time
    deadline = time.monotonic() + timeout
    last = None
    while time.monotonic() < deadline:
        r = c.get(f"/worlds/{wid}/status")
        if r.status_code == 404:                # 受付直後の一瞬＝world 行がまだ作られていない
            time.sleep(0.2)
            continue
        assert r.status_code == 200, r.text
        last = r.json()
        if last.get("running_progress") is None and last.get("last_run_status") is not None:
            return last
        time.sleep(0.2)
    raise AssertionError(f"取り込みが {timeout}s 以内に完了しませんでした: world={wid} last={last}")


def _wait_world_deleted(c, wid, *, timeout=60.0) -> None:
    """ING-3: DELETE /worlds/{wid} も即受付・削除は背景実行のため、world 行が消える
    （`GET /worlds/{wid}/status` が 404 になる）まで待つ（テスト専用）。
    """
    import time
    deadline = time.monotonic() + timeout
    last_status = None
    while time.monotonic() < deadline:
        r = c.get(f"/worlds/{wid}/status")
        last_status = r.status_code
        if r.status_code == 404:
            return
        time.sleep(0.2)
    raise AssertionError(f"削除が {timeout}s 以内に完了しませんでした: world={wid} last_status={last_status}")


def test_register_rebind_delete():
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        rel_a = "案件X/03_開発/01_ソース/FOOPROG.cbl"
        rel_b = "案件Y/03_開発/01_ソース/BARPROG.cbl"
        # 別案件＝新 world を参照元 a にバインドして取り込む（追加）
        r = worlds.register(W, str(pathlib.Path(a).resolve()))
        assert r["status"] in ("auto_published", "auto_published_with_flags")
        assert worlds.world_dir(W) == pathlib.Path(a).resolve()      # レジストリ binding が効く
        assert _names(drv) == ["FOOPROG"]                            # a の中身がグラフに
        assert {d["name"] for d in store.list_documents(W)} == {rel_a}   # 台帳も a の rel_path

        # 参照先変更＝旧を全削除して b から作り直し（FOOPROG 消え BARPROG だけ・台帳も）
        worlds.rebind(W, str(pathlib.Path(b).resolve()))
        assert worlds.world_dir(W) == pathlib.Path(b).resolve()
        assert _names(drv) == ["BARPROG"], _names(drv)               # 破棄→再作成（差分でない）
        assert {d["name"] for d in store.list_documents(W)} == {rel_b}   # 台帳: 旧 rel 消え新 rel だけ

        # world 完全削除（グラフ空・台帳空・レジストリ行なし）。参照元フォルダは残る。
        assert worlds.delete(W) is True
        assert _names(drv) == [] and store.get_world(W) is None
        assert store.list_documents(W) == []
        assert pathlib.Path(b).is_dir()                              # 外部フォルダは消さない
    finally:
        try:
            worlds.delete(W)
        except Exception:
            pass
        store.delete_world_row(W)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=W)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_delete_neo4j_success_then_pg_failure_self_heals():
    """Codex finding #5: delete が Neo4j 削除成功後に PG replace で失敗した窓。

    `_wipe_locked` は Neo4j 削除→台帳クリアの順（fail-closed）。PG replace が失敗すると例外が
    `worlds.delete` まで伝播し、registry 行削除（`store.delete_world_row`）に届かない＝行は残る
    （＝「PG クリア後～registry 削除前のクラッシュ窓」と同型: どちらも registry 行が生き残る）。
    旧実装はこの窓で last_sig に触れないため、次回 sync が「変更なし」と誤判定してグラフが
    空のまま恒久的に直らなかった。pre-invalidate 設計（secRV 再RV・2026-07-14）は `_wipe_locked`
    **冒頭**（Neo4j delete より前）で last_sig を無効化しておくので、途中のどこで失敗しても
    次回 sync が必ず再構築して自己修復する。
    """
    from sherpa.ingest import worker
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "DELFOO")
    wid = W + "_delete_pg_fail"
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        assert _names(drv, wid) == ["DELFOO"]
        sig0 = store.get_world(wid)["last_sig"]
        assert sig0

        orig_replace = store.replace_documents

        def _boom(world, rows):
            raise RuntimeError("pg replace fault (after neo4j delete commit)")

        store.replace_documents = _boom
        try:
            with pytest.raises(RuntimeError, match="pg replace fault"):
                worlds.delete(wid)
        finally:
            store.replace_documents = orig_replace

        # registry 行削除まで届かない＝行は残る（が Neo4j は既に空）
        row = store.get_world(wid)
        assert row is not None
        assert row["last_sig"] == ""                            # last_sig 無効化（旧実装は sig0 のまま残り穴になった）
        assert _names(drv, wid) == []                            # Neo4j は削除済み（空のまま）

        r = worker.sync(wid)                                     # 次回 sync＝自己修復（内容が戻る）
        assert r["changed"] is True
        assert r["status"] in ("auto_published", "auto_published_with_flags")
        assert _names(drv, wid) == ["DELFOO"]
        assert store.get_world(wid)["last_sig"] == sig0
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_wipe_invalidates_sig_before_neo4j_delete_order_pin():
    """secRV 再RV #5（順序 pin）: `_wipe_locked` は last_sig の無効化（pre-invalidate）を
    Neo4j delete より**前**に行う（旧実装は delete 成功後の best-effort だった）。
    呼び出し順を記録して固定する。"""
    from sherpa.ingest import world_neo4j
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "ORDERFOO")
    wid = W + "_wipe_order"
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        assert _names(drv, wid) == ["ORDERFOO"]

        calls = []
        orig_set_sig = store.set_world_sig
        orig_delete = world_neo4j.delete_world

        def _set_sig(world, sig, manifest=None):
            calls.append("set_world_sig")
            return orig_set_sig(world, sig, manifest=manifest)

        def _delete(*a, **k):
            calls.append("delete_world")
            return orig_delete(*a, **k)

        store.set_world_sig = _set_sig
        world_neo4j.delete_world = _delete
        try:
            assert worlds.delete(wid) is True
        finally:
            store.set_world_sig = orig_set_sig
            world_neo4j.delete_world = orig_delete

        assert calls == ["set_world_sig", "delete_world"]
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_wipe_pre_invalidate_failure_is_fail_closed_before_neo4j_delete():
    """secRV 再RV #5（fail-closed）: `_wipe_locked` 冒頭の pre-invalidate
    （`store.set_world_sig(world, "")`）が失敗したら、Neo4j delete に到達せず例外が伝播する。
    registry 行・last_sig・グラフはすべて無傷のまま残る（PG に無効化すら書けない状態で
    削除を開始しない）。"""
    from sherpa.ingest import world_neo4j
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "WIPEFAILFOO")
    wid = W + "_wipe_pre_invalidate_fail"
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        assert _names(drv, wid) == ["WIPEFAILFOO"]
        sig0 = store.get_world(wid)["last_sig"]
        assert sig0

        orig_delete = world_neo4j.delete_world
        delete_calls = {"n": 0}

        def _track_delete(*a, **k):
            delete_calls["n"] += 1
            return orig_delete(*a, **k)

        orig_set_sig = store.set_world_sig

        def _boom_set_sig(world, sig, manifest=None):
            raise RuntimeError("set_world_sig fault (pre-invalidate, PG断)")

        world_neo4j.delete_world = _track_delete
        store.set_world_sig = _boom_set_sig
        try:
            with pytest.raises(RuntimeError, match="pre-invalidate"):
                worlds.delete(wid)
        finally:
            world_neo4j.delete_world = orig_delete
            store.set_world_sig = orig_set_sig

        assert delete_calls["n"] == 0                             # Neo4j delete 未到達
        row = store.get_world(wid)
        assert row is not None                                    # registry 行は残る（delete が完遂していない）
        assert row["last_sig"] == sig0                             # 無効化が失敗＝last_sig は旧のまま
        assert _names(drv, wid) == ["WIPEFAILFOO"]                  # グラフも無傷
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_rebind_failure_preserves_derived_and_binding():
    """rebind 失敗時: バインドは旧 root に戻り、**旧派生物も保持**される（rv-full2 #1・fail-closed）。

    旧実装は再取込前に派生を rmtree していたため、失敗するとバインドは戻っても派生が消え DB と不整合だった。
    """
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(W, str(pathlib.Path(a).resolve()))
        der = worlds.derived_dir(W)
        marker = der / "semantic" / "_marker.txt"               # 旧派生に検出可能なファイルを置く
        marker.parent.mkdir(parents=True, exist_ok=True)
        marker.write_text("keep-me", encoding="utf-8")

        from sherpa.ingest import worker as _worker
        orig = _worker._run_locked                                # R3-S3: rebind は lock-free の _run_locked を直接呼ぶ

        def _boom(*aa, **kk):                                    # rebind 中の取り込みを失敗させる
            raise RuntimeError("boom")

        _worker._run_locked = _boom
        try:
            worlds.rebind(W, str(pathlib.Path(b).resolve()))
            assert False, "rebind は失敗するはず"
        except RuntimeError:
            pass
        finally:
            _worker._run_locked = orig

        assert worlds.world_dir(W) == pathlib.Path(a).resolve()  # バインドは旧 a に戻る
        assert marker.is_file() and marker.read_text(encoding="utf-8") == "keep-me"  # 旧派生は保持
        assert _names(drv) == ["FOOPROG"]                        # グラフも旧 a のまま（未変更）
    finally:
        try:
            worlds.delete(W)
        except Exception:
            pass
        store.delete_world_row(W)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=W)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_rebind_neo4j_committed_then_pg_replace_fails_restores_old_graph():
    """RV HIGH（2026-07-14）: rebind が **Neo4j load 成功（新 root コミット済）後に PG replace 段階で
    失敗**した場合、except 節の即時再構築で Neo4j が旧 root へ戻ること。

    旧実装（この修正前）は、失敗が Neo4j commit 後だと Neo4j＝新 root のまま残り、rebind が bind を
    旧へ戻すため `last_sig` 不変＝次回 sync も self-heal せず、registry/台帳/派生＝旧・Neo4j＝新 の
    **永続不整合**を作れた。上の test_rebind_failure_preserves_derived_and_binding は `_run_locked` を
    即時例外化＝Neo4j load 前に失敗するため、この経路を通らず検出できなかった。"""
    wid = "test_world_admin_neo4j_rollback"                       # 共有 W と衝突しない専用 world id（順序汚染回避）
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        assert _names(drv, wid) == ["FOOPROG"]                    # 旧 a のグラフ

        real_replace = store.replace_documents
        calls = {"n": 0}

        def _fail_first(world, rows):
            calls["n"] += 1
            if calls["n"] == 1:                                  # rebind の B に対する replace だけ失敗
                raise RuntimeError("pg replace fault (after neo4j commit)")   # ← Neo4j は既にコミット済
            return real_replace(world, rows)                     # rollback の A 即時再構築は成功させる

        store.replace_documents = _fail_first
        try:
            with pytest.raises(RuntimeError, match="pg replace fault"):
                worlds.rebind(wid, str(pathlib.Path(b).resolve()))
        finally:
            store.replace_documents = real_replace

        assert calls["n"] == 2                                    # rebind(B) 失敗 → rollback(A) 再構築の2回
        assert worlds.world_dir(wid) == pathlib.Path(a).resolve() # バインドは旧 a に戻る
        assert _names(drv, wid) == ["FOOPROG"]                    # ← 修正前は ["BARPROG"] が残り永続不整合
        assert store.get_world(wid)["last_sig"]                   # sig は復元済み（空でない＝即時再構築成功）
    finally:
        try:
            worlds.delete(wid)                                    # 派生 dir・台帳・Neo4j・registry を一括 wipe
        except Exception:
            pass
        try:
            store.delete_world_row(wid)
        except Exception:
            pass
        try:
            shutil.rmtree(worlds.derived_dir(wid), ignore_errors=True)   # 派生残骸を明示除去
        except Exception:
            pass
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_rebind_failure_preserves_original_exception_even_if_restore_fails():
    """RV HIGH round-2＋round-3（2026-07-14）: rebind 失敗時、**復元経路の二次例外で元例外を握り潰さない**＋
    bind 復元（同一 tx の bind＋sig 無効化）が失敗（PG 断継続）したら **復旧を一切足さない**（旧派生復元も
    旧 root 再構築もしない）＝bind=新のまま次回 sync が新へ収束。gate を外す mutation を落とすため、
    (a) 元例外が伝播 (b) 復旧の _run_locked が呼ばれない（calls==1） (c) bind が新 root のまま
    (d) 旧派生 A が active derived へ復元されていない（backup 側に残る）ことまで assert する。"""
    wid = "test_world_admin_exc_preserve"
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        der = worlds.derived_dir(wid)
        marker = der / "semantic" / "_marker_A.txt"             # 旧 root A の派生に検出可能なマーカー
        marker.parent.mkdir(parents=True, exist_ok=True)
        marker.write_text("A-derived", encoding="utf-8")
        backup_path = der.with_name("." + der.name + ".rebind-bak")

        from sherpa.ingest import worker as _worker
        orig_run = _worker._run_locked
        orig_restore = store.restore_bind_invalidate_sig
        calls = {"n": 0}

        def _primary(*aa, **kk):
            calls["n"] += 1
            raise ValueError("rebind-primary (取り込み失敗)")

        def _secondary(*aa, **kk):
            raise RuntimeError("restore-secondary (これが元例外を隠したら NG)")

        _worker._run_locked = _primary
        store.restore_bind_invalidate_sig = _secondary          # bind 復元自体を失敗させる（PG 断継続を模擬）
        try:
            with pytest.raises(ValueError, match="rebind-primary"):   # (a) RuntimeError ではなく元の ValueError
                worlds.rebind(wid, str(pathlib.Path(b).resolve()))
            assert calls["n"] == 1                              # (b) 復旧 rebuild は呼ばれない（gate で skip）
            assert worlds.world_dir(wid) == pathlib.Path(b).resolve()   # (c) bind は新 root のまま（復元失敗）
            assert backup_path.exists() and (backup_path / "semantic" / "_marker_A.txt").is_file()  # (d) 旧派生は backup 側
            assert not marker.exists()                          # (d) active derived へ A 派生は復元されていない
        finally:
            _worker._run_locked = orig_run
            store.restore_bind_invalidate_sig = orig_restore
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        try:
            store.delete_world_row(wid)
        except Exception:
            pass
        try:
            _der = worlds.derived_dir(wid)
            shutil.rmtree(_der, ignore_errors=True)
            shutil.rmtree(_der.with_name("." + _der.name + ".rebind-bak"),   # restore-fail 経路が残す退避 dir
                          ignore_errors=True)
        except Exception:
            pass
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_rebind_recovery_rebuild_failure_still_propagates_original_exception():
    """RV HIGH round-3（2026-07-14）: bind 復元が**成功**して復旧経路（旧 root 再構築）に入り、
    そこで別種例外が起きても、伝播するのは取り込み失敗の**元例外**である（復旧の _run_locked も
    guard 済み）。前テストは bind 復元失敗＝復旧スキップ経路を pin するので、こちらは復旧実行経路を pin。"""
    wid = "test_world_admin_recovery_exc"
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))

        from sherpa.ingest import worker as _worker
        orig_run = _worker._run_locked
        calls = {"n": 0}

        def _run(*aa, **kk):
            calls["n"] += 1
            if calls["n"] == 1:                                  # rebind 本体（B）＝元例外
                raise ValueError("rebind-primary (取り込み失敗)")
            raise RuntimeError("recovery-secondary (復旧の再構築失敗・元例外を隠したら NG)")  # 復旧（A）

        _worker._run_locked = _run                               # restore は本物＝bind 復元成功で復旧経路に入る
        try:
            with pytest.raises(ValueError, match="rebind-primary"):   # ← 復旧の RuntimeError ではなく元の ValueError
                worlds.rebind(wid, str(pathlib.Path(b).resolve()))
            assert calls["n"] == 2                                # 本体失敗 → 復旧再構築（bind 復元成功で実行）
        finally:
            _worker._run_locked = orig_run
    finally:
        for fn in (lambda: worlds.delete(wid), lambda: store.delete_world_row(wid),
                   lambda: shutil.rmtree(worlds.derived_dir(wid), ignore_errors=True)):
            try:
                fn()
            except Exception:
                pass
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_rebind_run_id_recovery_success_finalizes_once_with_rolled_back_reason():
    """`run_id` 指定時（受付run）は新root試行・（失敗時の）旧root復旧のどちらも
    非terminalな内部段として扱い、受付run自身の確定は最後に一度だけ行う。復旧（旧root再構築）が
    実際に成功していれば、その published_snapshot/source_doc_ids を温存したまま status=failed・
    reason=rebind_failed_rolled_back で確定する（復旧成功を「rebind成功」と誤認させない一方、
    NULL 上書きで Graph 件数 0 表示にもしない）。"""
    wid = "test_world_admin_rv2_recovery_ok"
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))

        from sherpa.ingest import worker as _worker
        orig_run = _worker._run_locked
        calls = {"n": 0}

        def _run(*aa, **kk):
            calls["n"] += 1
            if calls["n"] == 1:                # 新root（B）試行＝失敗させる
                raise ValueError("rebind-primary (取り込み失敗)")
            return orig_run(*aa, **kk)         # 復旧（旧root A）再構築は本物を通す＝実際に成功する

        run_row = store.start_ingest_run(wid, scan_root=None, created_by="admin")
        run_id = run_row["id"]
        _worker._run_locked = _run
        try:
            with pytest.raises(ValueError, match="rebind-primary"):
                worlds.rebind(wid, str(pathlib.Path(b).resolve()), run_id=run_id)
            assert calls["n"] == 2
        finally:
            _worker._run_locked = orig_run

        rec = store.get_latest_run_summary(wid)
        assert rec["id"] == run_id and rec["status"] == "failed"
        reasons = [f.get("reason") for f in (rec["extraction_snapshot"] or {}).get("flags", [])]
        assert "rebind_failed_rolled_back" in reasons
        pub = store.get_latest_published_run_summary(wid)          # 復旧（旧root A）の反映が温存されている
        assert pub is not None and pub["published_snapshot"]       # NULL 上書きなら None/空になる（#5）
        full = store.list_ingest_runs(wid, limit=1)[0]
        assert full["id"] == run_id and full["source_doc_ids"]      # 空 [] へ上書きされていない
    finally:
        for fn in (lambda: worlds.delete(wid), lambda: store.delete_world_row(wid),
                   lambda: shutil.rmtree(worlds.derived_dir(wid), ignore_errors=True)):
            try:
                fn()
            except Exception:
                pass
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_rebind_run_id_recovery_failure_uses_rollback_failed_reason():
    """復旧（bind 復元＋旧root再構築）の**成功を確認できない**場合は
    `rebind_failed_rolled_back` ではなく別理由（`rebind_rollback_failed`）で確定する
    ——復旧の再構築自体が失敗/例外なら「戻せた」とは言えない。"""
    wid = "test_world_admin_rv2_recovery_fail"
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))

        from sherpa.ingest import worker as _worker
        orig_run = _worker._run_locked
        calls = {"n": 0}

        def _run(*aa, **kk):
            calls["n"] += 1
            raise RuntimeError(f"boom-{calls['n']}")   # 新root試行・復旧再構築ともに失敗させる

        run_row = store.start_ingest_run(wid, scan_root=None, created_by="admin")
        run_id = run_row["id"]
        _worker._run_locked = _run
        try:
            with pytest.raises(RuntimeError, match="boom-1"):
                worlds.rebind(wid, str(pathlib.Path(b).resolve()), run_id=run_id)
            assert calls["n"] == 2                     # 新root試行＋復旧再構築の2回とも呼ばれる
        finally:
            _worker._run_locked = orig_run

        rec = store.get_latest_run_summary(wid)
        assert rec["id"] == run_id and rec["status"] == "failed"
        reasons = [f.get("reason") for f in (rec["extraction_snapshot"] or {}).get("flags", [])]
        assert "rebind_rollback_failed" in reasons
        assert "rebind_failed_rolled_back" not in reasons
    finally:
        for fn in (lambda: worlds.delete(wid), lambda: store.delete_world_row(wid),
                   lambda: shutil.rmtree(worlds.derived_dir(wid), ignore_errors=True)):
            try:
                fn()
            except Exception:
                pass
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_rebind_backup_move_failure_rolls_back_bind_and_self_heals():
    """Codex finding #6: rebind の退避（旧派生を `.rebind-bak` へ `os.replace`）が失敗した場合。

    旧実装はこの `os.replace` を try 外で呼んでいたため、失敗すると bind 更新（`store.upsert_world`
    による新 root への切替）だけが残り、ロールバック（`restore_bind_invalidate_sig`／即時再構築）が
    一切走らない＝bind=新 root なのに Neo4j/台帳/派生は旧 root のまま、という不整合を作れた
    （`rebind` の「取り込み失敗時は旧状態へ一貫復元」契約に反する）。退避を try 内に移したことで、
    この失敗も他の rebind 失敗と同じロールバック経路（bind を旧へ戻し last_sig を無効化 →
    旧 root から即時再構築）に必ず入る。
    """
    wid = "test_world_admin_backup_escape"
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "FOOPROG")
    _mk(b, "案件Y", "BARPROG")
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        assert _names(drv, wid) == ["FOOPROG"]

        import os
        real_replace = os.replace

        def _boom_replace(src, dst):
            if str(dst).endswith(".rebind-bak"):                 # 退避の os.replace(der, backup) だけ狙い撃ち
                raise OSError("boom: backup escape simulated (fault injection)")
            return real_replace(src, dst)

        os.replace = _boom_replace
        try:
            with pytest.raises(OSError, match="boom"):
                worlds.rebind(wid, str(pathlib.Path(b).resolve()))
        finally:
            os.replace = real_replace

        assert worlds.world_dir(wid) == pathlib.Path(a).resolve()   # bind は旧 a に戻る（新 b のまま残らない）
        assert _names(drv, wid) == ["FOOPROG"]                       # グラフも旧 a（自己修復のリビルド成功）
        assert store.get_world(wid)["last_sig"]                      # sig は復元済み（空でない＝即時再構築成功）
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        try:
            store.delete_world_row(wid)
        except Exception:
            pass
        try:
            shutil.rmtree(worlds.derived_dir(wid), ignore_errors=True)
        except Exception:
            pass
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_api_world_lifecycle():
    """API: POST /worlds（登録）→ /worlds（一覧）→ rebind → DELETE。

    同一フォルダ再登録は**冪等**（リラン/変更なし＝200）・別パスへの同名登録は 409・未知 rebind は 404。
    """
    from fastapi.testclient import TestClient
    from sherpa.api import app
    a, b = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "APIFOO")
    _mk(b, "案件Y", "APIBAR")
    c = TestClient(app)
    drv = driver()
    try:
        ok = c.post("/worlds", json={"world_id": W, "path": a, "label": "案件X"})
        assert ok.status_code == 202, ok.text
        assert ok.json()["run_id"] is not None
        _wait_ingest_idle(c, W)
        assert any(x["world_id"] == W for x in c.get("/worlds").json()["worlds"])
        assert _names(drv) == ["APIFOO"]
        # 同一フォルダの再登録は **冪等**（409 でなく 202＝チェック→リラン。即受付なので run_id/joined を返す）
        dup = c.post("/worlds", json={"world_id": W, "path": a})
        assert dup.status_code == 202, dup.text
        assert dup.json()["world_id"] == W
        _wait_ingest_idle(c, W)
        assert _names(drv) == ["APIFOO"]                            # 冪等再登録でグラフは不変
        # 同名 world_id を **別パス** に登録しようとするのは依然 409（参照先変更は rebind）
        assert c.post("/worlds", json={"world_id": W, "path": b}).status_code == 409
        # 不在パスは 422
        assert c.post("/worlds", json={"world_id": "nope2", "path": "/no/such/dir"}).status_code == 422
        # rebind → 破棄→再作成（ING-3: 即受付・再構築自体は背景実行）
        rb = c.post(f"/worlds/{W}/rebind", json={"path": b})
        assert rb.status_code == 202, rb.text
        assert rb.json()["run_id"] is not None
        _wait_ingest_idle(c, W)
        assert _names(drv) == ["APIBAR"]
        # 未知 world の rebind は 404
        assert c.post("/worlds/unknownw/rebind", json={"path": b}).status_code == 404
        # 削除（ING-3: 即受付・削除自体は背景実行）
        del_res = c.delete(f"/worlds/{W}")
        assert del_res.status_code == 202, del_res.text
        assert del_res.json()["run_id"] is not None
        _wait_world_deleted(c, W)
        assert _names(drv) == [] and store.get_world(W) is None
    finally:
        try:
            worlds.delete(W)
        except Exception:
            pass
        store.delete_world_row(W)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=W)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(b, ignore_errors=True)


def test_conflicts_and_fail_closed():
    """登録済みがあれば2本目は拒否（単一登録契約）／参照元が消えたら world_dir は fail-closed（None）。"""
    import shutil
    from sherpa import worlds as W2
    a, c = tempfile.mkdtemp(), tempfile.mkdtemp()
    _mk(a, "案件X", "DUPFOO")
    _mk(c, "案件Z", "GONE")
    w1, w2 = W + "_1", W + "_2"
    drv = driver()
    try:
        worlds.register(w1, str(pathlib.Path(a).resolve()))
        # 同じ world_id の再登録は WorldConflict（内容更新は refresh・参照先変更は rebind）
        try:
            worlds.register(w1, str(pathlib.Path(c).resolve())); assert False, "同名で例外が出るべき"
        except W2.WorldConflict:
            pass
        # 同一参照元を別 world_id で二重登録も WorldConflict
        try:
            worlds.register(w2, str(pathlib.Path(a).resolve())); assert False, "同一 root で例外が出るべき"
        except W2.WorldConflict:
            pass
        # 別 world_id・別 root でも2本目は拒否（登録元フォルダは全体で1本）
        try:
            worlds.register(w2, str(pathlib.Path(c).resolve())); assert False, "2本目で例外が出るべき"
        except W2.WorldConflict:
            pass
        # fail-closed: 登録後に参照元フォルダを消すと world_dir は None（fixtures/KB に落とさない）
        assert worlds.world_dir(w1) == pathlib.Path(a).resolve()
        shutil.rmtree(a, ignore_errors=True)
        assert worlds.world_dir(w1) is None                          # 参照元消失＝unavailable
    finally:
        for w in (w1, w2):
            try:
                worlds.delete(w)
            except Exception:
                pass
            store.delete_world_row(w)
            with drv.session() as s:
                s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=w)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)
        shutil.rmtree(c, ignore_errors=True)


def test_fs_browser_confined():
    """フォルダ選択 API: 許可ルート配下のサブフォルダだけ返し、範囲外・`..` は 403。"""
    from fastapi.testclient import TestClient
    from sherpa.api import app
    base = tempfile.mkdtemp()
    (pathlib.Path(base) / "proj" / "sub").mkdir(parents=True)
    os.environ["SHERPA_BROWSE_ROOTS"] = base
    try:
        c = TestClient(app)
        top = c.get("/fs/list").json()                              # path 空＝ルート直下
        assert any(e["name"] == "proj" for e in top["entries"]) and top["parent"] is None
        proj = c.get("/fs/list", params={"path": str(pathlib.Path(base) / "proj")}).json()
        assert any(e["name"] == "sub" for e in proj["entries"])
        assert proj["parent"] == str(pathlib.Path(base).resolve())  # 親はルート（その上へは行けない）
        assert c.get("/fs/list", params={"path": "/etc"}).status_code == 403          # 範囲外
        assert c.get("/fs/list", params={"path": str(pathlib.Path(base) / "..")}).status_code == 403
    finally:
        os.environ.pop("SHERPA_BROWSE_ROOTS", None)
        shutil.rmtree(base, ignore_errors=True)


def test_status_reports_indexed_and_skipped_office():
    """取り込み状況の正直化: txt はインデックス／壊れ Office は変換失敗／PDF は未対応／コード無しでグラフ0。"""
    from fastapi.testclient import TestClient
    from sherpa.api import app
    a = tempfile.mkdtemp()
    base = pathlib.Path(a) / "案件X" / "01_資料"
    base.mkdir(parents=True)
    (base / "メモ.txt").write_text("調査メモ\n", encoding="utf-8")
    (base / "資料.docx").write_bytes(b"PK\x03\x04dummy")      # 壊れOOXML→変換失敗
    (base / "表.xlsx").write_bytes(b"PK\x03\x04dummy")
    (base / "仕様.pdf").write_bytes(b"%PDF-1.4 dummy")        # PDF→未対応
    wid = "test_world_status"
    c = TestClient(app)
    drv = driver()
    try:
        assert c.post("/worlds", json={"path": str(pathlib.Path(a).resolve()), "world_id": wid}).status_code == 202
        s = _wait_ingest_idle(c, wid)
        from sherpa.ingest import office_md
        pdf_on = ".pdf" in office_md.convertible_exts()                      # PDF はバックエンド導入時のみ変換対象
        failed_notices = 3 if pdf_on else 2
        # MD化強化（2026-08-15 移植）: 内容抽出に失敗した文書も、locator/理由つきの「失敗の記録」
        # として索引する（黙って消えると「取り込んだのに検索に出ない」の原因が追えないため）。
        assert s["indexed"] == 1 + failed_notices
        assert s["by_doctype"].get("テキスト") == 1
        assert s["office_md"] == failed_notices
        assert s["office_failed"] == failed_notices                          # 壊れ docx/xlsx（＋PDF 有効なら壊れ pdf も）失敗
        assert s["skipped_office"] == (0 if pdf_on else 1)                   # バックエンド無しなら pdf は未対応
        assert s["graph_nodes"] == 0                                         # コード/抽出なし→グラフ空
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as ss:
            ss.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_office_indexed_and_searchable():
    """Office（xlsx）が取り込みで決定的MD化され、検索でヒット＋DLは原本に解決する（Phase1）。"""
    import openpyxl
    from fastapi.testclient import TestClient
    from sherpa.api import app
    from sherpa import documents
    from sherpa.grep_tool import grep_search
    a = tempfile.mkdtemp()
    d = pathlib.Path(a) / "案件X" / "01_資料"
    d.mkdir(parents=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["A2"] = "項目", "消費税率テスト値ZZZ"
    wb.save(d / "表.xlsx")
    rel = "案件X/01_資料/表.xlsx"
    wid = "test_world_office"
    c = TestClient(app)
    drv = driver()
    try:
        assert c.post("/worlds", json={"path": str(pathlib.Path(a).resolve()), "world_id": wid}).status_code == 202
        s = _wait_ingest_idle(c, wid)
        assert s["office_md"] == 1 and s["indexed"] == 1 and s["skipped_office"] == 0
        hits = grep_search("消費税率テスト値ZZZ", wid, max_hits=10)
        assert hits and hits[0]["doc_id"] == rel and hits[0]["ext"] == ".xlsx"   # 派生MD→元 xlsx の rel
        p = documents.resolve(rel, wid)                                          # DL は原本（.xlsx）に解決
        assert p and p.is_file() and p.suffix == ".xlsx"
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as ss:
            ss.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_sync_rebuilds_when_derived_missing():
    """RV High#1: ソース無変更でも派生MD が消えていれば sync は再ビルドする（後付け導入/派生削除の自己修復）。"""
    import shutil as _sh
    import openpyxl
    from sherpa.ingest import worker
    a = tempfile.mkdtemp()
    d = pathlib.Path(a) / "案件X"
    d.mkdir(parents=True)
    wb = openpyxl.Workbook()
    wb.active["A1"] = "派生再生成テスト"
    wb.save(d / "x.xlsx")
    wid = "test_world_derived"
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        der = worlds.derived_dir(wid)
        assert der.is_dir() and any(der.rglob("*.md"))          # 登録で派生MD ができる
        _sh.rmtree(der)                                         # 派生だけ消す（ソースは無変更）
        r = worker.sync(wid)
        assert r["changed"] is True                             # 無変更でも派生欠落で再ビルド
        assert der.is_dir() and any(der.rglob("*.md"))          # 復活
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as ss:
            ss.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        _sh.rmtree(a, ignore_errors=True)


def test_fs_subdirs_skips_inaccessible():
    """フォルダ一覧は 1 件の権限エラー（Windows システムファイル等）で全体が止まらない（per-entry skip・回帰）。"""
    import sherpa.api as api
    base = tempfile.mkdtemp()
    for n in ("aaa", "mmm", "zzz"):
        (pathlib.Path(base) / n).mkdir()
    (pathlib.Path(base) / "bbb_bad.tmp").write_text("x")        # "aaa" と "mmm" の間で is_dir が例外を投げる状況
    real_is_dir = pathlib.Path.is_dir

    def fake_is_dir(self):
        if self.name == "bbb_bad.tmp":
            raise PermissionError(13, "Permission denied")
        return real_is_dir(self)

    pathlib.Path.is_dir = fake_is_dir
    try:
        names = [e["name"] for e in api._subdirs(pathlib.Path(base))]
        assert names == ["aaa", "mmm", "zzz"], names           # 問題の1件はスキップし後続も全部出る
    finally:
        pathlib.Path.is_dir = real_is_dir
        shutil.rmtree(base, ignore_errors=True)


def test_post_world_without_id_autogenerates():
    """POST /worlds は world_id 省略可＝自動採番（UI はパスと名前だけ送る）。日本語名は label に。"""
    from fastapi.testclient import TestClient
    from sherpa.api import app
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "AUTOFOO")
    c = TestClient(app)
    drv = driver()
    wid = None
    try:
        r = c.post("/worlds", json={"path": str(pathlib.Path(a).resolve()), "label": "自動採番案件"})
        assert r.status_code == 202, r.text
        wid = r.json()["world_id"]
        assert wid
        _wait_ingest_idle(c, wid)
        match = next((x for x in c.get("/worlds").json()["worlds"] if x["world_id"] == wid), None)
        assert match is not None and match["label"] == "自動採番案件"
    finally:
        if wid:
            try:
                worlds.delete(wid)
            except Exception:
                pass
            store.delete_world_row(wid)
            with drv.session() as s:
                s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_refresh_change_detection():
    """変更検知: 変わってなければ no-op、フォルダにファイルを足すと sync が再取り込み。手動 refresh も同様。"""
    from sherpa.ingest import worker
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "SYNCA")
    wid = W + "_sync"
    drv = driver()
    try:
        worlds.register(wid, str(pathlib.Path(a).resolve()))
        assert _names(drv, wid) == ["SYNCA"]
        assert worker.sync(wid)["changed"] is False                 # 変更なし＝no-op
        _mk(a, "案件X", "SYNCB")                                    # フォルダに直接ファイル追加
        r = worker.sync(wid)
        assert r["changed"] is True and set(_names(drv, wid)) == {"SYNCA", "SYNCB"}   # 検知して再取り込み
        # 手動 refresh API: もう変更なし（ING-3＝受付処理自身が O(1) で run 行を確保する
        # ため run_id は必ず非 null——無変化でもその run は `auto_published` として terminal 化
        # される〔unchanged も同じ run を terminal 化する契約〕・joined=False のまま）。
        from fastapi.testclient import TestClient
        from sherpa.api import app
        c = TestClient(app)
        rr = c.post(f"/worlds/{wid}/refresh")
        assert rr.status_code == 202, rr.text
        payload = rr.json()
        assert payload["run_id"] is not None and payload["joined"] is False
        _wait_ingest_idle(c, wid)
        assert set(_names(drv, wid)) == {"SYNCA", "SYNCB"}
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_diff_and_idempotent_register():
    """差分チェック（read-only・書き込みなし）と「このフォルダを登録」の冪等性（未登録→登録／登録済み→リラン）。"""
    from fastapi.testclient import TestClient
    from sherpa.api import app
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "DIFFA")
    root = str(pathlib.Path(a).resolve())
    wid = "test_world_diff"
    c = TestClient(app)
    drv = driver()
    try:
        # 1) 未登録フォルダの差分＝全ファイルが added・registered False、かつ **登録されない**（read-only）
        d0 = c.post("/worlds/diff", json={"path": root}).json()
        assert d0["registered"] is False and d0["indexed"] == 0 and len(d0["added"]) >= 1
        assert not any(x["root_path"] == root for x in c.get("/worlds").json()["worlds"])
        # 2) 登録（ING-3＝即受付・背景実行の完了を待ってから以降の差分チェックへ進む）
        r = c.post("/worlds", json={"path": root, "world_id": wid, "label": "差分案件"})
        assert r.status_code == 202 and r.json()["world_id"] == wid, r.text
        _wait_ingest_idle(c, wid)
        # 3) 登録直後の差分＝なし
        d1 = c.post("/worlds/diff", json={"path": root}).json()
        assert d1["registered"] is True and d1["added"] == [] and d1["changed"] == [] and d1["removed"] == []
        # 4) ファイルを足すと差分に出る（まだ反映しない＝グラフは DIFFA のまま）
        _mk(a, "案件X", "DIFFB")
        d2 = c.get(f"/worlds/{wid}/diff").json()
        assert len(d2["added"]) == 1 and any("DIFFB" in x for x in d2["added"])
        assert _names(drv, wid) == ["DIFFA"]
        # 5) 「このフォルダを登録」再押下＝冪等リラン（受付→背景で反映）
        r2 = c.post("/worlds", json={"path": root})
        assert r2.status_code == 202 and r2.json()["world_id"] == wid, r2.text
        _wait_ingest_idle(c, wid)
        assert set(_names(drv, wid)) == {"DIFFA", "DIFFB"}
        # 6) もう差分なし
        d3 = c.get(f"/worlds/{wid}/diff").json()
        assert d3["added"] == [] and d3["changed"] == [] and d3["removed"] == []
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_diff_robust_and_id_mismatch():
    """RV: 明細未保存でも署名一致なら差分なし＋次回sync でバックフィル（High）／既存 root に別 world_id 登録は 409（Med#1）。"""
    from fastapi.testclient import TestClient
    from sherpa.api import app
    from sherpa.ingest import worker
    a = tempfile.mkdtemp()
    _mk(a, "案件X", "ROBUSTA")
    root = str(pathlib.Path(a).resolve())
    wid = "test_world_robust"
    c = TestClient(app)
    drv = driver()
    try:
        r = c.post("/worlds", json={"path": root, "world_id": wid, "label": "堅牢案件"})
        assert r.status_code == 202, r.text
        _wait_ingest_idle(c, wid)
        # 既存 root に **別 world_id** での登録要求は 409（すり替え拒否）
        assert c.post("/worlds", json={"path": root, "world_id": "other_id"}).status_code == 409
        # 明細を NULL に落としても（旧データ相当）、署名一致なら差分は「全件 added」にならず 0
        with store._connect() as cc:
            cc.execute("UPDATE worlds SET last_manifest=NULL WHERE world_id=%s", (wid,))
        d = c.get(f"/worlds/{wid}/diff").json()
        assert d["added"] == [] and d["changed"] == [] and d["removed"] == []
        # 次の同期で明細がバックフィルされる
        worker.sync(wid)
        assert store.get_world(wid).get("last_manifest")
    finally:
        try:
            worlds.delete(wid)
        except Exception:
            pass
        store.delete_world_row(wid)
        with drv.session() as s:
            s.run("MATCH (n:Entity {world_id:$w}) DETACH DELETE n", w=wid)
        drv.close()
        shutil.rmtree(a, ignore_errors=True)


def test_fixtures_not_referenced_without_flag():
    """C・本番非参照: SHERPA_USE_FIXTURES が無いと world_dir('v1') は架空 golden(fixtures/corpus) に落ちない。

    ＝fixtures はフラグでのみ参照される（本番＝フラグ無し→不活性）。レジストリ未登録の v1 を前提。
    """
    from fastapi import HTTPException
    from sherpa.agents import _kb_hint                        # Codex/agentic プロンプトの参照面も固定（rv-full2 MED）
    from sherpa.api import _require_world
    saved = os.environ.pop("SHERPA_USE_FIXTURES", None)
    o_get = store.get_world
    store.get_world = lambda *a, **k: None                    # v1 は未登録（レジストリに無い）
    try:
        assert not worlds._fixtures()                        # フラグ無し
        assert worlds.world_dir("v1") is None                # data/kb/v1 も無い→None（fixtures に落ちない・None に解決）
        # Codex プロンプトの資料案内も非参照（フラグ無しは data/kb のみ・fixtures を出さない）。
        assert "fixtures/corpus" not in _kb_hint("v1") and "data/kb" in _kb_hint("v1")
        try:                                                 # エンドポイント world ガードも未解決を 404 で弾く（Neo4j 直読み防止）
            _require_world("v1")
            assert False, "未解決 world は 404 のはず"
        except HTTPException as e:
            assert e.status_code == 404
        os.environ["SHERPA_USE_FIXTURES"] = "1"              # フラグ有り → 従来どおり fixtures に解決
        d2 = worlds.world_dir("v1")
        assert d2 is not None and "fixtures/corpus/v1" in str(d2)
        assert "fixtures/corpus/v1" in _kb_hint("v1")        # フラグ有りのときだけ fixtures を案内
        _require_world("v1")                                 # フラグ有りなら通る（例外を投げない）
    finally:
        store.get_world = o_get
        if saved is not None:
            os.environ["SHERPA_USE_FIXTURES"] = saved
        else:
            os.environ.pop("SHERPA_USE_FIXTURES", None)
