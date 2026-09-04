"""L8（複数アーム観測の並存＋LLM再構築の入力・§8.2）の単体/実往復テスト。

正典: `docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §8.2。実VLM/実LLM 呼び出しは
一切発生しない（`vision_arm.resolve_vlm`/`vision_arm._vlm_read` を monkeypatch）。固定する契約:

1. canonical が読めている要素へ第二アーム（VLM）が走らないこと（`ocr_router` の raster 候補選定を
   再利用＝picture 等 metadata_only の要素だけが対象）。
2. 観測が「AI観測」ラベル付きレコードとして rag.md に出ること（既存の `_ai_observation_records`
   契約に乗る）。
3. `llm_render` が観測レコードを成形対象外にすること（`test_llm_render.py` 側で別途固定・ここでは
   実生成された rag.md のマーカー行が一致することだけ確認する）。
4. VLM 不可の構成（vision 未有効／実効不可）では何も起きないこと（既定コストゼロ）。
"""
from __future__ import annotations

import hashlib
import os
import shutil
from pathlib import Path

os.environ.setdefault("SHERPA_USE_FIXTURES", "1")

import openpyxl
import pytest

from sherpa.ingest import ai_observation, evidence_ir as IR, llm_render, ocr_router, office_md
from sherpa.ingest.arms import vision_arm

_ROOT = Path(__file__).resolve().parents[2]
_XLSX_WITH_PICTURE = _ROOT / "fixtures" / "eval" / "deprecation_markers" / "inputs" / "DEP-XLSX-MARKERS.xlsx"


# ---- 1. vision_arm.build_asset_observations（単体・IR/ルート決定を手組みで固定） ------------------

def _asset_hash(data: bytes) -> str:
    return "sha256:" + hashlib.sha256(data).hexdigest()


def _picture_ir(asset_hash: str) -> IR.EvidenceIR:
    source = IR.EvidenceSource(file_type="xlsx", content_hash="sha256:" + "11" * 32)
    locator = IR.Locator(part="xl/drawings/drawing1.xml", sheet="対象")
    coverage_id = IR.make_coverage_id("element", "picture", locator)
    coverage = IR.CoverageItem(
        coverage_id=coverage_id, scope="element", detected_kind="picture",
        locator=locator, status="metadata_only",
        content_basis="pixel_only", reason_code="image_content_uninterpreted",
        parser_id="test", detail={},
    )
    element = IR.EvidenceElement(
        element_id="pic1", type="picture", parent_id=None, order=1, value=None,
        locator=locator,
        coverage_id=coverage_id, extension={"asset_sha256": asset_hash},
    )
    return IR.EvidenceIR(
        schema_version=IR.EVIDENCE_IR_SCHEMA_VERSION, parser_profile=IR.EVIDENCE_PARSER_PROFILE,
        source=source, elements=[element], coverage=[coverage],
    )


def _asset_decision(asset_hash: str, rel_path: str = "img.png") -> ocr_router.OCRRouteDecision:
    return ocr_router.OCRRouteDecision(
        route_input_id="ocr-input:" + hashlib.sha256(rel_path.encode()).hexdigest()[:24],
        target_evidence_id="pic1", input_kind="asset", status="selected",
        reason_code="evidence_raster_asset", priority=100,
        asset_sha256=asset_hash, asset_rel_path=rel_path, media_type="image/png",
        pixel_size=[10, 10],
    )


def _page_render_decision() -> ocr_router.OCRRouteDecision:
    return ocr_router.OCRRouteDecision(
        route_input_id="ocr-input:page1", target_evidence_id="pic1", input_kind="page_render",
        status="selected", reason_code="scan_page_render_fallback", priority=90,
        media_type="image/png",
        page_render={**ocr_router.PAGE_RENDER_PROFILE, "page_1_based": 1},
    )


def test_build_asset_observations_returns_none_when_vlm_unusable(monkeypatch, tmp_path):
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: None)

    def _boom(*a, **kw):
        raise AssertionError("VLM が実効不可なら画像を読んではいけない")

    monkeypatch.setattr(vision_arm, "_vlm_read", _boom)
    data = b"\x89PNG\r\n\x1a\n" + b"0" * 16
    (tmp_path / "img.png").write_bytes(data)
    ah = _asset_hash(data)
    result = vision_arm.build_asset_observations(
        _picture_ir(ah), decisions=[_asset_decision(ah)], asset_root=tmp_path)
    assert result is None


def test_build_asset_observations_skips_page_render_decisions(monkeypatch, tmp_path):
    """page_render は vision アーム本体（PDF全体の視覚読み取り）の単独担当＝補足観測では二重に読まない。"""
    data = b"\x89PNG\r\n\x1a\n" + b"0" * 16
    (tmp_path / "img.png").write_bytes(data)
    ah = _asset_hash(data)
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: {"provider": "ollama", "model": "qwen2.5vl"})
    calls = []

    def _fake_read(path, cfg, timeout):
        calls.append(str(path))
        return "読み取り結果"

    monkeypatch.setattr(vision_arm, "_vlm_read", _fake_read)
    result = vision_arm.build_asset_observations(
        _picture_ir(ah), decisions=[_page_render_decision(), _asset_decision(ah)], asset_root=tmp_path)
    assert result is not None
    assert len(calls) == 1 and calls[0].endswith("img.png")   # page_render分は読まれていない
    assert len(result.observations) == 1


def test_build_asset_observations_skips_empty_vlm_response(monkeypatch, tmp_path):
    data = b"\x89PNG\r\n\x1a\n" + b"0" * 16
    (tmp_path / "img.png").write_bytes(data)
    ah = _asset_hash(data)
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: {"provider": "ollama", "model": "qwen2.5vl"})
    monkeypatch.setattr(vision_arm, "_vlm_read", lambda *a, **kw: "   ")
    result = vision_arm.build_asset_observations(
        _picture_ir(ah), decisions=[_asset_decision(ah)], asset_root=tmp_path)
    assert result is None


def test_build_asset_observations_builds_valid_answer_grade_set(monkeypatch, tmp_path):
    """confidence/use_for_answerは既存の使用可否ルール（MIN_ANSWER_CONFIDENCE）を満たす値で固定。"""
    data = b"\x89PNG\r\n\x1a\n" + b"0" * 16
    (tmp_path / "img.png").write_bytes(data)
    ah = _asset_hash(data)
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: {"provider": "ollama", "model": "qwen2.5vl"})
    monkeypatch.setattr(vision_arm, "_vlm_read", lambda *a, **kw: "手書きメモの書き起こし")
    ir = _picture_ir(ah)
    result = vision_arm.build_asset_observations(ir, decisions=[_asset_decision(ah)], asset_root=tmp_path)
    assert result is not None
    assert ai_observation.validation_errors(result, ir=ir) == []
    obs = result.observations[0]
    assert obs.kind == "summary"
    assert obs.text == "手書きメモの書き起こし"
    assert obs.confidence == vision_arm.VISION_OBSERVATION_CONFIDENCE
    assert obs.confidence >= ai_observation.MIN_ANSWER_CONFIDENCE
    assert obs.use_for_answer is True
    assert ai_observation.answer_observations(result) == [obs]   # 使用可否ルールを実際に満たす


# ---- 2. office_md._build_vlm_observation_set（vision アーム有効化のゲート） -----------------------

def test_office_md_gate_skips_when_vision_arm_not_enabled(monkeypatch, tmp_path):
    from sherpa.ingest import arms as _arms
    monkeypatch.setattr(_arms, "enabled_arm_names", lambda: ["ooxml", "pdf_text"])

    def _boom():
        raise AssertionError("vision が無効なら resolve_vlm を呼んではいけない")

    monkeypatch.setattr(vision_arm, "resolve_vlm", _boom)
    ah = _asset_hash(b"x")
    result = office_md._build_vlm_observation_set(_picture_ir(ah), "a.xlsx", tmp_path)
    assert result is None


def test_office_md_gate_skips_when_vlm_unusable(monkeypatch, tmp_path):
    from sherpa.ingest import arms as _arms
    monkeypatch.setattr(_arms, "enabled_arm_names", lambda: ["ooxml", "pdf_text", "vision"])
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: None)
    ah = _asset_hash(b"x")
    result = office_md._build_vlm_observation_set(_picture_ir(ah), "a.xlsx", tmp_path)
    assert result is None


def test_office_md_gate_skips_when_no_assets_present(monkeypatch, tmp_path):
    from sherpa.ingest import arms as _arms
    monkeypatch.setattr(_arms, "enabled_arm_names", lambda: ["ooxml", "pdf_text", "vision"])
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: {"provider": "ollama", "model": "qwen2.5vl"})
    ah = _asset_hash(b"x")
    result = office_md._build_vlm_observation_set(_picture_ir(ah), "a.xlsx", tmp_path)  # 空ディレクトリ
    assert result is None


def test_office_md_gate_delegates_to_vision_arm_with_selected_asset_decisions_only(monkeypatch, tmp_path):
    from sherpa.ingest import arms as _arms
    monkeypatch.setattr(_arms, "enabled_arm_names", lambda: ["ooxml", "pdf_text", "vision"])
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: {"provider": "ollama", "model": "qwen2.5vl"})
    data = b"\x89PNG\r\n\x1a\n" + b"0" * 16
    (tmp_path / "img.png").write_bytes(data)
    ah = _asset_hash(data)
    ir = _picture_ir(ah)

    captured = {}

    def _fake_build(passed_ir, *, decisions, asset_root):
        captured["ir"] = passed_ir
        captured["decisions"] = decisions
        captured["asset_root"] = asset_root
        return "sentinel"

    monkeypatch.setattr(vision_arm, "build_asset_observations", _fake_build)
    result = office_md._build_vlm_observation_set(ir, "a.xlsx", tmp_path)
    assert result == "sentinel"
    assert captured["ir"] is ir
    assert captured["asset_root"] == tmp_path
    assert len(captured["decisions"]) == 1
    assert captured["decisions"][0].status == "selected"
    assert captured["decisions"][0].input_kind == "asset"


# ---- 3. 実往復（office_md.build_derived）: 4つの受け入れ条件を実ファイルで固定 ---------------------

@pytest.fixture
def _vlm_stub(monkeypatch):
    """VLM を usable にし、呼び出しを記録するfixture（呼ばれなければ calls は空のまま）。"""
    calls: list[str] = []

    def _fake_read(image_path, cfg, timeout):
        calls.append(str(image_path))
        return "手書きで「テスト観測」と書かれている"

    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: {
        "provider": "ollama", "model": "qwen2.5vl", "cloud_allowed": False,
        "ollama_url": "http://localhost:11434",
    })
    monkeypatch.setattr(vision_arm, "_vlm_read", _fake_read)
    return calls


def _built_rag_md(monkeypatch, tmp_path, source: Path, *, arms: str = "ooxml,pdf_text,vision") -> tuple[str, Path]:
    monkeypatch.setenv("SHERPA_ARMS", arms)
    wd = tmp_path / "world"
    wd.mkdir()
    shutil.copy(source, wd / source.name)
    dmd = tmp_path / "derived"
    rep = office_md.build_derived(wd, dmd)
    assert not rep.get("error") and rep["rag_failed"] == 0, rep
    rag_path = dmd.parent / "rag" / f"{source.name}.rag.md"
    return rag_path.read_text(encoding="utf-8"), dmd


@pytest.mark.skipif(not _XLSX_WITH_PICTURE.is_file(), reason="fixture が無い環境")
def test_ai_observation_appears_as_labeled_record_when_vlm_usable(monkeypatch, tmp_path, _vlm_stub):
    """契約2: 観測が「AI観測」ラベル付きレコードとしてrag.mdに出る。"""
    rag, _ = _built_rag_md(monkeypatch, tmp_path, _XLSX_WITH_PICTURE)
    assert "採用AI観測Set:" in rag
    assert "AI観測生成元: ollama/qwen2.5vl（local）" in rag
    assert llm_render._AI_OBSERVATION_BODY_MARKER in rag
    assert "手書きで「テスト観測」と書かれている" in rag
    assert _vlm_stub, "VLM が呼ばれていない"


@pytest.mark.skipif(not _XLSX_WITH_PICTURE.is_file(), reason="fixture が無い環境")
def test_no_observation_when_vision_arm_not_enabled(monkeypatch, tmp_path, _vlm_stub):
    """契約4: VLM不可（vision が有効アームに無い）の構成では何も起きない（既定コストゼロ）。"""
    rag, _ = _built_rag_md(monkeypatch, tmp_path, _XLSX_WITH_PICTURE, arms="ooxml,pdf_text")
    assert llm_render._AI_OBSERVATION_BODY_MARKER not in rag
    assert not _vlm_stub, "vision 無効なのに VLM が呼ばれている"


@pytest.mark.skipif(not _XLSX_WITH_PICTURE.is_file(), reason="fixture が無い環境")
def test_no_observation_when_vlm_unresolvable(monkeypatch, tmp_path):
    """契約4: vision は有効でも VLM が実効利用不可（resolve_vlm=None）なら何も起きない。"""
    monkeypatch.setattr(vision_arm, "resolve_vlm", lambda: None)

    def _boom(*a, **kw):
        raise AssertionError("実効不可なのに画像を読んではいけない")

    monkeypatch.setattr(vision_arm, "_vlm_read", _boom)
    rag, _ = _built_rag_md(monkeypatch, tmp_path, _XLSX_WITH_PICTURE)
    assert llm_render._AI_OBSERVATION_BODY_MARKER not in rag


def test_no_vlm_call_when_canonical_fully_readable(monkeypatch, tmp_path, _vlm_stub):
    """契約1: canonical が読めている要素（画像を含まない素のxlsx）へ第二アームが走らない。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "項目"
    ws["B1"] = "値"
    ws["A2"] = "対象システム"
    ws["B2"] = "BETA契約管理システム"
    plain = tmp_path / "plain.xlsx"
    wb.save(plain)
    rag, _ = _built_rag_md(monkeypatch, tmp_path, plain)
    assert llm_render._AI_OBSERVATION_BODY_MARKER not in rag
    assert not _vlm_stub, "画像の無い文書で VLM が呼ばれている"


@pytest.mark.skipif(not _XLSX_WITH_PICTURE.is_file(), reason="fixture が無い環境")
def test_llm_render_skips_observation_record_from_real_generated_rag_md(monkeypatch, tmp_path, _vlm_stub):
    """契約3: 実生成された rag.md を llm_render に通しても、観測レコードは本文が変わらない
    （保護行/原値検証の実データでの確認・canonical record 側は成形される）。"""
    rag, _ = _built_rag_md(monkeypatch, tmp_path, _XLSX_WITH_PICTURE)
    rag = llm_render.stamp_rule_only(rag) if llm_render.needs_llm_pass(rag) else rag
    from sherpa.ingest import graph_extract

    def _fake_complete(system, user, cfg, timeout=None):
        # 保護行/引用値をすべて素通しし、末尾に一言足すだけの無害な成形（検証を必ず通す）。
        body = user.split("次のレコードを整形してください:\n\n", 1)[1].split("\n\n---\n参考情報")[0]
        return __import__("json").dumps({"text": body + "\n（成形済み）"})

    monkeypatch.setattr(graph_extract, "complete_json", _fake_complete)
    cfg = {"provider": "openai", "model": "gpt-5.5"}
    result = llm_render.format_document("v1", "a.xlsx", rag, cfg, {})
    assert result is not None
    # 観測本文（マーカー行〜「Canonical Evidenceの値ではなく」まで）は一字一句そのまま残る。
    assert "手書きで「テスト観測」と書かれている" in result.markdown
    assert "AI画像観測（原本確定値ではない）\n観測内容: 手書きで「テスト観測」と書かれている\n（成形済み）" not in result.markdown
    assert result.llm_count >= 1   # 少なくとも canonical 側のどれかは成形されている
