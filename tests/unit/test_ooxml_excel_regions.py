"""Excel（.xlsx）表候補抽出（`sherpa/ingest/ooxml/excel.py::regions()` 系）の単体テスト。

`regions()` は値グリッド・結合セル情報・背景色情報を受け取る純関数のため、ほとんどのケースは
openpyxl を経由せず素の Python データ構造だけで検証する。`filled_cells()`（openpyxl ワークシート消費層）と
実ファイルを通した end-to-end 経路（`ooxml_arm._build_xlsx_ir`）は、最小限の openpyxl ワークブックで
別途検証する。
"""
from __future__ import annotations

import json
import pathlib
import time
import zipfile

import openpyxl
import pytest
from openpyxl.styles import GradientFill, PatternFill
from openpyxl.styles.colors import Color

from sherpa.ingest import office_md
from sherpa.ingest.arms import ooxml_arm
from sherpa.ingest.ooxml import excel


def _grid(rows: int, cols: int, values: dict[tuple[int, int], object]) -> list[list]:
    """1-based座標→値の疎な dict から `regions()` 向けの `ws_values` グリッド（0-basedリストのリスト）を作る。"""
    g: list[list] = [[None] * cols for _ in range(rows)]
    for (r, c), v in values.items():
        g[r - 1][c - 1] = v
    return g


# ---- 背景色付きセルの占有化 ----

def test_background_fill_bridges_adjacent_value_cells():
    grid = _grid(1, 3, {(1, 1): "A", (1, 3): "B"})
    without = excel.regions(grid, cap_rows=1, cap_cols=3)
    assert len(without) == 2                           # 橋渡しが無ければ非連結のまま（従来どおり）

    with_fill = excel.regions(grid, cap_rows=1, cap_cols=3, filled={(1, 2)})
    assert len(with_fill) == 1
    rg = with_fill[0]
    assert rg.range == "A1:C1"
    assert rg.cells == frozenset({(1, 1), (1, 2), (1, 3)})
    assert rg.value_cell_count == 2                    # 橋渡しセル自体は値を持たないため数えない


def test_filled_beyond_cap_is_ignored():
    grid = _grid(1, 3, {(1, 1): "A", (1, 3): "B"})
    out = excel.regions(grid, cap_rows=1, cap_cols=2, filled={(1, 2)})
    assert len(out) == 1                               # cap_cols=2 なので (1,3) 自体が走査対象外


# ---- 結合セルの非anchor（継続セル）の占有化 ----

def test_merge_continuation_bridges_adjacent_value_cluster():
    # A1:C1 結合（anchor=(1,1) に値、継続セル=(1,2)/(1,3) は空）の直下にデータ行（2行目）が続く、
    # よくある「結合ヘッダ＋データ行」レイアウト。継続セルが占有されて初めて2行分がきれいな矩形になる。
    grid = _grid(2, 3, {(1, 1): "見出し", (2, 1): "a", (2, 2): "b", (2, 3): "c"})
    merged = {(1, 1): {"anchor": (1, 1), "row_span": 1, "column_span": 3},
              (1, 2): {"anchor": (1, 1), "row_span": 1, "column_span": 3},
              (1, 3): {"anchor": (1, 1), "row_span": 1, "column_span": 3}}

    without = excel.regions(grid, cap_rows=2, cap_cols=3)
    assert len(without) == 2                           # 結合情報が無ければ従来どおり分裂（既知の限界）

    with_merge = excel.regions(grid, cap_rows=2, cap_cols=3, merged=merged)
    assert len(with_merge) == 1
    rg = with_merge[0]
    assert rg.range == "A1:C2"
    assert rg.cells == frozenset({(1, 1), (1, 2), (1, 3), (2, 1), (2, 2), (2, 3)})
    assert rg.value_cell_count == 4                    # 継続セル2つは値を持たないため数えない


def test_fully_blank_merge_without_fill_creates_no_region():
    grid = _grid(2, 2, {})
    merged = {(1, 1): {"anchor": (1, 1), "row_span": 1, "column_span": 2},
              (1, 2): {"anchor": (1, 1), "row_span": 1, "column_span": 2}}
    out = excel.regions(grid, cap_rows=2, cap_cols=2, merged=merged)
    assert out == []                                   # 値も塗りも無い装飾用の結合だけでは表候補にしない


def test_merge_with_fill_on_continuation_occupies_whole_span():
    # anchor は空・値なしだが、継続セルに背景色がある → 結合範囲全体を占有にする。
    grid = _grid(1, 2, {})
    merged = {(1, 1): {"anchor": (1, 1), "row_span": 1, "column_span": 2},
              (1, 2): {"anchor": (1, 1), "row_span": 1, "column_span": 2}}
    out = excel.regions(grid, cap_rows=1, cap_cols=2, merged=merged, filled={(1, 2)})
    assert len(out) == 1
    assert out[0].cells == frozenset({(1, 1), (1, 2)})


# ---- 空行区切り（負の対照: 誤って連結しないことの確認）----

def test_blank_row_separator_keeps_tables_separate():
    grid = _grid(3, 2, {(1, 1): "a", (1, 2): "b", (3, 1): "c", (3, 2): "d"})   # 行2は空行区切り
    out = excel.regions(grid, cap_rows=3, cap_cols=2)
    assert len(out) == 2
    assert {rg.range for rg in out} == {"A1:B1", "A3:B3"}


# ---- ヒストグラム法の最大矩形反復抽出（隣接表・L字の癒着解消）----

def test_split_avoids_phantom_blank_cells_in_l_shaped_cluster():
    """表 A（5行×3列）と表 B（3行×3列）が列境界で接触し1つの連結成分になるが、
    外接矩形1つのままだと非占有セル（右下の空白）を巻き込む。分割後はどの領域の外接矩形も
    非占有セルを含まないこと（＝空セルの巻き込み無し）と、完全性（セルの総数が変わらないこと）を確認する。
    """
    values: dict[tuple[int, int], object] = {}
    for r in range(1, 6):
        for c in range(1, 4):
            values[(r, c)] = f"A{r}{c}"
    for r in range(1, 4):
        for c in range(4, 7):
            values[(r, c)] = f"B{r}{c}"
    grid = _grid(5, 6, values)

    out = excel.regions(grid, cap_rows=5, cap_cols=6)
    assert len(out) >= 2                               # 1つの外接矩形のままでは無い（癒着解消）

    occupied = set(values.keys())
    union: set[tuple[int, int]] = set()
    for rg in out:
        for r in range(rg.min_row, rg.max_row + 1):
            for c in range(rg.min_col, rg.max_col + 1):
                assert (r, c) in occupied, f"region {rg.range} が非占有セル ({r},{c}) を巻き込んでいる"
        union |= rg.cells
    assert union == occupied                           # 完全性（silent-drop ゼロ）


def test_split_is_deterministic_across_repeated_calls():
    values = {(r, c): "v" for r in range(1, 7) for c in range(1, 7) if (r + c) % 3 != 0}
    grid = _grid(6, 6, values)
    first = excel.regions(grid, cap_rows=6, cap_cols=6)
    second = excel.regions(grid, cap_rows=6, cap_cols=6)
    assert first == second
    assert [rg.range for rg in first] == [rg.range for rg in second]


def test_split_cap_falls_back_without_losing_cells():
    """1本の背骨（行1全列）から多数の「指」が垂れ下がる、非矩形性の強い連結成分。
    最大矩形の反復抽出回数の上限（`_MAX_RECT_SPLITS`）に達しても、完全性（セルの総数）は保たれる
    ことと、実行時間が有界であることを確認する（計算量の安全弁）。
    """
    n_fingers = 50
    width = n_fingers * 2 - 1
    values: dict[tuple[int, int], object] = {(1, c): "spine" for c in range(1, width + 1)}
    for i in range(n_fingers):
        col = 1 + i * 2
        for r in (2, 3, 4):
            values[(r, col)] = f"finger{i}"
    grid = _grid(4, width, values)

    started = time.monotonic()
    out = excel.regions(grid, cap_rows=4, cap_cols=width)
    elapsed = time.monotonic() - started

    assert elapsed < 5.0, f"分割の反復回数上限に達した場合でも高速に完了すること（実測 {elapsed:.2f}s）"
    union: set[tuple[int, int]] = set()
    for rg in out:
        union |= rg.cells
    assert union == set(values.keys())                 # 反復回数の上限到達後もセルは失われない


# ---- 表候補スコア（付与のみ・抑制はしない）----

def test_score_tapers_for_small_regions_and_saturates_for_dense_ones():
    singleton = _grid(1, 1, {(1, 1): "x"})
    out = excel.regions(singleton, cap_rows=1, cap_cols=1)
    rg = out[0]
    assert rg.value_cell_count == 1 and rg.density == 1.0
    expected = 1.0 * min(1.0, 1 / excel._SCORE_MIN_CELLS)
    assert rg.score == pytest.approx(expected)
    assert rg.score < 1.0                              # 最小非空セル数未満は下駄を履かせて割り引く

    dense = _grid(3, 3, {(r, c): "x" for r in range(1, 4) for c in range(1, 4)})
    out2 = excel.regions(dense, cap_rows=3, cap_cols=3)
    rg2 = out2[0]
    assert rg2.value_cell_count == 9 and rg2.density == 1.0
    assert rg2.score == pytest.approx(1.0)             # 最小非空セル数以上では密度がそのままスコアになる


def test_score_reflects_sparse_density_when_component_too_large_to_decompose():
    """外接矩形面積が分割コスト上限（`_MAX_RECT_DECOMPOSE_CELLS`）を超える連結成分は分割せず単一の
    外接矩形のまま返す（巨大シートの計算量安全弁）ため、密度が1.0未満のまま観測できる。
    """
    rows = 100
    cols = excel._MAX_RECT_DECOMPOSE_CELLS // rows + 1  # 面積が上限を必ず超えるよう動的に決める（HM1で上限値を変更してもテストが追随）
    values: dict[tuple[int, int], object] = {(1, c): "h" for c in range(1, cols + 1)}
    values.update({(r, 1): "v" for r in range(1, rows + 1)})
    grid = _grid(rows, cols, values)

    started = time.monotonic()
    out = excel.regions(grid, cap_rows=rows, cap_cols=cols)
    elapsed = time.monotonic() - started

    assert elapsed < 5.0
    assert len(out) == 1                               # 分割コスト上限を超えるため単一の外接矩形のまま
    rg = out[0]
    assert rg.min_row == 1 and rg.max_row == rows and rg.min_col == 1 and rg.max_col == cols
    assert rg.value_cell_count == len(values)
    assert 0.0 < rg.density < 1.0
    assert rg.score == pytest.approx(rg.density)       # 最小非空セル数を大きく超えるため頭打ちのみ


# ---- 巨大シート: 単一の密な表は分割コストを払わず高速に1領域のまま ----

def test_huge_dense_single_table_bypasses_decomposition_and_stays_one_region():
    rows, cols = 2000, 30                               # 面積 60,000 > _MAX_RECT_DECOMPOSE_CELLS
    values = {(r, c): "v" for r in range(1, rows + 1) for c in range(1, cols + 1)}
    grid = _grid(rows, cols, values)

    started = time.monotonic()
    out = excel.regions(grid, cap_rows=rows, cap_cols=cols)
    elapsed = time.monotonic() - started

    assert elapsed < 5.0
    assert len(out) == 1
    assert out[0].value_cell_count == len(values)
    assert out[0].score == pytest.approx(1.0)


# ---- filled_cells(): openpyxl ワークシートからの背景色抽出 ----

def test_filled_cells_detects_pattern_fill_and_respects_caps():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B2"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    ws["E5"].fill = PatternFill(fill_type="solid", fgColor="00FF00")

    got = excel.filled_cells(ws, cap_rows=10, cap_cols=10)
    assert got == {(2, 2), (5, 5)}

    capped = excel.filled_cells(ws, cap_rows=3, cap_cols=3)
    assert capped == {(2, 2)}                          # cap 外の (5,5) は対象外


# ---- end-to-end: ooxml_arm._build_xlsx_ir 経由の配線確認 ----

def test_build_xlsx_ir_wires_background_fill_bridge_and_score(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws["A1"] = "左"
    ws["C1"] = "右"
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor="FFFF00")   # 値なしの橋渡しセル
    p = tmp_path / "a.xlsx"
    wb.save(p)

    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    tables = [e for e in doc.elements if e.type == "table"]
    assert len(tables) == 1                            # 背景色の橋渡しで1領域に連結（filled 配線の確認）
    tbl = tables[0]
    assert isinstance(tbl.source_map.get("score"), float)
    texts = {(c.row, c.column): c.text for c in tbl.cells}
    assert texts[(1, 1)] == "左" and texts[(1, 3)] == "右"


def test_build_xlsx_ir_wires_merge_continuation_bridge(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws["A1"] = "見出し"
    ws.merge_cells("A1:C1")                            # 継続セルは B1/C1（値なし）
    ws["A2"], ws["B2"], ws["C2"] = "a", "b", "c"        # 結合ヘッダ直下のデータ行
    p = tmp_path / "a.xlsx"
    wb.save(p)

    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    tables = [e for e in doc.elements if e.type == "table"]
    assert len(tables) == 1                            # 結合継続セルの橋渡しで1領域に連結（merged 配線の確認）


# ---- ING-1: 静かな部分抽出の検知（宣言行数 vs 実際に値が入っていた行数）----

def _sheet_source_map(doc):
    sheets = [e for e in doc.elements if e.type == "sheet"]
    assert len(sheets) == 1
    return sheets[0].source_map


def test_build_xlsx_ir_flags_partial_extraction_when_declared_rows_far_exceed_extracted(tmp_path):
    """宣言行数（openpyxl `ws.max_row`）が実際に値の入った行より極端に多ければ疑いに計上する。

    スタイルだけ（値なし）を遠い行へ適用すると、Excel の「使用範囲」が実データより広く申告される
    （現実にもよくある：書式だけ残って内容が消えたセル）——`ws.max_row` はこの遠い行まで伸びるが、
    実際に値が入っている行は1行だけのまま。cap による打切り（自己申告）ではないので
    `sheet_sm["truncated"]` は立たない。
    """
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws["A1"] = "唯一の値"
    ws.cell(row=150, column=1).font = Font(bold=True)     # 値なし・スタイルだけの遠い行（宣言行数を伸ばす）
    p = tmp_path / "a.xlsx"
    wb.save(p)

    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    sm = _sheet_source_map(doc)
    assert sm.get("truncated") is not True
    assert sm.get("partial_extraction_suspected") is True
    assert sm.get("declared_rows") == 150
    assert sm.get("extracted_rows") == 1


def test_build_xlsx_ir_does_not_flag_normal_dense_sheet(tmp_path):
    """通常の（宣言行数が小さい、または実際に値が詰まっている）シートは疑いに計上しない。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    for r in range(1, 11):
        ws.cell(row=r, column=1, value=f"row{r}")
    p = tmp_path / "a.xlsx"
    wb.save(p)

    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    sm = _sheet_source_map(doc)
    assert "partial_extraction_suspected" not in sm


def test_build_xlsx_ir_does_not_flag_when_cap_truncation_already_self_declared(tmp_path, monkeypatch):
    """H2 の予算打切り（`sheet_sm["truncated"]`＝自己申告）は正常なので、たとえ宣言/抽出の比率が
    疑いの閾値を満たしても部分抽出の疑いには計上しない（区別する・ING-1 裁定4）。"""
    from openpyxl.styles import Font

    from sherpa.ingest.ooxml import excel
    monkeypatch.setattr(excel, "sheet_truncated", lambda *a, **kw: True)   # cap 打切り済みとして扱う

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws["A1"] = "唯一の値"
    ws.cell(row=150, column=1).font = Font(bold=True)
    p = tmp_path / "a.xlsx"
    wb.save(p)

    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    sm = _sheet_source_map(doc)
    assert sm.get("truncated") is True
    assert "partial_extraction_suspected" not in sm


def test_build_xlsx_ir_does_not_flag_sparse_but_genuine_real_fixture():
    """`JPX-007.xlsx`（宣言501行・非空はA1と501行目の2行のみ）は、最終非空行が
    宣言終端(501)に達しているため部分抽出の疑いに計上しない——単純な行数比だけで判定すると
    2/501 ≈ 0.4% は閾値を大きく下回り誤検知していた（実データが宣言の最後まで正しく到達している
    疎な業務ファイルと、宣言終端の手前で黙って途切れた壊れ方を区別する）。
    """
    p = _EVAL_XLSX_DIR / "JPX-007.xlsx"
    if not p.is_file():
        pytest.skip("実 fixture が無い環境")
    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    sheets = [e for e in doc.elements if e.type == "sheet" and e.source_map.get("sheet") == "境界"]
    assert len(sheets) == 1
    sm = sheets[0].source_map
    assert sm.get("truncated") is not True
    assert "partial_extraction_suspected" not in sm


# ---- 実ファイルでの回帰（silent-drop ゼロ）----

_EVAL_XLSX_DIR = (pathlib.Path(__file__).resolve().parent.parent.parent
                  / "fixtures" / "eval" / "excel_ja" / "inputs")


def _small_real_xlsx_fixtures(max_bytes: int = 60_000) -> list[pathlib.Path]:
    if not _EVAL_XLSX_DIR.is_dir():
        return []
    return sorted(p for p in _EVAL_XLSX_DIR.glob("*.xlsx") if p.stat().st_size <= max_bytes)


def _expected_nonblank_cells(path: pathlib.Path) -> dict[tuple[str, int, int], str]:
    """`_build_xlsx_ir` と独立に、ファイルの非空セルを直接読んで期待値を作る（オラクル）。"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    try:
        out: dict[tuple[str, int, int], str] = {}
        for ws in wb.worksheets:
            cap_rows = excel.effective_cap_rows(ws.max_column)
            max_row = min(ws.max_row or 1, cap_rows + 1)
            max_col = min(ws.max_column or 1, excel.DEFAULT_CAP_COLS + 1)
            for r, row in enumerate(
                    ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True),
                    start=1):
                if r > cap_rows:
                    continue
                for c, v in enumerate(row, start=1):
                    if c > excel.DEFAULT_CAP_COLS:
                        continue
                    if v is not None and str(v).strip() != "":
                        out[(ws.title, r, c)] = str(v)
        return out
    finally:
        wb.close()


@pytest.mark.parametrize("path", _small_real_xlsx_fixtures(), ids=lambda p: p.name)
def test_real_fixture_no_silent_drop(path):
    """実在の Excel ファイルで、シート上の非空セルが1つ残らずいずれかの table 要素へ現れることを確認する
    （`regions()` の占有判定/分割を変えても、以前拾えていたセルが拾えなくなるケースが無いことの回帰検証）。
    """
    doc = ooxml_arm._build_xlsx_ir(path)
    assert doc is not None
    actual: dict[tuple[str, int, int], str] = {}
    for el in doc.elements:
        if el.type != "table":
            continue
        sheet = el.source_map["sheet"]
        for cell in el.cells or []:
            if cell.text:
                actual[(sheet, cell.row, cell.column)] = cell.text

    expected = _expected_nonblank_cells(path)
    missing = set(expected) - set(actual)
    assert not missing, f"{path.name}: 取りこぼしたセル {sorted(missing)[:10]}"
    mismatched = {k for k in expected if k in actual and actual[k] != expected[k]}
    assert not mismatched, f"{path.name}: 値不一致 {sorted(mismatched)[:10]}"


# ---- XLSX_EXTRACTOR_VERSION が document_ir/evidence_ir の署名に含まれる契約の固定 ----

def _sig_component(sig: str, key: str) -> str | None:
    for part in sig.split(";"):
        if part.startswith(key + "="):
            return part[len(key) + 1:]
    return None


def test_xlsx_extractor_version_included_in_document_and_evidence_sig():
    doc_sig = office_md._current_document_ir_sig()
    evidence_sig = office_md._current_evidence_ir_sig()
    assert _sig_component(doc_sig, "xlsx") == ooxml_arm.XLSX_EXTRACTOR_VERSION
    assert _sig_component(evidence_sig, "xlsx") == ooxml_arm.XLSX_EXTRACTOR_VERSION


def test_upgrade_from_v1_signature_regenerates_v2_region_structure(tmp_path):
    """v1（旧 XLSX_EXTRACTOR_VERSION）で構築済みの world を模し、drift 判定→refresh 関数の直接呼び出しで
    Document/Evidence IR が現行版へ更新される**成果物そのもの**を検証する（`worker.sync` への配線は
    別レーン＝`ooxml_arm.XLSX_EXTRACTOR_VERSION` docstring 参照。ここでは refresh 関数を直接呼ぶ）。

    署名文字列の一致だけでは「マーカーだけ書き換わって中身は古いまま」という取り違えを検出できない
    ため、`.document.json` を明示的に壊れた内容へ置き換えてから refresh し、v2 固有の領域構造
    （隣接2表の癒着解消による分割・`score` フィールド）を伴って正しく再生成されることを確認する。
    """
    world = tmp_path / "world"
    world.mkdir()
    derived = tmp_path / "derived"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    # 隣接する2表（列境界で接触・癒着ケース）: v2 は regions() の分割で2つの table に分かれる。
    for r in range(1, 6):
        for c in range(1, 4):
            ws.cell(row=r, column=c, value=f"A{r}{c}")
    for r in range(1, 4):
        for c in range(4, 7):
            ws.cell(row=r, column=c, value=f"B{r}{c}")
    wb.save(world / "a.xlsx")

    office_md.build_derived(world, derived)             # 現行(v2)コードで完全な派生一式を構築
    # §8.1 三階層＝.document.json/.evidence.json は ir 層（derived＝md 層の兄弟）。
    derived_ir = derived.parent / "ir"
    doc_path = derived_ir / "a.xlsx.document.json"
    evidence_path = derived_ir / "a.xlsx.evidence.json"
    original_doc_json = doc_path.read_text(encoding="utf-8")
    original_evidence_json = evidence_path.read_text(encoding="utf-8")
    v2_tables = [e for e in json.loads(original_doc_json)["elements"] if e["type"] == "table"]
    assert len(v2_tables) == 2                           # 前提の確認: v2 は癒着を2領域に分割する
    assert all("score" in t["source_map"] for t in v2_tables)

    doc_marker = derived / office_md._DOCUMENT_IR_SIG_MARKER
    evidence_marker = derived / office_md._EVIDENCE_IR_SIG_MARKER
    current_doc_sig = doc_marker.read_text(encoding="utf-8")
    current_evidence_sig = evidence_marker.read_text(encoding="utf-8")
    xlsx_component = f"xlsx={ooxml_arm.XLSX_EXTRACTOR_VERSION}"
    assert xlsx_component in current_doc_sig and f"{xlsx_component};" in current_evidence_sig

    old_doc_sig = current_doc_sig.replace(xlsx_component, "xlsx=xlsx-ooxml-v1")
    old_evidence_sig = current_evidence_sig.replace(f"{xlsx_component};", "xlsx=xlsx-ooxml-v1;")
    doc_marker.write_text(old_doc_sig, encoding="utf-8")          # v1 世代のマーカーを模す
    evidence_marker.write_text(old_evidence_sig, encoding="utf-8")
    doc_path.write_text('{"corrupted": "v1-era placeholder"}', encoding="utf-8")   # 中身も古いまま壊す
    evidence_path.write_text('{"corrupted": "v1-era placeholder"}', encoding="utf-8")

    assert office_md.document_ir_sig_drift(derived) is True
    assert office_md.evidence_ir_sig_drift(derived) is True

    doc_result = office_md.refresh_document_ir(world, derived)   # sync を経由せず直接呼ぶ
    assert doc_result["document_ir_failed"] == 0
    evidence_result = office_md.refresh_evidence_ir(world, derived)
    assert evidence_result["evidence_ir_failed"] == 0

    assert office_md.document_ir_sig_drift(derived) is False     # 再生成後は drift 解消
    assert office_md.evidence_ir_sig_drift(derived) is False
    assert doc_marker.read_text(encoding="utf-8") == current_doc_sig
    assert evidence_marker.read_text(encoding="utf-8") == current_evidence_sig

    regenerated_doc_json = doc_path.read_text(encoding="utf-8")
    assert regenerated_doc_json == original_doc_json              # 壊した内容ではなく v2 の中身が復元される
    regen_tables = [e for e in json.loads(regenerated_doc_json)["elements"] if e["type"] == "table"]
    assert len(regen_tables) == 2                                  # 癒着解消の分割構造が再生成される
    assert all("score" in t["source_map"] for t in regen_tables)   # v2 固有のフィールドが復元される
    assert evidence_path.read_text(encoding="utf-8") == original_evidence_json


# ---- GradientFill で AttributeError にならないこと（グラデーションは占有扱い）----

def test_gradient_fill_is_treated_as_occupied_without_crash():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"].fill = GradientFill(stop=(Color(rgb="FFFF0000"), Color(rgb="FF00FF00")))
    got = excel.filled_cells(ws, cap_rows=10, cap_cols=10)   # 例外を投げないこと自体が確認事項
    assert got == {(2, 2)}


def test_gradient_fill_all_white_stops_is_not_occupied():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["B2"].fill = GradientFill(stop=(Color(rgb="FFFFFFFF"), Color(auto=True)))
    got = excel.filled_cells(ws, cap_rows=10, cap_cols=10)
    assert got == set()


# ---- 結合範囲の非anchorだけに値がある異常OOXML（既知の限界・クラッシュしない）----

def _inject_raw_cell_value(xlsx_path: pathlib.Path, out_path: pathlib.Path, row_xml_needle: str,
                          extra_cell_xml: str) -> None:
    """`xlsx_path` の `sheet1.xml` へ raw XML レベルでセルを1つ追加した `out_path` を書き出す
    （openpyxl 経由では表現できない異常な OOXML を模すためのテスト専用ヘルパ）。
    """
    with zipfile.ZipFile(xlsx_path) as zf:
        contents = {n: zf.read(n) for n in zf.namelist()}
    sheet_xml = contents["xl/worksheets/sheet1.xml"].decode("utf-8")
    assert row_xml_needle in sheet_xml
    contents["xl/worksheets/sheet1.xml"] = sheet_xml.replace(
        row_xml_needle, row_xml_needle[: -len("</row>")] + extra_cell_xml + "</row>").encode("utf-8")
    with zipfile.ZipFile(out_path, "w") as zf:
        for name, data in contents.items():
            zf.writestr(name, data)


def test_merge_non_anchor_raw_value_is_dropped_without_crash(tmp_path):
    """結合範囲（A1:B1）の非anchor（B1）だけに raw XML レベルで値がある異常 OOXML。Excel 仕様上
    非anchorセルの値は無効で、openpyxl の通常ロードの時点で失われる（既知の限界・`merged_map()`
    docstring 参照）。ここではクラッシュしないこと、および anchor（塗りあり）の範囲が領域として
    出力されることだけを固定する。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws.merge_cells("A1:B1")
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")   # anchor に塗りのみ（値は空）
    src = tmp_path / "a.xlsx"
    wb.save(src)

    malformed = tmp_path / "b.xlsx"
    _inject_raw_cell_value(
        src, malformed, '<row r="1"><c r="A1" s="1" t="n"></c></row>',
        '<c r="B1" t="inlineStr"><is><t>異常値</t></is></c>')

    reloaded = openpyxl.load_workbook(malformed)
    assert reloaded.active["B1"].value is None               # openpyxl が非anchorの値を破棄する（受容記録）

    doc = ooxml_arm._build_xlsx_ir(malformed)                 # 例外を投げないこと自体が確認事項
    assert doc is not None
    tables = [e for e in doc.elements if e.type == "table"]
    assert any(t.source_map.get("range") == "A1:B1" for t in tables)


# ---- 背景色の定義（白/自動/テーマ白の除外・条件付き書式は対象外）の境界テスト ----

def test_fill_boundary_solid_white_is_not_occupied():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
    assert excel.filled_cells(ws, cap_rows=5, cap_cols=5) == set()


def test_fill_boundary_solid_color_is_occupied():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor="FFFF0000")
    assert excel.filled_cells(ws, cap_rows=5, cap_cols=5) == {(1, 1)}


def test_fill_boundary_theme_white_is_not_occupied(tmp_path):
    # テーマ解決（`_resolve_theme_lt1_rgb`）は wb.loaded_theme（実ファイルの生テーマXML）を読むため、
    # 保存前のワークブックのままでは解決不能＝占有側に倒れてしまう。save+reload で実ファイル相当にする。
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0, tint=0.0))
    p = tmp_path / "a.xlsx"
    wb.save(p)
    reloaded = openpyxl.load_workbook(p)
    assert excel.filled_cells(reloaded.active, cap_rows=5, cap_cols=5) == set()


def test_fill_boundary_none_is_not_occupied():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値はあるが塗りは無し"
    assert excel.filled_cells(ws, cap_rows=5, cap_cols=5) == set()


# ---- シート全体での分割予算（連結成分が多数ある場合の合計コスト頭打ち）----

def test_sheet_wide_split_budget_bounds_many_comb_components():
    """1シートに独立した櫛形の連結成分（`test_split_cap_falls_back_without_losing_cells` と同型・
    50本指）が100個ある場合、連結成分1つあたりの安全弁（`_MAX_RECT_DECOMPOSE_CELLS`/`_MAX_RECT_SPLITS`）
    だけでは合計コストが頭打ちにならない（実測: シート全体予算を外すと5,100 Region）。
    シート全体の予算（`_MAX_RECT_SPLITS_PER_SHEET`/`_MAX_REGIONS_PER_SHEET`）が働けば大幅に少ない
    Region数で高速に完了し、かつ完全性（セルの総数）が保たれることを確認する。
    """
    n_fingers, gap, n_side = 50, 1, 10                    # 10x10=100 個の独立した櫛形連結成分
    comb_width = n_fingers * 2 - 1
    comb_rows = 4
    values: dict[tuple[int, int], object] = {}
    for bi in range(n_side):
        for bj in range(n_side):
            r0 = bi * (comb_rows + gap)
            c0 = bj * (comb_width + gap)
            for c in range(1, comb_width + 1):            # 背骨（1行）
                values[(r0 + 1, c0 + c)] = "spine"
            for i in range(n_fingers):
                col = c0 + 1 + i * 2
                for r in (2, 3, 4):
                    values[(r0 + r, col)] = f"finger{i}"
    rows = n_side * (comb_rows + gap)
    cols = n_side * (comb_width + gap)
    grid = _grid(rows, cols, values)

    started = time.monotonic()
    out = excel.regions(grid, cap_rows=rows, cap_cols=cols)
    elapsed = time.monotonic() - started

    assert elapsed < 5.0, f"シート全体の分割予算が働かず遅い（実測 {elapsed:.2f}s）"
    union: set[tuple[int, int]] = set()
    for rg in out:
        union |= rg.cells
    assert union == set(values.keys())                    # 完全性（silent-drop ゼロ）
    # 予算が無ければ100成分×最大51領域=5,100 Region（実測確認済み）。予算が働けば数百個規模で頭打ちになる。
    assert len(out) < 1000, f"シート全体の分割予算が働いていない可能性（Region数={len(out)}）"


# ---- 反復回数の予算内でもフォールバック断片数だけで領域数上限を超えうるケース ----

def test_two_row_comb_exceeds_region_budget_and_is_capped():
    """2行の櫛形（背骨1行＋指1行）で指が500本あると、反復回数上限（32回）は32個の断片しか消費できず、
    残り468本が個別断片としてフォールバックに回る（region数上限が無ければ501領域になる＝実測確認済み）。
    `_MAX_REGIONS_PER_SHEET` 適用後はその上限以内に収まり、かつ完全性（セルの総数）が保たれることを
    確認する。
    """
    n_fingers = 500
    width = n_fingers * 2 - 1
    values: dict[tuple[int, int], object] = {(1, c): "spine" for c in range(1, width + 1)}
    for i in range(n_fingers):
        values[(2, 1 + i * 2)] = f"finger{i}"
    grid = _grid(2, width, values)

    started = time.monotonic()
    out = excel.regions(grid, cap_rows=2, cap_cols=width)
    elapsed = time.monotonic() - started

    assert elapsed < 5.0, f"region数上限が働かず遅い（実測 {elapsed:.2f}s）"
    assert len(out) <= excel._MAX_REGIONS_PER_SHEET, f"region数上限が効いていない（実測 {len(out)}）"
    assert any(rg.split_budget_exhausted for rg in out)    # 予算切れフォールバックが実際に発生した証跡
    union: set[tuple[int, int]] = set()
    for rg in out:
        union |= rg.cells
    assert union == set(values.keys())                     # 完全性（silent-drop ゼロ）


# ---- 予算切れフォールバックの consumer 契約（source_map への伝播）----

def test_split_budget_exhausted_propagates_to_table_source_map(tmp_path):
    """予算切れフォールバックで生成された `table:N` だけに `split_budget_exhausted: true` が付き、
    ヒストグラム抽出で綺麗に切り出せた通常の `table:N` にはキー自体が無い（局所性・`truncated` と
    同じ「False の時はキーを書かない」規約）ことを確認する。"""
    n_fingers = 500
    width = n_fingers * 2 - 1
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    for c in range(1, width + 1):
        ws.cell(row=1, column=c, value="s")
    for i in range(n_fingers):
        ws.cell(row=2, column=1 + i * 2, value=f"f{i}")
    # 櫛形コンポーネントとは別の、離れた小さな2x2ブロック（H2で列上限を Excel 実仕様値
    # （16,384）へ引き上げた後も、櫛形全体が単一の集中成分になり「通常側」が消える事態を避ける
    # ため独立成分を明示的に用意する＝この2x2は必ずヒストグラム抽出1回で綺麗に切り出せる）。
    ws.cell(row=10, column=1, value="x")
    ws.cell(row=10, column=2, value="y")
    ws.cell(row=11, column=1, value="z")
    ws.cell(row=11, column=2, value="w")
    p = tmp_path / "a.xlsx"
    wb.save(p)

    doc = ooxml_arm._build_xlsx_ir(p)
    assert doc is not None
    tables = [e for e in doc.elements if e.type == "table"]
    assert len(tables) <= excel._MAX_REGIONS_PER_SHEET
    assert any(t.source_map.get("split_budget_exhausted") is True for t in tables)     # fallback 側
    assert any("split_budget_exhausted" not in t.source_map for t in tables)           # 通常側（負のassert）
    # キーがある場合は必ず True（False を明示的に書く経路が無いことの確認・truncated と同じ規約）。
    assert all(t.source_map.get("split_budget_exhausted", True) is True for t in tables)


# ---- インデックスパレットの白(1/9)・テーマ白の負tint(暗色化)境界 ----

def test_fill_boundary_indexed_white_and_negative_tint_theme():
    """標準インデックスパレットの白（1・9、実測でどちらも白に解決する重複エントリ）は非占有。
    テーマ背景1（`theme=0`）でも負の tint（Excel の tint モデルで暗色化を意味する）が付くと、もはや
    白ではなく暗いグレー相当の見た目になるため占有として扱う。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=1))    # 標準パレットの白(1)
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=9))    # 標準パレットの白(9・重複)
    ws["A3"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0, tint=-0.5))  # 暗色化した背景1
    got = excel.filled_cells(ws, cap_rows=5, cap_cols=5)
    assert got == {(3, 1)}                                  # A1/A2（白）は非占有、A3（暗色化）だけ占有


# ---- Region数の契約: len(regions(...)) <= max(連結成分数, _MAX_REGIONS_PER_SHEET) ----

def test_checkerboard_many_isolated_components_all_stay_unsplit():
    """市松模様で500個の孤立した1セル連結成分ができる場合、各成分は最低1件の Region を出す
    （silent-drop ゼロのため成分数を下回れない）契約により、`_MAX_REGIONS_PER_SHEET`（256）を
    超えて500 Region になる（バグではなく契約どおり）。全領域が単一セルのまま（未分割）であることと
    完全性（セルの総数）が保たれることを確認する。
    """
    cols = 500
    values: dict[tuple[int, int], object] = {}
    for c in range(1, cols + 1):
        values[(1, c)] = "a" if c % 2 == 1 else None
        if c % 2 == 0:
            values[(2, c)] = "b"
    values = {k: v for k, v in values.items() if v is not None}
    grid = _grid(2, cols, values)

    started = time.monotonic()
    out = excel.regions(grid, cap_rows=2, cap_cols=cols)
    elapsed = time.monotonic() - started

    assert elapsed < 5.0
    assert len(out) == 500                                 # 連結成分数（256を超える・契約どおり）
    assert all(len(rg.cells) == 1 for rg in out)            # すべて未分割（1セルのまま）
    union: set[tuple[int, int]] = set()
    for rg in out:
        union |= rg.cells
    assert union == set(values.keys())                      # 完全性（silent-drop ゼロ）


def test_few_components_total_stays_within_region_budget():
    """連結成分数（10個）が `_MAX_REGIONS_PER_SHEET`（256）より少ない場合、そのうち1つが大きく
    断片化しうる成分（500本指の櫛）であっても、合計 Region 数は256を超えない（1つの成分が予算を
    独占して他の保証枠まで食い潰さない）ことを確認する。
    """
    n_fingers = 500
    width = n_fingers * 2 - 1
    values: dict[tuple[int, int], object] = {(1, c): "spine" for c in range(1, width + 1)}
    for i in range(n_fingers):
        values[(2, 1 + i * 2)] = f"finger{i}"
    gap_col = width + 10
    for k in range(9):                                      # 残り9個の孤立した単純成分
        values[(4, gap_col + k * 3)] = f"iso{k}"
    rows, cols = 5, gap_col + 9 * 3 + 5
    grid = _grid(rows, cols, values)

    out = excel.regions(grid, cap_rows=rows, cap_cols=cols)

    assert len(out) <= excel._MAX_REGIONS_PER_SHEET
    union: set[tuple[int, int]] = set()
    for rg in out:
        union |= rg.cells
    assert union == set(values.keys())                      # 完全性（silent-drop ゼロ）


# ---- 白判定: RGB/indexed 色にも tint（負なら暗色化＝占有）を適用 ----

def test_negative_tint_on_rgb_and_indexed_white_bridges_without_splitting():
    """`_is_white_color` は白判定の対象（RGB直接色・インデックスパレット色）を問わず、負の tint
    （暗色化）なら白として扱わない。橋渡しセルとして使った場合、正しく占有されて分裂しないことを
    確認する（indexed=1 と rgb=FFFFFFFF の両方で）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "左"
    ws["C1"] = "右"
    ws["B1"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=1, tint=-0.5))
    got = excel.filled_cells(ws, cap_rows=5, cap_cols=5)
    assert got == {(1, 2)}                                  # 暗色化された白は占有（橋渡しに使える）

    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2["A1"] = "左"
    ws2["C1"] = "右"
    ws2["B1"].fill = PatternFill(fill_type="solid", fgColor=Color(rgb="FFFFFFFF", tint=-0.5))
    got2 = excel.filled_cells(ws2, cap_rows=5, cap_cols=5)
    assert got2 == {(1, 2)}

    for target_ws in (ws, ws2):
        out = excel.regions(
            _grid(1, 3, {(1, 1): "左", (1, 3): "右"}), cap_rows=1, cap_cols=3,
            filled=excel.filled_cells(target_ws, cap_rows=1, cap_cols=3))
        assert len(out) == 1                                # 分裂しない


# ---- ワークブック固有の palette/theme 解決（custom indexedColors・theme lt1） ----

def test_custom_indexed_and_theme_palette_resolved_as_occupied(tmp_path):
    """標準パレットでは白（1・9）のインデックス色、標準テーマでは白（背景1=lt1）のテーマ色でも、
    ワークブックがそれらをカスタム定義で赤へ上書きしていれば、実際の定義（`wb._colors`/
    `wb.loaded_theme`）を解決して占有として扱うことを確認する。
    """
    import re as _re
    from openpyxl.writer.theme import theme_xml

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=1))     # 標準なら白だがカスタムで赤
    ws["A2"].fill = PatternFill(fill_type="solid", fgColor=Color(indexed=9))     # 同上
    ws["A3"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0))       # 標準なら白だがカスタムで赤

    colors = list(wb._colors)
    colors[1] = "00FF0000"
    colors[9] = "00FF0000"
    wb._colors = colors
    custom_theme = _re.sub(r"<a:lt1>.*?</a:lt1>", '<a:lt1><a:srgbClr val="FF0000"/></a:lt1>',
                           theme_xml, flags=_re.S)
    wb.loaded_theme = custom_theme.encode("utf-8")

    p = tmp_path / "a.xlsx"
    wb.save(p)
    reloaded = openpyxl.load_workbook(p).active

    got = excel.filled_cells(reloaded, cap_rows=5, cap_cols=5)
    assert got == {(1, 1), (2, 1), (3, 1)}                   # 標準パレットの白判定を上書きして全部占有


# ---- 領域数予算の「全部か無か」判定: 部分的な分割が予算を超えないこと ----

def test_leading_l_shape_component_folds_to_single_bbox_within_budget():
    """256個の連結成分（先頭がL字型3セル、残り255個は孤立1セル）がある場合、
    `_MAX_REGIONS_PER_SHEET`（256）とちょうど成分数が一致するため各成分の予算は1件ずつになる。
    L字（非矩形・自然な分割には最低2領域必要）は「予算1件に対し自然な分割が2件必要」なので
    分割そのものを諦めて1つの外接矩形へ畳み、合計が257領域に膨らまないことを確認する
    （以前は「反復1回＋残りをまとめた1件」で合計2件になり、合計257・後続成分の予算が負になる
    バグがあった）。
    """
    values: dict[tuple[int, int], object] = {(1, 1): "a", (1, 2): "b", (2, 2): "c"}   # L字（非矩形）
    n_iso = 255
    for i in range(n_iso):
        values[(4, 10 + i * 2)] = f"iso{i}"                 # 列を2つおきに離し非連結を保証
    rows, cols = 4, 10 + n_iso * 2 + 2
    grid = _grid(rows, cols, values)

    out = excel.regions(grid, cap_rows=rows, cap_cols=cols)

    assert len(out) == 256                                  # 257 に膨らまない（連結成分数のまま）
    l_shape = [rg for rg in out if len(rg.cells) == 3]
    assert len(l_shape) == 1 and l_shape[0].split_budget_exhausted is True
    union: set[tuple[int, int]] = set()
    for rg in out:
        union |= rg.cells
    assert union == set(values.keys())                       # 完全性（silent-drop ゼロ）


# ---- テーマXML解析は xml.etree で正規に行う（壊れたXMLは解決不能＝占有側）----

def test_malformed_theme_xml_missing_closing_tag_is_occupied(tmp_path):
    """`loaded_theme` の閉じタグが欠落した壊れた XML は `ET.ParseError` になり、正規表現の部分一致では
    誤って白を受理しかねない箇所でも、xml.etree によるパースなら確実に「解決不能＝占有側」に倒れる
    ことを確認する。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0))
    p = tmp_path / "a.xlsx"
    wb.save(p)
    reloaded = openpyxl.load_workbook(p)
    good_theme = reloaded.loaded_theme.decode("utf-8")
    assert "<a:lt1>" in good_theme and "</a:lt1>" in good_theme

    broken_theme = good_theme.replace("</a:lt1>", "")        # 閉じタグを欠落させて XML を壊す
    reloaded.loaded_theme = broken_theme.encode("utf-8")
    with pytest.raises(Exception):
        excel.ET.fromstring(reloaded.loaded_theme)            # 壊れている前提の確認（ParseError系）

    got = excel.filled_cells(reloaded.active, cap_rows=5, cap_cols=5)
    assert got == {(1, 1)}                                    # 解決不能→占有側（白として受理しない）


# ---- テーマ色の解決はワークブックあたり1回だけキャッシュする ----

def test_theme_resolution_is_cached_per_workbook(tmp_path, monkeypatch):
    """同一ワークブック内に多数のテーマ色セルがあっても、`loaded_theme` の XML パース
    （`_resolve_theme_lt1_rgb`）はワークブックあたり1回しか呼ばれないことを確認する
    （`filled_cells()` の入口で構築する `ColorResolver` がキャッシュする）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    n_cells = 50
    for i in range(n_cells):
        ws.cell(row=1, column=i + 1).fill = PatternFill(fill_type="solid", fgColor=Color(theme=0))
    p = tmp_path / "a.xlsx"
    wb.save(p)
    reloaded = openpyxl.load_workbook(p).active

    calls = []
    original = excel._resolve_theme_lt1_rgb

    def counting(wb_arg):
        calls.append(1)
        return original(wb_arg)

    monkeypatch.setattr(excel, "_resolve_theme_lt1_rgb", counting)
    got = excel.filled_cells(reloaded, cap_rows=5, cap_cols=n_cells + 1)

    assert got == set()                                       # 標準テーマの白 = 非占有（前提の確認）
    assert len(calls) == 1                                    # 50セル分あっても解決は1回だけ


# ---- 白判定の値そのものを厳密検証する（末尾一致だけでは不正値を誤受理する回帰の再発防止） ----

def test_is_white_hex_rejects_malformed_values():
    """`_is_white_hex` は文字列全体の形式（6桁 or ARGB8桁の16進）を厳密に検証する。
    `str.endswith("FFFFFF")` だけの判定だと `"garbageFFFFFF"` のような不正な値まで白として
    誤受理してしまう（xml.etree 化で正規表現の6桁 HEX 完全一致検証が失われた回帰・実際に踏んだ）。
    """
    assert excel._is_white_hex("FFFFFF") is True
    assert excel._is_white_hex("ffffff") is True
    assert excel._is_white_hex("00FFFFFF") is True              # ARGB（先頭2桁がアルファ）
    assert excel._is_white_hex("garbageFFFFFF") is False        # 6桁超・16進以外の文字混入
    assert excel._is_white_hex("GGFFFFFF") is False             # 8桁だが先頭2桁が16進でない
    assert excel._is_white_hex("FFFF") is False                 # 短すぎる
    assert excel._is_white_hex(None) is False
    assert excel._is_white_hex(123456) is False


def test_theme_xml_garbage_value_is_not_white_and_is_occupied(tmp_path):
    """テーマXMLの `sysClr` の `lastClr` に `"garbageFFFFFF"` のような不正な値が入っていても、
    末尾一致だけで白と誤受理せず、解決不能＝占有側として扱われることを確認する。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0))
    p = tmp_path / "a.xlsx"
    wb.save(p)
    reloaded = openpyxl.load_workbook(p)
    good_theme = reloaded.loaded_theme.decode("utf-8")
    assert 'lastClr="FFFFFF"' in good_theme
    garbage_theme = good_theme.replace('lastClr="FFFFFF"', 'lastClr="garbageFFFFFF"')
    reloaded.loaded_theme = garbage_theme.encode("utf-8")

    got = excel.filled_cells(reloaded.active, cap_rows=5, cap_cols=5)
    assert got == {(1, 1)}                                       # 不正値は白と確認できない→占有側


# ---- ColorResolver はワークブックあたり1つに揃える（シートをまたいだ使い回し） ----

def test_color_resolver_reused_across_sheets_via_build_xlsx_ir(tmp_path, monkeypatch):
    """複数シートを持つワークブックを `_build_xlsx_ir` で処理する場合、`ColorResolver` は
    ワークブックあたり1つだけ構築されてシート間で使い回されるため、`loaded_theme` の XML パース
    （`_resolve_theme_lt1_rgb`）は2シート分あっても1回しか呼ばれないことを確認する
    （シートごとに新規構築すると2回呼ばれてしまう＝実際に踏んだ非効率）。
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "シート1"
    ws1["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0))
    ws2 = wb.create_sheet("シート2")
    ws2["A1"].fill = PatternFill(fill_type="solid", fgColor=Color(theme=0))
    p = tmp_path / "a.xlsx"
    wb.save(p)

    calls = []
    original = excel._resolve_theme_lt1_rgb

    def counting(wb_arg):
        calls.append(1)
        return original(wb_arg)

    monkeypatch.setattr(excel, "_resolve_theme_lt1_rgb", counting)
    doc = ooxml_arm._build_xlsx_ir(p)

    assert doc is not None
    assert len(calls) == 1                                       # 2シート分あっても解決は1回だけ


# ---- 巨大な結合範囲の座標展開はクリップしてから行う（cap 適用前の全展開を防ぐ）----
#
# `ws.merge_cells()`（openpyxl の通常 API）自体が結合範囲内の全セルを `MergedCell` へ変換する処理を
# 持つため、実際に "A1:XFD1048576" 級の結合を `merge_cells()` で作ると、テストしたい不具合の再現より
# 先にテスト自体が長時間かかってしまう（実測: A1:Z20000＝52万セルの結合作成だけで約6秒）。
# `openpyxl.worksheet.merge.MergedCellRange` を直接構築して `ws.merged_cells.ranges` へ加えることで、
# 実際のセル変換を経ずに「Excel の絶対上限まで宣言された結合」を安価に再現する。

def _inject_declared_merge(ws, range_str: str):
    """`ws.merge_cells()` を経由せず、宣言だけの結合範囲を安価に注入する（テスト専用ヘルパ）。"""
    from openpyxl.worksheet.merge import MergedCellRange
    mcr = MergedCellRange(ws, range_str)
    ws.merged_cells.ranges.add(mcr)
    return mcr


def test_clip_merge_enumeration_bounds_caps_area_not_just_each_axis():
    """`_clip_merge_enumeration_bounds` は各軸を cap_rows/cap_cols でクリップするだけでなく、
    面積（積）も `DEFAULT_CAP_CELLS` 以内へ追加でクリップする（片方の cap だけ大きい場合に
    積が安全弁を超えて残ることを防ぐ）。"""
    # cap_rows=100万・cap_cols=16384 のまま両軸クリップだけだと 100万×16384 ≈ 164億のまま残ってしまう。
    max_row, max_col = excel._clip_merge_enumeration_bounds(
        1, 1, 1_048_576, 16_384, cap_rows=1_000_000, cap_cols=16_384)
    area = max_row * max_col
    assert area <= excel.DEFAULT_CAP_CELLS
    assert max_col == 16_384                                     # 列側は cap_cols のまま


def test_merged_map_clips_huge_declared_range_bounded_time_and_size():
    """A1:XFD1048576（Excel絶対上限）まで宣言された結合でも、座標展開前に
    cap_rows×cap_cols でクリップされ、現実的な時間・件数で完走する
    （172億座標を辞書化しようとする回帰の再発防止）。宣言どおりの row_span/column_span は
    クリップされずそのまま保持する（`expand_regions_for_merges` 側の別のクランプに委ねるため）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    _inject_declared_merge(ws, "A1:XFD1048576")

    start = time.monotonic()
    merges = excel.merged_map(ws, cap_rows=100, cap_cols=100)
    elapsed = time.monotonic() - start

    assert elapsed < 2.0
    assert len(merges) == 100 * 100                                # cap を超えて展開していない
    assert merges[(1, 1)]["row_span"] == 1_048_576                 # 宣言どおりの値は保持
    assert merges[(1, 1)]["column_span"] == 16_384
    assert (101, 1) not in merges and (1, 101) not in merges       # cap の外は展開されない


def test_filled_cells_clips_huge_declared_range_bounded_time():
    """`filled_cells()` の `anchor_of` 事前構築も同じクリップを使う
    （`merged_map()` と同じ回帰の再発防止・`anchor_of` は塗り解決のための補助構造）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    _inject_declared_merge(ws, "A1:XFD1048576")

    start = time.monotonic()
    excel.filled_cells(ws, cap_rows=100, cap_cols=100)
    elapsed = time.monotonic() - start

    assert elapsed < 2.0


# ---- HM1: `_MAX_RECT_DECOMPOSE_CELLS` 引き上げ（JPX-021.xlsx の縮退解消）----

def test_jpx021_large_component_no_longer_degrades_to_single_bbox():
    """JPX-021.xlsx「統合設計」シートは旧上限（5,000）では2連結成分（A30:Z289＝面積6,760・
    A291:N654＝面積5,096）がどちらも `split_budget_exhausted=True` の単一外接矩形へ縮退していた
    （`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §8.4 の L1 実測で唯一
    excel2md に負けていた箇所）。上限を20,000へ引き上げた現在、この2成分はどちらも複数矩形へ
    正しく分割され、`split_budget_exhausted` が消えることを実データで固定する。
    """
    path = _EVAL_XLSX_DIR / "JPX-021.xlsx"
    if not path.is_file():
        pytest.skip(f"fixture が無い環境: {path}")
    started = time.monotonic()
    doc = ooxml_arm._build_xlsx_ir(path)
    elapsed = time.monotonic() - started
    assert doc is not None
    assert elapsed < 5.0, f"分割変更後も高速に完了すること（実測 {elapsed:.2f}s）"

    sheet = next(e for e in doc.elements if e.type == "sheet" and e.source_map.get("sheet") == "統合設計")
    tables = [e for e in doc.elements if e.type == "table" and e.parent_id == sheet.element_id]
    exhausted = [t for t in tables if t.source_map.get("split_budget_exhausted")]
    assert exhausted == [], f"縮退が残っている: {[t.source_map.get('range') for t in exhausted]}"
    assert len(tables) > 27, "旧上限（5,000）では27 table だった（縮退2件込み）——分割後は増えるはず"


def test_max_rect_decompose_cells_raised_above_old_value():
    """上限値そのもの（5,000→20,000）を固定する回帰（値を戻す変更を検知する）。"""
    assert excel._MAX_RECT_DECOMPOSE_CELLS >= 20_000


# ---- HM1: 画像の存在（枚数）検出（`picture_counts_by_sheet`・人間向けMD注記用）----

def test_picture_counts_by_sheet_real_fixture():
    """DEP-XLSX-MARKERS.xlsx は「対象」シートに画像1枚のみ（「旧版」「内部退避」には無し）。
    0枚のシートはキー自体を持たない契約。"""
    path = (pathlib.Path(__file__).resolve().parents[2]
            / "fixtures" / "eval" / "deprecation_markers" / "inputs" / "DEP-XLSX-MARKERS.xlsx")
    if not path.is_file():
        pytest.skip(f"fixture が無い環境: {path}")
    with zipfile.ZipFile(path) as z:
        counts = excel.picture_counts_by_sheet(z)
    assert counts == {"対象": 1}


def test_picture_counts_by_sheet_no_drawing_part_returns_empty(tmp_path):
    """drawing を持たない通常の xlsx は空 dict（0枚のシートはキーを持たない契約の裏返し）。"""
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    p = tmp_path / "no_image.xlsx"
    wb.save(p)
    with zipfile.ZipFile(p) as z:
        assert excel.picture_counts_by_sheet(z) == {}


def test_picture_counts_by_sheet_counts_pictures_inside_group(tmp_path):
    """`xdr:grpSp`（グループ化図形）の中の画像も数える（トップレベル anchor だけでなく
    グループの子孫まで `.iter()` で拾えることの回帰）。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    p = tmp_path / "grouped.xlsx"
    wb.save(p)

    # openpyxl はグループ化図形の書き込みAPIを持たないため、drawing part を直接注入する。
    import shutil
    drawing_xml = (
        '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing" '
        'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">'
        '<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff>'
        '<xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>'
        '<xdr:ext cx="1" cy="1"/>'
        '<xdr:grpSp><xdr:nvGrpSpPr><xdr:cNvPr id="1" name="Group 1"/><xdr:cNvGrpSpPr/></xdr:nvGrpSpPr>'
        '<xdr:grpSpPr/>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="Image 1"/><xdr:cNvPicPr/></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip r:embed="rId1"/></xdr:blipFill><xdr:spPr/></xdr:pic>'
        '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="3" name="Image 2"/><xdr:cNvPicPr/></xdr:nvPicPr>'
        '<xdr:blipFill><a:blip r:embed="rId1"/></xdr:blipFill><xdr:spPr/></xdr:pic>'
        '</xdr:grpSp><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>')
    drawing_rels_xml = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        'relationships/image" Target="../media/image1.png"/></Relationships>')
    sheet_rels_xml = (
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/'
        'relationships/drawing" Target="../drawings/drawing1.xml"/></Relationships>')

    unpacked = tmp_path / "unpacked"
    with zipfile.ZipFile(p) as z:
        z.extractall(unpacked)
    (unpacked / "xl" / "drawings").mkdir(parents=True, exist_ok=True)
    (unpacked / "xl" / "drawings" / "drawing1.xml").write_text(drawing_xml, encoding="utf-8")
    (unpacked / "xl" / "drawings" / "_rels").mkdir(parents=True, exist_ok=True)
    (unpacked / "xl" / "drawings" / "_rels" / "drawing1.xml.rels").write_text(drawing_rels_xml, encoding="utf-8")
    (unpacked / "xl" / "media").mkdir(parents=True, exist_ok=True)
    (unpacked / "xl" / "media" / "image1.png").write_bytes(b"\x89PNG\r\n\x1a\n")
    (unpacked / "xl" / "worksheets" / "_rels").mkdir(parents=True, exist_ok=True)
    (unpacked / "xl" / "worksheets" / "_rels" / "sheet1.xml.rels").write_text(sheet_rels_xml, encoding="utf-8")
    sheet1 = unpacked / "xl" / "worksheets" / "sheet1.xml"
    sheet1.write_text(
        sheet1.read_text(encoding="utf-8").replace(
            "</worksheet>",
            '<drawing xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" '
            'r:id="rId1"/></worksheet>'),
        encoding="utf-8")

    repacked = tmp_path / "grouped_with_image.xlsx"
    with zipfile.ZipFile(repacked, "w", zipfile.ZIP_DEFLATED) as zw:
        for f in unpacked.rglob("*"):
            if f.is_file():
                zw.write(f, f.relative_to(unpacked))
    shutil.rmtree(unpacked)

    with zipfile.ZipFile(repacked) as z:
        counts = excel.picture_counts_by_sheet(z)
    assert counts == {"Sheet": 2}
