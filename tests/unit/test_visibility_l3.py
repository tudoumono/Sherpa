"""可視性・廃止表現の全形式展開（L3・docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md §2）の単体テスト。

対象: `evidence_spike.py`（xlsx の図形/画像による覆い判定・pptx の既存判定の Evidence IR への移送）、
`ooxml/excel.py::strike_cells`／`ooxml/word.py::strike_runs`（取り消し線・D-1）、
`arms/ooxml_arm.py`（`strike:N` 要素の採番・`_docx_floating_anchor_facts` の浮動図形と段落の関連付け・D-2）。

意味の断定（「廃止」等）はしない契約（可視性表現の共通思想）のため、本テストも
`visibility`/`extension["visibility_reason"]`/`extension["occluded_by"]`/`extension["covered_by_text"]`/
`extension["floating_anchors"]`（幾何的事実・前面要素の生テキスト）だけを検証し、自然文の断定は検証しない。
"""
from __future__ import annotations

import pathlib
import re
import tempfile
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
import pytest
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, TwoCellAnchor
from openpyxl.styles import Font

from sherpa.ingest import document_ir, evidence_ir, evidence_render, evidence_spike, office_md
from sherpa.ingest.arms import ooxml_arm
from sherpa.ingest.ooxml import excel, word

_ROOT = pathlib.Path(__file__).resolve().parents[2]
_STAMP_ASSET = _ROOT / "fixtures/eval/excel_ja/assets/廃止スタンプ.png"
_EXCEL_JA_INPUTS = _ROOT / "fixtures/eval/excel_ja/inputs"

_XDR_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"


def _zip_entries(path: pathlib.Path) -> dict[str, bytes]:
    with zipfile.ZipFile(path) as z:
        return {name: z.read(name) for name in z.namelist()}


def _write_zip(path: pathlib.Path, entries: dict[str, bytes]) -> None:
    with zipfile.ZipFile(path, "w") as z:
        for name, data in entries.items():
            z.writestr(name, data)


def _xlsx_with_patched_drawing(tmp_path: pathlib.Path, cells: dict[str, object], drawing_xml: str) -> pathlib.Path:
    """openpyxlでセル値だけのxlsxを保存し、drawing部分だけ手組みXMLへ差し替える。

    xdr:sp（図形）はopenpyxlのAPIでは組み立てられないため、`ws.add_image`でContent_Types/rels一式を
    openpyxlに配線させてから、そのdrawing1.xmlの中身だけ差し替える（画像用relationshipは未参照のまま
    残るが、evidence_spike側はdrawing xmlのノードが実際に参照するrelationshipしか辿らないため実害無い）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for coord, value in cells.items():
        ws[coord] = value
    img = XLImage(str(_STAMP_ASSET))
    img.anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=0, colOff=0, row=0, rowOff=0),
        to=AnchorMarker(col=1, colOff=0, row=1, rowOff=0),
    )
    ws.add_image(img)
    p = tmp_path / "shape.xlsx"
    wb.save(p)
    entries = _zip_entries(p)
    drawing_name = next(n for n in entries if re.fullmatch(r"xl/drawings/drawing\d+\.xml", n))
    entries[drawing_name] = drawing_xml.encode("utf-8")
    _write_zip(p, entries)
    return p


def _sp_drawing_xml(*, solid_fill: bool) -> str:
    fill = (
        '<a:solidFill><a:srgbClr val="FF0000"/></a:solidFill>' if solid_fill else "<a:noFill/>"
    )
    return f"""<?xml version="1.0"?>
<xdr:wsDr xmlns:xdr="{_XDR_NS}" xmlns:a="{_A_NS}">
 <xdr:twoCellAnchor editAs="oneCell">
  <xdr:from><xdr:col>0</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>0</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:from>
  <xdr:to><xdr:col>1</xdr:col><xdr:colOff>0</xdr:colOff><xdr:row>1</xdr:row><xdr:rowOff>0</xdr:rowOff></xdr:to>
  <xdr:sp>
   <xdr:nvSpPr><xdr:cNvPr id="1" name="覆い図形"/><xdr:cNvSpPr/></xdr:nvSpPr>
   <xdr:spPr>{fill}</xdr:spPr>
  </xdr:sp>
  <xdr:clientData/>
 </xdr:twoCellAnchor>
</xdr:wsDr>"""


# ---- xlsx: 図形/画像による覆い（occlusion） ----

def test_xlsx_picture_occludes_covered_cells_with_real_stamp_asset(tmp_path):
    """`廃止スタンプ.png`（実アセット）がセル範囲を覆うと、そのセルが hidden になり前面画像情報が残る。"""
    assert _STAMP_ASSET.exists(), "廃止スタンプ.png fixture asset is missing"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws["A1"] = "見出し"
    ws["A2"] = "廃止された値"
    ws["B2"] = "廃止された値2"
    ws["C2"] = "覆われない値"
    img = XLImage(str(_STAMP_ASSET))
    img.anchor = TwoCellAnchor(
        editAs="oneCell",
        _from=AnchorMarker(col=0, colOff=0, row=1, rowOff=0),
        to=AnchorMarker(col=2, colOff=0, row=2, rowOff=0),
    )
    ws.add_image(img)
    p = tmp_path / "stamp.xlsx"
    wb.save(p)

    ir = evidence_spike.extract(p)
    assert evidence_ir.validation_errors(ir) == []
    cells_by_range = {el.locator.cell_range: el for el in ir.elements if el.type == "cell"}

    covered_a2 = cells_by_range["A2"]
    assert covered_a2.visibility == "hidden"
    assert covered_a2.value == "廃止された値"                          # 値は変えない
    assert covered_a2.extension["visibility_reason"] == "occluded_by_picture"
    occluded_by = covered_a2.extension["occluded_by"]
    assert occluded_by["kind"] == "picture"

    covered_b2 = cells_by_range["B2"]
    assert covered_b2.visibility == "hidden"

    uncovered_c2 = cells_by_range["C2"]
    assert uncovered_c2.visibility == "visible"

    uncovered_a1 = cells_by_range["A1"]
    assert uncovered_a1.visibility == "visible"


def test_xlsx_solid_fill_shape_occludes_cell(tmp_path):
    """塗りつぶし図形（`a:solidFill`）もセルを覆う（画像に限らない）。"""
    p = _xlsx_with_patched_drawing(tmp_path, {"A1": "廃止対象"}, _sp_drawing_xml(solid_fill=True))
    ir = evidence_spike.extract(p)
    assert evidence_ir.validation_errors(ir) == []
    cell = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A1")
    assert cell.visibility == "hidden"
    assert cell.extension["visibility_reason"] == "occluded_by_shape"
    assert cell.extension["occluded_by"]["kind"] == "shape"


def test_xlsx_nofill_shape_does_not_occlude_cell(tmp_path):
    """`a:noFill` の図形は覆い扱いにしない（pptx の A5 と同じ判定基準）。"""
    p = _xlsx_with_patched_drawing(tmp_path, {"A1": "そのまま見える"}, _sp_drawing_xml(solid_fill=False))
    ir = evidence_spike.extract(p)
    cell = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A1")
    assert cell.visibility == "visible"
    assert "visibility_reason" not in cell.extension


def test_xlsx_cell_outside_drawing_range_stays_visible(tmp_path):
    """図形の cell_range に含まれないセルは覆い判定の対象外のまま。"""
    p = _xlsx_with_patched_drawing(
        tmp_path, {"A1": "覆われる", "B1": "覆われない"}, _sp_drawing_xml(solid_fill=True))
    ir = evidence_spike.extract(p)
    a1 = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A1")
    b1 = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "B1")
    assert a1.visibility == "hidden"
    assert b1.visibility == "visible"


# ---- xlsx: 取り消し線 ----

def test_excel_strike_cells_direct():
    """`excel.strike_cells` の純関数契約: strikeフォント かつ 値ありのセルだけを返す。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "通常"
    ws["A2"] = "取り消し線あり"
    ws["A2"].font = Font(strike=True)
    ws["A3"].font = Font(strike=True)                 # 値なし=出さない
    out = excel.strike_cells(ws)
    assert out == [{"cell": "A2", "row": 2, "column": 1, "text": "取り消し線あり"}]


def test_xlsx_strike_cells_become_strike_text_elements(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws["A1"] = "通常値"
    ws["A2"] = "廃止された項目"
    ws["A2"].font = Font(strike=True)
    p = tmp_path / "strike.xlsx"
    wb.save(p)

    document = ooxml_arm._build_xlsx_ir(p)
    strikes = [e for e in document.elements if e.type == "strike_text"]
    assert len(strikes) == 1
    assert strikes[0].text == "廃止された項目"
    assert strikes[0].visibility == "visible"          # 取り消し線は可視性を変えない幾何的事実
    assert strikes[0].status == "active"
    assert strikes[0].source_map == {"sheet": "一覧", "cell": "A2"}

    ir = evidence_spike.extract(p, legacy_ir=document)
    assert evidence_ir.validation_errors(ir) == []
    ev_strike = next(el for el in ir.elements if el.type == "strike_text")
    assert ev_strike.value == "廃止された項目"
    assert ev_strike.visibility == "visible"
    # 独立要素だけでなく、対応する cell 要素自身にも構造的に反映される（`_field_piece` は cell を読む）。
    ev_cell = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A2")
    assert ev_cell.visibility == "visible"              # 取り消し線は本文が読める状態のまま
    assert ev_cell.extension["visibility_reason"] == "strike"


def test_xlsx_hidden_row_marks_cell_hidden(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws["A1"] = "通常行"
    ws["A2"] = "非表示行の値"
    ws.row_dimensions[2].hidden = True
    p = tmp_path / "hidden_row.xlsx"
    wb.save(p)

    ir = evidence_spike.extract(p)
    assert evidence_ir.validation_errors(ir) == []
    a1 = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A1")
    a2 = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A2")
    assert a1.visibility == "visible"
    assert a2.visibility == "hidden"
    assert a2.extension["visibility_reason"] == "hidden_row"


def test_xlsx_hidden_column_marks_cell_hidden(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "一覧"
    ws["A1"] = "通常列"
    ws["B1"] = "非表示列の値"
    ws.column_dimensions["B"].hidden = True
    p = tmp_path / "hidden_col.xlsx"
    wb.save(p)

    ir = evidence_spike.extract(p)
    a1 = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "A1")
    b1 = next(el for el in ir.elements if el.type == "cell" and el.locator.cell_range == "B1")
    assert a1.visibility == "visible"
    assert b1.visibility == "hidden"
    assert b1.extension["visibility_reason"] == "hidden_column"


# ---- docx: 取り消し線（D-1） ----

_DOCX_NS = 'xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main"'


def _docx_zip(tmp_path: pathlib.Path, body_xml: str) -> pathlib.Path:
    doc_xml = f"""<?xml version="1.0"?>
<w:document {_DOCX_NS}>
 <w:body>{body_xml}</w:body>
</w:document>"""
    p = tmp_path / "a.docx"
    _write_zip(p, {"word/document.xml": doc_xml.encode("utf-8")})
    return p


def test_word_strike_runs_direct():
    root = ET.fromstring(f"""<w:p {_DOCX_NS}>
 <w:r><w:t>通常</w:t></w:r>
 <w:r><w:rPr><w:strike/></w:rPr><w:t>単取消線</w:t></w:r>
 <w:r><w:rPr><w:dstrike/></w:rPr><w:t>二重取消線</w:t></w:r>
 <w:r><w:rPr><w:strike w:val="0"/></w:rPr><w:t>打消し済み</w:t></w:r>
</w:p>""")
    assert word.strike_runs(root) == ["単取消線", "二重取消線"]


def test_docx_strike_runs_flow_to_evidence_ir_and_stay_visible(tmp_path):
    body = f"""<w:p>
 <w:r><w:t>通常本文</w:t></w:r>
 <w:r><w:rPr><w:strike/></w:rPr><w:t>取り消し線本文</w:t></w:r>
 <w:r><w:rPr><w:vanish/></w:rPr><w:t>隠し文字本文</w:t></w:r>
 <w:del w:id="1"><w:r><w:delText>削除本文</w:delText></w:r></w:del>
</w:p>"""
    p = _docx_zip(tmp_path, body)

    document = ooxml_arm._build_docx_ir(p)
    strike_els = [e for e in document.elements if e.type == "strike_text"]
    assert len(strike_els) == 1
    assert strike_els[0].text == "取り消し線本文"
    assert strike_els[0].visibility == "visible"
    assert strike_els[0].visibility_reason == "strike"
    assert strike_els[0].status == "active"

    ir = evidence_spike.extract(p, legacy_ir=document)
    assert evidence_ir.validation_errors(ir) == []
    by_type = {el.type: el for el in ir.elements if el.type in {"strike_text", "hidden_text", "deleted_text"}}
    assert by_type["strike_text"].value == "取り消し線本文"
    assert by_type["strike_text"].visibility == "visible"
    assert by_type["strike_text"].extension["visibility_reason"] == "strike"
    # 既存の隠し文字/削除本文（D-1で新規追加した取り消し線と同じ経路）が壊れていないことの回帰確認。
    assert by_type["hidden_text"].visibility == "hidden"
    assert by_type["deleted_text"].lifecycle == "deleted"


def test_docx_strike_inside_table_cell(tmp_path):
    """表セル内の取り消し線も取りこぼさない（`_append_paragraph_extras` の cell_map 経路・RV High #1 と同じ設計）。"""
    body = f"""<w:tbl>
 <w:tr><w:tc><w:p><w:r><w:rPr><w:strike/></w:rPr><w:t>セル内取消線</w:t></w:r></w:p></w:tc></w:tr>
</w:tbl>"""
    p = _docx_zip(tmp_path, body)
    document = ooxml_arm._build_docx_ir(p)
    strike_els = [e for e in document.elements if e.type == "strike_text"]
    assert len(strike_els) == 1
    assert strike_els[0].text == "セル内取消線"
    assert strike_els[0].source_map.get("table_index") is not None


# ---- pptx: 既存判定（occluded/covered_by_text）の Evidence IR への移送 ----

_PPTX_NS = (
    'xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" '
    'xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main"'
)


def _pptx_zip(tmp_path: pathlib.Path, shapes_xml: str, name: str = "occ.pptx") -> pathlib.Path:
    slide_xml = f"""<?xml version="1.0"?>
<p:sld {_PPTX_NS}>
 <p:cSld><p:spTree>
  {shapes_xml}
 </p:spTree></p:cSld>
</p:sld>"""
    p = tmp_path / name
    _write_zip(p, {"ppt/slides/slide1.xml": slide_xml.encode("utf-8")})
    return p


def _sp_text(text: str, x=0, y=0, cx=4000000, cy=2000000) -> str:
    return f"""<p:sp>
  <p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm></p:spPr>
  <p:txBody><a:p><a:r><a:t>{text}</a:t></a:r></a:p></p:txBody>
 </p:sp>"""


def _pic(x=0, y=0, cx=4000000, cy=2000000) -> str:
    return f"""<p:pic><p:spPr><a:xfrm><a:off x="{x}" y="{y}"/><a:ext cx="{cx}" cy="{cy}"/></a:xfrm></p:spPr></p:pic>"""


def test_pptx_occlusion_reaches_both_evidence_representations(tmp_path):
    """`_build_pptx_ir` が確定した occluded 判定が、adapter経由・raw walk経由の両方のEvidence要素に
    同じ理由で伝わる（`adapted_by_slot` 経由の1回だけの計算・二重実装ではない証拠）。
    """
    p = _pptx_zip(tmp_path, _sp_text("画像に隠れる") + _pic())
    ir = evidence_spike.extract(p)
    assert evidence_ir.validation_errors(ir) == []
    shapes = [el for el in ir.elements if el.type == "shape" and el.value == "画像に隠れる"]
    assert len(shapes) == 2                             # adapter由来1件 + raw walk由来1件
    for shape in shapes:
        assert shape.visibility == "hidden"
        assert shape.extension.get("visibility_reason") == "occluded"


def test_pptx_covered_by_text_reaches_both_representations_with_front_text(tmp_path):
    """前面テキストによる重ね（取り消し線的表現）は visibility を変えないが、前面要素の生テキストへ
    `extension["covered_by_text"]` 経由で辿れる（意味の断定はしない・前面テキストをそのまま持つだけ）。
    """
    p = _pptx_zip(tmp_path, _sp_text("下敷きテキスト") + _sp_text("廃止"), name="cov.pptx")
    ir = evidence_spike.extract(p)
    assert evidence_ir.validation_errors(ir) == []
    covered_shapes = [el for el in ir.elements if el.type == "shape" and el.value == "下敷きテキスト"]
    assert len(covered_shapes) == 2
    for shape in covered_shapes:
        assert shape.visibility == "visible"            # covered_by_textはhiddenにしない（意味の断定はしない）
        front = shape.extension.get("covered_by_text")
        assert front is not None
        assert front["text"] == "廃止"


def test_pptx_occlusion_shares_single_ratio_constant_not_reimplemented(tmp_path, monkeypatch):
    """`office_md._OCCLUSION_RATIO` を差し替えると Evidence IR 側の occluded 判定も追随する
    （同じ定数・同じ計算を1箇所だけで行っている証拠。閾値だけ evidence_spike 側に別定義があれば
    このテストは追随せず失敗する）。
    """
    # 50%重なり（デフォルト閾値0.9では非occluded）。
    p = _pptx_zip(
        tmp_path,
        _sp_text("半分隠れる") + _pic(x=2000000),
        name="partial.pptx",
    )
    ir_default = evidence_spike.extract(p)
    shape_default = next(el for el in ir_default.elements if el.type == "shape" and el.value == "半分隠れる"
                          and el.extension.get("origin") == "document-ir-v2-adapter")
    assert shape_default.visibility == "visible"

    monkeypatch.setattr(office_md, "_OCCLUSION_RATIO", 0.3)
    ir_lowered = evidence_spike.extract(p)
    shape_lowered = next(el for el in ir_lowered.elements if el.type == "shape" and el.value == "半分隠れる"
                          and el.extension.get("origin") == "document-ir-v2-adapter")
    assert shape_lowered.visibility == "hidden"
    assert shape_lowered.extension.get("visibility_reason") == "occluded"


def test_xlsx_occlusion_shares_single_ratio_constant_not_reimplemented(tmp_path, monkeypatch):
    """xlsx の覆い判定も `office_md._OCCLUSION_RATIO` 1箇所を読む（pptx 側の同名テストの xlsx 版）。

    しきい値を evidence_spike 側に複製すると、片方だけ変えたときに pptx と xlsx で判定が黙って
    食い違う。差し替えに追随することで「定義は1箇所」を固定する。
    """
    assert evidence_spike._occlusion_ratio() == office_md._OCCLUSION_RATIO
    monkeypatch.setattr(office_md, "_OCCLUSION_RATIO", 0.3)
    assert evidence_spike._occlusion_ratio() == 0.3


# ---- xlsx: グループ（`xdr:grpSp`）内図形の個別要素化 ----
#
# 実測（2026-09-02時点）: `_xlsx_objects` はグループ直下しか見ておらず、JPX-014/020/099 では
# フロー図の各ノード（`入力ノード`/`処理ノード`/`処理コネクタ`等）が1個の `type=group` 要素へ
# 連結された1本の文字列として潰れ、ノード名で検索できなかった（pptx の `walk()` は再帰していた
# ため非対称だった）。

def _real_fixture_elements(name: str):
    p = _EXCEL_JA_INPUTS / f"{name}.xlsx"
    if not p.is_file():
        pytest.skip(f"fixture が無い環境: {p}")
    return evidence_spike.extract(p)


def test_xlsx_group_children_become_individual_elements_real_fixture():
    """JPX-014（実データ）: グループ内の`入力ノード`/`処理ノード`が個別要素として出る
    （1個の連結文字列 `顧客ID入力顧客登録処理` に潰れない）。"""
    ir = _real_fixture_elements("JPX-014")
    assert evidence_ir.validation_errors(ir) == []
    group = next(el for el in ir.elements if el.type == "group")
    children = [el for el in ir.elements if el.parent_id == group.element_id]
    child_values = {el.value for el in children}
    assert "顧客ID入力" in child_values
    assert "顧客登録処理" in child_values
    # 束ねられた1本の文字列としては存在しない。
    assert not any(el.value == "顧客ID入力顧客登録処理" for el in children)


def test_xlsx_group_connector_endpoints_are_scoped_to_connector_real_fixture():
    """JPX-014（実データ）: `start_object_id`/`end_object_id` はグループ内のコネクタ要素にだけ付き、
    グループ自身の属性にはならない（従来はグループが子孫コネクタの接続先を誤って持っていた）。"""
    ir = _real_fixture_elements("JPX-014")
    group = next(el for el in ir.elements if el.type == "group")
    assert group.extension.get("start_object_id") is None
    assert group.extension.get("end_object_id") is None
    connector = next(el for el in ir.elements if el.type == "connector" and el.parent_id == group.element_id)
    assert connector.extension.get("start_object_id") is not None
    assert connector.extension.get("end_object_id") is not None


@pytest.mark.parametrize("name", ["JPX-014", "JPX-020", "JPX-099"])
def test_xlsx_group_children_render_without_coverage_gap(name):
    """グループ内図形が増えても`evidence_render.validation_errors()`が実xlsxの往復で空のまま
    （citation完全性: 増えた要素がどこかのchunkへcitationされないと `element_coverage_missing` で
    文書ごとRAG表現を作れなくなる契約・pptxで過去に実際に起きた事故と同じ形）。"""
    p = _EXCEL_JA_INPUTS / f"{name}.xlsx"
    if not p.is_file():
        pytest.skip(f"fixture が無い環境: {p}")
    ir = evidence_spike.extract(p)
    result = evidence_render.render(ir, source_name=str(p))
    assert evidence_render.validation_errors(ir, result) == []


def test_xlsx_top_level_connector_still_extracted_real_fixture():
    """JPX-021（実データ）: トップレベル（グループ外）のコネクタは従来どおり取れる
    （グループ再帰を入れてもトップレベル図形の抽出が壊れない）。"""
    ir = _real_fixture_elements("JPX-021")
    connector = next(
        el for el in ir.elements
        if el.type == "connector" and el.locator.object_id == 32
    )
    assert connector.extension.get("start_object_id") == 30
    assert connector.extension.get("end_object_id") == 31


def test_group_children_do_not_claim_geometry_of_parent_anchor(tmp_path):
    """グループ子孫は幾何関係（覆い判定・overlaps）を主張しない（検収是正）。

    子の cell_range は親 anchor の継承（近似）なので、その矩形で覆いを判定すると
    「グループ内の小さな塗りつぶし図形がグループ全域を覆う」という事実でない幾何になる
    （フロー図の下のセルが誤って hidden になる）。子要素は検索向けに個別に出すが、
    幾何関係はトップレベル（anchor 矩形が自分自身のもの）に限る。
    """
    from pathlib import Path as _P
    src = _P("fixtures/eval/excel_ja/inputs/JPX-014.xlsx")
    ir = evidence_spike.extract(src)
    groups = {el.element_id for el in ir.elements if el.type == "group"}
    children = {el.element_id for el in ir.elements if el.parent_id in groups}
    assert children                                   # グループ子孫が個別要素として出ている（Z の成果・回帰）
    # 子孫を occluder とする hidden セルが無い
    for el in ir.elements:
        if el.type == "cell" and el.visibility == "hidden":
            occ = el.extension.get("occluded_by") or {}
            assert occ.get("element_id") not in children
    # 子孫を source とする overlaps 関係も無い
    for rel in ir.relations:
        if rel.type == "overlaps":
            assert rel.source_id not in children


# ---- docx: 浮動図形と段落の関連付け（D-2） ----

_WP_DRAWING_NS = 'xmlns:wp="http://schemas.openxmlformats.org/drawingml/2006/wordprocessingDrawing"'


def _docx_floating_anchor_body(*, behind_doc: str, shape_name: str = "廃止スタンプ", shape_text: str = "廃止") -> str:
    return f"""<w:p>
 <w:r><w:t>対象機能の説明本文</w:t></w:r>
 <w:r><w:drawing {_WP_DRAWING_NS}>
  <wp:anchor behindDoc="{behind_doc}" relativeHeight="1">
   <wp:docPr id="1" name="{shape_name}"/>
   <wp:extent cx="100" cy="100"/>
   <w:t>{shape_text}</w:t>
  </wp:anchor>
 </w:drawing></w:r>
</w:p>"""


def test_docx_floating_anchor_paragraph_association(tmp_path):
    """浮動図形の親段落を辿った関連付けの事実（`floating_anchors`）が document-ir・Evidence IR の両方に
    載り、幾何的な覆い断定（visibility の変更）はしない（D-2）。Word はフロー配置で段落の実座標が
    レイアウト計算前に確定しないため、xlsx/pptx と異なり幾何交差判定はできない、という制約がある。
    """
    body = _docx_floating_anchor_body(behind_doc="0")
    p = _docx_zip(tmp_path, body)

    document = ooxml_arm._build_docx_ir(p)
    para = next(e for e in document.elements if e.type == "paragraph")
    assert para.source_map["floating_anchors"] == [
        {"name": "廃止スタンプ", "text": "廃止", "behind_doc": False},
    ]
    assert para.visibility == "visible"                 # 幾何断定はしない

    ir = evidence_spike.extract(p, legacy_ir=document)
    assert evidence_ir.validation_errors(ir) == []
    para_ev = next(e for e in ir.elements if e.type == "paragraph")
    assert para_ev.extension["floating_anchors"] == [
        {"name": "廃止スタンプ", "text": "廃止", "behind_doc": False},
    ]
    assert para_ev.visibility == "visible"               # 幾何断定はしない（本テストの中心的な固定点）

    # 浮動図形自身も幾何断定でhiddenにされていない
    floating = next(el for el in ir.elements if el.type in {"floating_object", "floating_textbox"})
    assert floating.visibility != "hidden"

    # 既存のEMU座標overlaps（`_docx_host_objects`の浮動図形どうしの幾何関係）とは別物（キー名で区別）。
    # paragraphはbboxを持たないためoverlaps対象にもならない。
    assert not any(
        rel.type == "overlaps" and para_ev.element_id in (rel.source_id, rel.target_id)
        for rel in ir.relations
    )

    result = evidence_render.render(ir, source_name=str(p))
    assert evidence_render.validation_errors(ir, result) == []


def test_docx_floating_anchor_behind_doc_distinguished(tmp_path):
    """`behindDoc`（背面配置）とそれ以外（前面）の別が XML から区別して残る（D-2）。"""
    body = _docx_floating_anchor_body(behind_doc="1", shape_name="背面画像", shape_text="")
    p = _docx_zip(tmp_path, body)
    document = ooxml_arm._build_docx_ir(p)
    para = next(e for e in document.elements if e.type == "paragraph")
    assert para.source_map["floating_anchors"] == [{"name": "背面画像", "behind_doc": True}]


def test_docx_floating_anchor_absent_when_no_drawing(tmp_path):
    """浮動図形が無い段落には `floating_anchors` キー自体が付かない（存在しない関連付けでノイズを
    増やさない）。"""
    body = "<w:p><w:r><w:t>通常の段落</w:t></w:r></w:p>"
    p = _docx_zip(tmp_path, body)
    document = ooxml_arm._build_docx_ir(p)
    para = next(e for e in document.elements if e.type == "paragraph")
    assert "floating_anchors" not in para.source_map
