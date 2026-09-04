"""Office→決定的MD 変換（office_md）の単体テスト（OOXML 直パース・DB不要）。

docx/pptx は最小 OOXML を zip で組んで検証（外部ライブラリ不要の経路）。xlsx は openpyxl で実ファイルを作る。
PDF/旧形式/壊れファイルは None（未対応）であることも確認。
"""
from __future__ import annotations

import pathlib
import shutil
import tempfile
import zipfile

from pypdf import PdfWriter

from sherpa import json_io
from sherpa.ingest import office_md

_DOCX_XML = """<?xml version="1.0"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
 <w:body>
  <w:p><w:pPr><w:pStyle w:val="Heading1"/></w:pPr><w:r><w:t>タイトル見出し</w:t></w:r></w:p>
  <w:p><w:r><w:t>本文テキストABC</w:t></w:r></w:p>
  <w:tbl><w:tr><w:tc><w:p><w:r><w:t>セル1</w:t></w:r></w:p></w:tc>
   <w:tc><w:p><w:r><w:t>セル2</w:t></w:r></w:p></w:tc></w:tr></w:tbl>
 </w:body>
</w:document>"""

_PPTX_SLIDE = """<?xml version="1.0"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
       xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
 <p:cSld><p:spTree><a:t>スライド本文XYZ</a:t></p:spTree></p:cSld>
</p:sld>"""


def _zip(path, entries: dict):
    with zipfile.ZipFile(path, "w") as z:
        for name, data in entries.items():
            z.writestr(name, data)


def _blank_pdf(path: pathlib.Path) -> None:
    writer = PdfWriter()
    writer.add_blank_page(width=100, height=100)
    with path.open("wb") as stream:
        writer.write(stream)


def test_docx_to_md_heading_body_table():
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "a.docx"
    _zip(p, {"word/document.xml": _DOCX_XML})
    md = office_md.to_markdown(p)
    assert md is not None
    assert "# タイトル見出し" in md            # 見出しスタイル→#
    assert "本文テキストABC" in md             # 本文
    assert "| セル1 | セル2 |" in md           # 表→パイプ表


_DOCX_MERGED_NESTED_XML = """<?xml version="1.0"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
 <w:body>
  <w:tbl>
   <w:tr>
    <w:tc><w:tcPr><w:gridSpan w:val="2"/></w:tcPr><w:p><w:r><w:t>見出し結合</w:t></w:r></w:p></w:tc>
   </w:tr>
   <w:tr>
    <w:tc><w:tcPr><w:vMerge w:val="restart"/></w:tcPr><w:p><w:r><w:t>縦結合</w:t></w:r></w:p></w:tc>
    <w:tc><w:p><w:r><w:t>値2</w:t></w:r></w:p></w:tc>
   </w:tr>
   <w:tr>
    <w:tc><w:tcPr><w:vMerge/></w:tcPr><w:p/></w:tc>
    <w:tc><w:p><w:r><w:t>値3</w:t></w:r></w:p></w:tc>
   </w:tr>
  </w:tbl>
  <w:tbl>
   <w:tr>
    <w:tc>
     <w:p><w:r><w:t>外側セル</w:t></w:r></w:p>
     <w:tbl><w:tr><w:tc><w:p><w:r><w:t>ネスト値</w:t></w:r></w:p></w:tc></w:tr></w:tbl>
    </w:tc>
   </w:tr>
  </w:tbl>
 </w:body>
</w:document>"""


def test_docx_to_md_merged_cells_and_nested_table():
    """H2（正典 §3.2）: `_docx_table_walk` が解決した row_span/column_span を R5（値を継続セルへ複製）で
    展開し、ネスト表はパイプ表の直後に小見出し付きで続ける（パイプ表はセル内に表を持てないため）。"""
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "merged.docx"
    _zip(p, {"word/document.xml": _DOCX_MERGED_NESTED_XML})
    md = office_md.to_markdown(p)
    assert md is not None
    assert "| 見出し結合 | 見出し結合 |" in md        # gridSpan=2 の値を両セルへ複製
    assert "| 縦結合 | 値2 |" in md                    # vMerge restart（起点）
    assert "| 縦結合 | 値3 |" in md                    # vMerge 継続セルへ起点の値を複製
    assert "外側セル" in md and "ネスト値" in md
    assert "#### ネスト表（1行1列）" in md


def test_pptx_to_md_slide_text():
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "a.pptx"
    _zip(p, {"ppt/slides/slide1.xml": _PPTX_SLIDE})
    md = office_md.to_markdown(p)
    assert md is not None and "## スライド 1" in md and "スライド本文XYZ" in md


def test_xlsx_to_md_values():
    """H2（`docs/proposals/2026-08-28-人間向けMDの刷新.md` §3.1）: シート丸ごと1枚のパイプ表ではなく、
    `regions()` が検出した表候補ごとに `### {セル範囲}` の小見出し＋パイプ表を出す。"""
    import openpyxl
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "a.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "シート1"
    ws["A1"], ws["B1"] = "項目", "値"
    ws["A2"], ws["B2"] = "売上", "消費税率10%"
    wb.save(p)
    md = office_md.to_markdown(p)
    assert md is not None
    assert "## シート「シート1」" in md
    assert "### A1:B2" in md
    assert "| 項目 | 値 |" in md and "| 売上 | 消費税率10% |" in md


def test_xlsx_to_md_merged_cell_and_multiple_regions():
    """H2: 結合セルは値を継続セルへ複製する（R5・正典 §10 裁定#2）。同一シート内の癒着していない
    複数の表候補は、それぞれ独立した `### {セル範囲}` 小見出しになる（正典 §3.1）。"""
    import openpyxl
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "b.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "台帳"
    ws["A1"] = "見出し結合"
    ws.merge_cells("A1:B1")
    ws["A2"], ws["B2"] = "行", "値"
    ws["D1"], ws["E1"] = "甲", "乙"                    # 列 C を空けて別の連結成分にする
    ws["D2"], ws["E2"] = "丙", "丁"
    wb.save(p)
    md = office_md.to_markdown(p)
    assert md is not None
    assert "## シート「台帳」" in md
    assert "### A1:B2" in md and "### D1:E2" in md
    assert "| 見出し結合 | 見出し結合 |" in md         # R5: 結合セルの値を継続セルへ複製
    assert "| 行 | 値 |" in md
    assert "| 甲 | 乙 |" in md and "| 丙 | 丁 |" in md
    assert md.index("A1:B2") < md.index("D1:E2")       # 出現順は regions() の (min_row, min_col) 順


def test_unsupported_and_broken_return_none():
    d = tempfile.mkdtemp()
    pdf = pathlib.Path(d) / "a.pdf"
    pdf.write_bytes(b"%PDF-1.4 dummy")
    assert office_md.to_markdown(pdf) is None        # PDF バックエンド無 or 壊れ → 未対応
    bad = pathlib.Path(d) / "b.docx"
    bad.write_bytes(b"not a zip")
    assert office_md.to_markdown(bad) is None         # 壊れ docx（非zip）


# ---- PDF（テキスト層）: バックエンド未導入でも配線・整形をスタブで検証 ----

def test_pdf_normalize_deterministic():
    n = office_md._normalize_pdf_text
    assert n("a  \n\n\n b \r\nc") == "a\n\n b\nc"     # 行末空白除去・連続空行→1・CRLF正規化
    assert n("   ") == "" and n("") == ""


def test_pdf_md_assembly_stubbed():
    """H2 は PDF を対象外のまま据え置く（正典 §3.4・§8 裁定#6）。バイト完全一致で固定し、
    将来 `_pdf_md` へ意図しない変更が紛れ込んだ時に確実に検知する。"""
    o_b, o_p = office_md._pdf_backend, office_md._pdf_pages
    office_md._pdf_backend = lambda: "pypdf"
    office_md._pdf_pages = lambda p: ["ページ1の本文  ", "", "  二枚目\nの本文 "]   # 2枚目は空
    try:
        md = office_md.to_markdown(pathlib.Path("x.pdf"))
        assert md == "## ページ 1\n\nページ1の本文\n\n## ページ 3\n\n二枚目\nの本文"
    finally:
        office_md._pdf_backend, office_md._pdf_pages = o_b, o_p


def test_pdf_all_empty_returns_none_stubbed():
    o_b, o_p = office_md._pdf_backend, office_md._pdf_pages
    office_md._pdf_backend = lambda: "pypdf"
    office_md._pdf_pages = lambda p: ["", "   ", "\f"]      # スキャン画像/暗号化＝本文ゼロ
    try:
        assert office_md.to_markdown(pathlib.Path("x.pdf")) is None   # 本文ゼロ→未対応（OCR要）
    finally:
        office_md._pdf_backend, office_md._pdf_pages = o_b, o_p


def test_convertible_exts_tracks_backend():
    o_b = office_md._pdf_backend
    try:
        office_md._pdf_backend = lambda: None
        assert ".pdf" not in office_md.convertible_exts() and not office_md.pdf_available()
        office_md._pdf_backend = lambda: "pypdf"
        assert ".pdf" in office_md.convertible_exts() and office_md.pdf_available()
    finally:
        office_md._pdf_backend = o_b


def test_build_derived_pdf_buckets():
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    _blank_pdf(src / "doc.pdf")
    der = pathlib.Path(d) / "derived"
    o_b, o_p = office_md._pdf_backend, office_md._pdf_pages
    try:
        office_md._pdf_backend = lambda: None                # バックエンド無 → unsupported
        rep = office_md.build_derived(src, der)
        assert rep["unsupported"] == 1 and rep["converted"] == 0
        office_md._pdf_backend = lambda: "pypdf"              # バックエンド有＋本文あり → converted
        office_md._pdf_pages = lambda p: ["税率10%の説明"]
        rep = office_md.build_derived(src, der)
        assert rep["converted"] == 1 and rep["unsupported"] == 0
        assert "税率10%" in (der / "doc.pdf.md").read_text(encoding="utf-8")
    finally:
        office_md._pdf_backend, office_md._pdf_pages = o_b, o_p


def test_check_partial_extraction_flags_large_source_with_tiny_md(tmp_path):
    """ING-1: 原本サイズが十分大きい（1MiB超）のに生成MDが極端に小さければ疑いに計上する
    （docx/pdf 等にも効く粗い網・拡張子は問わない）。"""
    rp = tmp_path / "big.docx"
    rp.write_bytes(b"x" * (office_md._PARTIAL_SIZE_MIN_SOURCE_BYTES + 1))
    out: list[dict] = []
    office_md._check_partial_extraction(rp, "少しだけ", "big.docx", None, out)
    assert out == [{"doc": "big.docx", "basis": "size_ratio",
                    "source_bytes": office_md._PARTIAL_SIZE_MIN_SOURCE_BYTES + 1, "md_bytes": len("少しだけ".encode("utf-8"))}]


def test_check_partial_extraction_ignores_small_source(tmp_path):
    """原本が下限未満なら、MD が小さくても疑いに計上しない（小さい原本はMDも小さくて正常）。"""
    rp = tmp_path / "small.docx"
    rp.write_bytes(b"x" * 100)
    out: list[dict] = []
    office_md._check_partial_extraction(rp, "", "small.docx", None, out)
    assert out == []


def test_check_partial_extraction_ignores_large_source_with_proportionate_md(tmp_path):
    """原本が大きくても、生成MDも十分な量あれば疑いに計上しない（通常の大きい文書）。"""
    rp = tmp_path / "big.docx"
    rp.write_bytes(b"x" * (office_md._PARTIAL_SIZE_MIN_SOURCE_BYTES + 1))
    out: list[dict] = []
    office_md._check_partial_extraction(rp, "本文" * 10000, "big.docx", None, out)
    assert out == []


def test_check_partial_extraction_truncated_sheet_does_not_hide_other_sheet_suspicion(tmp_path):
    """1シートの自己申告打切り（`truncated`）は文書全体の `size_ratio` 判定だけを省略し、
    別シートの `partial_extraction_suspected` 走査は継続する（1枚の打切りが他シートの疑いまで
    消してしまうと、両状態を持つ最小入力で出力が空になっていた）。"""
    from types import SimpleNamespace
    rp = tmp_path / "big.xlsx"
    rp.write_bytes(b"x" * (office_md._PARTIAL_SIZE_MIN_SOURCE_BYTES + 1))
    document = SimpleNamespace(elements=[
        SimpleNamespace(type="sheet", source_map={"sheet": "A", "truncated": True}),
        SimpleNamespace(type="sheet", source_map={"sheet": "B", "partial_extraction_suspected": True,
                                                   "declared_rows": 100, "extracted_rows": 1}),
    ])
    out: list[dict] = []
    # md は極小（truncated が無ければ size_ratio が先に1件だけ計上して return してしまう入力）。
    office_md._check_partial_extraction(rp, "少しだけ", "big.xlsx", document, out)
    assert out == [{"doc": "big.xlsx", "basis": "xlsx_row_ratio", "declared_rows": 100, "extracted_rows": 1}]


def test_build_derived_broken_docx_is_fail_closed_not_sealed_as_success():
    """document-ir 構築に失敗した docx/xlsx は `document_ir_failed` へ計上され、
    `.document_ir_sig` を「全件成功」で確定しない（次回 sync の drift 検知で再試行できるように
    する＝失敗を成功マーカーで封印しない・fail-closed）。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    (src / "broken.docx").write_bytes(b"not a zip")          # BadZipFile→document is None
    der = pathlib.Path(d) / "derived"
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert rep["document_ir_failed"] >= 1
    assert any(f["doc"] == "broken.docx" and f["reason"].startswith("document_ir_failed:")
              for f in rep["document_ir_failures"])
    assert office_md.document_ir_sig_drift(der) is True      # 「全件成功」を偽っていない


def test_build_derived_docx_without_body_element_publishes_failed_notice():
    """`<w:body>` 要素自体が無い docx（例外は起きず構造的に None＝`document_ir_failed:malformed_structure`）は、
    Evidence 側の独立な再抽出（`_extract_canonical_evidence`）がより緩い判定で「成功」してしまうことがあり、
    その成否に notice の要否を委ねると `{rel}.md` が一切書かれず文書が台帳・grep から消える。
    IR 失敗の理由から直接 failed notice を組み立てて確実に発行することを固定する。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    docx_xml = ('<?xml version="1.0"?>'
                '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                "</w:document>")               # <w:body> が無い＝ _build_docx_ir が例外無しで None を返す
    _zip(src / "nobody.docx", {"word/document.xml": docx_xml})
    der = pathlib.Path(d) / "derived"
    rep = office_md.build_derived(src, der)
    assert rep["document_ir_failed"] >= 1
    assert any(f["doc"] == "nobody.docx" and f["reason"] == "document_ir_failed:malformed_structure"
              for f in rep["document_ir_failures"])
    assert rep["published_notice_count"] == 1
    md_path = der / "nobody.docx.md"
    assert md_path.is_file()                                  # notice が確実に発行される（消えない）
    meta = json_io.read_json(der / "nobody.docx.md.meta.json", default=None)
    assert meta is not None and meta["arm"] == "evidence_notice"
    assert "reason_code=source_parse_failed" in meta["notes"]
    assert (der.parent / "ir" / "nobody.docx.evidence.json").is_file()
    assert (der.parent / "rag" / "nobody.docx.rag.md").is_file()


def test_build_derived_oversized_office_file_skips_conversion_with_failed_notice(monkeypatch):
    """MEM-1: 入口サイズガード（`SHERPA_OFFICE_FILE_CAP_BYTES` 相当）を超えるファイルは、
    変換（openpyxl 等のフルロード）を一切試みず「サイズ超過」で失敗記録する（黙って落とさない）。
    上限未満の他ファイルは通常どおり変換される（1件の超過が他ファイルを巻き込まない）。"""
    import openpyxl
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "big.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    wb.save(p)
    _zip(src / "normal.docx", {"word/document.xml": _DOCX_XML})   # 上限未満＝通常どおり変換される
    der = pathlib.Path(d) / "derived"
    monkeypatch.setattr(office_md, "_OFFICE_FILE_CAP_BYTES", p.stat().st_size - 1)   # 確実に超過させる
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert {"doc": "big.xlsx", "reason": "size_exceeded"} in rep["conversion_failures"]
    md_path = der / "big.xlsx.md"
    assert md_path.is_file()                                     # notice が発行される（消えない）
    meta = json_io.read_json(der / "big.xlsx.md.meta.json", default=None)
    assert meta is not None and meta["arm"] == "evidence_notice"
    assert "reason_code=size_exceeded" in meta["notes"]
    assert rep["converted"] == 1                                  # normal.docx は通常どおり変換される
    assert (der / "normal.docx.md").is_file()


def test_build_derived_xlsx_cell_count_exceeded_skips_conversion_with_failed_notice(monkeypatch):
    """MEM-2: st_size ガード（MEM-1）は圧縮後サイズしか見ないため、圧縮率の高い xlsx（数万行の一覧表等）
    は素通りしうる。開封前（openpyxl フルロード前）に zip 内 `<dimension ref="..."/>` だけを見て
    セル数上限を超えたら「セル数超過」で失敗記録する（黙って落とさない）。上限未満の他ファイルは
    通常どおり変換される（1件の超過が他ファイルを巻き込まない）。"""
    import openpyxl
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "wide.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    ws["J20"] = "値"          # dimension は A1:J20（10列×20行=200セル）に確定する
    wb.save(p)
    _zip(src / "normal.docx", {"word/document.xml": _DOCX_XML})   # 上限未満＝通常どおり変換される
    der = pathlib.Path(d) / "derived"
    monkeypatch.setattr(office_md, "_XLSX_CELL_CAP", 100)          # 200セル > 100 で確実に超過させる
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert {"doc": "wide.xlsx", "reason": "cell_count_exceeded"} in rep["conversion_failures"]
    md_path = der / "wide.xlsx.md"
    assert md_path.is_file()                                       # notice が発行される（消えない）
    meta = json_io.read_json(der / "wide.xlsx.md.meta.json", default=None)
    assert meta is not None and meta["arm"] == "evidence_notice"
    assert "reason_code=cell_count_exceeded" in meta["notes"]
    assert rep["converted"] == 1                                   # normal.docx は通常どおり変換される
    assert (der / "normal.docx.md").is_file()


def test_build_derived_xlsx_within_cell_cap_converts_normally(monkeypatch):
    """MEM-2: セル数が上限以下の xlsx は誤検知せず通常どおり変換される（境界＝ちょうど上限は超過扱いしない）。"""
    import openpyxl
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "ok.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    ws["B2"] = "値2"          # dimension は A1:B2（2列×2行=4セル）
    wb.save(p)
    der = pathlib.Path(d) / "derived"
    monkeypatch.setattr(office_md, "_XLSX_CELL_CAP", 4)             # ちょうど境界＝超過ではない
    rep = office_md.build_derived(src, der)
    assert rep["converted"] == 1
    assert rep["failed"] == 0
    assert (der / "ok.xlsx.md").is_file()


def test_xlsx_estimated_cell_count_returns_none_when_dimension_missing():
    """MEM-2: `<dimension>` タグの無い/壊れたシート原本は見積不能＝None（fail-open）。
    ガード不可の場合は呼び出し側が通常の変換経路へそのまま流す契約。"""
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "no_dim.xlsx"
    _zip(p, {
        "xl/worksheets/sheet1.xml": (
            '<?xml version="1.0"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            '<sheetData><row r="1"><c r="A1" t="inlineStr"><is><t>値</t></is></c></row></sheetData>'
            "</worksheet>"
        ),
    })
    assert office_md._xlsx_estimated_cell_count(p) is None


def test_build_derived_docx_uncompressed_size_exceeded_skips_conversion_with_failed_notice(monkeypatch):
    """MEM-2: docx/pptx にも同型の圧縮爆弾リスクがある。zip の非圧縮サイズ合計
    （`SHERPA_OFFICE_UNCOMPRESSED_CAP_BYTES` 相当）を超えるファイルは変換を試みず
    「展開後サイズ超過」で失敗記録する。上限未満の他ファイルは通常どおり変換される。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "huge.docx"
    padded_xml = _DOCX_XML.replace("本文テキストABC", "本文テキストABC" + "パディング" * 2000)
    _zip(p, {"word/document.xml": padded_xml})
    _zip(src / "normal.docx", {"word/document.xml": _DOCX_XML})   # 上限未満＝通常どおり変換される
    der = pathlib.Path(d) / "derived"
    normal_bytes = office_md._office_uncompressed_total_bytes(src / "normal.docx")
    huge_bytes = office_md._office_uncompressed_total_bytes(p)
    cap = (normal_bytes + huge_bytes) // 2                          # 両者の中間＝normalは通し・hugeは超過させる
    monkeypatch.setattr(office_md, "_OFFICE_UNCOMPRESSED_CAP_BYTES", cap)
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert {"doc": "huge.docx", "reason": "uncompressed_size_exceeded"} in rep["conversion_failures"]
    md_path = der / "huge.docx.md"
    assert md_path.is_file()
    meta = json_io.read_json(der / "huge.docx.md.meta.json", default=None)
    assert meta is not None and meta["arm"] == "evidence_notice"
    assert "reason_code=uncompressed_size_exceeded" in meta["notes"]
    assert rep["converted"] == 1
    assert (der / "normal.docx.md").is_file()


def test_build_derived_legacy_materialized_uncompressed_size_exceeded(monkeypatch):
    """MEM-2: 旧形式（.doc等）を①OOXMLへ前段変換した後の materialized ファイルにも非圧縮サイズ
    上限を適用する（原本 .doc 自体は小さくても変換後の docx が巨大になりうるケースへの備え）。"""
    from sherpa.ingest.arms import legacy_convert
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "old.doc"
    p.write_bytes(b"legacy-binary-not-a-real-doc")   # 中身は問わない（ensure_ooxml を直接 mock する）
    materialized_dir = pathlib.Path(d) / "materialized"; materialized_dir.mkdir()
    materialized = materialized_dir / "old.docx"
    _zip(materialized, {"word/document.xml": _DOCX_XML})
    monkeypatch.setattr(legacy_convert, "legacy_exts", lambda: {".doc"})
    monkeypatch.setattr(legacy_convert, "ensure_ooxml", lambda src, rel, cache_root: (materialized, []))
    der = pathlib.Path(d) / "derived"
    cap = office_md._office_uncompressed_total_bytes(materialized) - 1   # 確実に超過させる
    monkeypatch.setattr(office_md, "_OFFICE_UNCOMPRESSED_CAP_BYTES", cap)
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert {"doc": "old.doc", "reason": "uncompressed_size_exceeded"} in rep["conversion_failures"]
    md_path = der / "old.doc.md"
    assert md_path.is_file()
    meta = json_io.read_json(der / "old.doc.md.meta.json", default=None)
    assert meta is not None and meta["arm"] == "evidence_notice"
    assert "reason_code=uncompressed_size_exceeded" in meta["notes"]


def test_xlsx_dimension_missing_falls_back_to_actual_cell_count_exceeded(monkeypatch):
    """RV是正#2: `<dimension>` 欠落は「見積不能」であって「安全」ではない——dimension は書き手の
    自己申告にすぎないため、欠落/不正な原本を fail-open で素通りさせず、シート XML の `<c ` 実数を
    ストリーミングでカウントして cap 超過を検出する（`_xlsx_estimated_cell_count` 単体の fail-open
    契約はそのまま・置換されるのは呼び出し側の「見積不能→素通り」という挙動）。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "no_dim_wide.xlsx"
    cells = "".join(f'<c r="A{i + 1}" t="inlineStr"><is><t>v</t></is></c>' for i in range(50))
    _zip(p, {
        "xl/worksheets/sheet1.xml": (
            '<?xml version="1.0"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f"<sheetData>{cells}</sheetData>"
            "</worksheet>"
        ),
    })
    assert office_md._xlsx_estimated_cell_count(p) is None      # 前提: dimension無し＝見積不能
    der = pathlib.Path(d) / "derived"
    monkeypatch.setattr(office_md, "_XLSX_CELL_CAP", 10)         # 実セル50 > 10 で確実に超過させる
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert {"doc": "no_dim_wide.xlsx", "reason": "cell_count_exceeded"} in rep["conversion_failures"]
    meta = json_io.read_json(der / "no_dim_wide.xlsx.md.meta.json", default=None)
    assert meta is not None and "reason_code=cell_count_exceeded" in meta["notes"]


def test_xlsx_actual_cell_count_stops_early_past_cap():
    """`_xlsx_actual_cell_count` は cap 超過が確定した時点で残りを読まずに打ち切る（正確な総数は
    不要・超過の有無だけが要件）。ここでは cap を小さくし、返り値が実際のセル数（100）ではなく
    cap 超過の事実だけを示す値（cap を上回っている）であることを確認する。"""
    d = tempfile.mkdtemp()
    p = pathlib.Path(d) / "many.xlsx"
    cells = "".join(f'<c r="A{i + 1}" t="inlineStr"><is><t>v</t></is></c>' for i in range(100))
    _zip(p, {
        "xl/worksheets/sheet1.xml": (
            '<?xml version="1.0"?>'
            '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
            f"<sheetData>{cells}</sheetData>"
            "</worksheet>"
        ),
    })
    result = office_md._xlsx_actual_cell_count(p, cap=5)
    assert result is not None and result > 5


def test_build_derived_legacy_materialized_cell_count_exceeded(monkeypatch):
    """RV是正#3: 旧形式（.xls等）を①OOXMLへ前段変換した後の materialized ファイルには従来
    非圧縮サイズガードしか適用されておらず、セル数ガードが欠けていた（原本 .xls 自体は小さくても
    変換後の xlsx が巨大なセル数になりうるケースへの備え・MEM-2 の xlsx セル数ガードと同型を
    materialized 側にも適用する）。"""
    import openpyxl
    from sherpa.ingest.arms import legacy_convert
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "old.xls"
    p.write_bytes(b"legacy-binary-not-a-real-xls")   # 中身は問わない（ensure_ooxml を直接 mock する）
    materialized_dir = pathlib.Path(d) / "materialized"; materialized_dir.mkdir()
    materialized = materialized_dir / "old.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    ws["J20"] = "値"          # dimension は A1:J20（10列×20行=200セル）に確定する
    wb.save(materialized)
    monkeypatch.setattr(legacy_convert, "legacy_exts", lambda: {".xls"})
    monkeypatch.setattr(legacy_convert, "ensure_ooxml", lambda src, rel, cache_root: (materialized, []))
    der = pathlib.Path(d) / "derived"
    monkeypatch.setattr(office_md, "_XLSX_CELL_CAP", 100)          # 200セル > 100 で確実に超過させる
    rep = office_md.build_derived(src, der)
    assert rep["failed"] == 1
    assert {"doc": "old.xls", "reason": "cell_count_exceeded"} in rep["conversion_failures"]
    md_path = der / "old.xls.md"
    assert md_path.is_file()
    meta = json_io.read_json(der / "old.xls.md.meta.json", default=None)
    assert meta is not None and meta["arm"] == "evidence_notice"
    assert "reason_code=cell_count_exceeded" in meta["notes"]


def test_build_derived_conv_cache_hit_does_not_bypass_size_guard(monkeypatch):
    """RV是正#4: 入口ガード群は CONV-CACHE のキャッシュ照合より前に評価する。原本 mtime/size・
    変換パイプライン署名が不変のまま上限だけ引き下げられた（＝ガードが後から強化された）状況でも、
    既存キャッシュのヒットでガードを迂回して復元されてはならない（毎回この原本を通すたび改めて
    ガードにかかる）。"""
    import openpyxl
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "wide.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    ws["J20"] = "値"          # dimension は A1:J20（200セル）
    wb.save(p)
    der = pathlib.Path(d) / "derived"
    rep1 = office_md.build_derived(src, der)
    assert rep1["converted"] == 1                                    # 1回目は上限内で普通に変換される
    cache_root = office_md._conv_cache_root_for(der)
    assert (cache_root / "wide.xlsx.key.json").is_file()             # キャッシュ済み

    monkeypatch.setattr(office_md, "_XLSX_CELL_CAP", 100)             # 200セル > 100 に引き下げ（署名は不変）
    rep2 = office_md.build_derived(src, der)
    assert rep2["converted"] == 0                                     # キャッシュ復元で素通りしていない
    assert rep2["failed"] == 1
    assert {"doc": "wide.xlsx", "reason": "cell_count_exceeded"} in rep2["conversion_failures"]


def test_conv_cache_skips_store_when_evidence_write_fails_then_recovers(monkeypatch):
    """RV是正#5: `converted` 判定された rel でも Evidence の一時書込（OSError）で
    `evidence_ir_failed` が増えた回はキャッシュへ保存しない。原本が変わらないまま書込要因が
    解消した次回は（キャッシュミスのまま）フル実変換が再試行され正しく回復する——もし保存されて
    いたら、2回目もキャッシュヒットで失敗delta（`evidence_ir_failed=1`）がそのまま再生され、
    実際には直っているのに失敗が焼き付いたままになってしまう。"""
    from sherpa.ingest import evidence_ir
    import openpyxl
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    p = src / "a.xlsx"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(p)
    der = pathlib.Path(d) / "derived"

    orig_write = evidence_ir.write_json_atomic
    should_fail = {"v": True}

    def flaky_write(path, data):
        if should_fail["v"]:
            raise OSError("simulated evidence.json write failure")
        return orig_write(path, data)
    monkeypatch.setattr(evidence_ir, "write_json_atomic", flaky_write)

    rep1 = office_md.build_derived(src, der)
    assert rep1["evidence_ir_failed"] == 1
    assert rep1["converted"] == 1                          # md自体はEvidence書込失敗と独立に成功扱い
    cache_root = office_md._conv_cache_root_for(der)
    assert not (cache_root / "a.xlsx.key.json").exists()   # 失敗deltaを含む結果はキャッシュされない

    should_fail["v"] = False                                # 一時要因（ディスク逼迫等）が解消したとする
    rep2 = office_md.build_derived(src, der)
    assert rep2["evidence_ir_failed"] == 0                  # キャッシュに焼き付いていたら1のまま再生される
    assert rep2["converted"] == 1
    assert (cache_root / "a.xlsx.key.json").is_file()       # 今回は成功のみ＝正しくキャッシュされる


def test_conv_cache_lookup_rejects_entry_with_failed_delta(tmp_path):
    """RV是正#5（防御的二重チェック）: rep_delta に失敗カウンタが残っている実体は、store側の
    新チェックより前の版で書かれた古いキャッシュだったとしても復元しない。"""
    cache_root = tmp_path / "_conv_cache"
    content_dir = cache_root / "a.xlsx.d"
    (content_dir / "md").mkdir(parents=True)
    (content_dir / "md" / "a.xlsx.md").write_text("dummy", encoding="utf-8")
    meta_path = cache_root / "a.xlsx.key.json"
    json_io.write_json_atomic(meta_path, {
        "key": "k1",
        "rep_delta": {
            "document_ir_generated": 1, "document_ir_failed": 0,
            "evidence_ir_generated": 0, "evidence_ir_failed": 1,
            "rag_generated": 0, "rag_failed": 0,
        },
    })
    assert office_md._conv_cache_lookup(cache_root, "a.xlsx", "k1") is None


def test_xlsx_extractor_version_bump_triggers_document_ir_and_human_md_drift(monkeypatch):
    """DOCX/XLSX の抽出規則（span/vMerge・列上限等）を変えて抽出器版を上げたら、
    `document_ir_sig_drift`（→ evidence/rag への連鎖）と `human_md_sig_drift`（→ `{rel}.md` の
    選択的再生成）の両方が発火する。document-ir は human MD と rag の共通の抽出元であり、抽出器版が
    片方の drift だけ据え置くと、人間向け MD と rag.md が別々の抽出結果を参照する
    split-brain になるため。"""
    import openpyxl
    from sherpa.ingest.arms import ooxml_arm

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    wb.save(src / "a.xlsx")
    der = pathlib.Path(d) / "derived"

    rep = office_md.build_derived(src, der)
    assert rep["converted"] == 1
    assert office_md.document_ir_sig_drift(der) is False
    assert office_md.human_md_sig_drift(src, der) is False

    monkeypatch.setattr(ooxml_arm, "XLSX_EXTRACTOR_VERSION", "xlsx-ooxml-vX-test-bump")
    assert office_md.document_ir_sig_drift(der) is True
    assert office_md.human_md_sig_drift(src, der) is True


def test_build_derived_reports_actual_candidate_progress():
    """未対応・破損を含め、候補文書を1件処理するごとに実数を報告する。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    _blank_pdf(src / "unsupported.pdf")
    (src / "broken.docx").write_bytes(b"not a zip")
    der = pathlib.Path(d) / "derived"
    observed: list[tuple[int, int]] = []
    original_backend = office_md._pdf_backend
    try:
        office_md._pdf_backend = lambda: None
        office_md.build_derived(src, der, progress=lambda processed, total: observed.append((processed, total)))
    finally:
        office_md._pdf_backend = original_backend

    assert observed[0] == (0, 2)
    assert observed[-1] == (2, 2)
    assert [processed for processed, _total in observed] == [0, 1, 2]


def test_arms_sig_drift_marker():
    """RV High: PDF バックエンドの導入/除去を派生 marker（`.arms_sig`）で検知し、署名同一でも作り直す。

    旧 `.pdf_backend` を一般化した `.arms_sig`（有効アーム＋バックエンド署名）で判定する（A1）。
    """
    o_b, o_p = office_md._pdf_backend, office_md._pdf_pages
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    _blank_pdf(src / "doc.pdf")
    der = pathlib.Path(d) / "derived"
    try:
        office_md._pdf_backend = lambda: None                 # バックエンド無で派生ビルド
        office_md.build_derived(src, der)
        # 署名は legacy 変換（W0）＋ VLM（vision）の実効可用性も含む（既定 none＝
        # 挙動不変）。document-ir 版（docir）は arms_sig に含まない（DOC-IR-001.5・修正3＝別マーカー
        # `.document_ir_sig` で管理・`es_index._arms_config_sig()` 経由の全 world ES 再索引誘発を避けるため）。
        # tesseract 直の `ocr` アーム撤去（2026-07-08）で `;ocr=` 成分は無くなった（署名フォーマット変更）。
        assert (der / office_md._ARMS_SIG_MARKER).read_text(encoding="utf-8") == "arms=ooxml,pdf_text;pdf=none;legacy=none;vlm=none"
        assert office_md.arms_sig_drift(der) is False          # 同状態＝drift 無し（無限ループしない）
        office_md._pdf_backend = lambda: "pypdf"               # 後からバックエンド導入
        assert office_md.arms_sig_drift(der) is True           # drift＝再ビルド要
        office_md._pdf_pages = lambda p: ["税率10%の説明"]
        office_md.build_derived(src, der)                      # 再ビルドで marker 更新＋PDF変換
        assert (der / office_md._ARMS_SIG_MARKER).read_text(encoding="utf-8") == "arms=ooxml,pdf_text;pdf=pypdf;legacy=none;vlm=none"
        assert office_md.arms_sig_drift(der) is False
        assert "税率10%" in (der / "doc.pdf.md").read_text(encoding="utf-8")
    finally:
        office_md._pdf_backend, office_md._pdf_pages = o_b, o_p


def test_arms_sig_drift_on_old_format_marker():
    """RV Low（2026-07-08）: tesseract 撤去前の旧フォーマット（`;ocr=` 成分入り）の `.arms_sig` marker は
    現行署名と一致しない＝drift=True で1回だけ全再ビルドされる（意図した挙動を固定・読み替え変更の見逃し防止）。"""
    o_b = office_md._pdf_backend
    d = tempfile.mkdtemp()
    der = pathlib.Path(d) / "derived"
    der.mkdir(parents=True)
    try:
        office_md._pdf_backend = lambda: None
        # 撤去前の実フォーマット（⑤ 時点）: ;ocr= 成分を含む旧署名を marker に置く。
        (der / office_md._ARMS_SIG_MARKER).write_text(
            "arms=ooxml,pdf_text;pdf=none;legacy=none;md=none;ocr=none;vlm=none", encoding="utf-8")
        assert office_md.arms_sig_drift(der) is True           # 旧フォーマット＝drift（再ビルド誘発）
        office_md._write_arms_sig_marker(der)                  # 再ビルド相当で marker が現行に更新
        assert office_md.arms_sig_drift(der) is False          # 以後は安定（ループしない）
    finally:
        office_md._pdf_backend = o_b


# ---- sidecar 欠落検知（`rag_sidecars_missing`）は生成時マニフェスト（`{rel}.derived.json`）の
#      記録をそのまま照合する。どの sidecar が書かれるかは実行時条件（空/image-only・有効 arm・
#      legacy backend 到達性・raster 等）で決定的に変わるため、原本の拡張子だけからは導出できない
#      ——生成した側が実際に書いた sidecar をそのまま書き残す設計をここで固定する。----

def test_rag_sidecars_missing_empty_ooxml_is_not_falsely_flagged():
    """本文が空の docx は legacy `.md` を持たない（`to_markdown` が実質空を返す）が、
    Evidence/RAG は生成される。マニフェストが実態どおり `.md` を含まなければ欠落として扱わない。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    docx_xml = ('<?xml version="1.0"?>'
                '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                "<w:body></w:body></w:document>")
    _zip(src / "empty.docx", {"word/document.xml": docx_xml})
    rep = office_md.build_derived(src, der)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
    assert not (der / "empty.docx.md").is_file()                # legacy MDは正当に無い
    assert (der.parent / "ir" / "empty.docx.evidence.json").is_file()
    manifest = json_io.read_json(der.parent / "ir" / "empty.docx.derived.json", default=None)
    assert manifest is not None and ".md" not in manifest["sidecars"]
    assert office_md.rag_sidecars_missing(src, der) is False


def test_rag_sidecars_missing_disabled_arm_extension_is_not_falsely_flagged(monkeypatch):
    """`ooxml` arm が無効化されている環境では xlsx/docx/pptx は何も生成されない（`unsupported`）。
    マニフェストは空リストとして記録され、欠落扱いにはならない。"""
    import openpyxl

    monkeypatch.setenv("SHERPA_ARMS", "pdf_text")     # ooxml を含まない構成
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "内容"
    wb.save(src / "a.xlsx")
    rep = office_md.build_derived(src, der)
    assert rep["converted"] == 0 and rep["unsupported"] == 1
    manifest = json_io.read_json(der.parent / "ir" / "a.xlsx.derived.json", default=None)
    assert manifest == {"schema": office_md._DERIVED_MANIFEST_SCHEMA_VERSION, "sidecars": []}
    assert office_md.rag_sidecars_missing(src, der) is False


def test_rag_sidecars_missing_legacy_backend_off_notice_tracked(monkeypatch):
    """legacy backend（LibreOffice 等）が利用不可な環境でも、`.doc` は
    `legacy_backend_unavailable` の通知として `.md`/`.md.meta.json`/`.evidence.json`/`.rag.md`/
    `.rag_chunks.jsonl` の5点が揃って書かれる。マニフェストはこの5点をそのまま記録し、
    どれか1つでも外部要因で消えれば欠落として検知する。"""
    from sherpa.ingest.arms import legacy_convert

    monkeypatch.setattr(legacy_convert, "legacy_exts", lambda: set())   # backend未導入を強制
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    (src / "old.doc").write_bytes(b"legacy binary stub")
    rep = office_md.build_derived(src, der)
    assert rep["published_notice_count"] == 1
    assert (der / "old.doc.md").is_file()
    manifest = json_io.read_json(der.parent / "ir" / "old.doc.derived.json", default=None)
    assert manifest is not None and set(manifest["sidecars"]) == {
        ".md", ".md.meta.json", ".evidence.json", ".rag.md", ".rag_chunks.jsonl"}
    assert office_md.rag_sidecars_missing(src, der) is False

    (der.parent / "rag" / "old.doc.rag_chunks.jsonl").unlink()
    assert office_md.rag_sidecars_missing(src, der) is True


def test_rag_sidecars_missing_raster_image_tracked():
    """単体 PNG/JPEG は Evidence/RAG が生成される（`_generate_evidence` を直接呼ぶ raster 経路）。
    マニフェストへ5点が記録され、`.evidence.json` の外部削除を検知できる。"""
    from PIL import Image

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    Image.new("RGB", (4, 4), color="red").save(src / "scan.png")
    rep = office_md.build_derived(src, der)
    assert rep["converted"] == 1 and rep["failed"] == 0
    manifest = json_io.read_json(der.parent / "ir" / "scan.png.derived.json", default=None)
    assert manifest is not None and set(manifest["sidecars"]) == {
        ".md", ".md.meta.json", ".evidence.json", ".rag.md", ".rag_chunks.jsonl"}
    assert office_md.rag_sidecars_missing(src, der) is False

    (der.parent / "ir" / "scan.png.evidence.json").unlink()
    assert office_md.rag_sidecars_missing(src, der) is True


def test_rag_sidecars_missing_normal_pdf_md_meta_deletion_detected():
    """テキスト層のある通常 PDF（image-only ではない）は `.md`/`.md.meta.json` も含めて生成される。
    これらだけが外部要因で消えても（`.evidence.json` は残っていても）欠落として検知する。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    _blank_pdf(src / "note.pdf")
    o_b, o_p = office_md._pdf_backend, office_md._pdf_pages
    try:
        office_md._pdf_backend = lambda: "pypdf"
        office_md._pdf_pages = lambda p: ["本文テキスト"]
        rep = office_md.build_derived(src, der)
    finally:
        office_md._pdf_backend, office_md._pdf_pages = o_b, o_p
    assert rep["converted"] == 1
    assert (der / "note.pdf.md").is_file()
    assert office_md.rag_sidecars_missing(src, der) is False

    (der / "note.pdf.md").unlink()
    (der / "note.pdf.md.meta.json").unlink()
    assert office_md.rag_sidecars_missing(src, der) is True


def test_rag_sidecars_missing_asset_directory_deletion_detected():
    """`{rel}.assets/` の個々のファイル削除・ディレクトリ全体の削除のどちらも欠落として検知する
    （raster PNG は必ず1件の content-addressed asset を持つ）。"""
    from PIL import Image

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    Image.new("RGB", (4, 4), color="blue").save(src / "scan.png")
    office_md.build_derived(src, der)
    manifest = json_io.read_json(der.parent / "ir" / "scan.png.derived.json", default=None)
    assert manifest is not None and manifest.get("assets")
    assert office_md.rag_sidecars_missing(src, der) is False

    asset_files = list((der.parent / "rag" / "scan.png.assets").iterdir())
    assert len(asset_files) == 1
    asset_files[0].unlink()                                        # 個々のファイルだけ削除
    assert office_md.rag_sidecars_missing(src, der) is True

    office_md.build_derived(src, der)                               # 健全な状態へ作り直す
    assert office_md.rag_sidecars_missing(src, der) is False
    shutil.rmtree(der.parent / "rag" / "scan.png.assets")           # ディレクトリ自体を削除
    assert office_md.rag_sidecars_missing(src, der) is True


def test_refresh_evidence_ir_regenerates_empty_ooxml_instead_of_deleting():
    """空 OOXML（legacy `.md` を持たない正当なケース）の `.evidence.json`/`.rag.md`/
    `.rag_chunks.jsonl`/manifest は、`refresh_evidence_ir()` を呼んでも「対象外」として
    削除されず、実際に再生成される（`seen` の判定基準が `.md` の有無に依存しない回帰確認）。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    docx_xml = ('<?xml version="1.0"?>'
                '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
                "<w:body></w:body></w:document>")
    _zip(src / "empty.docx", {"word/document.xml": docx_xml})
    rep = office_md.build_derived(src, der)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
    assert not (der / "empty.docx.md").is_file()
    assert (der.parent / "ir" / "empty.docx.evidence.json").is_file()

    rep2 = office_md.refresh_evidence_ir(src, der)
    assert rep2["evidence_ir_failed"] == 0 and rep2["rag_failed"] == 0
    assert rep2["evidence_ir_generated"] == 1 and rep2["rag_generated"] == 1   # 削除ではなく再生成
    assert (der.parent / "ir" / "empty.docx.evidence.json").is_file()
    assert (der.parent / "rag" / "empty.docx.rag.md").is_file()
    assert (der.parent / "rag" / "empty.docx.rag_chunks.jsonl").is_file()

    rep3 = office_md.refresh_evidence_ir(src, der)                            # 2回目も安定して再生成
    assert rep3["evidence_ir_failed"] == 0 and rep3["rag_failed"] == 0
    assert (der.parent / "ir" / "empty.docx.evidence.json").is_file()


def test_human_md_sig_drift_and_refresh_touch_only_the_md_asset(monkeypatch):
    """`asset_versions.human_md` の食い違いは `refresh_human_md` が
    `{rel}.md` だけを選択的に再生成する（`.document.json`/`.evidence.json`/`.rag.md`/
    `.rag_chunks.jsonl`・`.document_ir_sig`/`.evidence_ir_sig`/`.rag_sig` はいずれも無変更のまま）。
    マニフェストの `schema` は v1（`_DERIVED_MANIFEST_SCHEMA_VERSION` は上げない・全再構築は誘発しない）。
    """
    import openpyxl

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "値"
    wb.save(src / "a.xlsx")
    der = pathlib.Path(d) / "derived"

    rep = office_md.build_derived(src, der)
    assert rep["converted"] == 1
    manifest = json_io.read_json(der.parent / "ir" / "a.xlsx.derived.json", default=None)
    assert manifest["schema"] == office_md._DERIVED_MANIFEST_SCHEMA_VERSION == "derived-sidecar-manifest-v1"
    assert manifest["asset_versions"]["human_md"] == office_md._current_human_md_sig()
    assert office_md.human_md_sig_drift(src, der) is False    # 直後は追随済み

    before_evidence = (der.parent / "ir" / "a.xlsx.evidence.json").read_bytes()
    before_doc_ir_sig = (der / office_md._DOCUMENT_IR_SIG_MARKER).read_text(encoding="utf-8")
    before_evidence_ir_sig = (der / office_md._EVIDENCE_IR_SIG_MARKER).read_text(encoding="utf-8")

    monkeypatch.setattr(office_md, "_current_human_md_sig", lambda: "bumped-human-md-version")
    assert office_md.human_md_sig_drift(src, der) is True
    rep2 = office_md.refresh_human_md(src, der)
    assert rep2["human_md_generated"] == 1 and rep2["human_md_failed"] == 0
    assert office_md.human_md_sig_drift(src, der) is False   # 追随した

    manifest2 = json_io.read_json(der.parent / "ir" / "a.xlsx.derived.json", default=None)
    assert manifest2["asset_versions"]["human_md"] == "bumped-human-md-version"
    assert manifest2["schema"] == "derived-sidecar-manifest-v1"     # schema は不変＝全再構築は誘発しない

    # rag/evidence/document_ir 側は一切触れていない（内容・drift マーカーとも無変更）。
    assert (der.parent / "ir" / "a.xlsx.evidence.json").read_bytes() == before_evidence
    assert (der / office_md._DOCUMENT_IR_SIG_MARKER).read_text(encoding="utf-8") == before_doc_ir_sig
    assert (der / office_md._EVIDENCE_IR_SIG_MARKER).read_text(encoding="utf-8") == before_evidence_ir_sig
    assert office_md.document_ir_sig_drift(der) is False
    assert office_md.evidence_ir_sig_drift(der) is False
    assert office_md.rag_sig_drift(der) is False


def test_refresh_human_md_generates_md_for_previously_md_less_rel():
    """空 xlsx（旧世代＝`.md`/`asset_versions` を持たないマニフェスト）は、`.md` の有無で対象を
    絞り込まないため移行対象から漏れず、`refresh_human_md` が初めて `.md` を生成する
    （新レンダラは表候補0件でも見出し＋注記の MD を返すため）。"""
    import openpyxl

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    wb = openpyxl.Workbook()
    wb.save(src / "empty.xlsx")                      # 値を書かない＝空シート
    der = pathlib.Path(d) / "derived"
    der.mkdir()
    # 旧世代のマニフェストを手で再現する（.md 無し・asset_versions 自体が無い）。
    # マニフェスト本体は ir 層（§8.1 三階層）に置く。
    der_ir = der.parent / "ir"
    der_ir.mkdir()
    json_io.write_json_atomic(
        der_ir / "empty.xlsx.derived.json",
        {"schema": office_md._DERIVED_MANIFEST_SCHEMA_VERSION, "sidecars": []})

    assert not (der / "empty.xlsx.md").is_file()
    assert office_md.human_md_sig_drift(src, der) is True

    rep = office_md.refresh_human_md(src, der)
    assert rep["human_md_generated"] == 1 and rep["human_md_failed"] == 0
    assert (der / "empty.xlsx.md").is_file()
    assert "値のあるセルが見つかりませんでした" in (der / "empty.xlsx.md").read_text(encoding="utf-8")
    assert office_md.human_md_sig_drift(src, der) is False   # 追随した＝以後は再評価しない


def test_human_md_sig_drift_and_refresh_are_noop_when_ooxml_arm_disabled(monkeypatch):
    """`ooxml` アームが無効化されている間は、`.md` sidecar の有無に依存しない今の判定方式でも
    docx/xlsx を評価対象にしない（`convertible_exts()` がこれらを「未対応」のまま扱う契約に従う）。
    無効化中に迂回して人間向け MD を新規生成してしまうのを防ぐ。"""
    import openpyxl

    from sherpa.ingest import arms as _arms

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"; src.mkdir()
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src / "a.xlsx")
    der = pathlib.Path(d) / "derived"
    der.mkdir()

    monkeypatch.setattr(_arms, "enabled_arm_names", lambda: ["pdf_text"])   # ooxml を含まない
    assert office_md.human_md_sig_drift(src, der) is False
    rep = office_md.refresh_human_md(src, der)
    assert rep == {"human_md_generated": 0, "human_md_failed": 0, "human_md_failures": []}
    assert not (der / "a.xlsx.md").is_file()
    assert not (der.parent / "ir" / "a.xlsx.derived.json").exists()

    monkeypatch.setattr(_arms, "enabled_arm_names", lambda: ["ooxml", "pdf_text"])   # 再有効化
    assert office_md.human_md_sig_drift(src, der) is True
    rep = office_md.refresh_human_md(src, der)
    assert rep["human_md_generated"] == 1 and rep["human_md_failed"] == 0


def test_human_md_partial_failure_keeps_es_meta_pending_until_fixed(monkeypatch, tmp_path):
    """1 rel だけ human_md 再生成に失敗している間は `es_index._human_md_config_sig` が現行版を
    確定させず pending センチネル（`_HUMAN_MD_PENDING_SENTINEL`）を返し続け（fail-closed）、
    その rel が直り、かつ ES の bulk 成功を確認できた（`confirm_human_md_es_sig`）次回だけ
    現行版へ進む——部分失敗のまま ES の meta を確定すると、後で直っても reindex の契機が
    失われるため。pending が `None` ではなくセンチネルなのは、meta にフィールド自体が無い
    （旧索引・None）ケースと区別するため（`es_index._human_md_config_sig` docstring 参照）。"""
    import openpyxl

    from sherpa import es_index
    from sherpa import worlds as worlds_mod
    from sherpa.ingest.arms import ooxml_arm

    wd = tmp_path / "world"; wd.mkdir()
    dmd = tmp_path / "derived"
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(wd / "a.xlsx")
    rep = office_md.build_derived(wd, dmd)
    assert rep["evidence_ir_failed"] == 0

    monkeypatch.setattr(worlds_mod, "world_dir", lambda w: wd)
    monkeypatch.setattr(worlds_mod, "derived_md_dir", lambda w: dmd)
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: False)
    # レンダラ/抽出器の版を上げたのと同じ状況（実運用の drift 発生源）を模す。
    monkeypatch.setattr(office_md, "_current_human_md_sig", lambda: "human-md-vNEW")

    should_fail = {"v": True}
    real_build_xlsx_ir = ooxml_arm._build_xlsx_ir

    def _flaky_build(p):
        if should_fail["v"]:
            raise RuntimeError("simulated ir build failure")
        return real_build_xlsx_ir(p)
    monkeypatch.setattr(ooxml_arm, "_build_xlsx_ir", _flaky_build)

    result = office_md.refresh_human_md(wd, dmd)
    assert result["human_md_failed"] == 1
    assert office_md.human_md_sig_drift(wd, dmd) is True
    # fail-closed: 現行版を確定させない（None ではなく明示のセンチネル）。
    assert es_index._human_md_config_sig("w") == es_index._HUMAN_MD_PENDING_SENTINEL

    should_fail["v"] = False                                    # 次回 sync で直ったとする
    result = office_md.refresh_human_md(wd, dmd)
    assert result["human_md_failed"] == 0
    assert office_md.human_md_sig_drift(wd, dmd) is False
    # render 側は直ったが、ES の bulk 成功をまだ確認していない（`.human_md_es_sig` 未確定）間は
    # 依然 pending のまま——bulk_errors 等で index 自体は作られても中身が古いままの可能性がある。
    assert es_index._human_md_config_sig("w") == es_index._HUMAN_MD_PENDING_SENTINEL

    # worker が ES の bulk 成功を確認した後にだけ呼ぶ確定ヘルパ（`.rag_sig` と同型のホールドバック）。
    assert office_md.confirm_human_md_es_sig(wd, dmd) is True
    assert es_index._human_md_config_sig("w") == "human-md-vNEW"   # 両方直った今回だけ現行版へ進む


def test_build_derived_blocks_publish_on_unhandled_exception():
    """belt-and-braces の外側 except（想定外の例外）で終わった rel が1件でもあると、
    `unhandled_failed` へ計上され `build_derived` は公開しない（`error` を返し derived
    ディレクトリ自体が作られない）。想定内の notice 縮退（`failed` カウンタ）とは区別する。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src / "a.xlsx")

    original = office_md._convert_with_arms

    def _boom(*args, **kwargs):
        raise RuntimeError("simulated unhandled crash")
    office_md._convert_with_arms = _boom
    try:
        rep = office_md.build_derived(src, der)
    finally:
        office_md._convert_with_arms = original

    assert rep["unhandled_failed"] == 1
    assert rep["error"] == "derived_incomplete:unhandled_failed=1"
    assert not der.is_dir()                                        # ステージングのみ・公開されない
    # 件数だけでなく対象ファイル名・理由（クラス名のみ・マスク済み）も明細に残る。
    assert rep["unhandled_failures"] == [{"doc": "a.xlsx", "reason": "unhandled_exception:RuntimeError"}]


def test_build_derived_blocks_publish_on_manifest_write_failure():
    """`{rel}.derived.json` の書込自体が失敗した場合も `unhandled_failed` へ計上され公開しない
    （書けなかったマニフェストを「対象外」と取り違えて次回の欠落検知から漏らさないため）。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active["A1"] = "x"
    wb.save(src / "a.xlsx")

    original = json_io.write_text_atomic

    def _boom(path, *args, **kwargs):
        if str(path).endswith(".derived.json"):
            raise OSError("simulated manifest write failure")
        return original(path, *args, **kwargs)
    office_md.json_io.write_text_atomic = _boom
    try:
        rep = office_md.build_derived(src, der)
    finally:
        office_md.json_io.write_text_atomic = original

    assert rep["unhandled_failed"] == 1
    assert rep["error"] == "derived_incomplete:unhandled_failed=1"
    assert not der.is_dir()
    assert rep["unhandled_failures"] == [{"doc": "a.xlsx", "reason": "manifest_write_failed"}]


def test_publish_failure_after_retire_rolls_back_old_derived_content(monkeypatch):
    """後半 rename（staging→target）が失敗した時点で、前半（target→retired）は既に完了している。
    ロールバックしないと derived root ごと消え、Office 文書が grep 不能になる——`_publish_staging`
    は retired→target へ即時ロールバックし、旧内容のまま残す。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active["A1"] = "old"
    wb.save(src / "a.xlsx")

    rep1 = office_md.build_derived(src, der)
    assert not rep1.get("error")
    old_md = (der / "a.xlsx.md").read_text(encoding="utf-8")
    assert "old" in old_md

    wb2 = openpyxl.Workbook()
    wb2.active["A1"] = "new"
    wb2.save(src / "a.xlsx")

    original_rename = pathlib.Path.rename

    def _boom_rename(self, target):
        if self.name.endswith(office_md._STAGING_SUFFIX):
            raise OSError("simulated staging->target rename failure")
        return original_rename(self, target)
    monkeypatch.setattr(pathlib.Path, "rename", _boom_rename)

    rep2 = office_md.build_derived(src, der)

    assert rep2["error"].startswith("derived_publish_failed:")
    assert der.is_dir()                                                     # derived root が消えない
    assert (der / "a.xlsx.md").read_text(encoding="utf-8") == old_md        # 旧内容のまま（rollback）
    assert not der.with_name(der.name + office_md._STAGING_SUFFIX).exists()  # staging は掃除済み
    assert not der.with_name(der.name + office_md._RETIRED_SUFFIX).exists()  # retired は残らない


def test_double_rename_failure_preserves_retired_generation(monkeypatch):
    """公開rename失敗時のロールバック自体も失敗し（唯一残る旧世代が retired に取り残され）、
    かつ次回 build の `_recover_interrupted_swap` による復旧も同じ理由で失敗する二重障害でも、
    retired を消してはいけない（消すと派生物が全消失する）。障害が解消すれば次の build で
    retired から復旧できる。"""
    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    import openpyxl
    wb = openpyxl.Workbook()
    wb.active["A1"] = "old"
    wb.save(src / "a.xlsx")

    rep1 = office_md.build_derived(src, der)
    assert not rep1.get("error")
    old_md = (der / "a.xlsx.md").read_text(encoding="utf-8")

    wb2 = openpyxl.Workbook()
    wb2.active["A1"] = "new"
    wb2.save(src / "a.xlsx")

    original_rename = pathlib.Path.rename
    retired = der.with_name(der.name + office_md._RETIRED_SUFFIX)

    def _boom_rename(self, target):
        # staging→target（公開）と retired→target（ロールバック／`_recover_interrupted_swap`の
        # 復旧のどちらも同じ呼び出し形）を両方失敗させ、「ロールバックも次回の復旧も失敗する」
        # 二重障害を再現する。
        if self.name.endswith(office_md._STAGING_SUFFIX) or self.name.endswith(office_md._RETIRED_SUFFIX):
            raise OSError("simulated rename failure")
        return original_rename(self, target)
    monkeypatch.setattr(pathlib.Path, "rename", _boom_rename)

    rep2 = office_md.build_derived(src, der)
    assert rep2["error"].startswith("derived_publish_failed:")
    assert retired.is_dir()                                              # rollback失敗でもretiredは残る
    assert (retired / "a.xlsx.md").read_text(encoding="utf-8") == old_md

    # 次回の build（同じ障害が続いている想定）: `_recover_interrupted_swap` の復旧renameも失敗する
    # →setup errorとして即座に打ち切り、`_publish_staging`（冒頭でretiredを無条件削除する）へは進まない。
    rep3 = office_md.build_derived(src, der)
    assert rep3["error"].startswith("derived_setup_failed:")
    assert retired.is_dir()                                              # 二重障害でも消失しない
    assert (retired / "a.xlsx.md").read_text(encoding="utf-8") == old_md
    assert not der.is_dir()                                              # target自体は未復旧（想定内）

    # 障害が解消すれば、唯一残っていた旧世代（retired）から復旧した上で最新原本を反映できる。
    monkeypatch.setattr(pathlib.Path, "rename", original_rename)
    rep4 = office_md.build_derived(src, der)
    assert not rep4.get("error")
    assert "new" in (der / "a.xlsx.md").read_text(encoding="utf-8")
    assert not retired.exists()


def test_write_derived_sidecar_manifest_returns_false_on_asset_iterdir_failure(monkeypatch):
    """`{rel}.assets/` の列挙（`iterdir()`）が失敗した場合も、例外を伝播させずFalseを返す
    （呼び出し元がbool失敗として一様に扱えるようにする・iterdirはtry節の中で呼ぶ契約）。"""
    from PIL import Image

    d = tempfile.mkdtemp()
    src = pathlib.Path(d) / "src"
    src.mkdir()
    der = pathlib.Path(d) / "derived"
    Image.new("RGB", (4, 4), color="green").save(src / "scan.png")
    office_md.build_derived(src, der)

    original_iterdir = pathlib.Path.iterdir

    def _boom_iterdir(self):
        if self.name == "scan.png.assets":
            raise OSError("simulated iterdir failure")
        return original_iterdir(self)
    monkeypatch.setattr(pathlib.Path, "iterdir", _boom_iterdir)

    assert office_md._write_derived_sidecar_manifest(
        der, der.parent / "rag", der.parent / "ir", "scan.png") is False
