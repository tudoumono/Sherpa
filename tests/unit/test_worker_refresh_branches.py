"""`worker.sync()` の evidence/rag drift 軽量再生成分岐（RAG-KV-001）を pin する。

実際の tmp world（本物の xlsx を `office_md.build_derived()` で派生生成）を使い、`worker.sync()`
を通して `office_md` の実装を実際に駆動する。monkeypatch するのは DB/Neo4j/ES（`store.*`／
`es_index.rag_es_enabled`／`es_index.index_world`／`es_index.needs_reindex`）と、全再構築本体
（`worker.run`／`worker._run_locked`）だけに限る。

分岐の優先順:
  ⓪ force=True または prev!=sig → 常に `run()`（全再構築・変更なし）。
  ① arms drift（`_derived_stale`）→ 常に `run()`（変更なし）。
  ②③④ ⓪①のどちらにも該当しない時だけ `_refresh_derived_representations()` を評価する:
     - sidecar 欠落（生成時マニフェストと現物の不一致）は drift の有無によらず常に先に確認する →
       `"needs_full_run"`。欠落検知〜全再構築（`_run_locked`）〜`.rag_sig` 削除は同一
       `store.world_lock` 区間で行う（公開 `run()` を呼ぶと自身の非再入 lock と衝突するため、
       lock-free 版の `_run_locked` を直接呼ぶ）。
     - evidence drift → `refresh_evidence_ir()` のみ（`refresh_rag()` は呼ばない）。
     - rag drift のみ → `refresh_rag()` のみ。
     - RAG_ES 有効時は refresh 成功後に `index_world(content_sig=sig)` を呼び、成功時だけ
       `write_rag_sig_marker()` で確定する（マーカー保留方式）。
     - drift 無し → 既存の backfill/ES `needs_reindex` 自己修復経路をそのまま維持する。
"""
from __future__ import annotations

import contextlib

import openpyxl
import pytest

from sherpa import es_index, store, worlds
from sherpa.ingest import office_md, worker


def _build_world(tmp_path):
    wd = tmp_path / "world"
    wd.mkdir()
    dmd = tmp_path / "derived"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws["A1"], ws["B1"] = "No", "内容"
    ws["A2"], ws["B2"] = 1, "サンプル内容"
    wb.save(wd / "a.xlsx")
    rep = office_md.build_derived(wd, dmd)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0
    return wd, dmd


def _bump_marker(dmd, name):
    marker = dmd / name
    marker.write_text(marker.read_text(encoding="utf-8") + ";simulated-version-bump", encoding="utf-8")


@pytest.fixture
def _stub(monkeypatch, tmp_path):
    """`sync()` を DB/Neo4j/ES 無しで駆動する。`office_md`/派生ファイルは実物のまま。"""
    wd, dmd = _build_world(tmp_path)
    calls: dict[str, list] = {"run": [], "index_world": [], "needs_reindex": [], "reflect_graph": []}

    # rv-s2-mention #1: 軽量再生成が成功すると `_reflect_graph_after_rag_rewrite`
    # （`build_world_graph`→`world_neo4j.load_world`）が呼ばれるようになった——本ファイルの
    # 対象は ES/マーカーの分岐配線であり Neo4j 反映自体の中身は対象外（実 Neo4j に触れない、
    # という本ファイルの docstring の前提を保つ）。呼ばれたことだけを記録する。
    monkeypatch.setattr(worker, "_reflect_graph_after_rag_rewrite",
                        lambda world: calls["reflect_graph"].append(world))

    monkeypatch.setattr(worker, "world_state", lambda world, **kw: ("sig", {}))
    monkeypatch.setattr(store, "get_world",
                         lambda world: {"last_sig": "sig", "last_manifest": {}, "last_doc_count": 0})
    monkeypatch.setattr(worker, "_derived_stale", lambda world: False)
    monkeypatch.setattr(worlds, "world_dir", lambda world: wd)
    monkeypatch.setattr(worlds, "derived_md_dir", lambda world: dmd)

    @contextlib.contextmanager
    def _noop_lock(world_id):
        yield
    monkeypatch.setattr(store, "world_lock", _noop_lock)

    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: False)

    def _index_world(world, content_sig=None, **kw):
        calls["index_world"].append(content_sig)
        return {"available": True, "indexed": 1, "chunks": 1}
    monkeypatch.setattr(es_index, "index_world", _index_world)

    def _needs_reindex(world, sig, **kw):
        calls["needs_reindex"].append(sig)
        return False
    monkeypatch.setattr(es_index, "needs_reindex", _needs_reindex)

    def _run(world, **kw):
        # `run()`（⓪①分岐）と `_run_locked()`（sidecar欠落フォールバック）のどちらから呼ばれても
        # 同じ形で記録する（呼び出しkwargsはreflect以外に created_by/scan_root を持つことがある）。
        calls["run"].append({"reflect": kw.get("reflect", True)})
        return {"status": "auto_published", "ledger": 0, "flags": []}
    monkeypatch.setattr(worker, "run", _run)
    monkeypatch.setattr(worker, "_run_locked", _run)

    return {"calls": calls, "wd": wd, "dmd": dmd}


# ---- ⓪①: 全再構築の優先分岐 ----

def test_sync_force_true_always_runs_full(_stub):
    """`force=True` は drift 状態と無関係に `run()`。"""
    res = worker.sync("w", force=True)
    assert _stub["calls"]["run"] == [{"reflect": True}]
    assert res["changed"] is True


def test_sync_prev_sig_mismatch_always_runs_full(_stub, monkeypatch):
    """`prev != sig` は drift 状態と無関係に `run()`。"""
    monkeypatch.setattr(store, "get_world", lambda world: {"last_sig": "old", "last_manifest": {},
                                                            "last_doc_count": 0})
    res = worker.sync("w")
    assert _stub["calls"]["run"] == [{"reflect": True}]
    assert res["changed"] is True


def test_sync_arms_drift_runs_full_via_derived_stale(_stub, monkeypatch):
    """`prev==sig`・`force=False`・arms drift（`_derived_stale`）→ `run()`。軽量refreshは評価しない。"""
    monkeypatch.setattr(worker, "_derived_stale", lambda world: True)
    old_rag_md = (_stub["dmd"].parent / "rag" / "a.xlsx.rag.md").read_text(encoding="utf-8")
    res = worker.sync("w")
    assert _stub["calls"]["run"] == [{"reflect": True}]
    assert (_stub["dmd"].parent / "rag" / "a.xlsx.rag.md").read_text(encoding="utf-8") == old_rag_md  # 軽量refreshは実行されない
    assert res["changed"] is True


# ---- ②③: evidence/rag drift の排他分岐（実 office_md を実際に駆動する） ----

def test_sync_rag_drift_only_regenerates_via_real_refresh_rag(_stub):
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".rag_sig")
    assert office_md.rag_sig_drift(dmd) is True
    res = worker.sync("w")
    assert office_md.rag_sig_drift(dmd) is False    # RAG_ES無効＝refresh_rag自身が確定する
    assert office_md.evidence_ir_sig_drift(dmd) is False   # evidence側は無関係のまま変化しない
    assert _stub["calls"]["run"] == []
    assert res["status"] == "unchanged" and res["changed"] is False
    assert _stub["calls"]["reflect_graph"] == ["w"]  # rv-s2-mention #1: rag.md書換え成功後にグラフも追随させる


def test_sync_evidence_drift_regenerates_via_real_refresh_evidence_ir_only(_stub):
    """evidence drift → `refresh_evidence_ir()` のみが走り、rag側も一緒に確定する
    （`refresh_evidence_ir()` は rag もまとめて再生成する契約・`refresh_rag()` は呼ばれない）。"""
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".evidence_ir_sig")
    _bump_marker(dmd, ".rag_sig")
    res = worker.sync("w")
    assert office_md.evidence_ir_sig_drift(dmd) is False
    assert office_md.rag_sig_drift(dmd) is False
    assert _stub["calls"]["run"] == []
    assert res["status"] == "unchanged" and res["changed"] is False


def test_sync_no_drift_keeps_existing_backfill_and_es_repair(_stub):
    """drift が何も無ければ軽量refreshは呼ばれず、既存の backfill/ES `needs_reindex` 経路が維持される。"""
    dmd = _stub["dmd"]
    old_rag_md = (dmd.parent / "rag" / "a.xlsx.rag.md").read_text(encoding="utf-8")
    res = worker.sync("w")
    assert (dmd.parent / "rag" / "a.xlsx.rag.md").read_text(encoding="utf-8") == old_rag_md
    assert _stub["calls"]["needs_reindex"] == ["sig"]        # 既存のES自己修復チェックは変わらず走る
    assert res["status"] == "unchanged" and res["changed"] is False
    assert _stub["calls"]["reflect_graph"] == []   # rag.md自体が書き換わっていない＝グラフ反映も不要


def test_sync_human_md_only_drift_still_reaches_es_repair_same_call(_stub, monkeypatch):
    """human_md drift だけが起きた場合、`refresh_human_md` が `"handled"` を返しても、同じ
    `sync()` 呼び出し内で ES の `needs_reindex` 自己修復まで到達する（RAG_ES OFF の world では
    legacy `{rel}.md` の中身が ES の索引元そのものであり、ここで打ち切ると ES が古いまま次回
    sync まで取り残されるため）。"""
    dmd = _stub["dmd"]
    monkeypatch.setattr(office_md, "_current_human_md_sig", lambda: "bumped-human-md-version")
    assert office_md.human_md_sig_drift(_stub["wd"], dmd) is True

    calls = {"index_world": []}
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: True)

    def _index_world(world, content_sig=None, **kw):
        calls["index_world"].append(content_sig)
        return {"available": True, "indexed": 1, "chunks": 1}
    monkeypatch.setattr(es_index, "index_world", _index_world)

    res = worker.sync("w")
    assert office_md.human_md_sig_drift(_stub["wd"], dmd) is False       # human_md 側は追随した
    assert calls["index_world"] == ["sig"]                               # 同じ sync 呼び出し内でESまで到達
    assert res["status"] == "unchanged" and res["changed"] is False


def test_sync_unchanged_es_repair_progress_throttled_and_dedupes_same_value(_stub, monkeypatch):
    """RV是正（rv-periphery #4・2026-09-05）: unchanged 分岐の ES 自己修復
    （`index_world_with_human_md_holdback` の progress コールバック）に、`_run_locked` 側の
    `_es_progress` と同じ「100件間隔の間引き＋同値抑止」（rv-periphery #3(c)）を適用する——
    旧実装は doc グループ（flush）単位の呼び出しをそのまま DB 書込みへ転送しており、間引きが
    掛かっていなかった。連続する同値（1,1／150,150／300,300）は2回目を書かない。"""
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: True)

    def _index_world_holdback(world, *, content_sig=None, run_id=None, progress=None):
        for done in (0, 1, 1, 2, 150, 150, 300, 300):     # es_index.index_world の実呼び出し列を模す
            progress(done, 300)
        return {"available": True, "indexed": 300, "chunks": 300}
    monkeypatch.setattr(worker, "index_world_with_human_md_holdback", _index_world_holdback)

    recorded: list[dict] = []
    monkeypatch.setattr(store, "update_ingest_run_progress",
                        lambda run_id, payload: recorded.append(payload))
    monkeypatch.setattr(store, "finish_ingest_run", lambda *a, **k: None)
    monkeypatch.setattr(worker.webhooks, "notify_run_terminal", lambda *a, **k: None)

    res = worker.sync("w", run_id=999)
    assert res["status"] == "unchanged"
    # `total == 300` で絞る＝直前の無条件なステージ遷移記録（`_progress("es_index", done=0,
    # total=None)`・本テストの対象外）を除外し、ホールドバックの progress コールバックが実際に
    # 書いた値だけを見る。1/1/2 は 100件間隔未満のため間引かれ、150/150・300/300 は同値の2回目が
    # 抑止される。
    done_values = [r["done"] for r in recorded if r["stage"] == "es_index" and r["total"] == 300]
    assert done_values == [0, 150, 300]


def test_sync_self_heal_success_confirms_human_md_es_marker(_stub, monkeypatch):
    """ES 自己修復（unchanged 分岐）の `index_world` が成功（bulk_errors 等の失敗が無い）したら、
    `office_md.confirm_human_md_es_sig` で `.human_md_es_sig` マーカーを確定し、続けて
    `es_index.confirm_human_md_meta` で ES 自身の `_meta.human_md_sig` も現行署名へ書き直す
    （RAG-KV の `.rag_sig` と同型のホールドバック方式・マーカーだけ確定して meta を放置すると
    次回 sync が meta の古い値を検知して無駄な再索引を繰り返し続ける・収束しない）。"""
    dmd = _stub["dmd"]
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: True)
    monkeypatch.setattr(es_index, "index_world",
                         lambda world, content_sig=None, **kw: {"available": True, "indexed": 1, "chunks": 1})
    meta_calls: list[str] = []
    monkeypatch.setattr(es_index, "confirm_human_md_meta", lambda world: meta_calls.append(world) or True)

    res = worker.sync("w")
    assert res["status"] == "unchanged"
    marker = dmd / ".human_md_es_sig"
    assert marker.is_file()
    assert marker.read_text(encoding="utf-8").strip() == office_md._current_human_md_sig()
    assert meta_calls == ["w"]


def test_sync_self_heal_bulk_errors_skips_marker_and_records_failure(_stub, monkeypatch):
    """ES 自己修復の `index_world` が部分失敗（`bulk_errors` 等）を返したら、`.human_md_es_sig`
    マーカーを確定せず（pending のまま次回 sync が自動的に再試行する）、失敗を `ingest_runs` へ
    記録する。直った次回は marker が確定する。"""
    dmd = _stub["dmd"]
    recorded: list[dict] = []
    monkeypatch.setattr(store, "add_ingest_run", lambda world, **kw: recorded.append({"world": world, **kw}))
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: True)

    should_fail = {"v": True}

    def _index_world(world, content_sig=None, **kw):
        if should_fail["v"]:
            return {"available": True, "indexed": 1, "chunks": 1, "error": "bulk_errors"}
        return {"available": True, "indexed": 1, "chunks": 1}
    monkeypatch.setattr(es_index, "index_world", _index_world)

    res = worker.sync("w")
    assert res["status"] == "unchanged"
    marker = dmd / ".human_md_es_sig"
    assert not marker.is_file()                          # 部分失敗＝確定しない
    assert len(recorded) == 1
    assert recorded[0]["status"] == "failed"
    assert recorded[0]["extraction_snapshot"]["error"] == "bulk_errors"

    should_fail["v"] = False                              # 次回 sync で直ったとする
    res = worker.sync("w")
    assert res["status"] == "unchanged"
    assert marker.is_file()                               # 今回は確定する
    assert len(recorded) == 1                              # 失敗記録は増えない（成功時は記録しない）


def test_sync_self_heal_failure_with_run_id_folds_into_same_run_not_a_new_one(_stub, monkeypatch):
    """`run_id`（受付run）指定時、unchanged 分岐の ES 自己修復失敗は別 run を作らず、
    受付 run 自身の終端（`finish_ingest_run`）へ畳み込む——`add_ingest_run`（新規行）は呼ばれない。"""
    new_runs: list[dict] = []
    monkeypatch.setattr(store, "add_ingest_run", lambda world, **kw: new_runs.append({"world": world, **kw}))
    finished: list[dict] = []
    monkeypatch.setattr(store, "finish_ingest_run",
                         lambda run_id, **kw: finished.append({"run_id": run_id, **kw}))
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: True)
    monkeypatch.setattr(es_index, "index_world",
                         lambda world, content_sig=None, **kw: {"available": True, "error": "bulk_errors"})

    res = worker.sync("w", run_id=999)
    assert res["status"] == "unchanged"
    assert new_runs == []                                  # 別 run は作られない
    assert len(finished) == 1 and finished[0]["run_id"] == 999
    assert finished[0]["status"] == "failed"
    reasons = [f.get("reason") for f in finished[0]["extraction_snapshot"].get("flags", [])]
    assert any(r and r.startswith("es_repair_failed") for r in reasons)


def test_sync_self_heal_success_with_run_id_finalizes_as_auto_published(_stub, monkeypatch):
    """対照ケース: ES 自己修復が成功すれば、受付 run は通常どおり `auto_published` で
    終端する（畳み込みの追加が成功パスを壊していないことの確認）。"""
    finished: list[dict] = []
    monkeypatch.setattr(store, "finish_ingest_run",
                         lambda run_id, **kw: finished.append({"run_id": run_id, **kw}))
    monkeypatch.setattr(es_index, "needs_reindex", lambda world, sig, **kw: True)
    monkeypatch.setattr(es_index, "index_world",
                         lambda world, content_sig=None, **kw: {"available": True, "indexed": 1, "chunks": 1})
    monkeypatch.setattr(es_index, "confirm_human_md_meta", lambda world: True)

    res = worker.sync("w", run_id=999)
    assert res["status"] == "unchanged"
    assert len(finished) == 1 and finished[0]["run_id"] == 999
    assert finished[0]["status"] == "auto_published"


def test_index_world_holdback_drops_marker_before_reindex_prevents_stale_confirmation(_stub, monkeypatch):
    """二段階更新の穴: 以前の成功で `.human_md_es_sig` が既に確定済みの状態から、（human_md とは
    無関係な別次元の変化などで）別の再索引が走り、その bulk が部分失敗しても、再索引前に
    マーカーを無効化しているため meta には確定値が残らず pending のまま——次回の呼び出しも
    reindex 対象であり続ける（複数回 sync しても、bulk が直るまで再索引され続けることの確認）。
    マーカーを先に落とさず `ensure_index` を呼ぶ（旧実装相当）と、bulk 前の時点で「pending で
    ない」現行版が meta へ書かれてしまい、bulk が部分失敗しても meta が確定値のまま固定される。"""
    wd, dmd = _stub["wd"], _stub["dmd"]
    assert office_md.confirm_human_md_es_sig(wd, dmd) is True   # 「以前の成功」を模す
    assert office_md.human_md_es_sig_drift(dmd) is False

    monkeypatch.setattr(es_index, "index_world",
                         lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 1, "chunks": 1, "error": "bulk_errors"})

    esr = worker.index_world_with_human_md_holdback("w", content_sig="sig")
    assert esr.get("error") == "bulk_errors"
    assert office_md.human_md_es_sig_drift(dmd) is True         # 無効化されたまま再確定されない

    # 2回目（次回 sync 相当）も bulk がまだ失敗していれば同様に pending のまま。
    esr2 = worker.index_world_with_human_md_holdback("w", content_sig="sig")
    assert esr2.get("error") == "bulk_errors"
    assert office_md.human_md_es_sig_drift(dmd) is True

    # bulk が直れば次回は確定する。
    monkeypatch.setattr(es_index, "index_world",
                         lambda world, content_sig=None, **kw: {"available": True, "indexed": 1, "chunks": 1})
    esr3 = worker.index_world_with_human_md_holdback("w", content_sig="sig")
    assert not esr3.get("error")
    assert office_md.human_md_es_sig_drift(dmd) is False


def test_derived_dir_missing_skips_sidecar_and_drift_check(_stub, monkeypatch):
    """`dmd.exists()` が False（text/code のみの world）なら sidecar/drift 判定自体をスキップする。"""
    empty_dir = _stub["dmd"].parent / "no-such-derived-dir"
    monkeypatch.setattr(worlds, "derived_md_dir", lambda world: empty_dir)
    res = worker.sync("w")
    assert _stub["calls"]["run"] == []
    assert res["status"] == "unchanged"


# ---- document_ir drift 連鎖（document_ir→evidence→rag・失敗分離とマーカー確定順） ----

def test_document_ir_partial_failure_still_cascades_to_evidence_and_rag(_stub, monkeypatch):
    """document_ir が失敗しても、それだけで evidence→rag への連鎖を打ち切らない（打ち切ると、
    その文書が直らない限り World 全体の RAG/ES まで永久に旧世代のまま固定されてしまう）。
    `.document_ir_sig` は world 単位で1つのマーカーしか持たないため、失敗が残る限り次回 sync も
    対象 OOXML 文書を全件（この文書だけでなく）再実行し、それに伴い evidence/rag（RAG_ES 有効時
    は ES 索引も）も document_ir drift が続く限り毎回再実行される（1文書だけを狙い撃つ選択的な
    再試行や、evidence/rag側の要否判定によるスキップは無い＝sync が既定でポーリング駆動でない
    ため、この冗長さを個別最適化する必要は無いと判断している）。"""
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".document_ir_sig")
    original_refresh_document_ir = office_md.refresh_document_ir

    def _failing(wd_, dmd_, **kw):
        return {"document_ir_generated": 0, "document_ir_failed": 1,
                "document_ir_failures": [{"doc": "a.xlsx", "reason": "build_failed:RuntimeError"}]}
    monkeypatch.setattr(office_md, "refresh_document_ir", _failing)

    res = worker.sync("w")
    assert office_md.document_ir_sig_drift(dmd) is True          # 失敗のまま＝次回sync再試行
    assert office_md.evidence_ir_sig_drift(dmd) is False         # それでもevidence/ragの連鎖は実行され成功する
    assert office_md.rag_sig_drift(dmd) is False
    assert res["status"] == "unchanged" and res["changed"] is False

    # 次回 sync: document_ir drift がまだ残っているため evidence/rag への連鎖も（既に確定済みで
    # あっても）もう一度再実行される。document_ir が復旧すれば今度こそ全体が収束する。
    monkeypatch.setattr(office_md, "refresh_document_ir", original_refresh_document_ir)
    worker.sync("w")
    assert office_md.document_ir_sig_drift(dmd) is False


def test_document_ir_marker_confirm_waits_for_downstream_holdback_success(_stub, monkeypatch):
    """document_ir 版だけが独立に上がった状況で RAG_ES=1 の時、evidence 側の `.rag_sig`
    ホールドバック削除（生成着手前の事前 unlink）が失敗すると、document_ir 自体の生成は
    成功していても document_ir マーカーを確定してはいけない——先に確定すると、次回 sync では
    document_ir drift が既に False になり、evidence/rag/ES の再試行の入口そのものを失って
    恒久的に旧世代のまま固定される。"""
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".document_ir_sig")
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    original_drop = office_md.drop_rag_sig_marker
    monkeypatch.setattr(office_md, "drop_rag_sig_marker", lambda dr: False)   # holdback unlink失敗を模す

    res = worker.sync("w")
    assert _stub["calls"]["index_world"] == []                    # evidence側が着手前に打ち切り＝ESも呼ばない
    assert office_md.document_ir_sig_drift(dmd) is True           # document_ir自体は成功してもマーカーは未確定
    assert office_md.evidence_ir_sig_drift(dmd) is False          # 触られていない（着手前に打ち切り）
    assert office_md.rag_sig_drift(dmd) is False                  # 同上
    assert res["status"] == "unchanged"

    # 次回 sync: holdback unlinkが復旧すれば document_ir drift が再検知されて連鎖が完走し収束する。
    monkeypatch.setattr(office_md, "drop_rag_sig_marker", original_drop)
    worker.sync("w")
    assert office_md.document_ir_sig_drift(dmd) is False
    assert office_md.rag_sig_drift(dmd) is False


def test_evidence_only_marker_missing_rag_marker_current_es_failure_retried(_stub, monkeypatch):
    """`.evidence_ir_sig` マーカーのみ欠落（`.rag_sig` マーカーは現在値と一致＝rag 自身は単体
    では drift していない）状態で RAG_ES=1 の `index_world` が失敗すると、ホールドバックの
    事前 unlink により `.rag_sig` も未確定へ戻り、次回 sync が rag 側だけを正しく再試行する
    （両マーカーが同時に drift していないと検知できない退行のケース）。"""
    dmd = _stub["dmd"]
    (dmd / ".evidence_ir_sig").unlink()
    assert office_md.evidence_ir_sig_drift(dmd) is True
    assert office_md.rag_sig_drift(dmd) is False                  # rag自身は現在値と一致（単体ではdriftしない）
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    monkeypatch.setattr(es_index, "index_world", lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 0, "chunks": 0, "error": "bulk_failed"})

    res = worker.sync("w")
    assert office_md.evidence_ir_sig_drift(dmd) is False           # evidence自体は生成成功→確定
    assert office_md.rag_sig_drift(dmd) is True                    # holdbackの事前unlinkで未確定へ戻る
    assert res["status"] == "unchanged"

    # 次回 sync: evidence側は既に確定済み＝rag driftのみとして再試行され、ESが復旧すれば収束する。
    monkeypatch.setattr(es_index, "index_world", lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 1, "chunks": 1})
    worker.sync("w")
    assert office_md.rag_sig_drift(dmd) is False


# ---- RAG_ES 接続（マーカー保留方式・実 index_world 呼び出しのみ mock） ----

def test_sync_refresh_success_with_rag_es_confirms_marker(_stub, monkeypatch):
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".rag_sig")
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    worker.sync("w")
    assert _stub["calls"]["index_world"] == ["sig"]           # content_sig明示・1回だけ
    assert office_md.rag_sig_drift(dmd) is False               # ES成功→workerが確定
    # human_md は RAG_ES の設定に関わらず ES 反映に影響しうるため、この bulk 成功
    # でも `.human_md_es_sig` を確定する（render 側は fixture の世界が既に追随済み）。
    assert (dmd / ".human_md_es_sig").is_file()
    assert (dmd / ".human_md_es_sig").read_text(encoding="utf-8").strip() == office_md._current_human_md_sig()


def test_sync_index_world_failure_leaves_marker_unconfirmed_then_retries(_stub, monkeypatch):
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".rag_sig")
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    monkeypatch.setattr(es_index, "index_world", lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 0, "chunks": 0, "error": "bulk_failed"})
    res = worker.sync("w")
    assert office_md.rag_sig_drift(dmd) is True                # マーカー保留・次回sync再試行
    assert res["status"] == "unchanged"

    # 次回 sync で ES が復旧すれば収束する（マーカー保留方式の自己修復を実際に2回のsyncで確認）
    monkeypatch.setattr(es_index, "index_world", lambda world, content_sig=None, **kw:
                         {"available": True, "indexed": 1, "chunks": 1})
    worker.sync("w")
    assert office_md.rag_sig_drift(dmd) is False


def test_sync_refresh_failure_skips_es_and_marker(_stub, monkeypatch):
    """refresh 自体が失敗（failure counter > 0）した場合は ES を呼ばずマーカーも確定しない。"""
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".rag_sig")
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    monkeypatch.setattr(office_md, "refresh_rag",
                         lambda wd_, dmd_, **kw: {"rag_generated": 0, "rag_failed": 1,
                                                   "rag_failures": [{"doc": "x", "reason": "write_failed"}]})
    worker.sync("w")
    assert _stub["calls"]["index_world"] == []
    assert office_md.rag_sig_drift(dmd) is True


def test_rag_es_off_refresh_confirms_marker_itself(_stub):
    """RAG_ES 無効時は `write_rag_sig_marker=False` を渡さない（既定 True のまま呼ぶ）契約を、
    実際に refresh_rag 自身がマーカーを確定することで確認する。"""
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".rag_sig")
    worker.sync("w")
    assert _stub["calls"]["index_world"] == []
    assert office_md.rag_sig_drift(dmd) is False


def test_holdback_unlink_failure_aborts_before_generation_and_keeps_marker(_stub, monkeypatch):
    """holdback（生成開始前の `.rag_sig` 削除）が実際に OSError で失敗すると、refresh は着手せず
    既存の `.rag_sig`／既存の `.rag.md` はどちらも無傷のまま残る（chmod で実際に unlink を失敗させる）。"""
    dmd = _stub["dmd"]
    _bump_marker(dmd, ".rag_sig")
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    old_marker = (dmd / ".rag_sig").read_text(encoding="utf-8")
    old_rag_md = (dmd.parent / "rag" / "a.xlsx.rag.md").read_text(encoding="utf-8")
    dmd.chmod(0o555)                     # ディレクトリ書込不可＝`.rag_sig`のunlinkがOSErrorで失敗する
    try:
        res = worker.sync("w")
    finally:
        dmd.chmod(0o755)
    assert _stub["calls"]["index_world"] == []                                    # 生成未着手＝ESも呼ばない
    assert (dmd / ".rag_sig").read_text(encoding="utf-8") == old_marker            # マーカー現状維持
    assert (dmd.parent / "rag" / "a.xlsx.rag.md").read_text(encoding="utf-8") == old_rag_md       # 既存成果物も無傷
    assert res["status"] == "unchanged"


# ---- sidecar 欠落フォールバック（実ファイルを削除/非regular file化して検知させる） ----

def _replace_with_directory(path):
    path.unlink()
    path.mkdir()


def _replace_with_broken_symlink(path):
    path.unlink()
    path.symlink_to(path.with_name("does-not-exist"))


def _sidecar_path(dmd, rel_stem, suffix):
    """§8.1 三階層のフォルダ分離: `.evidence.json` は ir 層、それ以外（`.md`/`.md.meta.json`）は
    dmd（md 層）そのもの。"""
    if suffix == ".evidence.json":
        return dmd.parent / "ir" / f"{rel_stem}{suffix}"
    return dmd / f"{rel_stem}{suffix}"


@pytest.mark.parametrize("suffix", [".md", ".md.meta.json", ".evidence.json"])
@pytest.mark.parametrize("mutate", [
    lambda p: p.unlink(),
    _replace_with_directory,
    _replace_with_broken_symlink,
], ids=["missing", "directory", "broken_symlink"])
def test_sidecar_missing_falls_back_to_full_run(_stub, suffix, mutate):
    """sidecar欠落（削除・ディレクトリ化・壊れたシンボリックリンク化のいずれか）を、`.evidence.json`
    だけでなく分岐②が前提にする `.md`／`.md.meta.json` についても検知したら、軽量refreshを使わず
    全再構築へフォールバックする（`rag_sidecars_missing` の存在判定は `is_file()`＝ディレクトリ／
    壊れた symlink も「無い」扱いになる契約）。"""
    mutate(_sidecar_path(_stub["dmd"], "a.xlsx", suffix))
    res = worker.sync("w")
    assert _stub["calls"]["run"] == [{"reflect": True}]
    assert res["changed"] is True


def test_sidecar_missing_fallback_runs_inside_same_lock_as_detection(_stub, monkeypatch):
    """欠落検知〜全再構築〜`.rag_sig`削除は同一 `store.world_lock` 区間で行う（`run()` を呼ぶと
    自身の非再入 lock と衝突するため、lock-free 版の `_run_locked` を lock 保持中に直接呼ぶ）。
    さらに検知（`rag_sidecars_missing`）→全再構築（`_run_locked`）→マーカー削除
    （`drop_rag_sig_marker`）の順序も固定する（順序が入れ替わると、検知結果を使わずに削除だけ
    先行する等の非一貫性を生みうる）。

    lock を一度解放してから全再構築を呼ぶ設計だと、その間隙に他プロセスの sync が割り込んで
    全再構築が重複したり、`.rag_sig` 削除だけが lock 外に取り残されたりする非原子性を生む。
    """
    (_stub["dmd"].parent / "ir" / "a.xlsx.evidence.json").unlink()
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    lock_active = {"value": False}
    events: list[str] = []

    @contextlib.contextmanager
    def _tracking_lock(world_id):
        lock_active["value"] = True
        try:
            yield
        finally:
            lock_active["value"] = False
    monkeypatch.setattr(store, "world_lock", _tracking_lock)

    orig_missing = office_md.rag_sidecars_missing

    def _missing(wd_, dmd_):
        events.append("detect")
        return orig_missing(wd_, dmd_)
    monkeypatch.setattr(office_md, "rag_sidecars_missing", _missing)

    run_seen_lock_active = []

    def _run_locked(world, **kw):
        events.append("run")
        run_seen_lock_active.append(lock_active["value"])   # 全再構築の実行中、lockはまだ保持されている
        _stub["calls"]["run"].append({"reflect": kw.get("reflect", True)})
        return {"status": "auto_published", "ledger": 0, "flags": []}
    monkeypatch.setattr(worker, "_run_locked", _run_locked)

    drop_seen_lock_active = []
    orig_drop = office_md.drop_rag_sig_marker

    def _drop(dr):
        events.append("drop")
        drop_seen_lock_active.append(lock_active["value"])  # `.rag_sig`削除もlock保持中に行う
        return orig_drop(dr)
    monkeypatch.setattr(office_md, "drop_rag_sig_marker", _drop)

    res = worker.sync("w")
    assert events == ["detect", "run", "drop"]              # 検知→全再構築→マーカー削除の順序を固定
    assert run_seen_lock_active == [True]
    assert drop_seen_lock_active == [True]
    assert lock_active["value"] is False                    # sync() 復帰後はlockを手放している
    assert _stub["calls"]["run"] == [{"reflect": True}]
    assert res["changed"] is True


def test_sidecar_missing_fallback_drops_rag_sig_when_rag_es_enabled(_stub, monkeypatch):
    (_stub["dmd"].parent / "ir" / "a.xlsx.evidence.json").unlink()
    monkeypatch.setattr(es_index, "rag_es_enabled", lambda: True)
    worker.sync("w")
    # run() 自体は mock なので派生物は復元されないが、`.rag_sig` は実際に drop_rag_sig_marker で
    # 削除される（次回 sync が分岐③へ確実に入るようにするため）。
    assert not (_stub["dmd"] / ".rag_sig").is_file()


def test_sidecar_missing_fallback_no_drop_when_rag_es_disabled(_stub):
    (_stub["dmd"].parent / "ir" / "a.xlsx.evidence.json").unlink()
    marker_present_before = (_stub["dmd"] / ".rag_sig").is_file()
    worker.sync("w")
    assert (_stub["dmd"] / ".rag_sig").is_file() == marker_present_before   # RAG_ES無効なら触らない


def test_sidecar_missing_detected_even_without_version_drift(_stub):
    """バージョン定数は不変（drift 無し）のまま sidecar だけが外部要因で欠落したケースでも、
    drift 判定に先立って検知され全再構築へフォールバックする。"""
    assert office_md.rag_sig_drift(_stub["dmd"]) is False
    assert office_md.evidence_ir_sig_drift(_stub["dmd"]) is False
    (_stub["dmd"].parent / "ir" / "a.xlsx.evidence.json").unlink()          # markerはそのまま・sidecarだけ欠落
    res = worker.sync("w")
    assert _stub["calls"]["run"] == [{"reflect": True}]
    assert res["changed"] is True
