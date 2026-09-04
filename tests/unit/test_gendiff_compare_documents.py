"""GEN-DIFF（`compare_documents`）の受け入れ条件テスト。

正典: `docs/proposals/2026-09-03-世代間diff比較.md` §3〜§8。ツールは grep と同格の素朴な決定的
diff——レコード同定・業務キー対応付け・要約はしない（agentic loop の LLM が diff テキストを読んで
行う）。ここでは決定的な入出力契約（`sherpa/compare_docs.py`）と、`agentic_search.run_tool`/
`openai_tools`/`gemini_tools`/`mcp_server._tool_defs()` への配線を、実ファイル（`office_md.
build_derived()` の実往復）で固定する。

**openpyxl 再保存の偽差分に注意**（CONV-CACHE レーンの教訓）: 世代間で「意図的に変えた値」以外の
差分が紛れ込まないよう、変更しない側（4期）は `shutil.copyfile` でバイトコピーし、変更する側
（5期）だけを openpyxl で読み込み→1セルだけ書き換えて保存する。両側とも同じベースラインからの
派生のため、rag.md の差分は「原本パス／SHA-256／チャンクアンカー（コンテンツハッシュ由来のため
世代が違えば必ず変わる）／意図的に変更したセル」に限られる——アサーションは全文一致ではなく
期待する追加/削除行が含まれることだけを見る（アンカー行の変化はノイズとして許容する）。
"""
from __future__ import annotations

import os
import shutil
import tempfile
from pathlib import Path

os.environ.setdefault("SHERPA_USE_FIXTURES", "1")

import openpyxl
import pytest

from sherpa import agentic_search, compare_docs, mcp_server, worlds
from sherpa.ingest import evidence_render, office_md

_WORLD = "gendiff-test"
_XLSX_NAME = "代理店手数料機能設計書.xlsx"
_STABLE_NAME = "旧仕様書.xlsx"
_CBL_NAME = "PROGRAM.cbl"
_CBL_TEXT = "       IDENTIFICATION DIVISION.\n       PROGRAM-ID. PROGRAM.\n"


def _make_wb(value: str) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "対象"
    ws["A1"] = "項目"
    ws["B1"] = "手数料率区分"
    ws["A2"] = "商品A"
    ws["B2"] = value
    return wb


@pytest.fixture(scope="module")
def _world_dir():
    """4期/5期の2世代を持つ world を実際に build_derived() へ通し、`(wd, derived_root)` を返す。"""
    d = Path(tempfile.mkdtemp())
    wd = d / "world"
    (wd / "4期").mkdir(parents=True)
    (wd / "5期").mkdir(parents=True)

    baseline = d / "baseline.xlsx"
    _make_wb("固定率").save(baseline)
    shutil.copyfile(baseline, wd / "4期" / _XLSX_NAME)          # バイトコピー＝偽差分を作らない
    work = d / "work.xlsx"
    shutil.copyfile(baseline, work)
    wb2 = openpyxl.load_workbook(work)
    wb2["対象"]["B2"] = "段階制"                                 # 意図的な改変（唯一の実質差分）
    wb2.save(wd / "5期" / _XLSX_NAME)

    stable = d / "stable.xlsx"
    _make_wb("変更なし").save(stable)
    shutil.copyfile(stable, wd / "4期" / _STABLE_NAME)
    shutil.copyfile(stable, wd / "5期" / _STABLE_NAME)

    (wd / "4期" / _CBL_NAME).write_text(_CBL_TEXT, encoding="utf-8")
    (wd / "5期" / _CBL_NAME).write_text(_CBL_TEXT, encoding="utf-8")

    derived_root = d / "derived"
    os.environ["SHERPA_DERIVED_DIR"] = str(derived_root)
    rep = office_md.build_derived(wd, worlds.derived_md_dir(_WORLD))
    assert rep["rag_failed"] == 0, rep

    # 表現バージョン不一致を検証するため、4期側の旧仕様書.rag.md だけヘッダの
    # RAG_RENDERER_VERSION を古い値へ機械的に書き換える（実際の旧バージョンでの再変換は不要——
    # §3 step4 は「現行と一致しない側があれば」を機械的な文字列比較で判定するだけ）。
    stale_rag = worlds.derived_rag_dir(_WORLD) / "4期" / f"{_STABLE_NAME}.rag.md"
    text = stale_rag.read_text(encoding="utf-8")
    patched = text.replace(evidence_render.RAG_RENDERER_VERSION, "evidence-rag-renderer-v1alpha1")
    assert patched != text
    stale_rag.write_text(patched, encoding="utf-8")

    return wd, derived_root


@pytest.fixture()
def _pinned_world(_world_dir):
    """`worlds.world_dir(_WORLD)` を本テスト用の一時ソースへ固定するコンテキスト。"""
    wd, _derived_root = _world_dir
    with worlds.pin_world_root(_WORLD, wd):
        yield _WORLD


# ---- status: comparable（正解付き fixture・§7）----

def test_comparable_explicit_pair_has_expected_diff(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {
        "left_doc_id": f"4期/{_XLSX_NAME}", "right_doc_id": f"5期/{_XLSX_NAME}"})
    assert result["status"] == "comparable"
    diff = result["diff"]
    assert "-手数料率区分: 「固定率」" in diff
    assert "+手数料率区分: 「段階制」" in diff
    assert "--- 4期/" in diff and "+++ 5期/" in diff


def test_comparable_compare_conditions_carries_sha_and_version(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {
        "left_doc_id": f"4期/{_XLSX_NAME}", "right_doc_id": f"5期/{_XLSX_NAME}"})
    cc = result["compare_conditions"]
    assert cc["left"]["doc_id"] == f"4期/{_XLSX_NAME}"
    assert cc["right"]["doc_id"] == f"5期/{_XLSX_NAME}"
    assert cc["left"]["sha256"] and cc["right"]["sha256"]
    assert cc["left"]["sha256"] != cc["right"]["sha256"]     # 中身が違う＝原本ハッシュも違う
    assert cc["left"]["renderer_version"] == evidence_render.RAG_RENDERER_VERSION
    assert result["notices"] == []                           # 表現バージョンは両側とも最新＝注記なし


def test_comparable_auto_discovery_exact_suffix_match(_pinned_world):
    """§4 step2: 世代を除いた相対 suffix が完全一致すれば1件に決まる（曖昧解消なし）。"""
    world = _pinned_world
    result = compare_docs.compare(world, {
        "source_doc_id": f"4期/{_XLSX_NAME}", "target_generation": "5期"})
    assert result["status"] == "comparable"
    assert result["compare_conditions"]["right"]["doc_id"] == f"5期/{_XLSX_NAME}"


# ---- status: unsupported（rag.md を持たない文書・§7）----

def test_unsupported_when_rag_md_missing(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {
        "left_doc_id": f"4期/{_CBL_NAME}", "right_doc_id": f"5期/{_CBL_NAME}"})
    assert result["status"] == "unsupported"
    assert result["left_doc_id"] == f"4期/{_CBL_NAME}"
    assert result["right_doc_id"] == f"5期/{_CBL_NAME}"
    assert "reason" in result


# ---- status: needs_disambiguation（対応文書が決まらない・§7）----

def test_needs_disambiguation_zero_candidates(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {
        "source_doc_id": f"4期/{_XLSX_NAME}", "target_generation": "6期"})   # 存在しない世代
    assert result["status"] == "needs_disambiguation"
    assert result["candidates"] == []
    assert result["source_doc_id"] == f"4期/{_XLSX_NAME}"
    assert result["target_generation"] == "6期"


# ---- 表現バージョン不一致（§3 step4・停止せず注記のみ）----

def test_version_mismatch_notice_does_not_block_diff(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {
        "left_doc_id": f"4期/{_STABLE_NAME}", "right_doc_id": f"5期/{_STABLE_NAME}"})
    assert result["status"] == "comparable"                  # 停止しない
    assert any("表現バージョンが古い" in n and "4期" in n for n in result["notices"])
    assert "evidence-rag-renderer-v1alpha1" in result["notices"][0]


# ---- 引数不備／封じ込め ----

def test_missing_args_returns_error(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {})
    assert "error" in result


def test_traversal_doc_id_falls_back_to_unsupported(_pinned_world):
    """封じ込め拒否＝rag.md が見つからないのと同じ扱い（情報を漏らさない・fail-closed）。"""
    world = _pinned_world
    result = compare_docs.compare(world, {
        "left_doc_id": "../evil.xlsx", "right_doc_id": f"5期/{_XLSX_NAME}"})
    assert result["status"] == "unsupported"


def test_scope_violation_returns_error(_pinned_world):
    world = _pinned_world
    result = compare_docs.compare(world, {
        "left_doc_id": f"4期/{_XLSX_NAME}", "right_doc_id": f"5期/{_XLSX_NAME}"},
        scope_paths=["4期"])
    assert result == {"error": "指定 doc_id は対象範囲外です"}


# ---- run_tool 配線（docs/出典・予算クリップ）----

def test_run_tool_comparable_adds_both_doc_ids_to_sources(_pinned_world):
    world = _pinned_world
    result, docs, _cites, _cards = agentic_search.run_tool(
        "compare_documents",
        {"left_doc_id": f"4期/{_XLSX_NAME}", "right_doc_id": f"5期/{_XLSX_NAME}"},
        world, None)
    assert result["status"] == "comparable"
    assert docs == {f"4期/{_XLSX_NAME}", f"5期/{_XLSX_NAME}"}


def test_run_tool_clips_diff_to_budget_and_reports_truncated(_pinned_world):
    world = _pinned_world
    result, _docs, _cites, _cards = agentic_search.run_tool(
        "compare_documents",
        {"left_doc_id": f"4期/{_XLSX_NAME}", "right_doc_id": f"5期/{_XLSX_NAME}"},
        world, None, tool_result_max_bytes=64)
    assert result["status"] == "comparable"
    assert result["truncated"] is True
    assert len(result["diff"].encode("utf-8")) <= 64


# ---- ツール登録（§5・4+1箇所のうちの3箇所を確認）----

def test_registered_in_openai_tools():
    names = [t["function"]["name"] for t in agentic_search.openai_tools(with_es=True, with_graph=True)]
    assert "compare_documents" in names


def test_registered_in_gemini_tools():
    names = [f["name"] for f in agentic_search.gemini_tools(with_es=True, with_graph=True)[0]["functionDeclarations"]]
    assert "compare_documents" in names


def test_registered_in_mcp_tool_defs(monkeypatch):
    monkeypatch.delenv("SHERPA_MCP_ASK_DISABLED", raising=False)
    names = [d["name"] for d in mcp_server._tool_defs()]
    assert "compare_documents" in names


def test_system_prompt_mentions_compare_documents():
    assert "compare_documents" in agentic_search.SYSTEM
