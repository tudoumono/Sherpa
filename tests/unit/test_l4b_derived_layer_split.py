"""L4b（派生物のフォルダ三分割・§8.1）の受け入れ条件を実配置で固定する。

正典: `docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §8.1（三階層＋フォルダ分離）。

既存の大半の単体テストは `worlds.derived_md_dir`/`derived_rag_dir`/`derived_ir_dir` を monkeypatch
で個別に差し替えている（office_md.build_derived への直接パス指定・grep_tool/agentic_search/
es_index/corpus_docs 側のテストダブル）。ここでは `worlds.world_dir` だけを差し替え、
`derived_*_dir` 系は**実物**（`SHERPA_DERIVED_DIR` 環境変数で tmp へ向けるだけ）のまま
grep/read_around/ES パス解決/corpus_docs の各消費者を通しで駆動し、新配置（`md/`・`rag/`・`ir/`
の3ディレクトリ分離）を実際に踏む「実配置を通す統合的なテスト」（受け入れ条件）とする。
"""
from __future__ import annotations

import shutil
import zipfile
from pathlib import Path

import openpyxl

from sherpa import agentic_search, corpus_docs, es_index, grep_tool, worlds
from sherpa.ingest import office_md

_DOCX_XML = """<?xml version="1.0"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
 <w:body><w:p><w:r><w:t>DOCX_NEEDLE 本文</w:t></w:r></w:p></w:body>
</w:document>"""

_PPTX_SLIDE = """<?xml version="1.0"?>
<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"
       xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">
 <p:cSld><p:spTree><a:t>PPTX_NEEDLE スライド本文</a:t></p:spTree></p:cSld>
</p:sld>"""


def _zip(path: Path, entries: dict) -> None:
    with zipfile.ZipFile(path, "w") as z:
        for name, content in entries.items():
            z.writestr(name, content)


def _real_world(monkeypatch, tmp_path, world_id: str):
    """`worlds.world_dir` だけ差し替え、`derived_*_dir` 系は実物のまま `SHERPA_DERIVED_DIR` 配下を使う。"""
    source = tmp_path / "kb" / world_id
    source.mkdir(parents=True)
    monkeypatch.setattr(worlds, "world_dir", lambda w: source if w == world_id else None)
    monkeypatch.setenv("SHERPA_DERIVED_DIR", str(tmp_path / "derived"))
    return source


def _build(world_id: str, source: Path) -> dict:
    """xlsx/docx/pptx の実往復。原本ごとに検索語（NEEDLE）を1件ずつ埋め込む。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"], ws["B1"] = "No", "内容"
    ws["A2"], ws["B2"] = 1, "XLSX_NEEDLE サンプル内容"
    wb.save(source / "a.xlsx")
    _zip(source / "b.docx", {"word/document.xml": _DOCX_XML})
    _zip(source / "c.pptx", {"ppt/slides/slide1.xml": _PPTX_SLIDE})

    rep = office_md.build_derived(source, worlds.derived_md_dir(world_id), world_sig="sig1")
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0, rep
    return rep


# ---- 受け入れ条件1: 3層への物理振り分け・md/ に rag/ir が生成されない ----

def test_real_placement_three_layers_physically_separated(monkeypatch, tmp_path):
    """実往復（xlsx/docx/pptx）で md/rag/ir の3層に正しく振り分けられ、`md/` に rag/ir 系ファイルが
    生成されないことを固定する。"""
    world_id = "l4b-split-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    _build(world_id, source)

    dmd = worlds.derived_md_dir(world_id)
    drag = worlds.derived_rag_dir(world_id)
    dir_ = worlds.derived_ir_dir(world_id)
    assert dmd.is_dir() and drag.is_dir() and dir_.is_dir()
    # md/rag/ir は derived root の兄弟（世界削除の一括伝播が3層とも消せる構造）。
    assert dmd.parent == drag.parent == dir_.parent == worlds.derived_dir(world_id)

    for rel in ("a.xlsx", "b.docx", "c.pptx"):
        assert (dmd / f"{rel}.md").is_file(), rel
        assert (dmd / f"{rel}.md.meta.json").is_file(), rel
        assert (drag / f"{rel}.rag.md").is_file(), rel
        assert (drag / f"{rel}.rag_chunks.jsonl").is_file(), rel
        assert (dir_ / f"{rel}.evidence.json").is_file(), rel
        assert (dir_ / f"{rel}.document.json").is_file(), rel
        assert (dir_ / f"{rel}.derived.json").is_file(), rel

    # md/ に rag/ir 系ファイルが生成されない（受け入れ条件の直接固定）。
    md_names = {p.name for p in dmd.rglob("*") if p.is_file()}
    rag_ir_suffixes = (".rag.md", ".rag_chunks.jsonl", ".evidence.json",
                      ".document.json", ".derived.json", ".ocr_route.json")
    assert not any(n.endswith(rag_ir_suffixes) for n in md_names), md_names
    # rag/ に md 専用・ir 専用ファイルが無い（legacy `.md` はあっても `.rag.md` ではない、で判定）。
    rag_names = {p.name for p in drag.rglob("*") if p.is_file()}
    md_ir_suffixes = (".md.meta.json", ".evidence.json", ".document.json",
                      ".derived.json", ".ocr_route.json")
    assert not any(n.endswith(md_ir_suffixes) for n in rag_names), rag_names
    assert not any(n.endswith(".md") and not n.endswith(".rag.md") for n in rag_names), rag_names
    # ir/ に rag 専用ファイルが無い。
    ir_names = {p.name for p in dir_.rglob("*") if p.is_file()}
    assert not any(n.endswith((".rag.md", ".rag_chunks.jsonl")) for n in ir_names), ir_names


# ---- 受け入れ条件2: grep（rag 優先／legacy フォールバックの両方）----

def test_real_placement_grep_search_rag_priority_and_legacy_fallback(monkeypatch, tmp_path):
    """grep_search は実配置（`worlds.derived_rag_dir`/`derived_md_dir` を monkeypatch しない）でも
    rag 優先・legacy フォールバックのどちらでも正しく1件だけヒットする（legacy 側は
    `grep_tool.rag_grep_enabled` を直接差し替えて模擬——TOGGLE-RM・2026-09-03 でグローバルな
    系統切替トグルは撤去済みのため env では OFF にできない）。"""
    world_id = "l4b-grep-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    _build(world_id, source)

    hits = grep_tool.grep_search("XLSX_NEEDLE", world=world_id)
    assert len(hits) == 1 and hits[0]["doc_id"] == "a.xlsx"

    monkeypatch.setattr(grep_tool, "rag_grep_enabled", lambda: False)
    hits_legacy = grep_tool.grep_search("XLSX_NEEDLE", world=world_id)
    assert len(hits_legacy) == 1 and hits_legacy[0]["doc_id"] == "a.xlsx"


# ---- 受け入れ条件3: read_doc/read_around（agentic_search）----

def test_real_placement_read_around_resolves_through_new_layout(monkeypatch, tmp_path):
    """`agentic_search.run_tool("read_around", ...)` が実配置（rag/md 別ディレクトリ）でも
    grep のヒットと同じ文書を精読できる。"""
    world_id = "l4b-read-around-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    _build(world_id, source)

    hits = grep_tool.grep_search("DOCX_NEEDLE", world=world_id)
    assert len(hits) == 1
    hit = hits[0]

    res, docs, _, _ = agentic_search.run_tool(
        "read_around", {"doc_id": hit["doc_id"], "line": hit["line"], "window": 2}, world_id, None)
    assert "error" not in res, res
    assert "DOCX_NEEDLE" in res["text"]
    assert "b.docx" in docs


# ---- 受け入れ条件4: ES のパス解決（rag 層）----

def test_real_placement_es_path_resolution_uses_rag_layer(monkeypatch, tmp_path):
    """`es_index.index_world` が実際に使う `_safe_rag_md_path`/`_safe_rag_chunks_path` を、
    monkeypatch していない実物の `worlds.derived_rag_dir` で呼んでも安全に解決できる。"""
    world_id = "l4b-es-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    _build(world_id, source)

    drag = worlds.derived_rag_dir(world_id)
    rag_path, reason = es_index._safe_rag_md_path(drag, "a.xlsx")
    assert reason is None and rag_path is not None and rag_path.is_file()
    chunks_path, reason2 = es_index._safe_rag_chunks_path(drag, "a.xlsx")
    assert reason2 is None and chunks_path is not None and chunks_path.is_file()


# ---- 受け入れ条件5: 台帳／グラフ抽出が読む corpus_docs の解決 ----

def test_real_placement_corpus_docs_include_rag_prefers_rag_layer(monkeypatch, tmp_path):
    """`corpus_docs.world_documents(include_rag=True)`（ES/グラフL層抽出が使う）の `md_path` は
    実物の `derived_rag_dir` を指し、`include_rag=False`（既定・台帳等）は `derived_md_dir` を指す。"""
    world_id = "l4b-corpus-docs-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    _build(world_id, source)

    rag_docs = {d["name"]: d for d in corpus_docs.world_documents(world_id, include_rag=True)}
    assert rag_docs["a.xlsx"]["md_path"] == str(worlds.derived_rag_dir(world_id) / "a.xlsx.rag.md")

    legacy_docs = {d["name"]: d for d in corpus_docs.world_documents(world_id)}
    assert legacy_docs["a.xlsx"]["md_path"] == str(worlds.derived_md_dir(world_id) / "a.xlsx.md")


# ---- 受け入れ条件6: 旧配置の残骸があっても新コードが壊れない ----

def test_old_layout_residue_does_not_crash_new_code(monkeypatch, tmp_path):
    """旧配置（`.rag.md` が `md/` に同居・`rag/`/`ir/` 自体が存在しない）の残骸があっても、新コードは
    例外を出さず「未生成」として扱う。`rag_sidecars_missing` が True を返す＝次回 sync の全再構築で
    新配置へ収束する自己修復の入口になる。"""
    world_id = "l4b-legacy-residue-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    (source / "old.docx").write_bytes(b"legacy binary stub")

    dmd = worlds.derived_md_dir(world_id)
    dmd.mkdir(parents=True)
    (dmd / "old.docx.md").write_text("legacy content NEEDLE_OLD", encoding="utf-8")
    (dmd / "old.docx.md.meta.json").write_text('{"arm":"ooxml"}', encoding="utf-8")
    (dmd / "old.docx.rag.md").write_text("## h\nrag content NEEDLE_OLD\n", encoding="utf-8")

    # rag/ir が存在しない旧世代でも例外を出さず True（欠落＝自己修復の全再構築対象）。
    assert office_md.rag_sidecars_missing(source, dmd) is True

    # grep はクラッシュせず、rag/ が無い（旧配置の残骸を新コードは探さない）ので legacy 側だけを
    # 1件ヒットする（旧配置の .rag.md 残骸が二重ヒットを作らない）。
    hits = grep_tool.grep_search("NEEDLE_OLD", world=world_id)
    assert len(hits) == 1
    assert hits[0]["doc_id"] == "old.docx"
    assert "legacy content" in hits[0]["text"]

    resolved = agentic_search._safe_doc_path(world_id, "old.docx")
    assert resolved is not None
    _root, lexical_rel, _rp = resolved
    assert lexical_rel == "old.docx.md"


# ---- 受け入れ条件7: 削除伝播が3層すべてを消す ----

def test_deletion_propagation_removes_all_three_layers(monkeypatch, tmp_path):
    """`shutil.rmtree(worlds.derived_dir(world))`（`worker._wipe_locked`／`worlds.register` の補償
    処理と同じ操作）が md/rag/ir の3層とも一括で消す（world 削除時の一括伝播）。"""
    world_id = "l4b-delete-world"
    source = _real_world(monkeypatch, tmp_path, world_id)
    _build(world_id, source)

    assert worlds.derived_md_dir(world_id).is_dir()
    assert worlds.derived_rag_dir(world_id).is_dir()
    assert worlds.derived_ir_dir(world_id).is_dir()

    shutil.rmtree(worlds.derived_dir(world_id), ignore_errors=True)

    assert not worlds.derived_md_dir(world_id).exists()
    assert not worlds.derived_rag_dir(world_id).exists()
    assert not worlds.derived_ir_dir(world_id).exists()
