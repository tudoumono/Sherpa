"""`excel_display.extract_cell_metadata` の出力固定テスト（MEM-1・2重フルロード是正のゴールデン）。

`extract_cell_metadata` は元々 `wb_formula`（data_only=False）と `wb_values`（data_only=True）を
**同時に**フルロードしていた（ピークメモリ2倍）。逐次化（`wb_formula` を読み終えてから閉じ、
その後 `wb_values` を開く）に是正するにあたり、出力が byte-identical であることをこのテストで
固定する。期待値（`_EXPECTED_JSON`）は**是正前のコード**を実際に実行して得た実測値
（`metadata_json()` の決定的直列化）であり、推測値ではない。

対象セル: formula（cache 有/無）・日付書式・素の整数・文字列・空セル（対象だが未使用の座標）。
"""
from __future__ import annotations

import weakref
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

from sherpa.ingest import excel_display


def _build_fixture_xlsx(dst: Path) -> Path:
    """formula(cache有/無)・日付書式・整数・文字列・空セルを含む最小 xlsx を組み立てる。

    openpyxl が自ら書いた formula には cached value（`<v>`）が付かないため、A3 は
    自然に「cache無し」になる。A4 は保存後に xlsx（zip）を直接いじって `<f>` の直後に
    `<v>84</v>` を差し込み、「cache有り」を作る（Excel で開いて再保存したのと同じ形）。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = 42
    ws["A2"] = "hello"
    ws["A3"] = "=A1+1"          # formula・cache無し
    ws["A4"] = "=A1*2"          # formula・cache有り（後段でXML注入）
    ws["A5"] = datetime(2024, 1, 15)
    ws["A5"].number_format = "yyyy-mm-dd"
    # A6 は触らない（targets には含めるが未使用セル＝空セルの経路を確認する）
    wb.save(dst)

    with zipfile.ZipFile(dst) as zin:
        names = zin.namelist()
        sheet_name = next(n for n in names if n.startswith("xl/worksheets/sheet") and n.endswith(".xml"))
        data = {n: zin.read(n) for n in names}

    xml = data[sheet_name].decode("utf-8")
    idx = xml.index('<c r="A4"')
    end_idx = xml.index("</c>", idx) + len("</c>")
    cell_xml = xml[idx:end_idx]
    f_close = cell_xml.index("</f>") + len("</f>")
    new_cell_xml = cell_xml[:f_close] + "<v>84</v>" + cell_xml[f_close:]
    data[sheet_name] = (xml[:idx] + new_cell_xml + xml[end_idx:]).encode("utf-8")

    with zipfile.ZipFile(dst, "w", zipfile.ZIP_DEFLATED) as zout:
        for n in names:
            zout.writestr(n, data[n])
    return dst


# 是正前のコード（同時2重フルロード版）を実際に実行して得た実測出力（推測値ではない）。
_EXPECTED_JSON = (
    '{"Sheet1!A1":{"cached_value":null,"calculation_status":"not_formula","display_reason":null,'
    '"display_source":"raw_value","display_status":"rendered","display_value":"42","formula":null,'
    '"number_format":"General","raw_value":42,"style_id":0,"typed_value":{"type":"integer","value":42}},'
    '"Sheet1!A2":{"cached_value":null,"calculation_status":"not_formula","display_reason":null,'
    '"display_source":"raw_value","display_status":"rendered","display_value":"hello","formula":null,'
    '"number_format":"General","raw_value":"hello","style_id":0,"typed_value":{"type":"string","value":"hello"}},'
    '"Sheet1!A3":{"cached_value":null,"calculation_status":"cache_missing","display_reason":"formula_cache_missing",'
    '"display_source":"excel-display-linux-v1","display_status":"unsupported","display_value":null,'
    '"formula":"=A1+1","number_format":"General","raw_value":null,"style_id":0,'
    '"typed_value":{"type":"blank","value":null}},'
    '"Sheet1!A4":{"cached_value":84,"calculation_status":"cached","display_reason":null,'
    '"display_source":"raw_value","display_status":"rendered","display_value":"84","formula":"=A1*2",'
    '"number_format":"General","raw_value":84,"style_id":0,"typed_value":{"type":"integer","value":84}},'
    '"Sheet1!A5":{"cached_value":null,"calculation_status":"not_formula","display_reason":null,'
    '"display_source":"excel-display-linux-v1","display_status":"rendered","display_value":"2024-01-15",'
    '"formula":null,"number_format":"yyyy-mm-dd","raw_value":45306,"style_id":2,'
    '"typed_value":{"type":"datetime","value":"2024-01-15T00:00:00"}},'
    '"Sheet1!A6":{"cached_value":null,"calculation_status":"not_formula","display_reason":null,'
    '"display_source":"raw_value","display_status":"rendered","display_value":"","formula":null,'
    '"number_format":"General","raw_value":null,"style_id":0,"typed_value":{"type":"blank","value":null}}}'
)


def test_extract_cell_metadata_matches_pre_refactor_golden(tmp_path):
    """逐次化リファクタ後も `extract_cell_metadata` の出力が是正前と完全一致すること。"""
    fixture = _build_fixture_xlsx(tmp_path / "fixture.xlsx")
    targets = {"Sheet1": {"A1", "A2", "A3", "A4", "A5", "A6"}}
    metadata = excel_display.extract_cell_metadata(fixture, targets)
    keyed = {f"{sheet}!{coord}": v for (sheet, coord), v in metadata.items()}
    assert excel_display.metadata_json(keyed) == _EXPECTED_JSON


def test_extract_cell_metadata_skips_sheet_missing_from_either_workbook(tmp_path):
    """`targets` に無いシート名を渡しても例外にならず空で返る（シート欠落チェックの契約）。"""
    fixture = _build_fixture_xlsx(tmp_path / "fixture2.xlsx")
    metadata = excel_display.extract_cell_metadata(fixture, {"NoSuchSheet": {"A1"}})
    assert metadata == {}


def test_extract_cell_metadata_frees_formula_workbook_before_second_pass(tmp_path, monkeypatch):
    """RV是正#1: `wb_formula.close()` だけでは openpyxl 内部の循環参照ゆえセル木が即解放されない
    前提のため、第2パス（`wb_values`）を開く**前**に第1パスのブックへの最後の参照を明示的に切り
    （`del`）、GCを1回回す（`gc.collect()`）。tracemalloc 等のヒープ実測はプロセス全体の割当状況に
    左右されflakyなため、より決定的な検証として「第2パスの `load_workbook` 呼び出し時点で第1パスの
    ブックが既に GC 到達不能（weakref 死亡）になっている」ことを直接確認する——ピーク時に生きている
    ブックが常に1冊であることの直接証拠になる。"""
    fixture = _build_fixture_xlsx(tmp_path / "fixture3.xlsx")
    orig_load_workbook = openpyxl.load_workbook
    state = {"calls": 0, "formula_wb_ref": None}

    def wrapped_load_workbook(*args, **kwargs):
        state["calls"] += 1
        if state["calls"] == 2:
            # 第2パスをロードする直前＝第1パスの del+gc.collect() が既に実行されているはず。
            assert state["formula_wb_ref"] is not None
            assert state["formula_wb_ref"]() is None, "第1パスのwb_formulaが第2パス開始時点でまだ生存"
        wb = orig_load_workbook(*args, **kwargs)
        if state["calls"] == 1:
            state["formula_wb_ref"] = weakref.ref(wb)
        return wb

    monkeypatch.setattr(openpyxl, "load_workbook", wrapped_load_workbook)
    metadata = excel_display.extract_cell_metadata(fixture, {"Sheet1": {"A1"}})
    assert state["calls"] == 2                       # 第1パス・第2パスの2回ロードされたことの確認
    assert metadata[("Sheet1", "A1")]["raw_value"] == 42
