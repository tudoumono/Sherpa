"""CONV-CACHE（per-file 変換結果の再ビルド跨ぎキャッシュ）の受け入れ条件を実往復で固定する。

背景: 実環境（10,000ファイル・1件30秒級）で取り込みの変換段（office_md）が完走できず、再実行が
毎回0から始まる障害への対応（正典 `docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §9・
CONV-CACHE 行）。`_build_derived_into_staging` は per-file ループの各 rel を
`(原本の resolved path, st_size, st_mtime_ns, 変換パイプライン署名)` キーでキャッシュし、ヒット時は
実変換（①アーム実行・Evidence/RAG 生成）を丸ごとスキップしてステージングへコピーするだけで済ませる。

ここでは `office_md.build_derived` を実際の xlsx フィクスチャで駆動し、`office_md._convert_with_arms`
（実変換1回＝1呼び出し）の呼び出し回数を数えることで「実変換をスキップできたか」を直接固定する。
"""
from __future__ import annotations

import shutil
import time
from pathlib import Path

import openpyxl
import pytest

from sherpa import worlds
from sherpa.ingest import office_md


def _new_world(tmp_path, monkeypatch, world_id: str) -> Path:
    """`SHERPA_DERIVED_DIR` を tmp へ向け、`world_id` 用の原本ディレクトリを用意する。

    `office_md.build_derived` は原本パスを直接受け取るため（`worlds.world_dir` 経由の解決を挟まない）、
    ここでは `derived_dir` が読む env だけ差し替えれば十分（`test_l4b_derived_layer_split.py` の
    `_real_world` と同じ発想だが、複数 world_id を同じ tmp_path で扱えるよう `world_dir` monkeypatch は
    省く）。
    """
    source = tmp_path / "kb" / world_id
    source.mkdir(parents=True)
    monkeypatch.setenv("SHERPA_DERIVED_DIR", str(tmp_path / "derived"))
    return source


def _write_fixture_files(source: Path, names: list[str]) -> None:
    """各 `name`（`*.xlsx`）に、ファイル名由来の一意な NEEDLE を1件ずつ埋め込む。"""
    for name in names:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"], ws["B1"] = "No", "内容"
        ws["A2"], ws["B2"] = 1, f"NEEDLE_{name}"
        wb.save(source / name)


def _snapshot_layers(world_id: str) -> dict:
    """world の md/rag/ir 3層（派生本体のみ・`_conv_cache`/`_legacy_cache` は含めない）を
    `{"md/a.xlsx.md": bytes, ...}` の形で丸ごと読む（byte-identical 比較用）。"""
    out: dict[str, bytes] = {}
    for layer, layer_dir in (
        ("md", worlds.derived_md_dir(world_id)),
        ("rag", worlds.derived_rag_dir(world_id)),
        ("ir", worlds.derived_ir_dir(world_id)),
    ):
        if not layer_dir.is_dir():
            continue
        for p in sorted(layer_dir.rglob("*")):
            if p.is_file():
                out[f"{layer}/{p.relative_to(layer_dir).as_posix()}"] = p.read_bytes()
    return out


def _wrap_convert_with_arms(monkeypatch):
    """`office_md._convert_with_arms`（実変換の入口・キャッシュヒット時は一切呼ばれない）を
    呼び出し記録付きに差し替える。素の実装は保持したまま呼ぶ（副作用は変えない）。"""
    calls: list[str] = []
    orig = office_md._convert_with_arms

    def counting(path, enabled):
        calls.append(Path(path).name)
        return orig(path, enabled)

    monkeypatch.setattr(office_md, "_convert_with_arms", counting)
    return calls


def test_cache_hit_byte_identical_and_rep_equivalent(monkeypatch, tmp_path):
    """2回目の build は実変換ゼロ（全件キャッシュヒット）で、派生3層・rep 集計とも1回目と完全一致する。"""
    world_id = "convcache-identical"
    source = _new_world(tmp_path, monkeypatch, world_id)
    _write_fixture_files(source, ["a.xlsx", "b.xlsx", "c.xlsx"])
    dmd = worlds.derived_md_dir(world_id)
    calls = _wrap_convert_with_arms(monkeypatch)

    rep1 = office_md.build_derived(source, dmd, world_sig="s1")
    assert rep1.get("error") is None, rep1
    assert sorted(calls) == ["a.xlsx", "b.xlsx", "c.xlsx"]     # 1回目は全件 実変換
    tree1 = _snapshot_layers(world_id)

    calls.clear()
    rep2 = office_md.build_derived(source, dmd, world_sig="s1")
    assert rep2.get("error") is None, rep2
    assert calls == []                                         # 2回目は実変換ゼロ（全件キャッシュヒット）
    tree2 = _snapshot_layers(world_id)

    assert tree1 == tree2                                       # 派生3層が byte-identical
    assert rep1 == rep2                                          # rep 集計（converted/document_ir_*等）も一致


def test_source_change_invalidates_only_that_file(monkeypatch, tmp_path):
    """原本 mtime/size が変わった1件だけキャッシュミスし、他は実変換されない。"""
    world_id = "convcache-source-change"
    source = _new_world(tmp_path, monkeypatch, world_id)
    _write_fixture_files(source, ["a.xlsx", "b.xlsx", "c.xlsx"])
    dmd = worlds.derived_md_dir(world_id)
    calls = _wrap_convert_with_arms(monkeypatch)

    office_md.build_derived(source, dmd, world_sig="s1")
    calls.clear()

    time.sleep(0.01)                        # st_mtime_ns を確実にずらす
    wb = openpyxl.Workbook()
    wb.active["A1"] = "CHANGED_NEEDLE"
    wb.save(source / "b.xlsx")

    office_md.build_derived(source, dmd, world_sig="s1")
    assert calls == ["b.xlsx"]               # a.xlsx/c.xlsx はキャッシュヒットで実変換されない


def test_pipeline_sig_change_invalidates_all_files(monkeypatch, tmp_path):
    """変換パイプライン署名（既存の各層版マーカーを束ねたもの）が変われば原本不変でも全件ミスする
    （現行の drift 全再ビルドと同じ挙動）。"""
    world_id = "convcache-sig-change"
    source = _new_world(tmp_path, monkeypatch, world_id)
    _write_fixture_files(source, ["a.xlsx", "b.xlsx"])
    dmd = worlds.derived_md_dir(world_id)
    calls = _wrap_convert_with_arms(monkeypatch)

    office_md.build_derived(source, dmd, world_sig="s1")
    calls.clear()

    # レンダラ/抽出器の版が上がった、というシナリオを模擬（既存の署名部品の1つだけを差し替える）。
    monkeypatch.setattr(office_md, "_current_human_md_sig", lambda: "renderer=test-bumped")
    office_md.build_derived(source, dmd, world_sig="s1")
    assert sorted(calls) == ["a.xlsx", "b.xlsx"]


def test_deleted_source_pruned_after_completed_build(monkeypatch, tmp_path):
    """完走した build の後、原本一覧に無くなった rel のキャッシュが剪定される。"""
    world_id = "convcache-prune"
    source = _new_world(tmp_path, monkeypatch, world_id)
    _write_fixture_files(source, ["a.xlsx", "b.xlsx"])
    dmd = worlds.derived_md_dir(world_id)

    office_md.build_derived(source, dmd, world_sig="s1")
    cache_root = office_md._conv_cache_root_for(dmd)
    assert (cache_root / "a.xlsx.key.json").is_file()
    assert (cache_root / "b.xlsx.key.json").is_file()

    (source / "b.xlsx").unlink()
    office_md.build_derived(source, dmd, world_sig="s1")

    assert (cache_root / "a.xlsx.key.json").is_file()            # 生きている rel は残る
    assert not (cache_root / "b.xlsx.key.json").exists()          # 消えた原本のキャッシュは剪定される
    assert not (cache_root / "b.xlsx.d").exists()


def test_mid_death_resume_reuses_cached_files_and_matches_full_rebuild(monkeypatch, tmp_path):
    """N件目で（プロセス強制終了相当の）例外→再実行で「先頭 N-1 件はキャッシュヒット・残りだけ実変換」、
    最終出力は一括実変換（比較用の別 world）と同一になる。

    プロセスの実強制終了はテストで直接再現できないため、1ファイル分の処理を包む
    `try: ... except OSError/Exception ...` に捕まらない `BaseException` を注入して代用する
    （`_build_derived_into_staging` の for ループ・`build_derived` の外側 `except BaseException` まで
    素通りし、`try/except Exception` による「1件失敗しても他は継続」という既存の fail-safe とは別の
    「呼び出し自体が完走しない」状況を作る）。
    """
    world_id = "convcache-resume"
    ref_world_id = "convcache-resume-ref"
    source = _new_world(tmp_path, monkeypatch, world_id)
    _write_fixture_files(source, ["a.xlsx", "b.xlsx", "c.xlsx", "d.xlsx"])
    dmd = worlds.derived_md_dir(world_id)

    class _SimulatedCrash(BaseException):
        pass

    orig = office_md._convert_with_arms
    crash_on = {"c.xlsx"}
    calls: list[str] = []

    def flaky(path, enabled):
        name = Path(path).name
        calls.append(name)
        if name in crash_on:
            raise _SimulatedCrash("simulated mid-build process death")
        return orig(path, enabled)

    monkeypatch.setattr(office_md, "_convert_with_arms", flaky)

    with pytest.raises(_SimulatedCrash):
        office_md.build_derived(source, dmd, world_sig="s1")
    assert calls == ["a.xlsx", "b.xlsx", "c.xlsx"]                # d.xlsx にはまだ到達していない

    cache_root = office_md._conv_cache_root_for(dmd)
    assert (cache_root / "a.xlsx.key.json").is_file()
    assert (cache_root / "b.xlsx.key.json").is_file()
    assert not (cache_root / "c.xlsx.key.json").exists()          # 例外を投げた本人はキャッシュされない
    assert not (cache_root / "d.xlsx.key.json").exists()

    crash_on.clear()                                              # 「再実行」＝原因が解消した状態を模擬
    calls.clear()
    rep = office_md.build_derived(source, dmd, world_sig="s1")
    assert rep.get("error") is None, rep
    assert calls == ["c.xlsx", "d.xlsx"]                          # a.xlsx/b.xlsx はキャッシュヒットで再変換されない

    monkeypatch.setattr(office_md, "_convert_with_arms", orig)    # 比較用の参照 build は素の実装で
    ref_source = _new_world(tmp_path, monkeypatch, ref_world_id)
    # `_write_fixture_files` を再度呼ぶと openpyxl が xlsx 内部（`docProps/core.xml` の作成日時）に
    # 保存時刻を埋め込むため、論理内容が同じでも原本バイト列（＝`source_content_hash`）がずれて
    # 比較が成立しない。原本ファイルをそのまま**バイトコピー**して同一性を保つ。
    for name in ("a.xlsx", "b.xlsx", "c.xlsx", "d.xlsx"):
        shutil.copyfile(source / name, ref_source / name)
    office_md.build_derived(ref_source, worlds.derived_md_dir(ref_world_id), world_sig="s1")

    assert _snapshot_layers(world_id) == _snapshot_layers(ref_world_id)   # 最終出力は一括実変換と同一
