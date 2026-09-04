"""`evidence_render.py` の key-value 直列化（RAG-KV-001）の文字列契約を pin する。

「〈ヘッダ〉は「値」である。」という自然文テンプレートを `〈ヘッダ〉: 「値」` の key-value 形式へ
変更した。この提案が対象とするテンプレート箇所（`_field_piece`／`_excel_value_note`／
`_context_summary_record`／`_context_prefix`／`_non_table_records`の長文分割／`_element_piece`
の可視性・状態ブロック）はどれも決定的な文字列組み立てであり、Evidence IR/Context IR の
バリデーションを経由しない直接呼び出しで契約を検証できる（`test_evidence_render_aliases.py`
と同じ流儀）。group repack・chunk リンク・バージョン確認は `office_md.build_derived()` を
使った実xlsx/docxの往復（DB非依存）で検証する。
"""
from __future__ import annotations

import json
import os
import re
import tempfile
import zipfile
from pathlib import Path

os.environ.setdefault("SHERPA_USE_FIXTURES", "1")

import openpyxl

from sherpa.ingest import context_ir as CIR
from sherpa.ingest import evidence_ir as IR
from sherpa.ingest import evidence_render as R
from sherpa.ingest import office_md


def _loc(**kw) -> IR.Locator:
    kw.setdefault("part", "xl/worksheets/sheet1.xml")
    return IR.Locator(**kw)


def _value_text(value) -> str:
    return R._value_text(value)


def _cell(eid: str, value, *, row: int, column: int, sheet: str = "明細", **extension) -> IR.EvidenceElement:
    return IR.EvidenceElement(
        element_id=eid, type="cell", parent_id=None, order=row * 100 + column, value=value,
        locator=_loc(sheet=sheet, extension={"row": row, "column": column}),
        coverage_id=f"cov:{eid}", extension=extension)


# ---- 1〜4: _field_piece ----

def test_field_piece_single_span_kv():
    cell = _cell("e1", "2026-09-01", row=2, column=1)
    piece = R._field_piece(cell, "納期", 0, 1, 0, len("2026-09-01"), "2026-09-01", [])
    assert piece["semantic"] == "納期: 「2026-09-01」"
    assert piece["markdown"] == "納期: 「2026-09-01」"


def test_field_piece_multi_span_kv_no_quotes():
    cell = _cell("e1", "長文本体", row=2, column=1)
    piece = R._field_piece(cell, "備考", 0, 2, 0, 4, "長文本体", [])
    assert piece["semantic"] == "備考（1/2）:\n長文本体"
    assert "「" not in piece["semantic"] and "」" not in piece["semantic"]


def test_field_piece_layout_note_continues_and_separate():
    cell = _cell("e1", "値", row=2, column=1)
    continues = R._field_piece(cell, "項目", 0, 1, 0, 1, "値", [{"type": "continues"}])
    separate = R._field_piece(cell, "項目", 0, 1, 0, 1, "値", [{"type": "separate"}])
    assert continues["semantic"] == "項目: 「値」（同じ縦結合ラベル内の前後記載と連続）"
    assert separate["semantic"] == "項目: 「値」（隣接行と左ラベルが異なる別項目）"


def test_field_piece_order_layout_before_excel_note():
    cell = _cell("e1", "45900", row=2, column=1, raw_value="45900", number_format="yyyy/mm/dd",
                 display_status="rendered", display_value="2025/09/10")
    piece = R._field_piece(cell, "期限", 0, 1, 0, 5, "45900", [{"type": "continues"}])
    layout_idx = piece["semantic"].index("（同じ")
    excel_idx = piece["semantic"].index("Excel原値")
    assert piece["semantic"].startswith("期限: 「45900」")
    assert layout_idx < excel_idx                 # field行→layout注記→Excelノートの順（§9.1-2）


# ---- 5: _context_summary_record ----

def _region(**kw) -> CIR.ContextRegion:
    kw.setdefault("region_id", "r1")
    kw.setdefault("table_id", "t1")
    kw.setdefault("sheet", "明細")
    kw.setdefault("start_column", 1)
    kw.setdefault("end_column", 3)
    kw.setdefault("start_row", 1)
    kw.setdefault("end_row", 1)
    kw.setdefault("title", None)
    kw.setdefault("header_row", 1)
    kw.setdefault("header_paths", ())
    kw.setdefault("mode", "grid")
    kw.setdefault("confidence", 1.0)
    return CIR.ContextRegion(**kw)


def _ctx(source_name="API詳細設計_顧客照会.xlsx", **kw) -> CIR.ContextIR:
    kw.setdefault("document_titles", {})
    return CIR.ContextIR(
        schema_version=CIR.CONTEXT_IR_SCHEMA_VERSION, analyzer_version="test",
        source_hash="sha256:" + "0" * 64, source_name=source_name, **kw)


def _ir(source_hash="sha256:" + "0" * 64, file_type="xlsx", elements=None) -> IR.EvidenceIR:
    return IR.EvidenceIR(
        schema_version=IR.EVIDENCE_IR_SCHEMA_VERSION, parser_profile="test",
        source=IR.EvidenceSource(file_type=file_type, content_hash=source_hash),
        elements=elements or [])


def test_context_summary_record_three_labels():
    region = _region(start_row=1, end_row=1, header_row=3)
    cells = [
        _cell("h1", "No", row=1, column=1),          # row == start_row → 領域見出し
        _cell("h2", "業務名", row=3, column=1),        # row == header_row → 列見出し
        _cell("h3", "備考欄", row=2, column=1),        # start_row と header_row の間 → 領域情報
    ]
    ir = _ir(elements=cells)
    ctx = _ctx()
    record = R._context_summary_record(ir, ctx, region, cells, None, {})
    texts = record["semantic_text"].splitlines()[1:]      # 先頭は出所行
    assert "領域見出し: 「No」" in texts
    assert "列見出し: 「業務名」" in texts
    assert "領域情報: 「備考欄」" in texts


# ---- 6〜8: _context_prefix ----

def test_context_prefix_section_level():
    region = _region()
    ctx = _ctx()
    assert R._context_prefix(ctx, region, None) == "出所: 原本「API詳細設計_顧客照会.xlsx」 / シート「明細」"


def test_context_prefix_record_level_single_key():
    region = _region()
    ctx = _ctx()
    key = CIR.RecordKey(label="No", value="1", evidence_id="e1")
    record = CIR.ContextRecord(record_id="cr1", region_id="r1", row=2, keys=(key,),
                                identifiers=(), identifier_mentions=(), confidence=1.0)
    prefix = R._context_prefix(ctx, region, record)
    assert prefix == "出所: 原本「API詳細設計_顧客照会.xlsx」 / シート「明細」 / No「1」"


def test_context_prefix_record_level_multi_key():
    region = _region()
    ctx = _ctx()
    keys = (
        CIR.RecordKey(label="No", value="1", evidence_id="e1"),
        CIR.RecordKey(label="区分 > 明細番号", value="A-1", evidence_id="e2"),
    )
    record = CIR.ContextRecord(record_id="cr1", region_id="r1", row=2, keys=keys,
                                identifiers=(), identifier_mentions=(), confidence=1.0)
    prefix = R._context_prefix(ctx, region, record)
    assert prefix == ("出所: 原本「API詳細設計_顧客照会.xlsx」 / シート「明細」"
                       " / No「1」、明細番号「A-1」")


def test_context_prefix_includes_section_region_and_common_header():
    """§2.2 の構成要素のうち §2 テストで未カバーだった「節」「領域」「区分」を確認する
    （原本／文書／シート・スライド・ページ／record key は既存テストでカバー済み）。"""
    region = _region(
        title="明細領域", section_path=("第1章", "明細"),
    )
    ctx = _ctx()
    prefix = R._context_prefix(ctx, region, None, common_header=("共通区分",))
    assert prefix == (
        "出所: 原本「API詳細設計_顧客照会.xlsx」 / シート「明細」"
        " / 節「第1章」 / 節「明細」 / 領域「明細領域」 / 区分「共通区分」"
    )


# ---- 9: 非表要素の長文分割（出所前置は維持） ----

def test_non_table_long_text_split_keeps_prefix():
    long_text = "あ" * 1500
    element = IR.EvidenceElement(
        element_id="e1", type="paragraph", parent_id=None, order=0, value=long_text,
        locator=IR.Locator(part="word/document.xml"), coverage_id="c1")
    ir = _ir(file_type="docx", elements=[element])
    ctx = _ctx(source_name="doc.docx")
    records = R._non_table_records(ir, ctx, {}, set())
    assert len(records) == 2
    assert records[0]["semantic_text"].startswith(
        "原本「doc.docx」の文書「doc」にあるparagraph内容（1/2）:\n")
    assert records[1]["semantic_text"].startswith(
        "原本「doc.docx」の文書「doc」にあるparagraph内容（2/2）:\n")
    assert "は次のとおりである" not in records[0]["semantic_text"]


def test_non_table_long_text_split_keeps_kv_tail_on_first_chunk():
    """1200文字超の非表要素で、状態KV（可視性/状態）は分割後の先頭chunkにだけ引き継がれる
    （本文だけのsemanticで上書きされて全chunkから失われることはない・重複を避けるため
    2つ目以降のchunkには意図的に含めない）。"""
    long_text = "あ" * 1500
    element = IR.EvidenceElement(
        element_id="e1", type="paragraph", parent_id=None, order=0, value=long_text,
        locator=IR.Locator(part="word/document.xml"), coverage_id="c1",
        visibility="hidden", lifecycle="deleted")
    ir = _ir(file_type="docx", elements=[element])
    ctx = _ctx(source_name="doc.docx")
    records = R._non_table_records(ir, ctx, {}, set())
    assert len(records) == 2
    assert records[0]["semantic_text"].endswith("可視性: 「hidden」 / 状態: 「deleted」")
    assert "可視性" not in records[1]["semantic_text"]     # 2件目以降は重複させない


def test_non_table_long_text_split_keeps_shape_fill_asset_on_first_chunk():
    """1200文字超の shape-fill 付き要素で、asset 説明（semantic）と画像リンク（markdown）は
    分割後の先頭chunkにだけ復元される（自己完結契約＝断片だけでも画像塗りassetの存在が分かる）。"""
    long_text = "あ" * 1500
    element = IR.EvidenceElement(
        element_id="e1", type="shape", parent_id=None, order=0, value=long_text,
        locator=IR.Locator(part="word/document.xml"), coverage_id="c1",
        extension={"name": "元図形",
                   "assets": [{"asset_role": "shape_fill", "asset_sha256": "abc123",
                               "media_part": "word/media/image1.png"}]})
    ir = _ir(file_type="docx", elements=[element])
    ctx = _ctx(source_name="doc.docx")
    records = R._non_table_records(ir, ctx, {}, set())
    assert len(records) == 2

    first_semantic, first_markdown = records[0]["semantic_text"], records[0]["markdown_text"]
    assert first_semantic.endswith(
        "この要素「元図形」には内容未解釈の画像塗りassetが1件存在し、各assetの原本bytesと参照先を保持している。")
    assert first_markdown.endswith(
        "![元図形の画像塗りasset 1/1（内容未解釈）](doc.docx.assets/abc123.png)\n\n"
        "画像塗りasset 1/1のSHA-256: abc123")

    second_semantic, second_markdown = records[1]["semantic_text"], records[1]["markdown_text"]
    assert "画像塗りasset" not in second_semantic          # 2件目以降には重複させない
    assert "画像塗りasset" not in second_markdown
    assert "![" not in second_markdown


# ---- 10〜11: _element_piece ----
# ---- 10a: 基底文（表以外の要素の本文行）のKV化・RAG-KV-002/提案C ----
# `_element_piece`の`exact`ありbranchは、旧「{document_context}に{type_label}「{exact}」がある。」
# という自然文テンプレートを持っていた（`_field_piece`側は既にKV化済みだったのに対する適用漏れ）。
# これを`_field_piece`/`_context_summary_record`と同じ「出所:」＋「{label}: 「値」」の2行KVへ揃える。

def test_element_piece_paragraph_matches_proposal_example():
    """`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md` §4 の実例どおりの出力を pin する。"""
    element = _element("e1", "paragraph", "対象システム: BETA契約管理システム")
    piece = R._element_piece(
        element, {"e1": element}, [], "業務フロー補足_契約.docx",
        ["文書「業務フロー補足_契約」"], include_native_metadata=True)
    assert piece["semantic"] == (
        "出所: 原本「業務フロー補足_契約.docx」 / 文書「業務フロー補足_契約」\n"
        "本文: 「対象システム: BETA契約管理システム」"
    )
    assert piece["markdown"] == piece["semantic"]
    assert "がある。" not in piece["semantic"]
    assert "である。" not in piece["semantic"]


def test_element_piece_content_line_uses_type_label_fallback_without_native_metadata():
    """`include_native_metadata=False`時は`{element.type}の文字列`ラベルのままKV化される
    （現行の条件分岐は踏襲し、KV化だけを適用する）。"""
    element = _element("e1", "paragraph", "値X")
    piece = R._element_piece(
        element, {"e1": element}, [], "x.docx", ["文書「x」"], include_native_metadata=False)
    assert piece["semantic"] == "出所: 原本「x.docx」 / 文書「x」\nparagraphの文字列: 「値X」"


def test_element_piece_notes_and_shape_type_labels_preserved_as_kv_field():
    """`type_label`（「発表者ノート」「図形」等）は捨てず、KVの field label として残る。"""
    notes = _element("e1", "notes", "発表者向け補足メモ")
    notes_piece = R._element_piece(
        notes, {"e1": notes}, [], "deck.pptx", ["スライド1"], include_native_metadata=True)
    assert notes_piece["semantic"] == "出所: 原本「deck.pptx」 / スライド1\n発表者ノート: 「発表者向け補足メモ」"

    shape = _element("e1", "shape", "元図形テキスト")
    shape_piece = R._element_piece(
        shape, {"e1": shape}, [], "doc.docx", ["見出し"], include_native_metadata=True)
    assert shape_piece["semantic"] == "出所: 原本「doc.docx」 / 見出し\n図形: 「元図形テキスト」"


def test_element_piece_source_line_joins_multi_part_section_path_with_slash():
    """複数階層のsection_pathは`_context_prefix`と同じ` / `区切りで連結される（`の`連結ではない）。"""
    element = _element("e1", "paragraph", "値Y")
    piece = R._element_piece(
        element, {"e1": element}, [], "doc.docx",
        ["文書「doc」", "節「第1章」"], include_native_metadata=True)
    assert piece["semantic"].startswith("出所: 原本「doc.docx」 / 文書「doc」 / 節「第1章」\n")


def test_element_piece_relation_text_on_separate_line():
    """relation叙述文は基底KV行と半角スペース連結ではなく独立行として続く
    （§9.1-2の順序＝基底文→relation叙述文は維持しつつ、区切りをExcelノート/状態KVと揃える）。"""
    target = _element("t1", "shape", "対象図形")
    source = _element("e1", "shape", "元図形")
    elements = {"t1": target, "e1": source}
    relation = IR.EvidenceRelation(relation_id="rel1", type="connects_to", source_id="e1",
                                    target_id="t1", evidence_ids=[], confidence=1.0)
    piece = R._element_piece(source, elements, [relation], "doc.docx", ["見出し"], include_native_metadata=True)
    assert piece["semantic"] == (
        "出所: 原本「doc.docx」 / 見出し\n"
        "図形: 「元図形」\n"
        "この要素は対象「対象図形」へ接続している。"
    )


def test_element_piece_value_not_altered_or_summarized():
    """値そのもの（改行・かぎ括弧を含む）は要約・言い換えされず、原値のまま`exact`とsemanticへ残る。"""
    original = "改行を含む\n値「かぎ括弧」付き"
    element = _element("e1", "shape", original)
    piece = R._element_piece(
        element, {"e1": element}, [], "x.pptx", ["スライド1"], include_native_metadata=True)
    assert piece["exact"] == original
    assert f"「{original}」" in piece["semantic"]

def _element(eid, etype, value, *, visibility="visible", lifecycle="active", **extension):
    return IR.EvidenceElement(
        element_id=eid, type=etype, parent_id=None, order=0, value=value,
        locator=IR.Locator(part="word/document.xml"), coverage_id=f"cov:{eid}",
        visibility=visibility, lifecycle=lifecycle, extension=extension)


def test_element_piece_visibility_state_kv():
    visibility_only = _element("e1", "paragraph", "本文A", visibility="hidden")
    lifecycle_only = _element("e2", "paragraph", "本文B", lifecycle="deleted")
    both = _element("e3", "paragraph", "本文C", visibility="hidden", lifecycle="deleted")
    elements = {e.element_id: e for e in (visibility_only, lifecycle_only, both)}

    p1 = R._element_piece(visibility_only, elements, [], "doc.docx", ["見出し"], include_native_metadata=True)
    p2 = R._element_piece(lifecycle_only, elements, [], "doc.docx", ["見出し"], include_native_metadata=True)
    p3 = R._element_piece(both, elements, [], "doc.docx", ["見出し"], include_native_metadata=True)

    assert p1["semantic"].endswith("可視性: 「hidden」")
    assert p2["semantic"].endswith("状態: 「deleted」")
    assert p3["semantic"].endswith("可視性: 「hidden」 / 状態: 「deleted」")


def test_element_piece_shape_fill_asset_description_is_independent_line():
    """shape-fill asset を持つ hidden/deleted 要素で、asset 説明文が状態KV行と同一行に連結
    されず、独立行として続く。"""
    element = _element("e1", "shape", "元図形", visibility="hidden", lifecycle="deleted", name="元図形",
                        assets=[{"asset_role": "shape_fill", "asset_sha256": "abc123",
                                 "media_part": "word/media/image1.png"}])
    piece = R._element_piece(element, {"e1": element}, [], "doc.docx", ["見出し"], include_native_metadata=True)
    state_idx = piece["semantic"].index("可視性: 「hidden」 / 状態: 「deleted」")
    tail = piece["semantic"][state_idx:]
    assert "\nこの要素「元図形」には内容未解釈の画像塗りassetが1件存在し、" in tail
    assert " この要素「元図形」には" not in tail            # 同一行連結（半角スペース区切り）にならない


def test_element_piece_order_relations_before_excel_note():
    target = _element("t1", "shape", "対象図形")
    source = _element("e1", "shape", "元図形", visibility="hidden",
                       raw_value="1", number_format="General", display_status="unsupported")
    elements = {"t1": target, "e1": source}
    relation = IR.EvidenceRelation(relation_id="rel1", type="connects_to", source_id="e1",
                                    target_id="t1", evidence_ids=[], confidence=1.0)
    piece = R._element_piece(source, elements, [relation], "doc.docx", ["見出し"], include_native_metadata=True)
    relation_idx = piece["semantic"].index("接続している")
    excel_idx = piece["semantic"].index("Excel表示状態")
    state_idx = piece["semantic"].index("可視性")
    assert relation_idx < excel_idx < state_idx    # 基底文→relation叙述文→Excelノート→状態（§9.1-2）


# ---- 12〜13: _excel_value_note ----

def test_excel_value_note_all_six_fields():
    metadata = {
        "raw_value": "45900", "number_format": "yyyy/mm/dd", "display_status": "rendered",
        "display_value": "2025/09/10", "formula": "=A1+1", "display_reason": "date_serial",
    }
    note = R._excel_value_note(metadata)
    assert note == (
        "\nExcel原値: 「45900」"
        "\nExcel書式: 「yyyy/mm/dd」"
        "\nExcel表示状態: rendered"
        "\nExcel表示値: 「2025/09/10」"
        "\nExcel数式: 「=A1+1」"
        "\nExcel表示理由: date_serial"
    )


def test_excel_value_note_display_value_material_even_if_equal_raw():
    metadata = {
        "raw_value": "10", "number_format": "General", "display_status": "rendered",
        "display_value": "10", "formula": "=SUM(A1:A2)", "display_reason": None,
    }
    note = R._excel_value_note(metadata)
    assert "Excel表示値: 「10」" in note        # raw と同値でも formula で material 判定→表示値行は出る


def test_excel_value_note_on_non_table_element():
    element = _element("e1", "formula", "10", raw_value="10", number_format="General",
                        display_status="rendered", display_value="10", formula="=SUM(A1:A2)")
    piece = R._element_piece(element, {"e1": element}, [], "a.xlsx", ["シート「明細」"], include_native_metadata=True)
    assert "Excel表示値: 「10」" in piece["semantic"]
    assert "Excel数式: 「=SUM(A1:A2)」" in piece["semantic"]


# ---- 16: スコープ外の叙述文は不変 ----

def test_out_of_scope_sentences_unchanged():
    picture = _element("e1", "picture", None, name="ロゴ")
    piece = R._element_piece(picture, {"e1": picture}, [], "doc.docx", ["見出しA"], include_native_metadata=True)
    assert piece["semantic"] == "原本「doc.docx」の見出しAに画像「ロゴ」が存在する。画像内容は未解釈である。"

    coverage = IR.CoverageItem(
        coverage_id="cov1", scope="shape", detected_kind="chart",
        locator=IR.Locator(part="ppt/slides/slide1.xml", slide=1), status="unsupported",
        content_basis="structured", reason_code="unsupported_shape", parser_id="p", detail={})
    ir = _ir(file_type="pptx")
    ir.coverage.append(coverage)
    records = R._coverage_notice_records(ir, "deck.pptx")
    assert records[0]["semantic_text"] == (
        "原本「deck.pptx」には、chartとして検知した内容があるが、変換結果は未対応である。"
        "理由コードはunsupported_shape、原本位置は"
        '{"part":"ppt/slides/slide1.xml","slide":1}である。内容を抽出済みとは扱わない。')


def test_coverage_notice_cell_count_exceeded_embeds_measured_values():
    """MEM-2: セル数超過 notice は「一報の質」として実測セル数と上限を平文に含める。"""
    coverage = IR.CoverageItem(
        coverage_id="cov1", scope="document", detected_kind="legacy_office_binary",
        locator=IR.Locator(part="source-file", object_id="legacy-office-source"), status="failed",
        content_basis="binary_only", reason_code="cell_count_exceeded", parser_id="p",
        detail={"measured_cells": 6203000, "cap_cells": 2000000})
    ir = _ir(file_type="xlsx")
    ir.coverage.append(coverage)
    records = R._coverage_notice_records(ir, "big.xlsx")
    assert "セル数が多すぎるため取り込み対象外である（上限2000000セル・このファイル約6203000セル）。" \
        in records[0]["semantic_text"]


def test_coverage_notice_uncompressed_size_exceeded_embeds_measured_values():
    """MEM-2: 展開後サイズ超過 notice は実測バイト数（MiB換算）と上限を平文に含める。"""
    coverage = IR.CoverageItem(
        coverage_id="cov1", scope="document", detected_kind="legacy_office_binary",
        locator=IR.Locator(part="source-file", object_id="legacy-office-source"), status="failed",
        content_basis="binary_only", reason_code="uncompressed_size_exceeded", parser_id="p",
        detail={"measured_bytes": 900 * 1024 * 1024, "cap_bytes": 500 * 1024 * 1024})
    ir = _ir(file_type="docx")
    ir.coverage.append(coverage)
    records = R._coverage_notice_records(ir, "huge.docx")
    assert "展開後サイズが大きすぎるため取り込み対象外である（上限500MiB・このファイル約900MiB）。" \
        in records[0]["semantic_text"]


def test_ai_observation_disclaimer_unchanged():
    """AI観測の免責文（`_markdown` の observation_set 節）は本提案のスコープ外＝完全不変（§2）。"""
    from sherpa.ingest import ai_observation

    observation_set = ai_observation.AIObservationSet(
        schema_version="test", source_content_hash="sha256:" + "0" * 64,
        canonical_generation_id="gen1", provider="openai", model="gpt-test",
        model_revision=None, execution_mode="batch", prompt_schema_version="v1",
        preprocessing_profile="p1", engine_profile_hash="h1", response_hash="r1",
        inputs=[], observations=[], observation_set_hash="obsset1",
    )
    ir = _ir()
    md = R._markdown(ir, "a.xlsx", [], {}, observation_set)
    assert "採用AI観測Set: obsset1" in md
    assert "AI観測生成元: openai/gpt-test（batch）" in md
    assert "AI観測は原本確定値ではない。" in md


# ---- 15/17/18/19/20: 実xlsx/docxの往復（build_derived）でformat共通経路・group repack・
#      chunkリンク・identifier overflow・バージョンを検証する ----

def _zip(path: Path, entries: dict):
    with zipfile.ZipFile(path, "w") as z:
        for name, data in entries.items():
            z.writestr(name, data)


def test_render_shared_path_across_docx_pptx_xlsx_and_pdf_tables():
    """`_field_piece`／`_context_summary_record` は docx/pptx/xlsx/pdf いずれの表レコード経路でも
    同一のkey-value形式を出す（region.mode==coordinate_fallback を含む共通コード経路・§5.1の
    「4形式共通経路」）。pdf は原本にネイティブな表マークアップを持たないため、ベクター罫線から
    grid を検出する専用経路（`evidence_spike._pdf_table_grids`/`_pdf_cell_slot`）で表化される
    （pdf の非表要素の長文分割経路は別テスト `test_pdf_non_table_long_text_split_via_real_pdf`
    で検証する）。"""
    from pypdf import PdfWriter
    from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject

    d = tempfile.mkdtemp()
    src = Path(d) / "src"
    src.mkdir()
    der = Path(d) / "derived"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws["A1"], ws["B1"] = "No", "内容"
    ws["A2"], ws["B2"] = 1, "サンプル内容"
    wb.save(src / "a.xlsx")

    docx_xml = (
        '<?xml version="1.0"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body><w:tbl>"
        "<w:tr><w:tc><w:p><w:r><w:t>項目</w:t></w:r></w:p></w:tc>"
        "<w:tc><w:p><w:r><w:t>値</w:t></w:r></w:p></w:tc></w:tr>"
        "<w:tr><w:tc><w:p><w:r><w:t>納期</w:t></w:r></w:p></w:tc>"
        "<w:tc><w:p><w:r><w:t>2026-09-01</w:t></w:r></w:p></w:tc></w:tr>"
        "</w:tbl></w:body></w:document>"
    )
    _zip(src / "doc.docx", {"word/document.xml": docx_xml})

    pptx_xml = (
        '<?xml version="1.0"?>'
        '<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
        ' xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
        "<p:cSld><p:spTree><p:graphicFrame>"
        '<p:nvGraphicFramePr><p:cNvPr id="2" name="Table 1"/></p:nvGraphicFramePr>'
        '<p:xfrm><a:off x="0" y="0"/><a:ext cx="1000" cy="1000"/></p:xfrm>'
        '<a:graphic><a:graphicData uri="http://schemas.openxmlformats.org/drawingml/2006/table">'
        '<a:tbl><a:tr h="370840">'
        "<a:tc><a:txBody><a:p><a:r><a:t>項目</a:t></a:r></a:p></a:txBody></a:tc>"
        "<a:tc><a:txBody><a:p><a:r><a:t>値</a:t></a:r></a:p></a:txBody></a:tc>"
        '</a:tr><a:tr h="370840">'
        "<a:tc><a:txBody><a:p><a:r><a:t>納期</a:t></a:r></a:p></a:txBody></a:tc>"
        "<a:tc><a:txBody><a:p><a:r><a:t>2026-09-01</a:t></a:r></a:p></a:txBody></a:tc>"
        "</a:tr></a:tbl></a:graphicData></a:graphic></p:graphicFrame></p:spTree></p:cSld></p:sld>"
    )
    _zip(src / "deck.pptx", {"ppt/slides/slide1.xml": pptx_xml})

    # pdf: ネイティブ表構造が無いため、2x2の罫線grid（縦横のm/l/S）と、セルごとに独立した
    # BT/ET＋絶対Tmで置いた文字列で表を作る（同一BT/ET内で複数Tjを使うとpypdfのtext_show_operations
    # が同一行内のテキストを1つのgroupへ結合してしまい、列ごとに分離できない＝実測で確認済み）。
    # Helveticaはラテン文字専用のためASCIIのラベル/値を使う（日本語文字は別途フォント埋込が必要）。
    writer = PdfWriter()
    page = writer.add_blank_page(width=500, height=300)
    font = DictionaryObject({
        NameObject("/Type"): NameObject("/Font"),
        NameObject("/Subtype"): NameObject("/Type1"),
        NameObject("/BaseFont"): NameObject("/Helvetica"),
    })
    font_ref = writer._add_object(font)
    page[NameObject("/Resources")] = DictionaryObject(
        {NameObject("/Font"): DictionaryObject({NameObject("/F1"): font_ref})})
    grid_ops = "\n".join(
        [f"{x} 100 m {x} 200 l S" for x in (50, 250, 450)]
        + [f"50 {y} m 450 {y} l S" for y in (100, 150, 200)]
    )
    cells = [(60, 175, "LABEL"), (260, 175, "VALUE"), (60, 125, "DUEDATE"), (260, 125, "20260901")]
    text_ops = "\n".join(f"BT /F1 10 Tf 1 0 0 1 {x} {y} Tm ({t}) Tj ET" for x, y, t in cells)
    content = f"{grid_ops}\n{text_ops}\n".encode("latin-1")
    stream = DecodedStreamObject()
    stream.set_data(content)
    page[NameObject("/Contents")] = writer._add_object(stream)
    with (src / "table.pdf").open("wb") as f:
        writer.write(f)

    orig_backend, orig_pages = office_md._pdf_backend, office_md._pdf_pages
    try:
        office_md._pdf_backend = lambda: "pypdf"
        office_md._pdf_pages = lambda p: ["ダミー本文"]   # legacy MD側の内容は本テストの対象外
        rep = office_md.build_derived(src, der)
    finally:
        office_md._pdf_backend, office_md._pdf_pages = orig_backend, orig_pages
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0

    der_rag = der.parent / "rag"          # §8.1 三階層＝.rag.md は der（md層）の兄弟
    xlsx_md = (der_rag / "a.xlsx.rag.md").read_text(encoding="utf-8")
    docx_md = (der_rag / "doc.docx.rag.md").read_text(encoding="utf-8")
    pptx_md = (der_rag / "deck.pptx.rag.md").read_text(encoding="utf-8")
    pdf_md = (der_rag / "table.pdf.rag.md").read_text(encoding="utf-8")
    for md in (xlsx_md, docx_md, pptx_md, pdf_md):
        assert "出所: 原本「" in md
        assert "は「" not in md and "である。" not in md.split("## ", 1)[-1]
    assert "領域見出し: 「No」" in xlsx_md and "No: 「1」" in xlsx_md
    assert "領域見出し: 「項目」" in docx_md and "項目: 「納期」" in docx_md and "値: 「2026-09-01」" in docx_md
    assert "領域見出し: 「項目」" in pptx_md and "項目: 「納期」" in pptx_md and "値: 「2026-09-01」" in pptx_md
    assert "領域見出し: 「LABEL」" in pdf_md and "領域見出し: 「VALUE」" in pdf_md
    assert "LABEL: 「DUEDATE」" in pdf_md and "VALUE: 「20260901」" in pdf_md


def test_docx_non_table_paragraph_is_key_value_end_to_end():
    """docxの表以外の要素（見出し・段落）が、実際の`office_md.build_derived`往復でも
    `本文「値」がある。`という旧「である」文ではなく`出所:`／`{種別}: 「値」`のKV形式で出ることを
    固定する（RAG-KV-002/提案C・§4の実例に対応する実ファイル経路）。"""
    d = tempfile.mkdtemp()
    src = Path(d) / "src"
    src.mkdir()
    der = Path(d) / "derived"
    docx_xml = (
        '<?xml version="1.0"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        "<w:body>"
        '<w:p><w:pPr><w:pStyle w:val="Heading1"/></w:pPr><w:r><w:t>業務フロー補足_契約</w:t></w:r></w:p>'
        "<w:p><w:r><w:t>対象システム: BETA契約管理システム</w:t></w:r></w:p>"
        "</w:body></w:document>"
    )
    _zip(src / "業務フロー補足_契約.docx", {"word/document.xml": docx_xml})
    rep = office_md.build_derived(src, der)
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0

    md = (der.parent / "rag" / "業務フロー補足_契約.docx.rag.md").read_text(encoding="utf-8")
    assert "出所: 原本「業務フロー補足_契約.docx」 / 文書「業務フロー補足_契約」" in md
    assert "本文: 「対象システム: BETA契約管理システム」" in md
    assert "がある。" not in md
    assert "である。" not in md

    chunks = _read_chunks(der, "業務フロー補足_契約.docx")
    bodies = "\n".join(c["body"] for c in chunks)
    assert "対象システム: BETA契約管理システム" in bodies    # citation本文側の原値は不変（本提案のスコープ外）


def test_pdf_non_table_long_text_split_via_real_pdf():
    """pdf は `_context_prefix` を通らず `_non_table_records` の長文分割経路のみ持つ（§3.1）。
    実PDF（pypdfの低レベルAPIで実テキスト層を書き込んだもの）を `build_derived` に通し、
    出所前置＋`:` へ置換された長文分割の新形式が実際に出ることを確認する。"""
    from pypdf.generic import DecodedStreamObject, DictionaryObject, NameObject
    from pypdf import PdfWriter

    d = tempfile.mkdtemp()
    src = Path(d) / "src"
    src.mkdir()
    der = Path(d) / "derived"

    writer = PdfWriter()
    page = writer.add_blank_page(width=2000, height=200)
    font = DictionaryObject({
        NameObject("/Type"): NameObject("/Font"),
        NameObject("/Subtype"): NameObject("/Type1"),
        NameObject("/BaseFont"): NameObject("/Helvetica"),
    })
    font_ref = writer._add_object(font)
    page[NameObject("/Resources")] = DictionaryObject(
        {NameObject("/Font"): DictionaryObject({NameObject("/F1"): font_ref})})
    long_text = "A" * 1500
    content = f"BT /F1 12 Tf 10 100 Td ({long_text}) Tj ET\n".encode("latin-1")
    stream = DecodedStreamObject()
    stream.set_data(content)
    page[NameObject("/Contents")] = writer._add_object(stream)
    with (src / "note.pdf").open("wb") as f:
        writer.write(f)

    orig_backend, orig_pages = office_md._pdf_backend, office_md._pdf_pages
    try:
        office_md._pdf_backend = lambda: "pypdf"
        office_md._pdf_pages = lambda p: ["ダミー本文"]   # legacy MD側の内容は本テストの対象外
        rep = office_md.build_derived(src, der)
    finally:
        office_md._pdf_backend, office_md._pdf_pages = orig_backend, orig_pages
    assert rep["evidence_ir_failed"] == 0 and rep["rag_failed"] == 0

    md = (der.parent / "rag" / "note.pdf.rag.md").read_text(encoding="utf-8")
    assert "原本「note.pdf」のページ1にあるpositioned_text内容（1/2）:\n" + long_text[:1200] in md
    assert "原本「note.pdf」のページ1にあるpositioned_text内容（2/2）:\n" + long_text[1200:] in md
    assert "は次のとおりである" not in md


def _build_wide_xlsx(ncols: int) -> tuple[Path, Path]:
    d = tempfile.mkdtemp()
    src = Path(d) / "src"
    src.mkdir()
    der = Path(d) / "derived"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    for c in range(1, ncols + 1):
        ws.cell(row=1, column=c, value=f"項目{c:03d}")
        ws.cell(row=2, column=c,
                value=f"値{c:03d}の内容テキストはグループ分割の閾値を試すために十分に長くしてある")
    wb.save(src / "big.xlsx")
    office_md.build_derived(src, der)
    return src, der


def _read_chunks(der: Path, rel: str) -> list[dict]:
    """`der` は md 層。`.rag_chunks.jsonl` は rag 層（`der` の兄弟）に物理配置される（§8.1 三階層）。"""
    return [json.loads(line) for line in (der.parent / "rag" / f"{rel}.rag_chunks.jsonl").read_text(
        encoding="utf-8").splitlines()]


def test_group_repack_at_max_group_chars_boundary_no_data_loss():
    """`logical_record_id` 単位で group 再編後の `body` 連結内容と `citations` の evidence_id
    集合を検証する（§3.7 の受け入れ条件どおり・chunk 単位ではなく論理レコード単位で束ねて比較）。"""
    _, der = _build_wide_xlsx(60)
    chunks = _read_chunks(der, "big.xlsx")
    table_chunks = [c for c in chunks if c["content_type"] == "table_record"]
    assert len({c["logical_record_id"] for c in table_chunks}) == 1
    assert table_chunks[0]["field_group_count"] > 1        # 1800字境界で group 分割が起きている
    table_chunks.sort(key=lambda c: c["field_group_index"])
    body = "".join(c["body"] for c in table_chunks)
    assert all(f"値{c:03d}の内容" in body for c in range(1, 61))   # 全60フィールドの値が欠落なく残る

    ir = IR.from_json_str((der.parent / "ir" / "big.xlsx.evidence.json").read_text(encoding="utf-8"))
    cell_ids = {el.element_id for el in ir.elements if el.type == "cell" and _value_text(el.value)}
    # ヘッダ行のセルは context_summary chunk 側に citation される（table_record ではない）ため、
    # 同一文書の全 chunk を対象に citation の evidence_id 集合を作る（§3.7 の受け入れ条件）。
    cited_ids = {citation["evidence_id"] for chunk in chunks for citation in chunk["citations"]}
    assert cell_ids <= cited_ids   # group分割後も全セルの値がどこかのchunkへcitationされている


def test_chunk_links_consistent_after_repack():
    """`previous_chunk_id`/`next_chunk_id` は全chunkのファイル順（record間も跨ぐ）、
    `sibling_chunk_ids` は同じ `logical_record_id` 内に限定される（group repack後も両立する）。"""
    _, der = _build_wide_xlsx(60)
    chunks = _read_chunks(der, "big.xlsx")          # JSONL の書込順＝chunk の確定順
    by_id = {c["chunk_id"]: c for c in chunks}
    for index, chunk in enumerate(chunks):
        expected_prev = chunks[index - 1]["chunk_id"] if index > 0 else None
        expected_next = chunks[index + 1]["chunk_id"] if index + 1 < len(chunks) else None
        assert chunk["previous_chunk_id"] == expected_prev
        assert chunk["next_chunk_id"] == expected_next
    table_chunks = [c for c in chunks if c["content_type"] == "table_record"]
    for chunk in table_chunks:
        siblings = set(chunk["sibling_chunk_ids"])
        expected_siblings = {
            c["chunk_id"] for c in table_chunks
            if c["logical_record_id"] == chunk["logical_record_id"]
        } - {chunk["chunk_id"]}
        assert siblings == expected_siblings
        for sibling_id in siblings:
            assert sibling_id in by_id


def _build_identifier_overflow_xlsx() -> tuple[Path, Path]:
    """各セルに1個ずつ識別子様トークン（`ANNN`）を持つ横長シートを作る（149列・ncols=150で実測固定）。

    見出しは `項NNN`（日本語＋数字）・値は `ANNN`（英字＋数字）とし、両者の字種を明確に分ける
    （見出しと値が英数字のみの近い形だと、region 検出が `coordinate_fallback` の別レコード扱いへ
    崩れ、狙った1論理レコードにならない実測結果を踏まえた選択）。`IDENTIFIER_MAX_MENTIONS_PER_CHUNK`
    ＝128 を1つの field group（1800字未満）内で超えさせ、かつ2つ目の field group（残り15件）も
    作る＝group 境界を跨いだ overflow の切り分けを1つの fixture で検証できる。
    """
    d = tempfile.mkdtemp()
    src = Path(d) / "src"
    src.mkdir()
    der = Path(d) / "derived"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "明細"
    ws.cell(row=1, column=1, value="No")
    ws.cell(row=2, column=1, value=1)
    for c in range(2, 151):
        ws.cell(row=1, column=c, value=f"項{c:03d}")
        ws.cell(row=2, column=c, value=f"A{c:03d}")
    wb.save(src / "ids.xlsx")
    office_md.build_derived(src, der)
    return src, der


def test_identifier_mentions_overflow_caps_at_limit_with_exact_value_set():
    """group1（列2〜135・134件）は上限128件（列2〜129のトークン）をちょうど保持し、超過6件
    （列130〜135）は `identifier_mention_overflow_count` として数える（値そのものは `body` に
    残る＝データは失われず、identifier索引だけが上限で切り詰められることを区別して確認する）。
    group2（列136〜150・15件）は上限未満なので overflow せず全件保持する。"""
    _, der = _build_identifier_overflow_xlsx()
    chunks = _read_chunks(der, "ids.xlsx")
    data_chunks = sorted((c for c in chunks if c["content_type"] != "context_summary"),
                         key=lambda c: c["field_group_index"])
    assert len(data_chunks) == 2
    group1, group2 = data_chunks
    assert group1["field_group_count"] == 2 and group2["field_group_count"] == 2

    kept1 = {m["value"] for m in group1["identifier_mentions"]}
    assert kept1 == {f"A{c:03d}" for c in range(2, 130)}       # 列2〜129の128件をちょうど保持
    assert group1["identifier_mention_overflow_count"] == 6     # 列130〜135の6件はoverflow
    assert group1["identifier_mention_count"] == len(kept1) + group1["identifier_mention_overflow_count"]
    for c in range(130, 136):                                   # overflowした値も body には残っている
        assert f"A{c:03d}" in group1["body"]

    kept2 = {m["value"] for m in group2["identifier_mentions"]}
    assert kept2 == {f"A{c:03d}" for c in range(136, 151)}      # 列136〜150の15件は上限未満で全件保持
    assert group2["identifier_mention_overflow_count"] == 0
    assert kept1.isdisjoint(kept2)                              # groupをまたいだ重複保持は無い


def test_chunk_versions_are_v1alpha10():
    """L9（R3・フロー図recordの追加）でv1alpha9→v1alpha10へbump。"""
    _, der = _build_wide_xlsx(4)
    chunks = _read_chunks(der, "big.xlsx")
    assert chunks
    for chunk in chunks:
        assert chunk["renderer_version"] == "evidence-rag-renderer-v1alpha10"
        assert chunk["chunker_version"] == "evidence-rag-chunker-v1alpha10"
    assert R.RAG_RENDERER_VERSION == "evidence-rag-renderer-v1alpha10"
    assert R.RAG_CHUNKER_VERSION == "evidence-rag-chunker-v1alpha10"


# ---- L4a: rag.mdの正本化（D1）・アンカー方式 ----

def _read_rag_md(der: Path, rel: str) -> str:
    """`der` は md 層。`.rag.md` は rag 層（`der` の兄弟）に物理配置される（§8.1 三階層）。"""
    return (der.parent / "rag" / f"{rel}.rag.md").read_text(encoding="utf-8")


def test_finalize_chunks_no_longer_has_search_text_field():
    """D1: jsonl（証跡サイドカー）はもう索引本文（旧`search_text`）を持たない。"""
    _, der = _build_wide_xlsx(2)
    chunks = _read_chunks(der, "big.xlsx")
    assert chunks
    assert all("search_text" not in chunk for chunk in chunks)


def test_markdown_anchors_precede_each_chunk_and_match_chunk_ids_1to1():
    """D1: rag.md の各chunk本文の直前に`<!-- chunk:{chunk_id} -->`が1行だけ出て、jsonlの
    全chunk_idと過不足なく1:1対応する。"""
    _, der = _build_wide_xlsx(2)
    chunks = _read_chunks(der, "big.xlsx")
    md = _read_rag_md(der, "big.xlsx")
    anchor_ids = re.findall(r"^<!-- chunk:(\S+) -->$", md, flags=re.MULTILINE)
    assert set(anchor_ids) == {c["chunk_id"] for c in chunks}
    assert len(anchor_ids) == len(set(anchor_ids))          # 重複無し
    for chunk in chunks:
        assert md.count(f"<!-- chunk:{chunk['chunk_id']} -->") == 1


def test_render_validation_errors_catch_anchor_chunk_mismatch():
    """`validation_errors()`はrag.mdのアンカー集合とjsonlのchunk_id集合の不一致を検出する
    （生成時点の自己検証・§受け入れ条件「1:1検証の破れ」の生成側の裏取り）。"""
    ir = _ir(file_type="docx")
    element = _element("e1", "paragraph", "本文A")
    result = R.RenderedEvidence(
        markdown="<!-- chunk:wrong-id -->\n本文A\n",
        chunks=[{"chunk_id": "rag-chunk:actual", "citations": [{"evidence_id": "e1"}],
                 "logical_record_id": "l1", "field_group_index": 1, "field_group_count": 1,
                 "sibling_chunk_ids": [], "source_rel_path": "doc.docx", "evidence_tier": "canonical",
                 "coverage_statuses": [], "has_unresolved_coverage": False, "needs_optional_vision": False}],
        coverage_summary={})
    ir.elements.append(element)
    errors = R.validation_errors(ir, result)
    assert "rag_md_anchor_mismatch" in errors


# ---- L4a: 可視性・廃止表現のKV直列化（OCC-1の表示側・提案書§2.3の裁定）----

def test_occlusion_kv_lines_maps_reason_to_plain_label():
    """生のreason文字列（`occluded_by_picture`等）はそのまま出さず、平文ラベルへ写像する。"""
    lines = R._occlusion_kv_lines({"visibility_reason": "occluded_by_picture"})
    assert lines == ["可視性: 「画像に覆われている」"]
    assert R._occlusion_kv_lines({"visibility_reason": "occluded_by_shape"}) == ["可視性: 「図形に覆われている」"]
    assert R._occlusion_kv_lines({"visibility_reason": "hidden_sheet"}) == ["可視性: 「シートが非表示」"]
    assert R._occlusion_kv_lines({"visibility_reason": "very_hidden"}) == ["可視性: 「シートが完全非表示」"]
    assert R._occlusion_kv_lines({"visibility_reason": "hidden_row"}) == ["可視性: 「行が非表示」"]
    assert R._occlusion_kv_lines({"visibility_reason": "hidden_column"}) == ["可視性: 「列が非表示」"]
    assert R._occlusion_kv_lines({"visibility_reason": "hidden_run"}) == ["可視性: 「非表示文字」"]
    assert R._occlusion_kv_lines({"visibility_reason": "hidden_slide"}) == ["可視性: 「スライドが非表示」"]
    assert R._occlusion_kv_lines({"visibility_reason": "hidden_slide_inherited"}) == [
        "可視性: 「非表示スライド内の要素」"]
    assert R._occlusion_kv_lines({"visibility_reason": "off_slide"}) == ["可視性: 「スライド範囲外」"]
    assert R._occlusion_kv_lines({"visibility_reason": "occluded"}) == ["可視性: 「図形に覆われている」"]


def test_occlusion_kv_lines_strike_is_boolean_presence():
    assert R._occlusion_kv_lines({"visibility_reason": "strike"}) == ["取り消し線: 「あり」"]


def test_occlusion_kv_lines_overlap_from_occluded_by_text():
    lines = R._occlusion_kv_lines({"occluded_by": {"kind": "shape", "text": "廃止"}})
    assert lines == ["重なり: 「廃止」"]


def test_occlusion_kv_lines_overlap_from_covered_by_text_matches_proposal_example():
    """`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md`§2.3の実例どおり。"""
    lines = R._occlusion_kv_lines({"covered_by_text": {"element_id": "shape:2", "text": "廃止"}})
    assert lines == ["重なり: 「廃止」"]


def test_occlusion_kv_lines_overlap_without_text_is_omitted():
    """occluded_by/covered_by_textにtextが無ければ重なり行は出さない（断定しない設計）。"""
    assert R._occlusion_kv_lines({"occluded_by": {"kind": "picture", "z_index": 3}}) == []


def test_occlusion_kv_lines_floating_anchors_behind_vs_front_doc():
    lines = R._occlusion_kv_lines({"floating_anchors": [
        {"behind_doc": True, "text": "透かし画像"},
        {"behind_doc": False, "name": "図形1"},
    ]})
    assert lines == ["背面図形: 「透かし画像」", "前面図形: 「図形1」"]


def test_occlusion_kv_lines_floating_anchor_without_name_or_text_is_omitted():
    assert R._occlusion_kv_lines({"floating_anchors": [{"behind_doc": True}]}) == []


def test_occlusion_kv_lines_unknown_reason_is_not_mapped():
    """マッピング表に無いreasonはこの関数からは何も出さない（呼び出し側の生値フォールバックに委ねる）。"""
    assert R._occlusion_kv_lines({"visibility_reason": "some_future_reason"}) == []


def test_occlusion_kv_lines_empty_extension_is_empty():
    assert R._occlusion_kv_lines({}) == []


def test_field_piece_occluded_by_picture_kv():
    cell = _cell("e1", "使用中", row=3, column=2, visibility_reason="occluded_by_picture",
                 occluded_by={"kind": "picture", "element_id": "shape:1", "z_order": 2})
    piece = R._field_piece(cell, "状態", 0, 1, 0, len("使用中"), "使用中", [])
    assert "可視性: 「画像に覆われている」" in piece["semantic"]
    assert "occluded_by_picture" not in piece["semantic"]   # 生enumは出さない
    assert "可視性: 「画像に覆われている」" in piece["exact"]


def test_field_piece_strike_kv():
    cell = _cell("e1", "廃止予定", row=2, column=2, visibility_reason="strike")
    piece = R._field_piece(cell, "状態", 0, 1, 0, len("廃止予定"), "廃止予定", [])
    assert "取り消し線: 「あり」" in piece["semantic"]


def test_field_piece_hidden_row_kv():
    cell = _cell("e1", "非表示行", row=4, column=2, visibility_reason="hidden_row")
    piece = R._field_piece(cell, "状態", 0, 1, 0, len("非表示行"), "非表示行", [])
    assert "可視性: 「行が非表示」" in piece["semantic"]


def test_field_piece_hidden_column_kv():
    cell = _cell("e1", "X02", row=3, column=3, visibility_reason="hidden_column")
    piece = R._field_piece(cell, "内部コード", 0, 1, 0, len("X02"), "X02", [])
    assert "可視性: 「列が非表示」" in piece["semantic"]


def test_field_piece_occlusion_kv_only_on_first_span():
    """長文分割時、可視性KVは先頭chunkにだけ付く（Excelノートと同じ流儀・重複を避ける）。"""
    cell = _cell("e1", "長文", row=2, column=2, visibility_reason="strike")
    p0 = R._field_piece(cell, "備考", 0, 2, 0, 2, "長文A", [])
    p1 = R._field_piece(cell, "備考", 1, 2, 2, 4, "長文B", [])
    assert "取り消し線" in p0["semantic"]
    assert "取り消し線" not in p1["semantic"]


def test_element_piece_hidden_run_uses_mapped_label_not_raw_enum():
    """L4a: visibility_reasonにマッピングがあれば、既存の生の`可視性: 「hidden」`ではなくラベルを出す
    （マッピングが無い場合の既存フォールバックは`test_element_piece_visibility_state_kv`で不変を確認済み）。"""
    element = _element("e1", "hidden_text", "内部メモ", visibility="hidden", visibility_reason="hidden_run")
    piece = R._element_piece(element, {"e1": element}, [], "doc.docx", ["見出し"], include_native_metadata=True)
    assert "可視性: 「非表示文字」" in piece["semantic"]
    assert "「hidden」" not in piece["semantic"]


def test_element_piece_strike_kv():
    element = _element("e1", "strike_text", "旧料金プラン", visibility_reason="strike")
    piece = R._element_piece(element, {"e1": element}, [], "doc.docx", ["見出し"], include_native_metadata=True)
    assert "取り消し線: 「あり」" in piece["semantic"]


def test_element_piece_covered_by_text_kv_matches_proposal_example():
    """`docs/proposals/2026-09-02-RAG表現の全形式展開と文脈保持.md`§2.3の実例どおり:
    前面テキスト「廃止」との重なりがそのまま出る（意味の断定はしない）。"""
    element = _element("e1", "shape", "旧料金体系: 月額1000円",
                        covered_by_text={"element_id": "shape:2", "text": "廃止"})
    piece = R._element_piece(element, {"e1": element}, [], "deck.pptx", ["スライド2"], include_native_metadata=True)
    assert "重なり: 「廃止」" in piece["semantic"]


def test_element_piece_off_slide_kv():
    element = _element("e1", "shape", "旧仕様メモ", visibility="hidden", visibility_reason="off_slide")
    piece = R._element_piece(element, {"e1": element}, [], "deck.pptx", ["スライド3"], include_native_metadata=True)
    assert "可視性: 「スライド範囲外」" in piece["semantic"]


def test_element_piece_floating_anchor_behind_doc_kv():
    element = _element("e1", "paragraph", "対象システム: BETA契約管理システム",
                        floating_anchors=[{"behind_doc": True, "name": "透かし画像"}])
    piece = R._element_piece(element, {"e1": element}, [], "doc.docx", ["見出し"], include_native_metadata=True)
    assert "背面図形: 「透かし画像」" in piece["semantic"]


def test_element_piece_no_native_metadata_suppresses_occlusion_kv():
    """include_native_metadata=False時は可視性・廃止KVも出さない（既存の状態KVゲートと同じ流儀）。"""
    element = _element("e1", "shape", "旧料金体系: 月額1000円",
                        covered_by_text={"element_id": "shape:2", "text": "廃止"})
    piece = R._element_piece(element, {"e1": element}, [], "deck.pptx", ["スライド2"], include_native_metadata=False)
    assert "重なり" not in piece["semantic"]


def test_element_piece_unmapped_reason_still_falls_back_to_raw_visibility():
    """マッピング表に無い状態（reason無し）は既存の生値フォールバックのまま（回帰確認）。"""
    element = _element("e1", "paragraph", "本文A", visibility="hidden")
    piece = R._element_piece(element, {"e1": element}, [], "doc.docx", ["見出し"], include_native_metadata=True)
    assert piece["semantic"].endswith("可視性: 「hidden」")
