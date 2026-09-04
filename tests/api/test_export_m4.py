"""出力 受け入れ（要 Neo4j）: Excel出力(R7)＋根拠DL(R6・パス基準)。AT-2/AT-3（鏡モデル）。"""
from __future__ import annotations

import pathlib

import pytest
from _world_setup import BILLGEN, SPEC, TAXCPY, TEST_WORLD_ID, ensure_v1
from openpyxl import load_workbook
from fastapi.testclient import TestClient

from sherpa.api import app
from sherpa.documents import resolve

ROOT = pathlib.Path(__file__).resolve().parents[2]
client = TestClient(app)
V = TEST_WORLD_ID   # 旧固定 'v1' から移行（2026-07-03 インシデント対応 HIGH#2・_world_setup.py 参照）


@pytest.fixture(autouse=True)
def _compat_mode(monkeypatch):
    """このファイルはログインせず直接叩く前提（compat モード）。"""
    monkeypatch.setenv("SHERPA_AUTH_DISABLED", "1")


def test_doc_resolve():
    """DL はパス基準（rel_path）。実在ソース→解決可。存在しない設計書 rel→None。トラバーサルは None。"""
    assert resolve(BILLGEN, V) is not None
    assert resolve(TAXCPY, V) is not None
    assert resolve("4期/02_設計/01_基本設計/未作成_NOEXIST.md", V) is None   # 実在しない設計書
    assert resolve("../etc/passwd", V) is None


def test_export_and_download():
    """S3（2026-09-04-グラフのソース正典化.md §4・K9-K11）: REALIZES 橋（業務語→コード）は撤去済みの
    ため、業務語「消費税率」は構造的な items[] ではなく presumed（grep 共起の推定）で見つかる。
    「判定」「method」列は K12 で撤去済み（全件同格・presumed だけ経路列に明示）。
    """
    ensure_v1()
    aid = client.post("/impact/run", json={"start": "消費税率", "world": V}).json()["analysis_id"]

    r = client.get(f"/impact/{aid}/export.xlsx")
    assert r.status_code == 200 and "spreadsheetml" in r.headers["content-type"]
    out = ROOT / "scratch_impact.xlsx"
    out.write_bytes(r.content)
    wb = load_workbook(out)
    ws = wb.active
    header = [c.value for c in ws[1]]
    assert header == ["種別", "対象", "経路", "根拠文書", "根拠DL"]
    rows = {row[1].value: row for row in ws.iter_rows(min_row=2)}
    assert "BILLGEN" in rows
    billgen = {ws[1][k].value: rows["BILLGEN"][k].value for k in range(len(header))}
    assert "△推定" in billgen["経路"] and "rel=" in billgen["根拠DL"]
    out.unlink()

    # 根拠DL（R6・AT-2・パス基準）: 実在ソースは DL 可、実在しない設計書は 404
    ok = client.get("/documents/download", params={"world": V, "rel": BILLGEN})
    assert ok.status_code == 200
    ng = client.get("/documents/download",
                    params={"world": V, "rel": "4期/02_設計/01_基本設計/未作成_NOEXIST.md"})
    assert ng.status_code == 404
