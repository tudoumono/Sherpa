"""出典の Office 原本DL 受け入れ（§1b・鏡モデル）。

回答末尾の出典から DL したとき、**写し（派生MD）でなく原本の Office バイナリ**が返ることを固定する。
鏡モデルでは原本は world 配下、写し（決定的MD）は `data/derived/{world}/md/{rel}.md` に分かれ、
出典の doc_id は常に**原本の rel_path**（grep は派生 `.md` を剥がす・ES は原本 rel で索引）。
`/documents/download?world&rel` は `documents.resolve`（パス基準・root 限定）で原本へ解決する。

Neo4j/ES 不要（一時 KB に Office を作り、実 worker と同じ `office_md.build_derived` で写しを作る）。
`/documents/download` は文書台帳（Postgres）の正準一致確認を持つため、`store.replace_documents`
（`worker._ledger_rows` と同じ行組み立て）で台帳も揃える——DB 到達は必要（テスト用 DB 分離
（`tests/conftest.py`）により専用 `sherpa_test` を使う・fixtures/corpus は汚さない）。
Office は openpyxl/zip で test 内生成する。DL エンドポイントの受け入れは**素の必須テスト**
（import 失敗は skip せず落とす）。runner は各テストファイルを別プロセスで実行（Makefile）＋
world id を PID で一意化＝env 変更・固定 world 名の衝突を避ける。
"""
from __future__ import annotations

import atexit
import os
import pathlib
import shutil
import tempfile
import threading
import time
import zipfile

import pytest

# 一時 KB / 派生領域を指す（未登録 world → KB へ落ちる・registry/DB 不要）。env は呼び出し時参照なので import 前後不問。
_KB = tempfile.mkdtemp(prefix="sherpa_kb_")
_DER = tempfile.mkdtemp(prefix="sherpa_der_")
os.environ["SHERPA_KB_DIR"] = _KB
os.environ["SHERPA_DERIVED_DIR"] = _DER
os.environ["SHERPA_DISABLE_EMBED"] = "1"
atexit.register(lambda: (shutil.rmtree(_KB, ignore_errors=True), shutil.rmtree(_DER, ignore_errors=True)))

from sherpa import corpus_docs, doc_ledger, store, worlds   # noqa: E402
from sherpa.ingest import office_md, worker                  # noqa: E402

W = f"dltest{os.getpid()}"          # PID 一意化＝registry/fixtures の同名 world と衝突しない（KB fallback を確実化）
# 原本（深い階層＝basename 命名の検証も兼ねる）。中身に固有トークンを入れて grep でヒットさせる。
XLSX_REL = "4期保守/03_定例作業/03_作業結果/作業結果報告_202606.xlsx"
DOCX_REL = "4期保守/02_障害対応/01_発生報告/障害報告_バッチ停止.docx"
PPTX_REL = "5期更改/02_設計/01_基本設計/基本設計_共通基盤.pptx"
XLSX_TOKEN = "棚卸照合OK_ZZ9"
DOCX_TOKEN = "夜間バッチ停止_QQ7"
PPTX_TOKEN = "共通基盤方針_PP3"

_DOCX_XML = (
    '<?xml version="1.0"?>'
    '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
    f'<w:body><w:p><w:r><w:t>{DOCX_TOKEN}</w:t></w:r></w:p></w:body></w:document>'
)
_PPTX_SLIDE = (
    '<?xml version="1.0"?>'
    '<p:sld xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main"'
    ' xmlns:p="http://schemas.openxmlformats.org/presentationml/2006/main">'
    f'<p:cSld><p:spTree><a:t>{PPTX_TOKEN}</a:t></p:spTree></p:cSld></p:sld>'
)

OFFICE = ((XLSX_REL, ".xlsx", XLSX_TOKEN, "spreadsheetml"),
          (DOCX_REL, ".docx", DOCX_TOKEN, "wordprocessingml"),
          (PPTX_REL, ".pptx", PPTX_TOKEN, "presentationml"))


def _build_world():
    """一時 KB に world の原本（xlsx/docx/pptx）を置き、worker と同じ build_derived で写しを作る。"""
    import openpyxl
    wd = pathlib.Path(_KB) / W
    xlsx = wd / XLSX_REL
    xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "結果"
    ws["A1"], ws["B1"] = "項目", "値"
    ws["A2"], ws["B2"] = "棚卸", XLSX_TOKEN
    wb.save(xlsx)
    docx = wd / DOCX_REL
    docx.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(docx, "w") as z:
        z.writestr("word/document.xml", _DOCX_XML)
    pptx = wd / PPTX_REL
    pptx.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(pptx, "w") as z:
        z.writestr("ppt/slides/slide1.xml", _PPTX_SLIDE)
    rep = office_md.build_derived(worlds.world_dir(W), worlds.derived_md_dir(W))
    store.replace_documents(W, worker._ledger_rows(W))   # /documents/download の台帳正準一致確認用
    return rep


_REP = _build_world()

# DL エンドポイントは infra 非依存＝必須テスト（import 失敗は握りつぶさず落とす）。
from fastapi.testclient import TestClient   # noqa: E402

from sherpa.api import app                  # noqa: E402

_CLIENT = TestClient(app)


@pytest.fixture(autouse=True)
def _compat_mode(monkeypatch):
    """このファイルはログインせず直接叩く前提（compat モード）。"""
    monkeypatch.setenv("SHERPA_AUTH_DISABLED", "1")


def test_setup_built_derived():
    assert worlds.world_dir(W) is not None and pathlib.Path(worlds.world_dir(W)).is_dir()
    assert _REP.get("error") is None and _REP["converted"] == 3          # xlsx + docx + pptx を写しに
    der = worlds.derived_md_dir(W)
    for rel, *_ in OFFICE:
        assert (der / (rel + ".md")).is_file(), rel


def test_ledger_lists_office_with_md_path():
    by = {d["name"]: d for d in corpus_docs.world_documents(W)}
    for rel, *_ in OFFICE:
        assert rel in by, rel
        assert by[rel]["branch"] == "office" and by[rel]["md_path"]      # 原本 rel・写しは別領域
    assert by[XLSX_REL]["top_scope"] == "4期保守"


def test_grep_citation_doc_id_is_original_not_md():
    """検索は写し（MD）を読むが、出典 doc_id は**原本 rel**（末尾 .md は剥がす）＝DL キーと一致。"""
    from sherpa import grep_tool
    for rel, suf, token, _mime in OFFICE:
        hits = grep_tool.grep_search(token, W)
        assert hits and hits[0]["doc_id"] == rel and hits[0]["doc_id"].endswith(suf), rel


def test_resolve_returns_binary_original():
    for rel, suf, *_ in OFFICE:
        p = doc_ledger.original_path(rel, W)
        assert p is not None and p.is_file() and p.suffix == suf
        assert p.read_bytes()[:2] == b"PK"                              # OOXML=zip＝原本バイナリ（MDテキストでない）
    assert doc_ledger.original_path("../etc/passwd", W) is None          # トラバーサル拒否
    assert doc_ledger.original_path("4期保守/未作成_NOEXIST.docx", W) is None


def test_download_endpoint_serves_office_original():
    for rel, _suf, token, mime in OFFICE:
        r = _CLIENT.get("/documents/download", params={"rel": rel, "world": W})
        assert r.status_code == 200, rel
        assert r.content[:2] == b"PK"                                  # 原本バイナリ（写しMDなら 'PK' で始まらない）
        assert mime in (r.headers.get("content-type") or ""), rel
        cd = r.headers.get("content-disposition") or ""
        assert "attachment" in cd                                      # 強制DL
        # 写し（MD）の中身とは別物であること（DL は原本でありMDテキストでない）
        md_text = (worlds.derived_md_dir(W) / (rel + ".md")).read_text(encoding="utf-8")
        assert token in md_text and r.content != md_text.encode("utf-8"), rel


def test_download_endpoint_supports_range_and_validators():
    """再開DL（Range）と検証ヘッダ（RV2 是正 #2・starlette.FileResponse 相当の挙動）。"""
    rel = XLSX_REL
    full = _CLIENT.get("/documents/download", params={"rel": rel, "world": W})
    assert full.status_code == 200
    assert full.headers.get("accept-ranges") == "bytes"
    etag = full.headers.get("etag")
    assert etag and full.headers.get("last-modified")

    partial = _CLIENT.get("/documents/download", params={"rel": rel, "world": W},
                          headers={"Range": "bytes=0-1"})
    assert partial.status_code == 206
    assert partial.content == full.content[:2]
    assert partial.headers.get("content-range") == f"bytes 0-1/{len(full.content)}"

    unsatisfiable = _CLIENT.get("/documents/download", params={"rel": rel, "world": W},
                                headers={"Range": f"bytes={len(full.content) + 100}-"})
    assert unsatisfiable.status_code == 416
    assert unsatisfiable.headers.get("content-range") == f"bytes */{len(full.content)}"


def test_download_endpoint_toctou_rejects_symlink_swapped_after_verification(monkeypatch):
    """検証（`doc_ledger.original_path`）が通った後で symlink に差し替えられていても、実配信は
    その `Path` を直接使わず、改めて `safe_open.open_file_nofollow_walk`（O_NOFOLLOW walk）で
    fd を取り直すため拒否する（RV1 是正 #3・検証〜配信間の TOCTOU 対策）。

    実際の競合タイミングは再現しない（決定的に待てないため）——`doc_ledger.original_path` を
    「検証は通ったことにする」よう差し替え、その裏で実ファイルを symlink にしておくことで、
    「検証済みの `Path` をそのまま信用しない」という配信側の contract だけを固定する。
    """
    rel = "4期保守/toctou_test.xlsx"
    outside_dir = pathlib.Path(tempfile.mkdtemp(prefix="sherpa_toctou_outside_"))
    secret = outside_dir / "secret.xlsx"
    secret.write_bytes(b"PK\x03\x04SECRET")
    link_path = pathlib.Path(_KB) / W / rel
    link_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        link_path.symlink_to(secret)
    except OSError as e:
        pytest.skip(f"symlink 非対応の環境: {e}")
    store.replace_documents(W, worker._ledger_rows(W) + [{"name": rel}])   # 台帳一致は素通しさせる
    try:
        monkeypatch.setattr(doc_ledger, "original_path", lambda r, w: link_path)
        r = _CLIENT.get("/documents/download", params={"rel": rel, "world": W})
        assert r.status_code == 404
    finally:
        store.replace_documents(W, worker._ledger_rows(W))
        link_path.unlink(missing_ok=True)
        shutil.rmtree(outside_dir, ignore_errors=True)


def test_download_endpoint_rejects_rel_missing_from_ledger():
    """台帳（Postgres `documents`）に無い rel は、ファイルが実在し filesystem 上は
    `doc_ledger.original_path` が解決できても 404（別名/列挙不能ディレクトリ対策・RV1 是正）。"""
    rel = XLSX_REL
    rows = [r for r in worker._ledger_rows(W) if r["name"] != rel]   # この rel だけ台帳から外す
    store.replace_documents(W, rows)
    try:
        assert doc_ledger.original_path(rel, W) is not None            # filesystem 単体では解決できる
        r = _CLIENT.get("/documents/download", params={"rel": rel, "world": W})
        assert r.status_code == 404
    finally:
        store.replace_documents(W, worker._ledger_rows(W))             # 後続テストへ影響しないよう復元


def test_download_rejects_traversal_and_missing():
    assert _CLIENT.get("/documents/download", params={"rel": "../etc/passwd", "world": W}).status_code == 404
    assert _CLIENT.get("/documents/download", params={"rel": "4期保守/NOPE.xlsx", "world": W}).status_code == 404


def test_download_endpoint_waits_for_exclusive_world_lock():
    """`world_lock_shared`（台帳確認〜fd の fstat 完了まで保持）は、rebind/削除/取り込みが使う
    排他ロック（`world_lock`）と直列化する——排他ロックが取れている間は DL が完了しない
    （RV2 是正 #1・台帳と root の世代整合）。"""
    rel = XLSX_REL
    acquired = threading.Event()
    release = threading.Event()

    def _hold_exclusive():
        with store.world_lock(W):
            acquired.set()
            release.wait(timeout=10)

    holder = threading.Thread(target=_hold_exclusive)
    holder.start()
    try:
        assert acquired.wait(timeout=5), "排他ロックを取得できなかった"
        result = {}

        def _download():
            result["r"] = _CLIENT.get("/documents/download", params={"rel": rel, "world": W})

        dl = threading.Thread(target=_download)
        t0 = time.monotonic()
        dl.start()
        time.sleep(0.3)
        assert dl.is_alive(), "排他ロックを待たずに DL が完了した（世代整合の直列化が効いていない）"
        release.set()
        dl.join(timeout=10)
        assert not dl.is_alive(), "DL がロック解放後も完了しなかった"
        assert time.monotonic() - t0 >= 0.3
        assert result["r"].status_code == 200
    finally:
        release.set()
        holder.join(timeout=10)
